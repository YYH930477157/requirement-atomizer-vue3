"""§3.2 + §五(M3) A/B runner 的离线契约测试（不跑真实 LLM——chain_runner 全部注入）。

覆盖方案 §5.5 测试矩阵 1-10：
1. 幻觉需求 → precision 门失败；
2. 中间 JSON 正确、最终 XLSX 正文为空 → 失败；
3. 行数相同但正文错列 → 失败；
4. 条件/例外/否定丢失 → 保存率失败；
5. 数值/单位/编码漂移 → 失败（替换=冲突失配；丢失=保存率 0）；
6. 一条输出不得匹配多条 truth 抬高 recall；
7. 缺任一必需阈值（如 precision）→ NO_GATE；
8. 任一文档 FAIL → overall FAIL；
9. stub/mixed/partial/产物缺失 → FAIL（FailureMatrixTests）；
10. A/B 两路唯一环境差异仍然只是 RATOMIZER_FUNCTIONAL_EXTRACT（EnvIsolationTests）。

复审修复（2026-08-15）：
- P1-3 保存率口径：分母 = 全部真值行的期望条目（未匹配真值计为未保存）——
  「整份文档的真值信息保存率」，与 recall 同口径的诚实分母；报告带
  ``preservation_denominator_scope = "all_truth_rows"`` 诊断键。
- P2-2 匹配算法：确定性贪心 → 自实现 Hungarian/Kuhn-Munkres 最大权重二分匹配
  （先最大化 TP，再最大化总匹配分，同分平局按 (truth_id, 行序) 稳定决胜）；
  含贪心次优反例（TP 更少 / 总权更低两个场景）与随机实例暴力对拍。
"""
from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

_REPO = Path(__file__).resolve().parents[1]
if str(_REPO / "tools") not in sys.path:
    sys.path.insert(0, str(_REPO / "tools"))

import ab_runner as ab  # noqa: E402

FIXTURE_TRUTH_PATH = _REPO / "golden_sets" / "ab_truth_m3_v1" / "functional_truth.jsonl"
FIXTURE_DOC_ID = "SYN-DLMS-AB-1"
TRUTH_SCHEMA_PATH = _REPO / "schemas" / "functional_truth.schema.json"

# 方案 §5.4 的全部 14 项必需阈值（测试全绿配置）。
FULL_THRESHOLDS = {
    "min_truth_precision": 0.99,
    "min_truth_recall": 0.99,
    "min_truth_f1": 0.99,
    "min_condition_preservation": 1.0,
    "min_exception_preservation": 1.0,
    "min_negation_preservation": 1.0,
    "min_number_preservation": 1.0,
    "min_unit_preservation": 1.0,
    "min_code_preservation": 1.0,
    "max_duplicate_rate": 0.0,
    "max_oversplit_rate": 0.0,
    "max_undersplit_rate": 0.0,
    "max_manual_action_estimate": 0,
    "max_final_row_growth_ratio": 1.5,
}

TEMPLATE_HEADERS = ("序号", "子模块", "描述", "需求", "说明、示例、注意事项", "客户需求章节")


def _parsed_dir(root: Path) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    (root / "blocks.jsonl").write_text(
        '{"block_id":"B1","section_path":["4.1"],"text":"The meter shall log events."}\n',
        encoding="utf-8")
    (root / "chunks.jsonl").write_text(
        '{"section_path":["4.1"],"heading":"4.1",'
        '"text":"The meter shall log events.","block_ids":["B1"]}\n', encoding="utf-8")
    return root


def _truth_row(truth_id: str, section: str, text: str, *, document_id: str = "*",
               conditions: tuple[str, ...] = (), exceptions: tuple[str, ...] = (),
               negations: tuple[str, ...] = (), numbers: tuple[str, ...] = (),
               units: tuple[str, ...] = (), codes: tuple[str, ...] = ()) -> dict:
    return {
        "truth_id": truth_id, "document_id": document_id, "section_id": section,
        "expected_text": text, "conditions": list(conditions),
        "exceptions": list(exceptions), "negations": list(negations),
        "numbers": list(numbers), "units": list(units), "codes": list(codes),
    }


def _produced_row(index: int, body: str, section: str = "4.1", sheet: str = "S") -> dict:
    """直接喂 match_truth_to_rows 的最终 XLSX 行（row_number = XLSX 行序，index 0 起）。"""
    return {"sheet": sheet, "row_number": index + 2, "body": body,
            "section": section, "module": "", "context": body}


def _fake_final_xlsx(out_dir: Path) -> None:
    """写 template-write 最终交付 软件需求列表-成文.xlsx（A/B 对比对象，旧口径 fixture）。"""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["模块", "需求"])
    sheet.append(["计量", "记录事件"])
    workbook.save(out_dir / ab.FINAL_XLSX)


def _final_xlsx_raw(out_dir: Path, headers: tuple[str, ...], data_rows: list[tuple]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(headers))
    for row in data_rows:
        sheet.append(list(row))
    workbook.save(out_dir / ab.FINAL_XLSX)


def _xlsx_from_truth(out_dir: Path, truth_rows: list[dict], *,
                     body_map=None, notes_map=None, extra_rows: tuple[tuple, ...] = (),
                     headers: tuple[str, ...] = TEMPLATE_HEADERS) -> None:
    """按真值行写最终 XLSX（默认行体=expected_text，说明列拼接全部期望修饰 → 保存率 1.0）。

    body_map/notes_map: ``callable(truth_row) -> str`` 覆盖默认，用于构造漂移/丢失场景。
    """
    def _default_notes(truth: dict) -> str:
        parts = [*truth["conditions"], *truth["exceptions"], *truth["negations"],
                 *truth["numbers"], *truth["units"], *truth["codes"]]
        return "; ".join(parts)

    data_rows = []
    for index, truth in enumerate(truth_rows, 1):
        body = body_map(truth) if body_map else truth["expected_text"]
        notes = notes_map(truth) if notes_map else _default_notes(truth)
        data_rows.append((index, "Metering", "", body, notes, truth["section_id"]))
    for offset, (body, notes, section) in enumerate(extra_rows, len(truth_rows) + 1):
        data_rows.append((offset, "Metering", "", body, notes, section))
    _final_xlsx_raw(out_dir, headers, data_rows)


def _write_b_product(out_dir: Path, *, execution_status: str = "ok",
                     route: str = "llm:test", conservation_ok: bool = True,
                     items: list[dict] | None = None) -> None:
    if items is None:
        items = [{
            "functional_requirement_id": "FRE-1",
            "objective": "The meter shall log events",
            "source_section": "4.1",
            "source_quote": "The meter shall log events.",
            "source_block_ids": ["B1"],
        }]
    (out_dir / "functional_requirements.json").write_text(json.dumps({
        "producer": "functional-extract-v1",
        "route_requested": "openai_compatible",
        "route": route,
        "execution_status": execution_status,
        "items": items,
        "conservation": {"ok": conservation_ok, "missing_block_ids": []},
    }, ensure_ascii=False), encoding="utf-8")


def _chain_runner_factory(*, b_ok: bool = True, b_status: str = "ok",
                          b_route: str = "llm:test", b_conservation: bool = True,
                          b_product: bool = True, a_ok: bool = True,
                          b_items: list[dict] | None = None,
                          b_xlsx=None, a_xlsx=None):
    """b_xlsx/a_xlsx: ``callable(out_dir)`` 写最终 XLSX（默认旧口径 _fake_final_xlsx）。"""
    def chain_runner(out_dir: Path, *, stages, route, template_path):
        if (out_dir.name == "B_direct"):
            if not a_ok:  # 复用开关控制 B 路链异常
                raise RuntimeError("B chain exploded")
            if not b_ok:
                raise RuntimeError("B chain exploded")
            if b_product:
                _write_b_product(
                    out_dir, execution_status=b_status, route=b_route,
                    conservation_ok=b_conservation, items=b_items)
        else:
            if not a_ok:
                raise RuntimeError("A chain exploded")
            _write_b_product(out_dir, route="llm:test")  # A 路也有产物（synthesis 家族）
        writer = b_xlsx if out_dir.name == "B_direct" else a_xlsx
        (writer or _fake_final_xlsx)(out_dir)
        return {"kind": "chain"}
    return chain_runner


class EnvIsolationTests(unittest.TestCase):
    def test_only_switch_differs_between_paths(self) -> None:
        """矩阵 10：A/B 两路唯一环境差异仍然只是 RATOMIZER_FUNCTIONAL_EXTRACT。"""
        seen: list[tuple[str, str, dict[str, str]]] = []

        def runner(out_dir: Path, *, stages, route, template_path):
            snapshot = {k: v for k, v in os.environ.items() if k.startswith("RATOMIZER_")}
            seen.append((out_dir.name, os.environ.get(ab.SWITCH_ENV, ""), snapshot))
            _write_b_product(out_dir)
            _fake_final_xlsx(out_dir)
            return {}

        sentinel = {"RATOMIZER_TEST_SENTINEL": "same-for-both-paths"}
        with mock.patch.dict(os.environ, sentinel, clear=False), \
                tempfile.TemporaryDirectory() as td:
            work = Path(td)
            report = ab.run_ab_for_document(
                _parsed_dir(work / "parsed"), route="openai_compatible",
                template_path=work / "t.xlsx", work_root=work / "ab",
                chain_runner=runner)
        self.assertEqual(seen[0][0], "A_atoms")
        self.assertEqual(seen[1][0], "B_direct")
        self.assertEqual(seen[0][1], "0")
        self.assertEqual(seen[1][1], "1")
        # 除开关外，其余 RATOMIZER_* 两路逐键一致
        a_snapshot = {k: v for k, v in seen[0][2].items() if k != ab.SWITCH_ENV}
        b_snapshot = {k: v for k, v in seen[1][2].items() if k != ab.SWITCH_ENV}
        self.assertEqual(a_snapshot, b_snapshot)
        self.assertEqual(a_snapshot.get("RATOMIZER_TEST_SENTINEL"), "same-for-both-paths")
        # 退出后恢复原值（默认未设）
        self.assertNotIn(ab.SWITCH_ENV, os.environ)
        # 无真值/阈值 → NO_GATE（门不完备不给 PASS）
        self.assertEqual(report["verdict"], "NO_GATE")


class FailureMatrixTests(unittest.TestCase):
    """矩阵 9：链异常/产物缺失/执行不完整/守恒未闭合/stub 降级一律 FAIL。"""

    def _run(self, **factory_kwargs) -> dict:
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            return ab.run_ab_for_document(
                _parsed_dir(work / "parsed"), route="openai_compatible",
                template_path=work / "t.xlsx", work_root=work / "ab",
                chain_runner=_chain_runner_factory(**factory_kwargs))

    def test_happy_path_is_no_gate_without_truth_and_thresholds(self) -> None:
        """缺真值集/阈值 → NO_GATE，不得 PASS（不能作翻转依据）。"""
        report = self._run()
        self.assertEqual(report["verdict"], "NO_GATE")
        self.assertEqual(len(report["missing_gates"]), 2)
        self.assertTrue(report["metrics"]["available"])
        self.assertEqual(report["metrics"]["functional_count"], 1)
        self.assertEqual(report["metrics"]["duplicate_rate"], 0.0)
        # M3：行级指标基于最终 XLSX
        self.assertEqual(report["metrics"]["b_final_xlsx_rows"], 1)
        self.assertEqual(report["metrics"]["final_xlsx"]["row_count"], 1)

    def test_chain_exception_fails(self) -> None:
        report = self._run(b_ok=False)
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("B chain exploded" in f for f in report["failures"]))

    def test_missing_product_fails(self) -> None:
        report = self._run(b_product=False)
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("无 functional_requirements.json" in f for f in report["failures"]))

    def test_execution_status_failed_fails(self) -> None:
        report = self._run(b_status="failed")
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("execution_status=failed" in f for f in report["failures"]))

    def test_mixed_status_fails(self) -> None:
        report = self._run(b_status="partial")
        self.assertEqual(report["verdict"], "FAIL")

    def test_unconserved_fails(self) -> None:
        report = self._run(b_conservation=False)
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("守恒未闭合" in f for f in report["failures"]))

    def test_stub_degradation_fails(self) -> None:
        report = self._run(b_route="stub", b_status="ok")
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("降级 stub" in f for f in report["failures"]))

    def test_a_path_failure_fails_too(self) -> None:
        report = self._run(a_ok=False)
        self.assertEqual(report["verdict"], "FAIL")

    def test_stub_route_rejected_upfront(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            with self.assertRaises(ValueError):
                ab.run_ab_for_document(
                    _parsed_dir(work / "parsed"), route="stub",
                    template_path=work / "t.xlsx", work_root=work / "ab",
                    chain_runner=_chain_runner_factory())


class TruthLoaderTests(unittest.TestCase):
    """§5.1 真值集格式：schema + 加载校验（坏行响亮报错）。"""

    def test_fixture_loads(self) -> None:
        rows = ab._load_truth(FIXTURE_TRUTH_PATH)
        self.assertEqual(len(rows), 6)
        self.assertEqual(rows[0]["truth_id"], "TRUTH-001")
        self.assertTrue(all(row["document_id"] == FIXTURE_DOC_ID for row in rows))
        # 列表键归一化补全
        self.assertEqual(rows[0]["conditions"], [])

    def test_fixture_conforms_to_schema(self) -> None:
        try:
            import jsonschema  # noqa: F401
        except ImportError:
            self.skipTest("jsonschema not installed")
        import jsonschema

        schema = json.loads(TRUTH_SCHEMA_PATH.read_text(encoding="utf-8"))
        rows = [json.loads(line) for line in
                FIXTURE_TRUTH_PATH.read_text(encoding="utf-8").splitlines() if line.strip()]
        for row in rows:
            jsonschema.validate(row, schema)  # 违例即抛错
        bad = _truth_row("X", "4.1", "text")
        del bad["truth_id"]
        with self.assertRaises(jsonschema.ValidationError):
            jsonschema.validate(bad, schema)

    def test_bad_rows_raise_loudly(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "truth.jsonl"
            # 缺必需键 + 行号定位
            path.write_text(json.dumps({"truth_id": "T1"}) + "\n", encoding="utf-8")
            with self.assertRaises(ValueError) as ctx:
                ab._load_truth(path)
            self.assertIn("missing required key", str(ctx.exception))
            self.assertIn(":1:", str(ctx.exception))
            # 重复 truth_id
            row = _truth_row("T1", "4.1", "a requirement")
            path.write_text(json.dumps(row) + "\n" + json.dumps(row) + "\n",
                            encoding="utf-8")
            with self.assertRaises(ValueError) as ctx:
                ab._load_truth(path)
            self.assertIn("duplicate truth_id", str(ctx.exception))
            # 列表键类型错
            bad = _truth_row("T1", "4.1", "a requirement")
            bad["numbers"] = "15"
            path.write_text(json.dumps(bad) + "\n", encoding="utf-8")
            with self.assertRaises(ValueError) as ctx:
                ab._load_truth(path)
            self.assertIn("must be a list", str(ctx.exception))


class FinalXlsxReaderTests(unittest.TestCase):
    """§5.3 最终 XLSX 读取：中英文表头别名、错误单元格、缺列、多 sheet。"""

    def _write_and_read(self, headers, data_rows) -> dict:
        with tempfile.TemporaryDirectory() as td:
            out_dir = Path(td)
            _final_xlsx_raw(out_dir, headers, data_rows)
            return ab._read_final_xlsx_rows(out_dir / ab.FINAL_XLSX)

    def test_chinese_template_headers_locate_columns(self) -> None:
        struct = self._write_and_read(
            TEMPLATE_HEADERS, [(1, "Metering", "", "The meter shall log events.", "n", "4.1")])
        self.assertTrue(struct["ok"])
        self.assertEqual(struct["row_count"], 1)
        row = struct["rows"][0]
        self.assertEqual(row["body"], "The meter shall log events.")
        self.assertEqual(row["section"], "4.1")
        self.assertEqual(row["module"], "Metering")
        self.assertIn("The meter shall log events.", row["context"])

    def test_english_headers_locate_columns(self) -> None:
        struct = self._write_and_read(
            ("No.", "Module", "Description", "Requirement", "Notes", "Source Section"),
            [(1, "Metering", "", "The meter shall log events.", "n", "4.1")])
        self.assertTrue(struct["ok"])
        self.assertEqual(struct["row_count"], 1)
        self.assertEqual(struct["rows"][0]["body"], "The meter shall log events.")
        self.assertEqual(struct["rows"][0]["section"], "4.1")

    def test_missing_body_column_reported(self) -> None:
        struct = self._write_and_read(
            ("序号", "子模块", "描述", "说明、示例、注意事项", "客户需求章节"),
            [(1, "Metering", "The meter shall log events.", "n", "4.1")])
        self.assertTrue(struct["ok"])
        self.assertEqual(struct["row_count"], 0)
        self.assertEqual(struct["missing_body_sheets"], [struct["sheets"][0]])
        self.assertEqual(len(struct["empty_body_rows"]), 1)

    def test_error_literal_cell_reported(self) -> None:
        struct = self._write_and_read(
            TEMPLATE_HEADERS, [(1, "Metering", "", "The meter shall log events.", "#REF!", "4.1")])
        self.assertTrue(struct["ok"])
        self.assertEqual(len(struct["unreadable_cells"]), 1)
        self.assertEqual(struct["row_count"], 1)  # 正文仍在，错误单元格计入失败明细

    def test_multiple_sheets_are_all_read(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as td:
            out_dir = Path(td)
            workbook = Workbook()
            sheet_a = workbook.active
            sheet_a.title = "计量需求"
            sheet_a.append(list(TEMPLATE_HEADERS))
            sheet_a.append([1, "Metering", "", "Row one body.", "", "4.1"])
            sheet_b = workbook.create_sheet("时钟需求")
            sheet_b.append(list(TEMPLATE_HEADERS))
            sheet_b.append([1, "Clock", "", "Row two body.", "", "4.2"])
            workbook.save(out_dir / ab.FINAL_XLSX)
            struct = ab._read_final_xlsx_rows(out_dir / ab.FINAL_XLSX)
        self.assertEqual(struct["row_count"], 2)
        self.assertEqual(sorted(struct["sheets"]), ["时钟需求", "计量需求"])


class TruthMatchingTests(unittest.TestCase):
    """§5.2 一对一匹配 + §5.3 保存率；方案 §5.5 矩阵 1-6。"""

    def _fixture_truth(self) -> list[dict]:
        return ab._load_truth(FIXTURE_TRUTH_PATH)

    def _fixture_items(self, truth: list[dict]) -> list[dict]:
        return [{"functional_requirement_id": f"FRE-{index}",
                 "objective": row["expected_text"],
                 "source_section": row["section_id"]}
                for index, row in enumerate(truth, 1)]

    def _run_fixture(self, *, truth=None, thresholds=None, b_xlsx=None, parsed_name=None):
        truth = self._fixture_truth() if truth is None else truth
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            return ab.run_ab_for_document(
                _parsed_dir(work / (parsed_name or FIXTURE_DOC_ID)),
                route="openai_compatible", template_path=work / "t.xlsx",
                work_root=work / "ab", truth_rows=truth,
                thresholds=FULL_THRESHOLDS if thresholds is None else thresholds,
                chain_runner=_chain_runner_factory(
                    b_items=self._fixture_items(truth),
                    b_xlsx=b_xlsx or (lambda out: _xlsx_from_truth(out, truth)),
                    a_xlsx=lambda out: _xlsx_from_truth(out, truth)))

    def test_round_trip_fixture_passes_full_gate(self) -> None:
        """全链绿灯：6 真值全命中、零 FP、保存率 1.0 → PASS（14 项阈值全达标）。"""
        report = self._run_fixture()
        self.assertEqual(report["verdict"], "PASS", report)
        self.assertEqual(report["threshold_violations"], [])
        metrics = report["metrics"]["truth_metrics"]
        self.assertEqual((metrics["tp"], metrics["fn"], metrics["fp"]), (6, 0, 0))
        self.assertEqual(metrics["precision"], 1.0)
        self.assertEqual(metrics["recall"], 1.0)
        self.assertEqual(metrics["f1"], 1.0)
        for kind in ("condition", "exception", "negation", "number", "unit", "code"):
            self.assertEqual(metrics["preservation"][kind]["rate"], 1.0, kind)
        # P1-3 口径诊断键 + P2-2 匹配算法标识（全命中时全真值行分母 = 匹配对分母）
        self.assertEqual(metrics["preservation_denominator_scope"], "all_truth_rows")
        self.assertEqual(metrics["matcher"], "hungarian-max-weight/v1")

    def test_matrix1_hallucination_rows_break_precision(self) -> None:
        """矩阵 1：真值全命中但额外幻觉需求 → precision 门失败。"""
        hallucination = ("The meter shall brew fresh coffee every morning.", "", "9.9")
        report = self._run_fixture(
            b_xlsx=lambda out: _xlsx_from_truth(out, self._fixture_truth(),
                                                extra_rows=(hallucination,)))
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("min_truth_precision" in v
                            for v in report["threshold_violations"]), report)
        metrics = report["metrics"]["truth_metrics"]
        self.assertEqual(metrics["fp"], 1)
        self.assertEqual(metrics["recall"], 1.0)

    def test_matrix2_empty_final_body_fails(self) -> None:
        """矩阵 2：中间 JSON 正确、最终 XLSX 正文为空 → 失败。"""
        report = self._run_fixture(
            b_xlsx=lambda out: _xlsx_from_truth(out, self._fixture_truth(),
                                                body_map=lambda truth: ""))
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("空正文行" in f for f in report["failures"]), report)

    def test_matrix3_wrong_column_fails(self) -> None:
        """矩阵 3：行数相同但正文写进描述列（缺需求列）→ 失败。"""
        truth = self._fixture_truth()

        def wrong_column(out_dir: Path) -> None:
            _final_xlsx_raw(
                out_dir,
                ("序号", "子模块", "描述", "说明、示例、注意事项", "客户需求章节"),
                [(index, "Metering", row["expected_text"], "", row["section_id"])
                 for index, row in enumerate(truth, 1)])

        report = self._run_fixture(b_xlsx=wrong_column)
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("缺需求正文列" in f for f in report["failures"]), report)

    def test_matrix4_modifier_loss_breaks_preservation(self) -> None:
        """矩阵 4：条件/例外/否定在最终 XLSX 丢失 → 六类保存率门全失败。"""
        report = self._run_fixture(
            b_xlsx=lambda out: _xlsx_from_truth(out, self._fixture_truth(),
                                                notes_map=lambda truth: ""))
        self.assertEqual(report["verdict"], "FAIL")
        violations = " | ".join(report["threshold_violations"])
        for key in ("min_condition_preservation", "min_exception_preservation",
                    "min_negation_preservation"):
            self.assertIn(key, violations)
        preservation = report["metrics"]["truth_metrics"]["preservation"]
        self.assertEqual(preservation["condition"]["rate"], 0.0)
        self.assertEqual(preservation["negation"]["expected"], 1)

    def test_matrix5_value_loss_breaks_number_unit_code_preservation(self) -> None:
        """矩阵 5a：数值/单位丢失（无替换值）→ 保存率 0，匹配仍成立。"""
        report = self._run_fixture(
            b_xlsx=lambda out: _xlsx_from_truth(out, self._fixture_truth(),
                                                notes_map=lambda truth: ""))
        violations = " | ".join(report["threshold_violations"])
        for key in ("min_number_preservation", "min_unit_preservation",
                    "min_code_preservation"):
            self.assertIn(key, violations)
        metrics = report["metrics"]["truth_metrics"]
        self.assertEqual(metrics["tp"], 6)  # 丢失≠冲突：行仍匹配，保存率惩罚
        self.assertEqual(metrics["preservation"]["number"]["rate"], 0.0)

    def test_matrix5_value_drift_conflict_unmatches(self) -> None:
        """矩阵 5b：数值/单位被替换（15 min → 30 s）→ 冲突 → 该真值失配，recall 掉。"""
        truth = self._fixture_truth()
        drifted_notes = {"TRUTH-002": "capture interval is configured to 30 s"}

        def notes(row: dict) -> str:
            if row["truth_id"] in drifted_notes:
                return drifted_notes[row["truth_id"]]
            parts = [*row["conditions"], *row["exceptions"], *row["negations"],
                     *row["numbers"], *row["units"], *row["codes"]]
            return "; ".join(parts)

        report = self._run_fixture(
            truth=truth, b_xlsx=lambda out: _xlsx_from_truth(out, truth, notes_map=notes))
        self.assertEqual(report["verdict"], "FAIL")
        metrics = report["metrics"]["truth_metrics"]
        self.assertEqual(metrics["unmatched_truth_ids"], ["TRUTH-002"])
        self.assertEqual(metrics["fn"], 1)
        self.assertTrue(any("min_truth_recall" in v
                            for v in report["threshold_violations"]), report)
        # P1-3：未匹配真值 TRUTH-002 的 15/min 进保存率分母（旧口径会空真 1.0）
        preservation = metrics["preservation"]
        self.assertEqual(preservation["unit"]["expected"], 2)  # kWh(T1) + min(T2)
        self.assertAlmostEqual(preservation["unit"]["rate"], 0.5)
        self.assertEqual(preservation["number"]["expected"], 1)  # 15(T2)
        self.assertEqual(preservation["number"]["rate"], 0.0)

    def test_matrix6_one_row_cannot_match_two_truths(self) -> None:
        """矩阵 6：一条合并输出同时覆盖两条真值 → 只算一次 TP，recall 不被抬高。"""
        truth = [
            _truth_row("T1", "4.1", "The meter shall log events."),
            _truth_row("T2", "4.1", "The meter shall log alarms."),
        ]

        def merged(out_dir: Path) -> None:
            _final_xlsx_raw(out_dir, TEMPLATE_HEADERS,
                            [(1, "Metering", "", "The meter shall log events and alarms.",
                              "", "4.1")])

        report = self._run_fixture(truth=truth, b_xlsx=merged,
                                   parsed_name="merged-doc")
        metrics = report["metrics"]["truth_metrics"]
        self.assertEqual(metrics["tp"], 1)
        self.assertEqual(metrics["fn"], 1)
        self.assertEqual(metrics["fp"], 0)
        self.assertEqual(metrics["recall"], 0.5)
        self.assertEqual(report["verdict"], "FAIL")

    def test_matching_stable_under_truth_permutation(self) -> None:
        """确定性：真值行输入顺序打乱，匹配结果逐对不变（含同分平局）。"""
        truth = [
            _truth_row("T-a", "4.1", "The meter shall log events."),
            _truth_row("T-b", "4.1", "The meter shall log events."),
            _truth_row("T-c", "4.2", "The meter shall send alarms."),
        ]
        rows = [{"sheet": "S", "row_number": i + 2,
                 "body": text, "section": section, "module": "", "context": text}
                for i, (text, section) in enumerate(
                    [("The meter shall log events.", "4.1"),
                     ("The meter shall send alarms.", "4.2")])]
        first = ab.match_truth_to_rows(truth, rows)
        second = ab.match_truth_to_rows(list(reversed(truth)), rows)
        key = lambda m: [(x["truth_id"], x["row_number"]) for x in m["matches"]]  # noqa: E731
        self.assertEqual(key(first), key(second))
        self.assertEqual(first["tp"], second["tp"])
        self.assertEqual(first["recall"], second["recall"])
        # 平局按 truth_id 排序：T-a 优先于 T-b
        self.assertEqual([m["truth_id"] for m in first["matches"]], ["T-a", "T-c"])

    # ---- P1-3：保存率分母 = 全部真值行（未匹配真值计为未保存） ----

    def test_p13_unmatched_truth_entries_count_as_unpreserved(self) -> None:
        """P1-3 直测：未匹配真值的期望条目进分母且计为未保存，匹配真值照常统计。"""
        truth = [
            _truth_row("T1", "4.1", "The meter shall log events.", units=("kWh",)),
            _truth_row("T2", "4.2",
                       "The meter shall store load profile data at the configured capture interval.",
                       conditions=("when the load survey profile is enabled",),
                       numbers=("15",), units=("min",)),
        ]
        rows = [_produced_row(0, "The meter shall log events. cumulative energy kWh",
                              section="4.1")]
        result = ab.match_truth_to_rows(truth, rows)
        self.assertEqual(result["tp"], 1)
        self.assertEqual(result["fn"], 1)
        self.assertEqual(result["unmatched_truth_ids"], ["T2"])
        self.assertEqual(result["preservation_denominator_scope"], "all_truth_rows")
        preservation = result["preservation"]
        # 旧口径（只统计匹配对）这些期望值为 0、rate 空真 1.0；新口径如实下降
        self.assertEqual(preservation["condition"]["expected"], 1)
        self.assertEqual(preservation["condition"]["preserved"], 0)
        self.assertEqual(preservation["condition"]["rate"], 0.0)
        self.assertEqual(preservation["number"]["expected"], 1)
        self.assertEqual(preservation["number"]["rate"], 0.0)
        # 单位：kWh（T1 已保存）+ min（T2 未匹配）→ 1/2，明显低于 1.0
        self.assertEqual(preservation["unit"]["expected"], 2)
        self.assertEqual(preservation["unit"]["preserved"], 1)
        self.assertAlmostEqual(preservation["unit"]["rate"], 0.5)

    def test_p13_unmatched_truth_lowers_preservation_in_report(self) -> None:
        """P1-3 反例（全链路）：真值含数值/条件、输出少一条该真值未匹配 → 保存率明显下降。"""
        truth = [
            _truth_row("T1", "4.1", "The meter shall log events."),
            _truth_row("T2", "4.2",
                       "The meter shall store load profile data at the configured capture interval.",
                       conditions=("when the load survey profile is enabled",),
                       numbers=("15",), units=("min",)),
        ]

        def only_first(out_dir: Path) -> None:
            _final_xlsx_raw(out_dir, TEMPLATE_HEADERS,
                            [(1, "Metering", "", truth[0]["expected_text"], "", "4.1")])

        report = self._run_fixture(truth=truth, b_xlsx=only_first,
                                   parsed_name="partial-doc")
        metrics = report["metrics"]["truth_metrics"]
        self.assertEqual(metrics["tp"], 1)
        self.assertEqual(metrics["fn"], 1)
        self.assertEqual(metrics["unmatched_truth_ids"], ["T2"])
        self.assertEqual(metrics["preservation_denominator_scope"], "all_truth_rows")
        preservation = metrics["preservation"]
        # 未匹配真值的期望条目全部进分母且未保存 → 0.0（旧口径为空真 1.0）
        self.assertEqual(preservation["condition"]["expected"], 1)
        self.assertEqual(preservation["condition"]["rate"], 0.0)
        self.assertEqual(preservation["number"]["expected"], 1)
        self.assertEqual(preservation["number"]["rate"], 0.0)
        self.assertEqual(preservation["unit"]["rate"], 0.0)
        # recall 与保存率同口径受罚 → FAIL
        self.assertEqual(report["verdict"], "FAIL")
        violations = " | ".join(report["threshold_violations"])
        self.assertIn("min_truth_recall", violations)
        self.assertIn("min_number_preservation", violations)

    # ---- P2-2：Hungarian 最大权重二分匹配（贪心次优反例 + 平局确定性） ----

    def test_p22_hungarian_beats_greedy_on_tp(self) -> None:
        """P2-2 反例（TP）：最高分对 (T1,row0)=1.0 贪心先占 → T2 无行可配（TP=1）；
        Hungarian 让 T1 让位 row1，T2 拿 row0 → TP=2。"""
        truth = [
            _truth_row("T1", "4.1",
                       "The meter shall record active energy in all four quadrants."),
            _truth_row("T2", "4.1",
                       "The meter shall record reactive energy in tariff mode."),
        ]
        rows = [
            _produced_row(0, "The meter shall record active energy and reactive energy "
                             "in four quadrants."),   # T1 覆盖 1.0，T2 覆盖 4/6
            _produced_row(1, "The meter shall store active energy readings."),  # 仅 T1 覆盖 0.5
        ]
        result = ab.match_truth_to_rows(truth, rows)
        self.assertEqual(result["tp"], 2)
        self.assertEqual(result["fn"], 0)
        self.assertEqual(result["recall"], 1.0)
        pairs = {(m["truth_id"], m["row_number"]) for m in result["matches"]}
        self.assertEqual(pairs, {("T1", 3), ("T2", 2)})  # T1→row1，T2→row0
        self.assertNotIn(("T1", 2), pairs)  # 贪心会拿走的全场最高分对被让出

    def test_p22_hungarian_maximizes_total_weight_equal_tp(self) -> None:
        """P2-2 反例（总权）：TP 同为 2 时取总权更高组合（贪心最高分先占得 1.8，
        Hungarian 得 0.95+0.90=1.85）。"""
        truth = [
            _truth_row("TA", "4.1", "alpha beta gamma delta epsilon zeta eta theta"),
            _truth_row("TB", "4.1", "kappa lambda sigma omega ion proton neutron electron"),
        ]
        rows = [
            _produced_row(0, "alpha beta gamma delta epsilon zeta eta theta "
                             "kappa lambda sigma omega ion proton neutron"),
            _produced_row(1, "alpha beta gamma delta epsilon zeta kappa lambda sigma omega"),
        ]
        result = ab.match_truth_to_rows(truth, rows)
        self.assertEqual(result["tp"], 2)
        pairs = {(m["truth_id"], m["row_number"]) for m in result["matches"]}
        self.assertEqual(pairs, {("TA", 3), ("TB", 2)})
        self.assertAlmostEqual(sum(m["score"] for m in result["matches"]), 1.85, places=6)

    def test_p22_same_score_tie_prefers_low_truth_id_then_row_order(self) -> None:
        """P2-2 平局：两条同文真值 × 两行同文，全部四对同分 → 字典序决胜：
        低 truth_id 取低行序（T-a→row0，T-b→row1）。"""
        truth = [
            _truth_row("T-b", "4.1", "The meter shall log events."),
            _truth_row("T-a", "4.1", "The meter shall log events."),
        ]
        rows = [_produced_row(0, "The meter shall log events."),
                _produced_row(1, "The meter shall log events.")]
        result = ab.match_truth_to_rows(truth, rows)
        self.assertEqual(result["tp"], 2)
        pairs = sorted((m["truth_id"], m["row_number"]) for m in result["matches"])
        self.assertEqual(pairs, [("T-a", 2), ("T-b", 3)])

    def test_match_truth_to_rows_empty_and_no_rows(self) -> None:
        """边界：空真值 × 空行 → 空真指标；有真值无行 → recall 0 且期望条目全未保存。"""
        empty = ab.match_truth_to_rows([], [])
        self.assertEqual(empty["tp"], 0)
        self.assertEqual(empty["precision"], 1.0)
        self.assertEqual(empty["recall"], 1.0)
        self.assertEqual(empty["preservation_denominator_scope"], "all_truth_rows")
        truth = [_truth_row("T1", "4.1", "The meter shall log events.", numbers=("15",))]
        no_rows = ab.match_truth_to_rows(truth, [])
        self.assertEqual((no_rows["tp"], no_rows["fn"]), (0, 1))
        self.assertEqual(no_rows["recall"], 0.0)
        self.assertEqual(no_rows["preservation"]["number"]["expected"], 1)
        self.assertEqual(no_rows["preservation"]["number"]["rate"], 0.0)


class BoundedTiebreakScorePriorityTests(unittest.TestCase):
    """三轮复审 P1-3：有界档决胜不得翻转总分优先级（乘法隔离层级）。"""

    def test_bounded_mode_total_score_strictly_maximized(self) -> None:
        """强制有界档 + 注入评分（含 1e-6 级差异）：返回匹配的总分必须等于
        暴力枚举的最大总分——旧加性编码会被名次加成翻转 1 个分数单位。"""
        from itertools import permutations

        eps = 1e-6
        # 评分矩阵：near-tie 场景，最优组合与若干名次更优组合相差恰 1 个分数单位
        score_map = {
            ("TRUTH-A", 0): 0.90, ("TRUTH-A", 1): 0.90 - eps, ("TRUTH-A", 2): 0.80,
            ("TRUTH-B", 0): 0.90 - eps, ("TRUTH-B", 1): 0.90, ("TRUTH-B", 2): 0.80,
            ("TRUTH-C", 0): 0.80, ("TRUTH-C", 1): 0.80, ("TRUTH-C", 2): 0.90,
        }
        truth = [
            {"truth_id": t, "document_id": "DOC", "section_id": "4.1",
             "expected_text": f"text {t}", "conditions": [], "exceptions": [],
             "negations": [], "numbers": [], "units": [], "codes": []}
            for t in ("TRUTH-A", "TRUTH-B", "TRUTH-C")
        ]
        rows = [
            {"sheet": "S", "row_number": i + 2, "context": "", "section": "4.1",
             "body": f"body {i}"}
            for i in range(3)
        ]

        def fake_pair(t: dict, row: dict) -> dict:
            score = score_map[(t["truth_id"], row["row_number"] - 2)]
            return {"eligible": True, "score": score, "coverage": 1.0,
                    "section_match": True, "conflict": False}

        best_total = max(
            sum(score_map[(t, perm[i])] for i, t in enumerate(
                ("TRUTH-A", "TRUTH-B", "TRUTH-C")))
            for perm in permutations(range(3))
        )
        self.assertGreater(best_total, 0)

        for forced_limit, expected_mode in ((4, "bounded-additive"), (4096, "lexicographic-geometric")):
            with self.subTest(mode=expected_mode):
                with mock.patch.object(ab, "_LEX_EXACT_EDGE_LIMIT", forced_limit),                         mock.patch.object(ab, "_evaluate_pair", side_effect=fake_pair):
                    result = ab.match_truth_to_rows(truth, rows)
                self.assertEqual(result["matcher_tiebreak"], expected_mode)
                self.assertEqual(result["tp"], 3)
                total = sum(m["score"] for m in result["matches"])
                self.assertAlmostEqual(total, best_total, places=9)


class TiebreakScaleSafetyTests(unittest.TestCase):
    """复审 P1-3 二轮：平局编码规模安全——几何位权不随候选边数无界增长。"""

    def _scale_result(self, n: int) -> dict:
        truth = [
            {"truth_id": f"TRUTH-{i:04d}", "document_id": "DOC", "section_id": "4.1",
             "expected_text": f"The controller shall perform task number {i}.",
             "conditions": [], "exceptions": [], "negations": [],
             "numbers": [], "units": [], "codes": []}
            for i in range(n)
        ]
        rows = [
            {"sheet": "Sheet1", "row_number": i + 2, "context": "", "section": "4.1",
             "body": f"The controller shall perform task number {i}."}
            for i in range(n)
        ]
        return ab.match_truth_to_rows(truth, rows)

    def test_bounded_mode_at_scale_full_tp_and_deterministic(self) -> None:
        result = self._scale_result(150)  # E = 22500 > 4096 → 有界加性档
        self.assertEqual(result["matcher_tiebreak"], "bounded-additive")
        self.assertEqual(result["tp"], 150)
        self.assertEqual(result["fn"], 0)
        self.assertEqual(result["fp"], 0)
        again = self._scale_result(150)
        self.assertEqual(
            [m["truth_id"] for m in result["matches"]],
            [m["truth_id"] for m in again["matches"]])

    def test_small_instance_keeps_exact_lexicographic_mode(self) -> None:
        result = self._scale_result(8)  # E = 64 ≤ 4096 → 几何严格字典序档
        self.assertEqual(result["matcher_tiebreak"], "lexicographic-geometric")
        self.assertEqual(result["tp"], 8)


class HungarianAlgorithmTests(unittest.TestCase):
    """P2-2 算法正确性：随机小实例与暴力枚举对拍（每行至多一列、每列至多一行）。"""

    def test_hungarian_matches_bruteforce_on_random_matrices(self) -> None:
        import itertools
        import random

        rng = random.Random(20260815)
        for _ in range(40):
            n = rng.randint(1, 5)
            m = rng.randint(1, 5)
            weights = [[rng.randint(0, 20) for _ in range(m)] for _ in range(n)]
            assignment = ab._hungarian_max_weight(weights)
            # 合法性：列不重复、索引在界内
            used_columns: set[int] = set()
            total = 0
            for row, column in enumerate(assignment):
                if column >= 0:
                    self.assertIn(column, range(m))
                    self.assertNotIn(column, used_columns, "一列被两行占用")
                    used_columns.add(column)
                    total += weights[row][column]
            # 暴力枚举所有部分匹配取最大总权
            best = 0
            for k in range(0, min(n, m) + 1):
                for row_subset in itertools.combinations(range(n), k):
                    for column_subset in itertools.combinations(range(m), k):
                        for permutation in itertools.permutations(column_subset):
                            best = max(best, sum(
                                weights[r][c]
                                for r, c in zip(row_subset, permutation)))
            self.assertEqual(total, best, f"weights={weights} assignment={assignment}")

    def test_hungarian_deterministic_repeat(self) -> None:
        """同一矩阵重复求解输出完全一致（纯函数，无输入顺序依赖）。"""
        weights = [[7, 2, 9], [4, 8, 3], [6, 5, 6]]
        first = ab._hungarian_max_weight(weights)
        for _ in range(5):
            self.assertEqual(ab._hungarian_max_weight(weights), first)


class ThresholdGateTests(unittest.TestCase):
    """§5.4 强制阈值（14 项全必需）+ 按文档独立配置；矩阵 7。"""

    def _healthy_run(self, *, truth=None, thresholds=None, parsed_name=None) -> dict:
        truth = ab._load_truth(FIXTURE_TRUTH_PATH) if truth is None else truth
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            return ab.run_ab_for_document(
                _parsed_dir(work / (parsed_name or FIXTURE_DOC_ID)),
                route="openai_compatible", template_path=work / "t.xlsx",
                work_root=work / "ab", truth_rows=truth, thresholds=thresholds,
                chain_runner=_chain_runner_factory(
                    b_items=[{"functional_requirement_id": f"FRE-{i}",
                              "objective": row["expected_text"],
                              "source_section": row["section_id"]}
                             for i, row in enumerate(truth, 1)],
                    b_xlsx=lambda out: _xlsx_from_truth(out, truth),
                    a_xlsx=lambda out: _xlsx_from_truth(out, truth)))

    def test_all_thresholds_met_passes(self) -> None:
        report = self._healthy_run(thresholds=dict(FULL_THRESHOLDS))
        self.assertEqual(report["verdict"], "PASS", report)
        self.assertEqual(report["threshold_violations"], [])

    def test_matrix7_missing_precision_threshold_is_no_gate(self) -> None:
        """矩阵 7：缺任一必需阈值（min_truth_precision）→ NO_GATE，不 PASS。"""
        partial = {k: v for k, v in FULL_THRESHOLDS.items() if k != "min_truth_precision"}
        report = self._healthy_run(thresholds=partial)
        self.assertEqual(report["verdict"], "NO_GATE")
        self.assertIn("missing required threshold: min_truth_precision",
                      report["missing_gates"])

    def test_threshold_violation_fails(self) -> None:
        # 真值 6 条但最终 XLSX 只有 4 行（丢 4.3/4.4）→ recall 4/6 < 0.99
        truth = ab._load_truth(FIXTURE_TRUTH_PATH)[:4]

        def partial_xlsx(out_dir: Path) -> None:
            _xlsx_from_truth(out_dir, truth)

        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            report = ab.run_ab_for_document(
                _parsed_dir(work / FIXTURE_DOC_ID), route="openai_compatible",
                template_path=work / "t.xlsx", work_root=work / "ab",
                truth_rows=ab._load_truth(FIXTURE_TRUTH_PATH),
                thresholds=dict(FULL_THRESHOLDS),
                chain_runner=_chain_runner_factory(
                    b_items=[{"functional_requirement_id": f"FRE-{i}",
                              "objective": row["expected_text"],
                              "source_section": row["section_id"]}
                             for i, row in enumerate(truth, 1)],
                    b_xlsx=partial_xlsx, a_xlsx=partial_xlsx))
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("min_truth_recall" in v
                            for v in report["threshold_violations"]))

    def test_truth_without_thresholds_is_no_gate(self) -> None:
        report = self._healthy_run(thresholds=None)
        self.assertEqual(report["verdict"], "NO_GATE")
        self.assertEqual(report["missing_gates"], ["thresholds not provided (--thresholds)"])

    def test_missing_truth_is_no_gate(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            report = ab.run_ab_for_document(
                _parsed_dir(work / "parsed"), route="openai_compatible",
                template_path=work / "t.xlsx", work_root=work / "ab",
                truth_rows=None, thresholds=dict(FULL_THRESHOLDS),
                chain_runner=_chain_runner_factory())
        self.assertEqual(report["verdict"], "NO_GATE")
        self.assertEqual(report["missing_gates"], ["truth set not provided (--truth)"])

    def test_truth_for_other_document_is_no_gate(self) -> None:
        truth = [_truth_row("T1", "4.1", "The meter shall log events.",
                            document_id="OTHER-DOC")]
        report = self._healthy_run(truth=truth)
        self.assertEqual(report["verdict"], "NO_GATE")
        self.assertTrue(any("no rows for document" in g for g in report["missing_gates"]))

    def test_unknown_threshold_key_flagged(self) -> None:
        thresholds = {**FULL_THRESHOLDS, "bogus_key": 1}
        report = self._healthy_run(thresholds=thresholds)
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("unknown threshold key" in v
                            for v in report["threshold_violations"]))

    def test_document_layer_overrides_default(self) -> None:
        """文档层覆盖默认层（两个方向都验证），层命中键如实入报告。"""
        hallucination = ("The meter shall brew fresh coffee every morning.", "", "9.9")
        truth = ab._load_truth(FIXTURE_TRUTH_PATH)
        loose_defaults = {**FULL_THRESHOLDS, "min_truth_precision": 0.5}
        thresholds = {
            **loose_defaults,
            "documents": {FIXTURE_DOC_ID: {"min_truth_precision": 0.99}},
        }
        hallucination_xlsx = lambda out: _xlsx_from_truth(  # noqa: E731
            out, truth, extra_rows=(hallucination,))
        # 文档层收紧 → 矩阵 1 场景 FAIL（默认层 0.5 会被误放过）
        report = self._run_with_xlsx(truth, thresholds, hallucination_xlsx)
        self.assertEqual(report["verdict"], "FAIL")
        self.assertTrue(any("min_truth_precision" in v
                            for v in report["threshold_violations"]))
        self.assertEqual(report["threshold_layering"]["document_key_matched"], FIXTURE_DOC_ID)
        self.assertEqual(report["thresholds"]["min_truth_precision"], 0.99)
        # 反向：文档层放宽（precision+f1）→ 同场景 PASS，证明覆盖生效
        relaxed = {
            **loose_defaults,
            "documents": {FIXTURE_DOC_ID: {"min_truth_precision": 0.5, "min_truth_f1": 0.9}},
        }
        report_relaxed = self._run_with_xlsx(truth, relaxed, hallucination_xlsx)
        self.assertEqual(report_relaxed["verdict"], "PASS", report_relaxed)

    def test_document_layer_without_default_needs_all_keys(self) -> None:
        """无默认层：文档层必须自带全部 14 键，缺键 → NO_GATE；齐全 → PASS。"""
        thirteen = {k: v for k, v in FULL_THRESHOLDS.items() if k != "min_truth_f1"}
        report = self._healthy_run(
            thresholds={"documents": {FIXTURE_DOC_ID: thirteen}})
        self.assertEqual(report["verdict"], "NO_GATE")
        self.assertIn("missing required threshold: min_truth_f1", report["missing_gates"])
        report_full = self._healthy_run(
            thresholds={"documents": {FIXTURE_DOC_ID: dict(FULL_THRESHOLDS)}})
        self.assertEqual(report_full["verdict"], "PASS", report_full)

    def test_sha256_document_key_layering(self) -> None:
        """documents 层与 truth document_id 都可用文档 sha256 键（blocks.jsonl 指纹）命中。"""
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            parsed = _parsed_dir(work / "sha-doc")
            sha_key = ab._document_keys(parsed)["sha256"]
            truth = [dict(row, document_id=sha_key)
                     for row in ab._load_truth(FIXTURE_TRUTH_PATH)]
            report = ab.run_ab_for_document(
                parsed, route="openai_compatible", template_path=work / "t.xlsx",
                work_root=work / "ab", truth_rows=truth,
                thresholds={"documents": {sha_key: dict(FULL_THRESHOLDS)}},
                chain_runner=_chain_runner_factory(
                    b_items=[{"functional_requirement_id": f"FRE-{i}",
                              "objective": row["expected_text"],
                              "source_section": row["section_id"]}
                             for i, row in enumerate(truth, 1)],
                    b_xlsx=lambda out: _xlsx_from_truth(out, truth),
                    a_xlsx=lambda out: _xlsx_from_truth(out, truth)))
        self.assertTrue(sha_key.startswith("sha256:"))
        self.assertEqual(report["verdict"], "PASS", report)
        self.assertEqual(report["threshold_layering"]["document_key_matched"], sha_key)

    def _run_with_xlsx(self, truth, thresholds, b_xlsx) -> dict:
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            return ab.run_ab_for_document(
                _parsed_dir(work / FIXTURE_DOC_ID), route="openai_compatible",
                template_path=work / "t.xlsx", work_root=work / "ab",
                truth_rows=truth, thresholds=thresholds,
                chain_runner=_chain_runner_factory(
                    b_items=[{"functional_requirement_id": f"FRE-{i}",
                              "objective": row["expected_text"],
                              "source_section": row["section_id"]}
                             for i, row in enumerate(truth, 1)],
                    b_xlsx=b_xlsx,
                    a_xlsx=lambda out: _xlsx_from_truth(out, truth)))


class OverallVerdictTests(unittest.TestCase):
    """矩阵 8：任一文档 FAIL → overall FAIL（不跨文档平均）。"""

    def test_any_document_fail_makes_overall_fail(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            d1 = _parsed_dir(work / "docA")
            d2 = _parsed_dir(work / "docB")
            truth_file = work / "truth.jsonl"
            truth_file.write_text(
                json.dumps(_truth_row("T1", "4.1", "The meter shall log events.")) + "\n",
                encoding="utf-8")
            thresholds_file = work / "thresholds.json"
            thresholds_file.write_text(json.dumps(FULL_THRESHOLDS), encoding="utf-8")
            original = ab.run_ab_for_document

            def make_chain(doc_name: str):
                def chain(out_dir: Path, *, stages, route, template_path):
                    _write_b_product(out_dir)
                    rows = [("The meter shall log events.", "", "4.1")]
                    if doc_name == "docB" and out_dir.name == "B_direct":
                        rows.append(("The meter shall brew coffee.", "", "9.9"))  # 幻觉行
                    _final_xlsx_raw(out_dir, TEMPLATE_HEADERS,
                                    [(i, "Metering", "", body, notes, section)
                                     for i, (body, notes, section) in enumerate(rows, 1)])
                    return {}
                return chain

            def patched_run(parsed_dir, **kwargs):
                kwargs.pop("chain_runner", None)
                return original(parsed_dir,
                                chain_runner=make_chain(Path(parsed_dir).name), **kwargs)

            try:
                ab.run_ab_for_document = patched_run
                exit_code = ab.main([
                    "--parsed-dir", str(d1), "--parsed-dir", str(d2),
                    "--template", str(work / "t.xlsx"),
                    "--truth", str(truth_file),
                    "--thresholds", str(thresholds_file),
                    "--out", str(work / "report.json"),
                ])
            finally:
                ab.run_ab_for_document = original
            self.assertEqual(exit_code, 2)
            saved = json.loads((work / "report.json").read_text(encoding="utf-8"))
            self.assertEqual(saved["overall_verdict"], "FAIL")
            self.assertEqual(saved["document_count"], 2)
            self.assertEqual(saved["documents"][0]["verdict"], "PASS")
            self.assertEqual(saved["documents"][1]["verdict"], "FAIL")


class CliContractTests(unittest.TestCase):
    def test_multi_doc_per_document_verdict(self) -> None:
        """逐份判定：任一份 FAIL 即整体 FAIL，明细保留。"""
        with tempfile.TemporaryDirectory() as td:
            work = Path(td)
            (work / "t.xlsx").touch()
            d1 = _parsed_dir(work / "d1")
            d2 = _parsed_dir(work / "d2")
            calls = {"n": 0}
            original = ab.run_ab_for_document

            def flaky(out_dir: Path, *, stages, route, template_path):
                calls["n"] += 1
                if calls["n"] == 4:  # 第二份（d2 的 B 路）炸——单份问题不被其他份掩盖
                    raise RuntimeError("second doc boom")
                _write_b_product(out_dir)
                _fake_final_xlsx(out_dir)
                return {}

            def patched_run(parsed_dir, **kwargs):
                kwargs.pop("chain_runner", None)
                return original(parsed_dir, chain_runner=flaky, **kwargs)

            try:
                ab.run_ab_for_document = patched_run
                exit_code = ab.main([
                    "--parsed-dir", str(d1), "--parsed-dir", str(d2),
                    "--template", str(work / "t.xlsx"),
                    "--out", str(work / "report.json"),
                ])
            finally:
                ab.run_ab_for_document = original
            self.assertEqual(exit_code, 2)
            saved = json.loads((work / "report.json").read_text(encoding="utf-8"))
            self.assertEqual(saved["overall_verdict"], "FAIL")
            self.assertEqual(saved["document_count"], 2)
            self.assertEqual(saved["documents"][0]["verdict"], "NO_GATE")
            self.assertEqual(saved["documents"][1]["verdict"], "FAIL")


if __name__ == "__main__":
    unittest.main()
