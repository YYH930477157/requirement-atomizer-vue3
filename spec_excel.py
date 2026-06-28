"""装配规格 → Excel（表计行业版，按功能域分 Sheet）。

版式移植自公司 requirement-analyst 技能的 generate_excel.py，使产出与该技能的
Excel 完全一致：总览 Sheet（统计 + 领域分布 + 缺口/冲突）+ 每个功能域一个 Sheet
（12 列需求表、优先级着色、冻结表头、自动筛选）。接收 assemble_spec 产出的 doc dict，
自带实现、不依赖外部技能目录，可随 PyInstaller 打包进 exe。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from text_normalize import formula_safe


# 表计行业 21 功能域（Sheet 顺序）
METERING_DOMAINS = [
    "计量", "时钟", "事件记录", "曲线", "需量", "费率", "结算", "状态字",
    "窃电", "电网质量", "预付费", "CIU", "门限范围", "Push", "显示",
    "升级", "负控", "节假日", "通信协议", "安全", "环境可靠性",
]

TYPE_LABELS = {
    "functional": "功能需求",
    "non_functional": "非功能需求",
    "constraint": "约束条件",
    "business_rule": "业务规则",
}

PRIORITY_COLORS = {
    "P0": ("FFFFFF", "C0392B"),  # 白字红底
    "P1": ("FFFFFF", "E67E22"),  # 白字橙底
    "P2": ("FFFFFF", "2980B9"),  # 白字蓝底
}

STATUS_LABELS = {
    "draft": "草稿",
    "confirmed": "已确认",
    "conflict": "冲突",
    "gap": "缺口",
}

OBJECT_MODEL_COLUMNS = [
    ("Index", 10),
    ("Object / Attribute Name", 48),
    ("Attribute Type", 26),
    ("Class", 10),
    ("Ver.", 8),
    ("SN", 10),
    ("OBIS Code / Default Value", 30),
    ("Public(16)", 14),
    ("Data Readout(3)", 16),
    ("Remote Management(1)", 22),
    ("Local Management(2)", 22),
]

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)

HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color="2C3E50")
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=11, color="5D6D7E")
STAT_FONT = Font(name="Microsoft YaHei", size=10, color="2C3E50")
CELL_FONT = Font(name="Microsoft YaHei", size=10)
CELL_FONT_BOLD = Font(name="Microsoft YaHei", size=10, bold=True)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
SUMMARY_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
OBJECT_HEADER_FILL = PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid")
OBJECT_GROUP_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
OBJECT_ROW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
ATTRIBUTE_NAME_FILL = PatternFill(start_color="FF00FFFF", end_color="FF00FFFF", fill_type="solid")

TABLE_COLUMNS = [
    ("ID", 12), ("标题", 30), ("类型", 12), ("优先级", 10), ("状态", 10),
    ("来源章节", 16), ("需求描述", 45), ("原文引用", 45), ("验收标准", 35),
    ("参数表", 30), ("依赖", 14), ("备注", 25),
]


def _domain(req: dict[str, Any], domain_set: set[str]) -> str:
    for label in req.get("labels", []):
        if label in domain_set:
            return label
    return "未分类"


def _format_threshold_table(tt: Any) -> str:
    if not tt or not isinstance(tt, dict):
        return ""
    cols = tt.get("columns", [])
    rows = tt.get("rows", [])
    if not cols or not rows:
        return ""
    lines = [tt.get("description", "")]
    for row in rows:
        lines.append(" | ".join(f"{c}: {v}" for c, v in zip(cols, row)))
    return "\n".join(lines)


def _style_header_row(ws: Any, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_priority_cell(cell: Any, priority: str) -> None:
    if priority in PRIORITY_COLORS:
        font_color, bg_color = PRIORITY_COLORS[priority]
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=font_color)
        cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def _write_summary_sheet(wb: Workbook, meta: dict, analysis: dict, domain_groups: dict, *, object_count: int = 0) -> None:
    ws = wb.active
    ws.title = "总览"
    ws.sheet_properties.tabColor = "2C3E50"

    ws.merge_cells("A1:H1")
    meter_label = {"electric": "电表", "water": "水表", "gas": "气表", "multi": "多表"}.get(
        meta.get("meter_type", ""), ""
    )
    doc_title = meta.get("source", "需求分析")
    if meter_label:
        doc_title = f"[{meter_label}] {doc_title}"
    ws["A1"].value = doc_title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"].value = (
        f"提取时间: {meta.get('extracted_at', '-')}    "
        f"目标标准: {', '.join(meta.get('target_standards', ['-']))}"
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 22

    row = 4
    stats = [
        ("需求总数", str(analysis.get("total_count", 0))),
        ("P0 (必须)", str(analysis.get("by_priority", {}).get("P0", 0))),
        ("P1 (应该)", str(analysis.get("by_priority", {}).get("P1", 0))),
        ("P2 (可延后)", str(analysis.get("by_priority", {}).get("P2", 0))),
    ]
    for i, (label, value) in enumerate(stats):
        col = 1 + i * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        label_cell = ws.cell(row=row, column=col)
        label_cell.value = f"{label}: {value}"
        label_cell.font = STAT_FONT
        label_cell.fill = SUMMARY_FILL
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = THIN_BORDER
        ws.cell(row=row, column=col + 1).fill = SUMMARY_FILL
        ws.cell(row=row, column=col + 1).border = THIN_BORDER
    ws.row_dimensions[row].height = 28

    row = 6
    ws.merge_cells(f"A{row}:H{row}")
    domain_title = "领域分布（行为/功能需求）"
    if object_count:
        domain_title += f"；另有 {object_count} 个 COSEM 对象定义见「COSEM Object Model」sheet（不在此重复计入）"
    ws.cell(row=row, column=1).value = domain_title
    ws.cell(row=row, column=1).font = Font(name="Microsoft YaHei", size=12, bold=True, color="2C3E50")
    ws.row_dimensions[row].height = 28

    row = 7
    headers = ["领域", "需求数", "P0", "P1", "P2", "功能", "非功能", "约束", "业务规则"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1).value = h
    _style_header_row(ws, row, len(headers))

    for domain in METERING_DOMAINS:
        if domain not in domain_groups:
            continue
        reqs = domain_groups.get(domain, [])
        row += 1
        values = [
            domain,
            len(reqs),
            sum(1 for r in reqs if r.get("priority") == "P0"),
            sum(1 for r in reqs if r.get("priority") == "P1"),
            sum(1 for r in reqs if r.get("priority") == "P2"),
            sum(1 for r in reqs if r.get("type") == "functional"),
            sum(1 for r in reqs if r.get("type") == "non_functional"),
            sum(1 for r in reqs if r.get("type") == "constraint"),
            sum(1 for r in reqs if r.get("type") == "business_rule"),
        ]
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=i + 1)
            cell.value = v
            cell.font = CELL_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            if i == 0:
                cell.font = CELL_FONT_BOLD
                cell.alignment = Alignment(vertical="center")

    gaps = analysis.get("gaps", [])
    if gaps:
        row += 2
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1).value = "检测到的缺口"
        ws.cell(row=row, column=1).font = Font(name="Microsoft YaHei", size=12, bold=True, color="C0392B")
        for gap in gaps:
            row += 1
            ws.merge_cells(f"A{row}:H{row}")
            ws.cell(row=row, column=1).value = f"[{gap.get('domain', '?')}] {gap.get('description', '')}"
            ws.cell(row=row, column=1).font = Font(name="Microsoft YaHei", size=10, color="C0392B")
            ws.cell(row=row, column=1).alignment = WRAP_ALIGN

    conflicts = analysis.get("conflicts", [])
    if conflicts:
        row += 2
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1).value = "检测到的冲突"
        ws.cell(row=row, column=1).font = Font(name="Microsoft YaHei", size=12, bold=True, color="E67E22")
        for c in conflicts:
            row += 1
            ws.merge_cells(f"A{row}:H{row}")
            ids = ", ".join(c.get("requirement_ids", []))
            ws.cell(row=row, column=1).value = f"[{ids}] {c.get('description', '')}"
            ws.cell(row=row, column=1).font = Font(name="Microsoft YaHei", size=10, color="E67E22")
            ws.cell(row=row, column=1).alignment = WRAP_ALIGN

    for i, w in enumerate([14, 10, 8, 8, 8, 8, 10, 10, 12]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def _write_domain_sheet(wb: Workbook, domain: str, reqs: list[dict]) -> None:
    ws = wb.create_sheet(title=domain[:31])
    has_p0 = any(r.get("priority") == "P0" for r in reqs)
    has_p1 = any(r.get("priority") == "P1" for r in reqs)
    ws.sheet_properties.tabColor = "C0392B" if has_p0 else ("E67E22" if has_p1 else "2980B9")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TABLE_COLUMNS))
    ws.cell(row=1, column=1).value = f"{domain} ({len(reqs)} 条需求)"
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    header_row = 3
    for i, (col_name, _) in enumerate(TABLE_COLUMNS):
        ws.cell(row=header_row, column=i + 1).value = col_name
    _style_header_row(ws, header_row, len(TABLE_COLUMNS))
    ws.row_dimensions[header_row].height = 24

    for row_idx, req in enumerate(reqs):
        r = header_row + 1 + row_idx
        criteria = req.get("acceptance_criteria", [])
        criteria_text = "\n".join(f"- {c}" for c in criteria) if criteria else ""
        deps = req.get("dependencies", [])
        deps_text = ", ".join(deps) if deps else ""
        values = [
            req.get("id", ""),
            req.get("title", ""),
            TYPE_LABELS.get(req.get("type", ""), req.get("type", "")),
            req.get("priority", ""),
            STATUS_LABELS.get(req.get("status", ""), req.get("status", "")),
            req.get("source_section", ""),
            req.get("description", ""),
            req.get("source_quote", ""),
            criteria_text,
            _format_threshold_table(req.get("threshold_table")),
            deps_text,
            req.get("notes", ""),
        ]
        for col_idx, val in enumerate(values):
            cell = ws.cell(row=r, column=col_idx + 1)
            cell.value = formula_safe(val)
            cell.font = CELL_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            if col_idx == 3:
                _style_priority_cell(cell, req.get("priority", ""))
        ws.row_dimensions[r].height = min(max(20, len(str(req.get("description", ""))) // 2), 120)

    for i, (_, width) in enumerate(TABLE_COLUMNS):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(TABLE_COLUMNS))}{header_row + len(reqs)}"


def _is_object_model_requirement(req: dict[str, Any]) -> bool:
    title = str(req.get("title") or "")
    quote = str(req.get("source_quote") or "")
    tt = req.get("threshold_table")
    return (
        "OBIS " in title
        and " / CL " in title
        and "COSEM object " in quote
        and isinstance(tt, dict)
        and "属性访问表" in str(tt.get("description") or "")
    )


def _split_access_row(row: list[Any]) -> dict[str, Any]:
    return {
        "index": row[0] if len(row) > 0 else "",
        "name": row[1] if len(row) > 1 else "",
        "type": row[2] if len(row) > 2 else "",
        "rc": row[3] if len(row) > 3 else "",
        "pc": row[4] if len(row) > 4 else "",
        "sc": row[5] if len(row) > 5 else "",
        "lc": row[6] if len(row) > 6 else "",
        "default": row[7] if len(row) > 7 else "",
    }


def _object_model_name(req: dict[str, Any]) -> str:
    title = str(req.get("title") or "")
    marker = " (OBIS "
    return title.split(marker, 1)[0].strip() if marker in title else title.strip()


def _object_model_class(req: dict[str, Any]) -> str:
    title = str(req.get("title") or "")
    if " / CL " not in title:
        return ""
    return title.rsplit(" / CL ", 1)[-1].rstrip(")").strip()


def _object_model_obis(req: dict[str, Any]) -> str:
    title = str(req.get("title") or "")
    if "(OBIS " not in title or " / CL " not in title:
        return ""
    return title.split("(OBIS ", 1)[-1].split(" / CL ", 1)[0].strip()


def _object_model_group(req: dict[str, Any]) -> str:
    section = str(req.get("source_section") or "").strip()
    labels = req.get("labels") or []
    if section:
        return section
    if labels:
        return str(labels[0])
    return "COSEM Objects"


def _style_object_model_row(ws: Any, row: int, fill: PatternFill, *, bold: bool = False) -> None:
    for col in range(1, len(OBJECT_MODEL_COLUMNS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(name="Microsoft YaHei", size=10, bold=bold, color="000000")
        cell.alignment = CENTER_ALIGN if col != 2 else Alignment(vertical="center")
        cell.border = THIN_BORDER


def _write_object_model_sheet(wb: Workbook, requirements: list[dict[str, Any]]) -> None:
    object_reqs = [req for req in requirements if _is_object_model_requirement(req)]
    if not object_reqs:
        return

    ws = wb.create_sheet(title="COSEM Object Model")
    ws.sheet_properties.tabColor = "00A6A6"
    for col_idx, (name, width) in enumerate(OBJECT_MODEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = name
        cell.fill = OBJECT_HEADER_FILL
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="000000")
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24

    row_idx = 2
    current_group = None
    for req in sorted(object_reqs, key=lambda r: (_object_model_group(r), _object_model_obis(r), _object_model_name(r))):
        group = _object_model_group(req)
        if group != current_group:
            current_group = group
            ws.cell(row=row_idx, column=2).value = group
            _style_object_model_row(ws, row_idx, OBJECT_GROUP_FILL, bold=True)
            row_idx += 1

        ws.cell(row=row_idx, column=2).value = formula_safe(_object_model_name(req))
        ws.cell(row=row_idx, column=4).value = formula_safe(_object_model_class(req))
        ws.cell(row=row_idx, column=7).value = formula_safe(_object_model_obis(req))
        _style_object_model_row(ws, row_idx, OBJECT_ROW_FILL, bold=True)
        row_idx += 1

        tt = req.get("threshold_table") or {}
        for raw in tt.get("rows") or []:
            attr = _split_access_row(list(raw))
            ws.cell(row=row_idx, column=1).value = formula_safe(attr["index"])
            ws.cell(row=row_idx, column=2).value = formula_safe(attr["name"])
            ws.cell(row=row_idx, column=3).value = formula_safe(attr["type"])
            ws.cell(row=row_idx, column=7).value = formula_safe(attr["default"])
            ws.cell(row=row_idx, column=8).value = formula_safe(attr["pc"])
            ws.cell(row=row_idx, column=9).value = formula_safe(attr["rc"])
            ws.cell(row=row_idx, column=10).value = formula_safe(attr["sc"])
            ws.cell(row=row_idx, column=11).value = formula_safe(attr["lc"])
            for col in range(1, len(OBJECT_MODEL_COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = CELL_FONT
                cell.alignment = CENTER_ALIGN if col != 2 else Alignment(vertical="center")
                cell.border = THIN_BORDER
            ws.cell(row=row_idx, column=2).fill = ATTRIBUTE_NAME_FILL
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OBJECT_MODEL_COLUMNS))}{max(1, row_idx - 1)}"


def write_xlsx(doc: dict[str, Any], output_path: Path) -> Path:
    """把装配好的需求 doc 写成表计行业版 Excel（与公司技能 generate_excel 同版式）。"""
    output_path = Path(output_path)
    requirements = doc.get("requirements", [])
    meta = doc.get("meta", {})
    analysis = doc.get("analysis", {})

    domain_set = set(METERING_DOMAINS)
    # 对象定义型需求集中在 COSEM Object Model sheet（含属性数据类型 + RC/PC/SC/LC 访问表 + 默认值），
    # 不在领域 sheet 重复出现——领域 sheet 只放行为/功能型需求，避免同一对象显示两遍。
    object_reqs = [req for req in requirements if _is_object_model_requirement(req)]
    domain_groups: dict[str, list[dict]] = {}
    for req in requirements:
        if _is_object_model_requirement(req):
            continue
        domain_groups.setdefault(_domain(req, domain_set), []).append(req)

    wb = Workbook()
    _write_summary_sheet(wb, meta, analysis, domain_groups, object_count=len(object_reqs))
    _write_object_model_sheet(wb, requirements)
    for domain in METERING_DOMAINS:
        if domain in domain_groups:
            _write_domain_sheet(wb, domain, domain_groups[domain])
    if "未分类" in domain_groups:
        _write_domain_sheet(wb, "未分类", domain_groups["未分类"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
