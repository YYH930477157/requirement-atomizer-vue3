from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text(
        "".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows),
        encoding="utf-8",
    )


def _semantic_case(case_id: str) -> dict:
    path = Path(__file__).resolve().parents[1] / "golden_sets" / "requirements_analysis_semantic_v1.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    return next(row["expected"] for row in payload["cases"] if row["id"] == case_id)


class StageLineageRegressionTests(unittest.TestCase):
    def test_requirements_analysis_reuse_invalidates_when_ai_input_changes(self) -> None:
        from desktop_tasks import (
            STAGE_REQUIRED_OUTPUTS,
            stage_input_fingerprint,
            stage_is_reusable,
            update_run_manifest,
        )

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _write_jsonl(out / "ai_requirements.jsonl", [{"ai_req_id": "AI-1", "title": "A"}])
            for name in STAGE_REQUIRED_OUTPUTS["requirements-analysis"]:
                (out / name).write_text("placeholder", encoding="utf-8")
            fingerprint = stage_input_fingerprint(out, "requirements-analysis", route="stub")
            update_run_manifest(
                out,
                "requirements-analysis",
                "ok",
                route="stub",
                outputs=STAGE_REQUIRED_OUTPUTS["requirements-analysis"],
                input_fingerprint=fingerprint,
            )
            self.assertTrue(stage_is_reusable(out, "requirements-analysis", route="stub"))

            _write_jsonl(out / "ai_requirements.jsonl", [{"ai_req_id": "AI-1", "title": "B"}])
            self.assertFalse(stage_is_reusable(out, "requirements-analysis", route="stub"))

    def test_functional_synthesis_fingerprint_tracks_llm_pipeline_config(self) -> None:
        from unittest import mock
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            config = out / "review_pipeline.yaml"
            _write_jsonl(out / "ai_requirements.jsonl", [{"ai_req_id": "AI-1", "title": "A"}])
            _write_jsonl(out / "ai_review_states.jsonl", [])
            config.write_text("model: first\n", encoding="utf-8")
            with mock.patch.object(desktop_tasks, "DEFAULT_PIPELINE_PATH", config):
                first = desktop_tasks.stage_input_fingerprint(
                    out, "functional-synthesis", route="openai_compatible")
                config.write_text("model: second\n", encoding="utf-8")
                second = desktop_tasks.stage_input_fingerprint(
                    out, "functional-synthesis", route="openai_compatible")

        self.assertNotEqual(first, second)

class ChainPreconditionRegressionTests(unittest.TestCase):
    def test_stub_request_can_reuse_valid_openai_ai_extract_output(self) -> None:
        from desktop_tasks import stage_input_fingerprint, stage_is_reusable, update_run_manifest

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _write_jsonl(out / "blocks.jsonl", [{"block_id": "B-1", "text": "The meter shall record data."}])
            _write_jsonl(out / "llm_review_results.jsonl", [])
            _write_jsonl(out / "review_states.jsonl", [])
            _write_jsonl(out / "ai_requirements.jsonl", [{"ai_req_id": "AI-1", "title": "记录数据"}])
            (out / "merged_spec_requirements.json").write_text("{}", encoding="utf-8")
            config = {"sample_ratio": None, "limit_sections": None}
            fingerprint = stage_input_fingerprint(out, "ai-extract", route="openai_compatible", config=config)
            update_run_manifest(
                out, "ai-extract", "ok", route="openai_compatible",
                outputs=["ai_requirements.jsonl", "merged_spec_requirements.json"],
                input_fingerprint=fingerprint, config=config,
            )

            self.assertTrue(stage_is_reusable(out, "ai-extract", route="stub", config=config))

    def test_stub_chain_allows_deterministic_annotation_export(self) -> None:
        from unittest import mock
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            with (
                mock.patch.object(desktop_tasks, "ai_extract_task", return_value={"written": ["merged_spec_requirements.json"]}),
                mock.patch.object(desktop_tasks, "export_annotation_html_task", return_value={"written": ["document_annotation.html"]}),
            ):
                payload = desktop_tasks.chain_task(
                    out, stages=["ai-extract", "export-annotation-html"], route="stub")

        self.assertEqual(payload["stages"], ["ai-extract", "export-annotation-html"])

    def test_stub_chain_rejects_ai_dependent_stages_without_ai_output(self) -> None:
        from desktop_tasks import chain_task

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            with self.assertRaisesRegex(ValueError, "openai_compatible.*AI 抽取"):
                chain_task(
                    out,
                    stages=["ai-extract", "functional-synthesis", "requirements-analysis"],
                    route="stub",
                )



class AnalysisFallbackRegressionTests(unittest.TestCase):
    def test_stub_analysis_preserves_upstream_delivery_baseline(self) -> None:
        from requirements_analysis import run_requirements_analysis

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _write_jsonl(out / "ai_requirements.jsonl", [{
                "ai_req_id": "AI-1",
                "title": "远程采集重要事件",
                "description": "系统须采集并远程上传重要事件。",
                "source_quote": "The system shall collect and remotely transmit significant events.",
                "source_block_ids": ["B-1"],
                "module": "事件记录",
                "dev_guidance": ["实现重要事件采集与上传接口"],
                "design_options": ["可采用异步发送队列解耦采集与上传"],
                "acceptance_criteria": ["产生重要事件后，远端能够收到对应事件记录"],
                "assumptions": ["远端通信链路由通信模块提供"],
            }])

            run_requirements_analysis(out, route="stub")
            item = json.loads((out / "engineering_analysis.json").read_text(encoding="utf-8"))["items"][0]

            self.assertEqual(item["software_requirement_text"], "系统须采集并远程上传重要事件。")
            self.assertEqual(item["developer_guidance"], ["实现重要事件采集与上传接口"])
            self.assertEqual(item["design_options"], ["可采用异步发送队列解耦采集与上传"])
            self.assertEqual(item["acceptance_criteria"], ["产生重要事件后，远端能够收到对应事件记录"])
            self.assertEqual(item["assumptions"], ["远端通信链路由通信模块提供"])


class OwnershipRegressionTests(unittest.TestCase):
    def test_synthesized_requirement_inherits_clarifications_from_source_atoms(self) -> None:
        from requirements_analysis import run_requirements_analysis

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _write_jsonl(out / "ai_requirements.jsonl", [{
                "ai_req_id": "AI-1", "title": "事件上报", "description": "上报事件。", "module": "事件",
            }])
            (out / "functional_requirements.json").write_text(json.dumps({
                "items": [{
                    "functional_requirement_id": "FREQ-1",
                    "source_ai_requirement_ids": ["AI-1"],
                    "title": "事件管理",
                    "description": "管理并上报事件。",
                    "module": "事件",
                }]
            }, ensure_ascii=False), encoding="utf-8")
            _write_jsonl(out / "clarification_answers.jsonl", [{
                "source_id": "AI-1",
                "question": "上报时限？",
                "answer": "事件产生后 30 秒内上报。",
                "adopted": True,
            }])

            run_requirements_analysis(out, route="stub")
            payload = json.loads((out / "engineering_analysis.json").read_text(encoding="utf-8"))

        self.assertIn("客户答复：事件产生后 30 秒内上报。", payload["items"][0]["notes"])
    def test_mobile_data_concentrator_is_co_design_not_hardware_only(self) -> None:
        from requirements_analysis_rules import classify_ownership

        result = classify_ownership({
            "title": "Dispositivo walk by",
            "description": "Device with mobile data concentrator function for walk-by and drive-by collection.",
            "module": "通信",
        })

        expected = _semantic_case("walk-by-data-concentrator")
        self.assertEqual(result["ownership"], expected["ownership"])


class DesignOptionRegressionTests(unittest.TestCase):
    def test_unsupported_implementation_choice_moves_to_design_options(self) -> None:
        import ai_extract

        section = {
            "section_id": "S1",
            "heading": "Data collection",
            "text": "The GdM shall collect, record locally and transmit the data remotely.",
            "block_ids": ["B-1"],
            "source_blocks": [{
                "block_id": "B-1",
                "text": "The GdM shall collect, record locally and transmit the data remotely.",
            }],
        }
        raw = [{
            "title": "采集、存储并远传数据",
            "description": "GdM 须采集数据、在本地记录并远程传输。",
            "source_quote": section["text"],
            "module": "通信",
            "type": "functional",
            "priority": "P1",
            "labels": ["数据采集"],
            "dev_guidance": ["采用 FIFO 队列解耦本地记录与远程发送"],
            "acceptance_criteria": [],
        }]

        result = ai_extract._process_raw_requirements(raw, section)[0]

        expected = _semantic_case("unsupported-buffer-choice")
        self.assertEqual(len(result["dev_guidance"]), expected["normative_guidance_count"])
        self.assertTrue(any(term in result["design_options"][0] for term in expected["design_options"]))
        self.assertNotIn("无依据条目已移入备注", result.get("notes", ""))


class SemanticBaselineRegressionTests(unittest.TestCase):
    def test_billing_period_definition_preserves_allowed_durations(self) -> None:
        import ai_extract

        expected = _semantic_case("billing-period-definition")
        source = (
            "A period of time that always begins on the first day of a month and ends on the first "
            "day of one or more subsequent months; it can be valid for 1, 2, 3, 4, 6, 12 months."
        )
        section = {
            "section_id": "3.24", "heading": "3.24 billing period", "text": source,
            "block_ids": ["B-BILLING-DEF"],
            "source_blocks": [{"block_id": "B-BILLING-DEF", "text": source}],
        }

        def chat(_system: str, _user: str) -> dict:
            return {"requirements": [{
                "title": "结算周期定义", "functional_key": expected["functional_key"],
                "description": "结算周期从每月第一天开始，可配置为 1、2、3、4、6 或 12 个月。",
                "source_quote": source, "module": "结算", "type": "business_rule",
                "priority": "P1", "labels": ["结算周期"], "acceptance_criteria": [],
                "dev_guidance": [],
            }]}

        requirements = ai_extract.extract_section(section, chat)

        self.assertEqual(len(requirements), expected["requirement_count"])
        self.assertEqual(requirements[0]["functional_key"], expected["functional_key"])
        self.assertEqual(requirements[0]["source_block_ids"], expected["source_block_ids"])
        delivery = requirements[0]["description"] + " " + requirements[0]["source_quote"]
        for value in expected["required_values"]:
            self.assertIn(value, delivery)


class SourceAnchorRegressionTests(unittest.TestCase):
    def test_source_quote_maps_to_exact_block_not_whole_section(self) -> None:
        import ai_extract

        section = {
            "section_id": "3",
            "heading": "Terms",
            "text": "3.23 unrelated\n3.24 billing period\nA period begins on the first day of a month.",
            "block_ids": ["B-1", "B-2", "B-3"],
            "source_blocks": [
                {"block_id": "B-1", "text": "3.23 unrelated"},
                {"block_id": "B-2", "text": "3.24 billing period"},
                {"block_id": "B-3", "text": "A period begins on the first day of a month."},
            ],
        }

        def chat(_system: str, _user: str) -> dict:
            return {"requirements": [{
                "title": "结算周期起始日",
                "description": "结算周期须从每月第一天开始。",
                "source_quote": "A period begins on the first day of a month.",
                "module": "结算",
                "type": "business_rule",
                "priority": "P1",
                "labels": ["结算周期"],
                "acceptance_criteria": [],
                "dev_guidance": [],
            }]}

        result = ai_extract.extract_section(section, chat)[0]

        self.assertEqual(result["anchor_block_id"], "B-3")
        self.assertEqual(result["source_block_ids"], ["B-3"])
        self.assertEqual(result["source_mapping"], "exact")


class FunctionalSynthesisReviewStateTests(unittest.TestCase):
    def test_rejected_ai_requirement_is_excluded_before_synthesis(self) -> None:
        from functional_synthesis import run_functional_synthesis

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _write_jsonl(out / "ai_requirements.jsonl", [
                {"ai_req_id": "AI-KEEP", "functional_key": "事件管理", "title": "保留", "module": "事件"},
                {"ai_req_id": "AI-DROP", "functional_key": "事件管理", "title": "拒绝", "module": "事件"},
            ])
            _write_jsonl(out / "ai_review_states.jsonl", [
                {"ai_req_id": "AI-DROP", "status": "rejected"},
            ])

            run_functional_synthesis(out)
            payload = json.loads((out / "functional_requirements.json").read_text(encoding="utf-8"))

        self.assertEqual(payload["source_requirements"], 2)
        self.assertEqual(payload["eligible_requirements"], 1)
        self.assertEqual(payload["items"][0]["source_ai_requirement_ids"], ["AI-KEEP"])

    def test_structured_function_fields_survive_analysis_and_excel(self) -> None:
        from functional_synthesis import run_functional_synthesis
        from openpyxl import load_workbook
        from requirements_analysis import run_requirements_analysis

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _write_jsonl(out / "ai_requirements.jsonl", [{
                "ai_req_id": "AI-1",
                "functional_key": "事件留存管理",
                "title": "配置事件留存",
                "description": "系统应保留事件30天。",
                "source_quote": "Events shall be retained for 30 days.",
                "source_section": "7.2",
                "source_block_ids": ["B-1"],
                "module": "事件记录",
                "preconditions": ["事件日志已启用"],
                "exceptions": ["存储故障时产生告警"],
                "related_dlms_objects": ["0-0:96.11.0.255"],
            }])

            run_functional_synthesis(out)
            run_requirements_analysis(out, route="stub")
            item = json.loads((out / "engineering_analysis.json").read_text(encoding="utf-8"))["items"][0]
            workbook = load_workbook(out / "software_requirements.xlsx", data_only=True)
            notes = str(next(workbook.active.iter_rows(min_row=2, values_only=True))[6] or "")

        self.assertEqual(item["objective"], "实现事件留存管理，并满足所有来源条款及适用变体。")
        self.assertEqual(item["behaviors"], ["系统应保留事件30天。"])
        self.assertEqual(item["preconditions"], ["事件日志已启用"])
        self.assertIn("30 天", item["data_constraints"])
        self.assertEqual(item["exceptions"], ["存储故障时产生告警"])
        self.assertIn("0-0:96.11.0.255", item["related_dlms_objects"])
        self.assertIn("功能目标：", notes)
        self.assertIn("前置条件：事件日志已启用", notes)
        self.assertIn("关联 DLMS 对象：0-0:96.11.0.255", notes)
    def test_ownership_override_survives_synthesis_and_analysis(self) -> None:
        from functional_synthesis import run_functional_synthesis
        from requirements_analysis import run_requirements_analysis

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _write_jsonl(out / "ai_requirements.jsonl", [{
                "ai_req_id": "AI-1", "title": "移动集中器", "description": "移动数据集中器设备。",
                "source_quote": "Device with mobile data concentrator function.", "module": "通信",
            }])
            _write_jsonl(out / "ai_review_states.jsonl", [{
                "ai_req_id": "AI-1", "status": "accepted", "ownership_override": "co_design",
            }])

            run_functional_synthesis(out)
            run_requirements_analysis(out, route="stub")
            item = json.loads((out / "engineering_analysis.json").read_text(encoding="utf-8"))["items"][0]

        self.assertEqual(item["ownership"], "co_design")

    def test_module_override_is_applied_before_synthesis(self) -> None:
        from functional_synthesis import run_functional_synthesis

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            _write_jsonl(out / "ai_requirements.jsonl", [
                {"ai_req_id": "AI-1", "title": "远程传输事件", "module": "其它"},
            ])
            _write_jsonl(out / "ai_review_states.jsonl", [
                {"ai_req_id": "AI-1", "status": "accepted", "module_override": "事件记录"},
            ])

            run_functional_synthesis(out)
            payload = json.loads((out / "functional_requirements.json").read_text(encoding="utf-8"))

        self.assertEqual(payload["items"][0]["module"], "事件记录")
        self.assertTrue(payload["items"][0]["functional_key"].startswith("事件记录:"))


class FunctionalSynthesisRegressionTests(unittest.TestCase):
    def test_cross_section_requirements_merge_by_functional_key(self) -> None:
        from functional_synthesis import synthesize_requirements

        requirements = [
            {
                "ai_req_id": "AI-1",
                "functional_key": "重要事件采集与远传",
                "title": "采集重要事件",
                "description": "设备须采集重要事件。",
                "source_quote": "The device shall collect significant events.",
                "source_section": "5 Event collection",
                "source_block_ids": ["B-1"],
                "module": "事件记录",
                "labels": ["重要事件"],
                "dev_guidance": ["实现重要事件采集入口"],
                "acceptance_criteria": ["重要事件发生后生成记录"],
            },
            {
                "ai_req_id": "AI-2",
                "functional_key": "重要事件采集与远传",
                "title": "远程传输重要事件",
                "description": "设备须将重要事件传输到中心。",
                "source_quote": "The device shall transmit significant events to the Center.",
                "source_section": "9 Remote communication",
                "source_block_ids": ["B-2"],
                "module": "事件记录",
                "labels": ["重要事件"],
                "dev_guidance": ["实现重要事件上传接口"],
                "acceptance_criteria": ["中心收到与本地一致的重要事件"],
            },
            {
                "ai_req_id": "AI-3",
                "functional_key": "结算周期管理",
                "title": "配置结算周期",
                "description": "设备须支持配置结算周期。",
                "source_quote": "The billing period shall be configurable.",
                "source_section": "10 Billing",
                "source_block_ids": ["B-3"],
                "module": "结算",
                "labels": ["结算周期"],
                "dev_guidance": [],
                "acceptance_criteria": [],
            },
        ]

        result = synthesize_requirements(requirements)

        self.assertEqual(len(result), 2)
        expected = _semantic_case("significant-event-cross-section")
        merged = next(row for row in result if row["functional_key"] == "重要事件采集与远传")
        self.assertEqual(len([merged]), expected["functional_requirement_count"])
        expected_ids = [value.replace("AI-EVENT-COLLECT", "AI-1").replace("AI-EVENT-TRANSMIT", "AI-2") for value in expected["source_ai_requirement_ids"]]
        self.assertEqual(merged["source_ai_requirement_ids"], expected_ids)
        self.assertEqual(merged["source_block_ids"], ["B-1", "B-2"])
        self.assertEqual(len(merged["source_quotes"]), 2)
        self.assertEqual(len(merged["developer_guidance"]), 2)
        self.assertEqual(len(merged["acceptance_criteria"]), 2)


if __name__ == "__main__":
    unittest.main()
