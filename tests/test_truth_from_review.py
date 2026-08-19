from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "tools"))

from truth_from_review import main as convert  # noqa: E402


def _make_xlsx(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["客户需求章节", "需求", "条件"])
    sheet.append(["4.2", "电表应支持事件记录读取。", "-25~+70 °C 环境下"])
    sheet.append(["6.1", "通信采用 OBIS 1-1:32.7.0 电压寄存器。", ""])
    sheet.append(["4.3", "", "空正文行应被跳过"])
    sheet.append(["4.4", "电表应支持 4 象限有功电能记录，单位 kWh。", ""])
    workbook.save(path)


class TruthFromReviewTests(unittest.TestCase):
    def test_xlsx_conversion_validates_and_writes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "人工核对.xlsx"
            out = Path(tmp) / "ws0" / "truth.jsonl"
            _make_xlsx(src)
            code = convert(["--input", str(src),
                            "--document-id", "abnt_nbr_16968",
                            "--output", str(out)])
            self.assertEqual(code, 0)
            rows = [json.loads(line) for line in
                    out.read_text(encoding="utf-8").splitlines() if line.strip()]
            self.assertEqual(len(rows), 3)   # 空正文行跳过
            from jsonschema import Draft202012Validator

            schema = json.loads((REPO_ROOT / "schemas" / "functional_truth.schema.json")
                                .read_text(encoding="utf-8"))
            validator = Draft202012Validator(schema)
            for row in rows:
                errors = list(validator.iter_errors(row))
                self.assertEqual(errors, [])
            self.assertTrue(all(r["document_id"] == "abnt_nbr_16968" for r in rows))
            self.assertEqual([r["truth_id"] for r in rows],
                             ["T-0001", "T-0002", "T-0003"])
            # 确定性抽取：OBIS 进 codes、数值进 numbers、单位进 units
            obis_row = next(r for r in rows if "OBIS" in r["expected_text"])
            self.assertIn("1-1:32.7.0", obis_row["codes"])
            env_row = rows[0]
            self.assertIn("25", env_row["numbers"])
            self.assertIn("70", env_row["numbers"])
            self.assertIn("°C".casefold(), [u.casefold() for u in env_row["units"]])
            # 条件文本进 conditions
            self.assertEqual(env_row["conditions"], ["-25~+70 °C 环境下"])

    def test_xlsx_missing_body_column_refuses(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "bad.xlsx"
            workbook = Workbook()
            workbook.active.append(["章节", "无关列"])
            workbook.active.append(["4.1", "x"])
            workbook.save(src)
            with self.assertRaises(SystemExit) as ctx:
                convert(["--input", str(src), "--document-id", "d",
                         "--output", str(Path(tmp) / "t.jsonl")])
            self.assertIn("缺需求正文列", str(ctx.exception))

    def test_json_conversion_with_explicit_keys(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "merged.json"
            out = Path(tmp) / "truth.jsonl"
            src.write_text(json.dumps({"items": [
                {"expected_text": "The meter shall log events.", "section_id": "4.1"},
                {"expected_text": "Voltage register OBIS 1-1:32.7.0.", "section_id": "6"},
            ]}, ensure_ascii=False), encoding="utf-8")
            code = convert(["--input", str(src), "--document-id", "*",
                            "--output", str(out), "--no-extract"])
            self.assertEqual(code, 0)
            rows = [json.loads(line) for line in
                    out.read_text(encoding="utf-8").splitlines() if line.strip()]
            self.assertEqual(len(rows), 2)
            self.assertTrue(all(r["numbers"] == [] and r["codes"] == []
                                for r in rows))   # --no-extract 留空
            self.assertEqual(rows[0]["section_id"], "4.1")

    def test_zero_rows_refuses_write(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "empty.xlsx"
            out = Path(tmp) / "t.jsonl"
            workbook = Workbook()
            workbook.active.append(["需求"])
            workbook.active.append([""])   # 唯一数据行为空
            workbook.save(src)
            with self.assertRaises(SystemExit) as ctx:
                convert(["--input", str(src), "--document-id", "d",
                         "--output", str(out)])
            self.assertIn("0 行", str(ctx.exception))
            self.assertFalse(out.exists())

    def test_cli_end_to_end(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "m.xlsx"
            out = Path(tmp) / "truth.jsonl"
            _make_xlsx(src)
            result = subprocess.run(
                [sys.executable, str(REPO_ROOT / "tools" / "truth_from_review.py"),
                 "--input", str(src), "--document-id", "abnt_nbr_16968",
                 "--output", str(out)],
                capture_output=True, text=True, encoding="utf-8", cwd=str(REPO_ROOT))
            self.assertEqual(result.returncode, 0, result.stderr[-400:])
            payload = json.loads(result.stdout)
            self.assertTrue(payload["ok"])
            self.assertEqual(payload["rows"], 3)
            self.assertTrue(out.is_file())


if __name__ == "__main__":
    unittest.main()
