"""Agent Phase 2 WP2：软件需求无依据富化字段强制"待澄清"的确定性规则测试。

冻结点：仅富化叙述字段（software_requirement_text/hardware_dependency/developer_guidance/
design_options/acceptance_criteria）；确定性 join 字段（id/归属/引句/模块）永不标；
只对"无依据"下手——有据字段逐字节不变（质量问题归专家审查,不归确定性层）。
"""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from requirements_analysis import (
    ANALYZE_PROMPT_VERSION,
    CLARIFY_MARK,
    UNFOUNDED_RULE_VERSION,
    _apply_llm_item,
    _enrich_key,
    run_requirements_analysis,
)


SOURCE = {
    "ai_req_id": "AI-1",
    "module": "事件记录",
    "source_quote": "The meter shall log tamper events.",
    "description": "",
    "requirement": "",
}
CTX = {"template_refs": "", "section_context": "", "doc_context": "",
       "exemplars": "", "answers": "", "siblings": ""}


def base_item(ownership: str = "software") -> dict:
    return {
        "analysis_id": "SRA-001",
        "ownership": ownership,
        "ownership_reason": "Matched software rule term: log",
        "ownership_source": "rule",
        "open_questions": [],
        "developer_guidance": [],
        "design_options": [],
        "acceptance_criteria": [],
        "hardware_dependency": "",
        "software_requirement_text": "The meter shall log tamper events.",
    }


class RejectedEnrichmentMarkingTests(unittest.TestCase):
    def test_fabricated_code_rejection_marks_unfounded_fields(self) -> None:
        """整体拒绝（编造编码）→ 无依据字段写"待澄清",不再静默以 base 文本充当正文。"""
        item = base_item()
        ok, issues = _apply_llm_item(
            item, SOURCE,
            {"software_requirement_text": "将事件写入 0-0:96.1.7 日志。",   # 换位 OBIS=编造
             "developer_guidance": ["写 0-0:96.1.7"]},
            CTX)

        self.assertFalse(ok)
        self.assertTrue(any("编造结构编码" in msg for msg in issues))
        self.assertEqual(item["software_requirement_text"], CLARIFY_MARK)
        self.assertEqual(item["developer_guidance"], [CLARIFY_MARK])
        self.assertEqual(item["design_options"], [CLARIFY_MARK])
        self.assertEqual(item["acceptance_criteria"], [CLARIFY_MARK])
        # 纯软件项 hardware_dependency 留空是设计语义,非"无依据"——不标
        self.assertEqual(item["hardware_dependency"], "")
        # 确定性 join 字段（归属/id）永不标
        self.assertEqual(item["ownership"], "software")
        self.assertEqual(item["analysis_id"], "SRA-001")

    def test_rejection_syncs_open_questions_per_marked_field(self) -> None:
        """每个"待澄清"同步一条 open_questions（内部核对受众）。"""
        item = base_item()
        _apply_llm_item(item, SOURCE, {"software_requirement_text": "写入 0-0:96.1.7。"}, CTX)

        questions = item["open_questions"]
        self.assertEqual(len(questions), 4)   # 正文 + 研发指引/设计候选/验收标准
        self.assertTrue(all("待澄清" in q for q in questions))
        self.assertTrue(all("内部核对" in q for q in questions))
        self.assertTrue(any("软件需求正文" in q for q in questions))

    def test_co_design_rejection_marks_hardware_dependency(self) -> None:
        """协同项的硬件依赖本应由此番富化产出——拒绝时同样标"待澄清"。"""
        item = base_item(ownership="co_design")
        _apply_llm_item(item, SOURCE, {"software_requirement_text": "写入 0-0:96.1.7。"}, CTX)

        self.assertEqual(item["hardware_dependency"], CLARIFY_MARK)
        self.assertTrue(any("硬件依赖" in q for q in item["open_questions"]))

    def test_grounded_base_list_fields_survive_rejection(self) -> None:
        """base 非空=源文有据内容——拒绝时逐字节保留（只对无依据下手,不毁有据内容）。"""
        item = base_item()
        item["developer_guidance"] = ["按源文：记录篡改事件"]
        _apply_llm_item(item, SOURCE, {"software_requirement_text": "写入 0-0:96.1.7。"}, CTX)

        self.assertEqual(item["developer_guidance"], ["按源文：记录篡改事件"])
        self.assertEqual(item["software_requirement_text"], CLARIFY_MARK)

    def test_empty_llm_item_marks_unfounded_fields(self) -> None:
        """富化未返回可采纳内容（整体回退 base 的另一条路径）→ 同样标"待澄清"。"""
        item = base_item()
        ok, issues = _apply_llm_item(item, SOURCE, {}, CTX)

        self.assertFalse(ok)
        self.assertEqual(item["software_requirement_text"], CLARIFY_MARK)
        self.assertTrue(any("待澄清" in q for q in item["open_questions"]))


class AcceptedEnrichmentFieldDowngradeTests(unittest.TestCase):
    def test_accepted_but_unfounded_number_field_marked(self) -> None:
        """接受但某字段证据校验降级（无据数字）→ 该字段"待澄清",其余有据字段原文保留。"""
        item = base_item()
        ok, issues = _apply_llm_item(
            item, SOURCE,
            {"software_requirement_text": "Log tamper events within 60 seconds.",   # 60 无据
             "developer_guidance": ["Log tamper events deterministically."]},
            CTX)

        self.assertTrue(ok)                       # 富化整体仍被采纳
        self.assertEqual(item["software_requirement_text"], CLARIFY_MARK)
        self.assertEqual(item["developer_guidance"], ["Log tamper events deterministically."])  # 有据保留
        self.assertEqual(item["analysis_source"], "llm")
        self.assertTrue(any("待澄清" in msg for msg in issues))
        self.assertTrue(any("软件需求正文" in q for q in item["open_questions"]))

    def test_grounded_enrichment_kept_byte_identical(self) -> None:
        """有据富化逐字节不变——正文/指引全部保留,open_questions 零新增。"""
        item = base_item()
        ok, _issues = _apply_llm_item(
            item, SOURCE,
            {"software_requirement_text": "Log tamper events deterministically.",
             "acceptance_criteria": ["Trigger a tamper event; it appears in the log."]},
            CTX)

        self.assertTrue(ok)
        self.assertEqual(item["software_requirement_text"], "Log tamper events deterministically.")
        self.assertEqual(item["acceptance_criteria"], ["Trigger a tamper event; it appears in the log."])
        self.assertEqual(item["open_questions"], [])

    def test_enum_step_markers_are_not_unfounded(self) -> None:
        """"1. 2. 3."步骤标号是格式归一不是编造数字（test18 判例）→ 字段保留不标。"""
        item = base_item()
        ok, _issues = _apply_llm_item(
            item, SOURCE,
            {"software_requirement_text": "1. 检测篡改；2. 记录事件；3. 上报。"},
            CTX)

        self.assertTrue(ok)
        self.assertEqual(item["software_requirement_text"], "1. 检测篡改；2. 记录事件；3. 上报。")
        self.assertEqual(item["open_questions"], [])

    def test_template_sourced_number_in_guidance_not_unfounded(self) -> None:
        """指引基线=源文∪模板注入——模板里的数值进指引不算"无依据"（有模板依据）。"""
        item = base_item()
        ctx = dict(CTX, template_refs="公司样本：事件环形缓冲 100 条")
        ok, _issues = _apply_llm_item(
            item, SOURCE,
            {"software_requirement_text": "Log tamper events.",
             "developer_guidance": ["公司通用做法：事件环形缓冲 100 条"]},
            ctx)

        self.assertTrue(ok)
        self.assertEqual(item["developer_guidance"], ["公司通用做法：事件环形缓冲 100 条"])
        self.assertEqual(item["open_questions"], [])

    def test_unfounded_number_in_guidance_list_marked(self) -> None:
        """研发指引含源文/模板均无据的数值 → 该字段整体标"待澄清"。"""
        item = base_item()
        ok, _issues = _apply_llm_item(
            item, SOURCE,
            {"software_requirement_text": "Log tamper events.",
             "developer_guidance": ["缓冲 4000 条事件"]},   # 4000 无据
            CTX)

        self.assertTrue(ok)
        self.assertEqual(item["developer_guidance"], [CLARIFY_MARK])
        self.assertEqual(item["software_requirement_text"], "Log tamper events.")   # 有据字段不动

    def test_base_fields_not_rescanned_on_accept(self) -> None:
        """接受路径只查本次 LLM 采纳的字段——base 值（源文派生）永不标。"""
        item = base_item()
        item["developer_guidance"] = ["源文派生指引含数字 42"]   # base 非空且含非源文数字
        ok, _issues = _apply_llm_item(
            item, SOURCE,
            {"software_requirement_text": "Log tamper events."},   # 只采纳正文
            CTX)

        self.assertTrue(ok)
        self.assertEqual(item["developer_guidance"], ["源文派生指引含数字 42"])   # base 不动


class VersionAndCacheFingerprintTests(unittest.TestCase):
    def test_prompt_version_bumped(self) -> None:
        self.assertEqual(ANALYZE_PROMPT_VERSION, "analyze-llm-v7")

    def test_enrich_key_covers_unfounded_rule_version(self) -> None:
        """确定性后处理（待澄清规则）版本必须进 analyze_enrich_cache 指纹（AGENTS.md 纪律）。"""
        req = {"source_quote": "q", "description": "d", "requirement": "r", "module": "m"}
        key_now = _enrich_key(req, "model-x")
        with patch("requirements_analysis.UNFOUNDED_RULE_VERSION", "analyze-unfounded-v0-hypothetical"):
            key_changed = _enrich_key(req, "model-x")
        self.assertNotEqual(key_now, key_changed)
        self.assertEqual(UNFOUNDED_RULE_VERSION, "analyze-unfounded-v2")


class IntegrationRenderAndClarificationLoopTests(unittest.TestCase):
    def _seed(self, tmp_path: Path) -> None:
        (tmp_path / "ai_requirements.jsonl").write_text(
            json.dumps({
                "ai_req_id": "AI-1",
                "description": "The meter shall log power-down events to OBIS 0-0:96.1.0.",
                "source_quote": "log power-down events at 0-0:96.1.0, keep 100 entries",
                "source_block_ids": ["B-1"],
                "module": "事件记录",
            }, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )

    def test_end_to_end_rejection_marks_and_renders(self) -> None:
        """全链路：编造拒绝 → engineering_analysis.json 待澄清 + open_questions 同步
        → xlsx 需求列/「待确认」通道原样透出 → 澄清报告（既有通道）收进条目。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed(tmp_path)

            def fake_chat(system: str, user: str) -> dict:
                return {"items": [{
                    "source_requirement_ids": ["AI-1"],
                    "software_requirement_text": "将事件写入 0-0:96.1.7 日志。",   # 换位 OBIS=编造
                    "developer_guidance": ["写 0-0:96.1.7"],
                }]}

            result = run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat)

            self.assertEqual(result["enriched"], 0)
            self.assertEqual(result["enrich_degraded"], 1)
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            item = payload["items"][0]
            self.assertEqual(item["software_requirement_text"], CLARIFY_MARK)
            self.assertEqual(item["developer_guidance"], [CLARIFY_MARK])
            self.assertTrue(any("待澄清" in q for q in item["open_questions"]))

            # 渲染透出：需求列=待澄清标注 + 兜底原始候选（2026-07-23 用户裁定：既要诚实
            # 标注也要可读内容;数据层字段仍恒为待澄清,渲染层才透出标注候选）
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path / "software_requirements.xlsx")
            ws = wb[wb.sheetnames[0]]
            header = [cell.value for cell in ws[1]]
            req_col = header.index("需求")
            notes_col = header.index("说明、示例、注意事项")
            req_cell = str(ws.cell(2, req_col + 1).value or "")
            self.assertIn(CLARIFY_MARK, req_cell)
            self.assertIn("未经依据校验", req_cell)
            self.assertIn("The meter shall log power-down events to OBIS 0-0:96.1.0.", req_cell)  # 兜底候选带标注
            notes = str(ws.cell(2, notes_col + 1).value or "")
            self.assertIn("待确认：内部核对·待澄清", notes)
            wb.close()

            # 澄清闭环：clarification_report 读 engineering_analysis.json 的通道已存在——
            # 待澄清 open_questions 以 analyze:open_question 信号进报告（内部核对受众文本随行）
            import clarification_report
            report_result = clarification_report.run_report(tmp_path)
            self.assertGreaterEqual(report_result.get("soft_questions", 0), 1)
            report = json.loads((tmp_path / "clarification_report.json").read_text(encoding="utf-8"))
            hits = [e for e in report.get("entries", []) if "待澄清" in str(e.get("question", ""))]
            self.assertTrue(hits)
            self.assertEqual(hits[0]["signal"], "analyze:open_question")

    def test_stub_route_never_marks(self) -> None:
        """无 LLM 富化（stub/默认关闭）→ base 字段逐字节保留——WP2 只管富化被拒/无据。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed(tmp_path)
            result = run_requirements_analysis(tmp_path, route="stub")
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))

        self.assertEqual(result["route"], "stub")
        item = payload["items"][0]
        self.assertEqual(item["software_requirement_text"],
                         "The meter shall log power-down events to OBIS 0-0:96.1.0.")
        self.assertEqual(item["open_questions"], [])


if __name__ == "__main__":
    unittest.main()
