"""澄清问题清单 + 就绪判定（确定性零 LLM）。

真实需求评审的标配产物是"问客户清单"——文档说要做 X 但没给限值、A 章与 B 章数值矛盾、
推导时不得不做的假设……这些疑问信号管线里全都有（open_questions / suspicion_reasons /
一致性报告 / assumptions / 富化降级），但散落各处没人汇总。本模块把它们聚合成一份
评审会直接可用的清单，并给出 READY / NEEDS WORK 就绪判定（吸收 BMAD 的
assumptions 契约 + FR 覆盖门模式，用代码实现它只用提示词说的事）。

分类沿用 hs-req 三分类 + 假设：模糊（说了但不可测/不逐字）、缺失（该有数值没给）、
矛盾（跨章数值发散/重复表述）、假设待确认（LLM 推导中记录的前提）。

用法：python -m clarification_report --out <atomizer 输出目录>
产物：clarification_questions.md / clarification_questions.xlsx / clarification_report.json
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
import time
from contextlib import contextmanager
from pathlib import Path
from threading import RLock
from typing import Any, Iterator

from clarification_check_states import (
    VALID_CHECK_ACTIONS,
    apply_clarification_check_actions_batch,
    read_clarification_check_states,
)
from io_utils import read_jsonl
from result_package import governed_artifact_path
from text_normalize import formula_safe

REPORT_JSON = "clarification_report.json"
REPORT_MD = "clarification_questions.md"
REPORT_XLSX = "clarification_questions.xlsx"
ANSWERS_FILE = "clarification_answers.jsonl"   # 评审会答复回灌（answer 列填写后 import 回来）
ANSWERS_LOCK = "clarification_answers.lock"
CLARIFICATION_REPORT_VERSION = "clarification/v8-param-row-aggregate"

_ANSWER_LOCKS: dict[Path, RLock] = {}
_ANSWER_LOCKS_GUARD = RLock()
_ANSWER_LOCK_TIMEOUT_S = 10.0
_ANSWER_LOCK_STALE_AFTER_S = 300.0
_ANSWER_REPLACE_ATTEMPTS = 20
_ANSWER_REPLACE_RETRY_DELAY_S = 0.05

CAT_AMBIGUOUS = "模糊"
CAT_MISSING = "缺失"
CAT_CONFLICT = "矛盾"
CAT_ASSUMPTION = "假设待确认"

# 信号分级（真实教训：v10 数据上模型自报 303 条假设 + 277 条 open_questions，清单膨胀到
# 612 条评审会没法用）。硬信号=确定性机器检出（矛盾/漏值/验收不可测/引用不逐字）——必答；
# 软信号=模型自报（assumptions/open_questions）——价值在"可见"，留档备查不算就绪门。
TIER_HARD = "必答"
TIER_SOFT = "参考"
# 遗漏候选（0714 批次一）：覆盖缺口的 requirement_like 段落——机器检出但**假阳性率高**
# （前言/引用书目/纯标题都会命中 requirement_like），量可达上百条。既不能进必答（挤爆就绪门
# READY_MAX_QUESTIONS）也不是"模型自报"，单列一档：独立 sheet/段落，审核员核对是漏抽还是噪声。
TIER_GAP = "遗漏候选"
# 必答再分受众（真实产物观察：56 条必答被"引用非逐字"占领——它是审核员在批注视图核的
# **内部核对项**，不是评审会问客户的问题。分开后评审会拿到的是十几条纯客户问题）
AUDIENCE_CUSTOMER = "问客户"
AUDIENCE_INTERNAL = "内部核对"

# blocker_level 与抽取需求自身的 priority(P0/P1/P2)是两套概念。前者只表示该澄清项
# 对当前规格可用性的阻塞程度，不回写需求优先级。
BLOCKER_BLOCKING = "blocking"
BLOCKER_IMPORTANT = "important"
MODULE_UNASSIGNED = "未归属"

# 稳定信号码到阻塞度的确定性映射。高风险结构字段、编码、数值和已锚定语义复核均为
# blocking；叙述性/流程性问题为 important。未知 suspicion 也会显式进入 important，
# 不再因新护栏增加一个 reason 就静默漏出澄清报告。
SIGNAL_BLOCKER_LEVELS = {
    "suspicion:code_drift": BLOCKER_BLOCKING,
    "suspicion:number_drift": BLOCKER_BLOCKING,
    "suspicion:漏值": BLOCKER_BLOCKING,
    "suspicion:value_pairing": BLOCKER_BLOCKING,
    "suspicion:table_text_mismatch": BLOCKER_BLOCKING,
    "suspicion:standard_ref": BLOCKER_BLOCKING,
    "suspicion:second_pass_semantic": BLOCKER_BLOCKING,
    "consistency:obis_values_differ": BLOCKER_BLOCKING,
    "synthesis:conflict_flag": BLOCKER_BLOCKING,
    "parse_audit:noise_char_ratio": BLOCKER_BLOCKING,
    "parse_audit:body_ratio": BLOCKER_BLOCKING,
    "suspicion:引用": BLOCKER_IMPORTANT,
    "suspicion:验收": BLOCKER_IMPORTANT,
    "suspicion:informative_source": BLOCKER_IMPORTANT,
    "suspicion:modal_strength": BLOCKER_IMPORTANT,
    "suspicion:self_check_added": BLOCKER_IMPORTANT,
    "suspicion:self_check_conflict": BLOCKER_IMPORTANT,
    "consistency:duplicate": BLOCKER_IMPORTANT,
    "consistency:uncovered": BLOCKER_IMPORTANT,
    "consistency:uncovered_overflow": BLOCKER_IMPORTANT,
    "consistency:compliance_uncovered": BLOCKER_BLOCKING,
    "consistency:compliance_uncovered_overflow": BLOCKER_BLOCKING,
    "analyze:open_question": BLOCKER_IMPORTANT,
    "analyze:assumption": BLOCKER_IMPORTANT,
    # WS4 弱词/可测性扫描（功能需求级）：弱词→模糊（important），验收不可测→缺失（blocking）
    "weakness:vague_word": BLOCKER_IMPORTANT,
    "weakness:untestable": BLOCKER_BLOCKING,
}

# WS4 弱词词典（Cap3）：内置词表 + 可选 YAML 覆盖（RATOMIZER_WEAK_WORDS_PATH，与 domain_packs 惯例一致）。
# 命中即汇入既有四分类的「模糊」类，不新建报告；LLM 标的 ambiguity 布尔保留原位、并列展示不合并。
# 刻意只用语义明确的整词（适当/尽快/灵活/酌情…），不含「相关/一般/通常」等高频泛词——
# 那些在精确规格里大量合法出现，作为弱词会造成整篇文档假阳性 NEEDS WORK（宁漏勿错）。
DEFAULT_WEAK_WORDS: tuple[str, ...] = (
    "适当", "尽快", "灵活", "酌情", "视情况", "根据实际情况", "相应的",
    "as appropriate", "in a timely manner", "flexible", "as needed",
    "etc", "and so on", "reasonable amount",
)


def _load_weak_words() -> tuple[str, ...]:
    """加载弱词词典：env YAML 覆盖 > 内置词表。YAML 缺失/损坏如实退内置，绝不伪造空词表。"""
    yaml_path = os.environ.get("RATOMIZER_WEAK_WORDS_PATH", "").strip()
    if not yaml_path:
        return DEFAULT_WEAK_WORDS
    try:
        path = Path(yaml_path).expanduser()
        text = path.read_text(encoding="utf-8")
    except OSError:
        return DEFAULT_WEAK_WORDS
    parsed: list[str] = list(DEFAULT_WEAK_WORDS)
    try:
        import yaml  # type: ignore[import-untyped]
        data = yaml.safe_load(text)
    except Exception:
        # PyYAML 缺失或 YAML 损坏：按行裸解析（每行一个词，# 起注释）
        for line in text.splitlines():
            line = line.split("#", 1)[0].strip()
            if line:
                parsed.append(line)
        return tuple(dict.fromkeys(parsed))
    if isinstance(data, dict):
        words = data.get("weak_words") or data.get("words") or []
    else:
        words = data or []
    if isinstance(words, (list, tuple)):
        for item in words:
            value = str(item or "").strip()
            if value:
                parsed.append(value)
    return tuple(dict.fromkeys(parsed))

_SUSPICION_POLICIES: dict[str, tuple[str, str, str, str, str]] = {
    # reason: (stable signal, category, audience, blocker_level, tier)
    "编码漂移": ("suspicion:code_drift", CAT_CONFLICT, AUDIENCE_INTERNAL, BLOCKER_BLOCKING, TIER_HARD),
    "数字漂移": ("suspicion:number_drift", CAT_CONFLICT, AUDIENCE_INTERNAL, BLOCKER_BLOCKING, TIER_HARD),
    "原文数值未带全": ("suspicion:漏值", CAT_MISSING, AUDIENCE_CUSTOMER, BLOCKER_BLOCKING, TIER_HARD),
    "引用非逐字": ("suspicion:引用", CAT_AMBIGUOUS, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD),
    "引用跨段": ("suspicion:引用", CAT_AMBIGUOUS, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_SOFT),
    "引用标点差异": ("suspicion:引用", CAT_AMBIGUOUS, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_SOFT),
    "验收不可测": ("suspicion:验收", CAT_AMBIGUOUS, AUDIENCE_CUSTOMER, BLOCKER_IMPORTANT, TIER_HARD),
    "资料性来源待核": ("suspicion:informative_source", CAT_AMBIGUOUS, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD),
    "资料性附录来源": ("suspicion:informative_source", CAT_AMBIGUOUS, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD),
    "情态升格待核": ("suspicion:modal_strength", CAT_CONFLICT, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD),
    "标准号待核": ("suspicion:standard_ref", CAT_CONFLICT, AUDIENCE_INTERNAL, BLOCKER_BLOCKING, TIER_HARD),
    "数值配对待核": ("suspicion:value_pairing", CAT_CONFLICT, AUDIENCE_INTERNAL, BLOCKER_BLOCKING, TIER_HARD),
    "表文数值不一致": ("suspicion:table_text_mismatch", CAT_CONFLICT, AUDIENCE_INTERNAL, BLOCKER_BLOCKING, TIER_HARD),
    "确定性合规兜底（LLM 未覆盖）": ("suspicion:compliance_fallback", CAT_MISSING, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD),
    "参数表行确定性展开": ("suspicion:param_row_expand", CAT_MISSING, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD),
    # WP-B 点解析：用户定点触发的单行/单块解析结果，先人工确认再转正（冻结口径）
    "用户定点解析": ("suspicion:spot_extract", CAT_AMBIGUOUS, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD),
    "自检补充（初抽遗漏）": ("suspicion:self_check_added", CAT_AMBIGUOUS, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD),
    "自检复核:描述与引句疑似矛盾": (
        "suspicion:self_check_conflict", CAT_CONFLICT, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD
    ),
}

# 就绪门默认阈值（NEEDS WORK 触发条件；试点后按真实分布调）
READY_MAX_QUESTIONS = 30
READY_MIN_COVERAGE = 60.0


def suspicion_policy(reason: str) -> tuple[str, str, str, str, str] | None:
    """公开只读访问 suspicion 路由表（agent_eval 判定器等外部消费方不得依赖私有名）。"""
    return _SUSPICION_POLICIES.get(reason)


def _hash_payload(payload: Any) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _entry(category: str, question: str, *, section: str = "", quote: str = "",
           source_id: str = "", signal: str = "", tier: str = TIER_HARD,
           audience: str = AUDIENCE_CUSTOMER, blocker_level: str | None = None,
           module: str = "", evidence: Any = None, subject_key: str = "") -> dict[str, Any]:
    """Build one versioned clarification subject without conflating it with req priority."""
    section_text = str(section or "")
    quote_text = str(quote or "")
    source_text = str(source_id or "")
    signal_text = str(signal or "")
    effective_blocker = blocker_level or SIGNAL_BLOCKER_LEVELS.get(signal_text, BLOCKER_IMPORTANT)
    evidence_payload = evidence if evidence is not None else {
        "section": section_text,
        "quote": quote_text,
        "question": str(question or ""),
    }
    evidence_fingerprint = _hash_payload(evidence_payload)
    clarification_id = "CLR-" + _hash_payload({
        "signal": signal_text,
        "identity": str(subject_key or source_text),
    })[:16]
    return {
        "clarification_id": clarification_id,
        "evidence_fingerprint": evidence_fingerprint,
        "category": category,
        "question": question,
        "section": section_text,
        "quote": quote_text[:200],
        "source_id": source_text,
        "signal": signal_text,
        "tier": tier,
        "audience": audience,
        "blocker_level": effective_blocker,
        "module": str(module or "").strip() or MODULE_UNASSIGNED,
        "state": "",
        "actor": "",
        "timestamp": "",
        "note": "",
    }


def _suspicion_question(reason: str, *, section: str, title: str) -> str:
    if reason == "原文数值未带全":
        return f"文档 §{section}「{title}」附近的数值清单未完整进入需求，请逐项核对并确认参数"
    if reason == "引用非逐字":
        return f"§{section}「{title}」的引用与原文不逐字一致，请核对该需求是否忠实原文"
    if reason == "验收不可测":
        return f"§{section}「{title}」的验收标准含不可测表述（如\"符合要求\"），请给出可判定的通过/失败条件"
    if reason == "编码漂移":
        return f"§{section}「{title}」含原文无据的 OBIS/事件码/编码并已被护栏拦截，请核对正确编码"
    if reason == "数字漂移":
        return f"§{section}「{title}」含原文未出现的数字，请核对数值来源并更正"
    if reason == "数值配对待核":
        return f"§{section}「{title}」含多档数值，请逐项核对数值与型号/条件的配对"
    if reason == "表文数值不一致":
        return f"§{section}「{title}」的正文/验收与阈值表数值不一致，请以原表为证据核对"
    if reason == "标准号待核":
        return f"§{section}「{title}」含本节原文未见的标准号，请核对标准归属"
    if reason in {"资料性来源待核", "资料性附录来源"}:
        return f"§{section}「{title}」来自资料性内容，请核对是否被误作强制需求"
    if reason == "情态升格待核":
        return f"§{section}「{title}」的约束强度已按原文软化，请核对 should/shall 等情态是否忠实"
    if reason == "自检补充（初抽遗漏）":
        return f"§{section}「{title}」由完整性自检补回，请优先核对是否为真实遗漏"
    if reason == "自检复核:描述与引句疑似矛盾":
        return f"§{section}「{title}」的描述与引句疑似矛盾，请按留痕证据核对"
    if reason.startswith("二遍复核:"):
        return f"§{section}「{title}」被二遍语义复核标记为“{reason.split(':', 1)[1]}”，请按双侧锚定证据核对"
    return f"§{section}「{title}」存在待核信号“{reason}”，请核对后处置"


def collect_questions(out_dir: Path) -> list[dict[str, Any]]:
    """聚合全链疑问信号（每个来源都可缺席——有什么聚什么，出处如实标注 signal）。"""
    entries: list[dict[str, Any]] = []
    closed_omission_blocks: set[str] = set()
    try:
        from omission_actions import omission_source_fingerprint, read_omission_states

        block_text = {
            str(block.get("block_id") or ""): str(block.get("text") or "")
            for block in read_jsonl(out_dir / "blocks.jsonl")
            if block.get("block_id")
        }
        for state in read_omission_states(out_dir).values():
            block_id = str(state.get("block_id") or "")
            if not block_id or block_id not in block_text:
                continue
            current_fingerprint = omission_source_fingerprint(block_id, block_text[block_id])
            if str(state.get("source_fingerprint") or "") != current_fingerprint:
                continue
            status = str(state.get("status") or "")
            if status in {"non_requirement", "resolved"}:
                closed_omission_blocks.add(block_id)
    except (OSError, TimeoutError, ValueError):
        # 缺少旧目录 sidecar 或锁暂时不可用时，保守保留候选，不静默消解。
        pass

    # ① 抽取层 suspicion（ai_requirements.jsonl）
    reqs_path = out_dir / "ai_requirements.jsonl"
    reqs = read_jsonl(reqs_path) if reqs_path.exists() else []
    from ai_review_actions import read_ai_review_states, source_ai_requirement_id
    review_states = read_ai_review_states(out_dir)
    module_by_source: dict[str, str] = {}
    for req in reqs:
        rid = source_ai_requirement_id(req)
        state = review_states.get(rid) or {}
        module_by_source[rid] = (
            str(state.get("module_override") or "").strip()
            or str(req.get("module") or "").strip()
            or MODULE_UNASSIGNED
        )

    def module_for_sources(source_ids: Any) -> str:
        if isinstance(source_ids, str):
            ids = [source_ids]
        elif isinstance(source_ids, (list, tuple, set)):
            ids = [str(value or "") for value in source_ids]
        else:
            ids = []
        modules = sorted({module_by_source.get(rid, "") for rid in ids if module_by_source.get(rid, "")})
        return " / ".join(modules) if modules else MODULE_UNASSIGNED

    requirement_occurrences: dict[tuple[str, tuple[str, ...]], int] = {}
    # 封堵三:PROW-DET 行级 suspicion(deterministic_fallback)按(表块,reason)聚合,
    # 免行级化后大量行级入口打爆 READY 门;LLM 需求 suspicion 仍逐条(个体语义)
    prow_groups: dict[tuple[str, str], dict[str, Any]] = {}
    for req in reqs:
        rid = source_ai_requirement_id(req)
        sec = str(req.get("source_section") or "")
        quote = str(req.get("source_quote") or "")
        title = str(req.get("title") or "")
        module = module_by_source.get(rid, MODULE_UNASSIGNED)
        anchor_key = (sec, tuple(str(value) for value in (req.get("source_block_ids") or [])))
        occurrence = requirement_occurrences.get(anchor_key, 0) + 1
        requirement_occurrences[anchor_key] = occurrence
        requirement_subject = "requirement:" + _hash_payload({
            "source_section": sec,
            "source_block_ids": list(anchor_key[1]),
            "occurrence": occurrence,
        })[:16]
        is_deterministic_row = str(req.get("source_mapping") or "") == "deterministic_fallback"
        for raw_reason in (req.get("suspicion_reasons") or []):
            reason = str(raw_reason or "").strip()
            if not reason:
                continue
            if reason.startswith("二遍复核:"):
                signal, category, audience, blocker, tier = (
                    "suspicion:second_pass_semantic",
                    CAT_CONFLICT,
                    AUDIENCE_INTERNAL,
                    BLOCKER_BLOCKING,
                    TIER_HARD,
                )
                reason_subject = f"{requirement_subject}:verify:{reason.split(':', 1)[1]}"
            else:
                policy = _SUSPICION_POLICIES.get(reason)
                if policy is None:
                    signal = "suspicion:other:" + hashlib.sha1(reason.encode("utf-8")).hexdigest()[:10]
                    category, audience, blocker, tier = CAT_AMBIGUOUS, AUDIENCE_INTERNAL, BLOCKER_IMPORTANT, TIER_HARD
                else:
                    signal, category, audience, blocker, tier = policy
                reason_subject = requirement_subject
            if is_deterministic_row:
                # 封堵三:行级确定性 suspicion 按表块聚合,不逐条 append
                block_ids = req.get("source_block_ids") or []
                block_id = str(block_ids[0]) if block_ids else ""
                key = (block_id, reason)
                grp = prow_groups.get(key)
                if grp is None:
                    grp = {"rows": [], "sec": sec, "module": module, "signal": signal,
                           "category": category, "audience": audience, "blocker": blocker,
                           "tier": tier, "subject": "param-row:" + _hash_payload(
                               {"block_id": block_id, "reason": reason})[:16]}
                    prow_groups[key] = grp
                grp["rows"].append({"row_index": req.get("source_row_index"),
                                    "source_quote": quote, "title": title})
                continue
            entries.append(_entry(
                category,
                _suspicion_question(reason, section=sec, title=title),
                section=sec,
                quote=quote,
                source_id=rid,
                signal=signal,
                tier=tier,
                audience=audience,
                blocker_level=blocker,
                module=module,
                evidence={
                    "reason": reason,
                    "source_section": sec,
                    "source_quote": quote,
                    "title": title,
                    "description": req.get("description") or "",
                    "sub_items": req.get("sub_items") or [],
                    "acceptance_criteria": req.get("acceptance_criteria") or [],
                    "threshold_table": req.get("threshold_table") or {},
                    "notes": req.get("notes") or "",
                },
                subject_key=reason_subject,
            ))

    # 封堵三:聚合 PROW-DET 行级 suspicion → 每表块每类一条汇总(明细挂 row_details)
    for (block_id, reason), grp in prow_groups.items():
        rows = grp["rows"]
        summary = f"表格 {block_id}（{grp['sec']}）：{reason} {len(rows)} 行待核"
        aggregate_entry = _entry(
            grp["category"],
            summary,
            section=grp["sec"],
            quote=rows[0].get("source_quote", "") if rows else "",
            source_id=block_id,
            signal=grp["signal"],
            tier=grp["tier"],
            audience=grp["audience"],
            blocker_level=grp["blocker"],
            module=grp["module"],
            evidence={"reason": reason, "table_block_id": block_id, "source_section": grp["sec"]},
            subject_key=grp["subject"],
        )
        aggregate_entry["row_details"] = rows   # 明细挂展开视图(xlsx/md 渲染读)
        entries.append(aggregate_entry)

    # ② 分析层 open_questions + assumptions（engineering_analysis.json）
    ana_path = out_dir / "engineering_analysis.json"
    if ana_path.exists():
        try:
            payload = json.loads(ana_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            payload = {}
        if not isinstance(payload, dict):
            raise ValueError("engineering_analysis.json must contain a JSON object")
        for item in payload.get("items") or []:
            sec = str(item.get("source_section") or "")
            rid = (item.get("source_requirement_ids") or [""])[0]
            quote = str(item.get("source_quote") or "")
            analysis_identity = str(
                item.get("analysis_id") or item.get("functional_requirement_id") or rid or sec
            )
            for question_index, q in enumerate(item.get("open_questions") or [], start=1):
                entries.append(_entry(CAT_MISSING, str(q), section=sec, quote=quote,
                                      source_id=str(rid), signal="analyze:open_question",
                                      tier=TIER_SOFT, module=module_for_sources(str(rid)),
                                      evidence={"question": str(q), "quote": quote, "section": sec},
                                      subject_key=f"{analysis_identity}:open-question:{question_index}"))
            for assumption_index, a in enumerate(item.get("assumptions") or [], start=1):
                entries.append(_entry(
                    CAT_ASSUMPTION, f"推导时假设：{a}——请确认该前提是否成立",
                    section=sec, quote=quote, source_id=str(rid), signal="analyze:assumption",
                    tier=TIER_SOFT, module=module_for_sources(str(rid)),
                    evidence={"assumption": str(a), "quote": quote, "section": sec},
                    subject_key=f"{analysis_identity}:assumption:{assumption_index}"))

    # ③ 一致性层（consistency_report.json）：跨章数值矛盾 + 重复表述
    con_path = out_dir / "consistency_report.json"
    if con_path.exists():
        try:
            con = json.loads(con_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            con = {}
        if not isinstance(con, dict):
            raise ValueError("consistency_report.json must contain a JSON object")
        for g in (con.get("obis_coreference") or []):
            if g.get("values_differ"):
                code = g.get("obis") or g.get("code") or ""   # 生产方写 obis；兼容早期 code 夹具
                members = [str(value or "") for value in (g.get("members") or [])]
                primary_source = members[0] if members else ""
                entries.append(_entry(
                    CAT_CONFLICT,
                    f"OBIS {code} 在多处被引用且数值不一致，请确认以哪一处为准",
                    section=" / ".join(str(value) for value in (g.get("sections") or [])),
                    source_id=primary_source,
                    signal="consistency:obis_values_differ",
                    blocker_level=BLOCKER_BLOCKING,
                    module=module_for_sources(members),
                    evidence={"obis": code, "members": members, "sections": g.get("sections") or []},
                    subject_key=f"obis:{code}"))
        for group_index, g in enumerate(con.get("duplicate_groups") or [], start=1):
            quote = str(g.get("source_quote") or "")
            members = [str(value or "") for value in (g.get("members") or [])]
            entries.append(_entry(
                CAT_CONFLICT,
                "同一原文语句被抽为多条需求，请确认是否合并或存在语义差异",
                section=" / ".join(str(value) for value in (g.get("sections") or [])),
                quote=quote, source_id=members[0] if members else "",
                signal="consistency:duplicate", module=module_for_sources(members),
                evidence={"quote": quote, "members": members, "sections": g.get("sections") or []},
                subject_key=f"duplicate:{group_index}"))
        # 覆盖缺口 → 遗漏候选（0714 批次一）：最直接的"漏需求"信号此前只有一个计数,
        # 30 条样本文本无处可看。兼容旧报表的裸字符串样本（无溯源字段按空处理）。
        coverage = con.get("coverage") or {}
        samples = coverage.get("uncovered_samples") or []
        candidate_ids = {
            str(value or "") for value in (coverage.get("uncovered_block_ids") or [])
            if str(value or "")
        }
        if not candidate_ids:
            # 兼容旧 consistency 报表：只对样本里能证明仍是候选的关闭项做减法。
            candidate_ids = {
                str(sample.get("block_id") or "")
                for sample in samples if isinstance(sample, dict) and sample.get("block_id")
            }
        closed_current_candidates = closed_omission_blocks.intersection(candidate_ids)
        for sample_index, s in enumerate(samples, start=1):
            if isinstance(s, dict):
                text = str(s.get("text") or "")
                bid = str(s.get("block_id") or "")
                sec = str(s.get("section") or "")
            else:
                text, bid, sec = str(s or ""), "", ""
            if not text.strip():
                continue
            if bid in closed_current_candidates:
                continue
            entries.append(_entry(
                CAT_MISSING,
                "该段疑似含需求但未被任何抽取需求覆盖——请核对是漏抽还是非需求文本（前言/引用/标题）",
                section=sec, quote=text, source_id=bid,
                signal="consistency:uncovered", tier=TIER_GAP, audience=AUDIENCE_INTERNAL,
                blocker_level=BLOCKER_IMPORTANT,
                evidence={"block_id": bid, "section": sec, "text": text},
                subject_key=bid or f"legacy-uncovered:{sample_index}"))
        total_uncovered = int(coverage.get("uncovered_count") or 0)
        effective_uncovered = max(0, total_uncovered - len(closed_current_candidates))
        listed_uncovered = sum(
            1 for sample in samples
            if not isinstance(sample, dict)
            or str(sample.get("block_id") or "") not in closed_current_candidates
        )
        if effective_uncovered > listed_uncovered and samples:
            # 无声截断禁令：报表样本有 30 条上限,超出部分必须留痕
            entries.append(_entry(
                CAT_MISSING,
                f"另有 {effective_uncovered - listed_uncovered} 条遗漏候选超出样本上限未列出"
                "（全量计数见 consistency_report.json coverage.uncovered_count）",
                signal="consistency:uncovered_overflow", tier=TIER_GAP,
                audience=AUDIENCE_INTERNAL, blocker_level=BLOCKER_IMPORTANT,
                evidence={"uncovered_count": effective_uncovered, "listed": listed_uncovered}))

        # 合规交付项单独计量：证书、法令、检定周期等漏项不会再稀释 core 覆盖率，但在
        # 人工判定“非需求”或补抽覆盖之前必须阻塞 READY。
        compliance_coverage = coverage.get("compliance") or {}
        compliance_samples = compliance_coverage.get("uncovered_samples") or []
        compliance_candidate_ids = {
            str(value or "")
            for value in (compliance_coverage.get("uncovered_block_ids") or [])
            if str(value or "")
        }
        if not compliance_candidate_ids:
            compliance_candidate_ids = {
                str(sample.get("block_id") or "")
                for sample in compliance_samples
                if isinstance(sample, dict) and sample.get("block_id")
            }
        closed_compliance = closed_omission_blocks.intersection(compliance_candidate_ids)
        for sample_index, sample in enumerate(compliance_samples, start=1):
            if isinstance(sample, dict):
                text = str(sample.get("text") or "")
                block_id = str(sample.get("block_id") or "")
                section = str(sample.get("section") or "")
            else:
                text, block_id, section = str(sample or ""), "", ""
            if not text.strip() or block_id in closed_compliance:
                continue
            entries.append(_entry(
                CAT_MISSING,
                "该段疑似合规交付义务但未被合规需求覆盖——请补抽，或核对后明确标为非需求",
                section=section,
                quote=text,
                source_id=block_id,
                signal="consistency:compliance_uncovered",
                tier=TIER_HARD,
                audience=AUDIENCE_INTERNAL,
                blocker_level=BLOCKER_BLOCKING,
                evidence={"block_id": block_id, "section": section, "text": text},
                subject_key=block_id or f"legacy-compliance-uncovered:{sample_index}",
            ))
        compliance_total_uncovered = int(compliance_coverage.get("uncovered_count") or 0)
        effective_compliance_uncovered = max(
            0, compliance_total_uncovered - len(closed_compliance)
        )
        listed_compliance_uncovered = sum(
            1 for sample in compliance_samples
            if not isinstance(sample, dict)
            or str(sample.get("block_id") or "") not in closed_compliance
        )
        if effective_compliance_uncovered > listed_compliance_uncovered and compliance_samples:
            entries.append(_entry(
                CAT_MISSING,
                f"另有 {effective_compliance_uncovered - listed_compliance_uncovered} 条合规漏项"
                "超出样本上限未列出（全量块 ID 见 consistency_report.json）",
                signal="consistency:compliance_uncovered_overflow",
                tier=TIER_HARD,
                audience=AUDIENCE_INTERNAL,
                blocker_level=BLOCKER_BLOCKING,
                evidence={
                    "uncovered_count": effective_compliance_uncovered,
                    "listed": listed_compliance_uncovered,
                },
            ))

    # 合成层冲突标记（C10，0710 评审）：确定性检出的"同一功能未限定参数冲突/归属覆盖冲突"
    # 此前只落 functional_requirements.json——评审会该裁的冲突在澄清清单上隐身
    synth_path = out_dir / "functional_requirements.json"
    if synth_path.exists():
        try:
            synth = json.loads(synth_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            synth = {}
        for item in (synth.get("items") or []) if isinstance(synth, dict) else []:
            title = str(item.get("title") or item.get("functional_key") or "")
            functional_identity = str(
                item.get("functional_requirement_id") or item.get("functional_key") or title
            )
            for flag_index, flag in enumerate(item.get("conflict_flags") or [], start=1):
                source_ids = [str(value or "") for value in (item.get("source_ai_requirement_ids") or [])]
                entries.append(_entry(
                    CAT_CONFLICT,
                    f"功能「{title[:40]}」存在合成冲突：{str(flag)[:140]}——请评审裁定以哪处为准",
                    source_id=source_ids[0] if source_ids else "",
                    signal="synthesis:conflict_flag", audience=AUDIENCE_INTERNAL,
                    blocker_level=BLOCKER_BLOCKING, module=module_for_sources(source_ids),
                    evidence={"title": title, "flag": str(flag), "source_ids": source_ids},
                    subject_key=f"{functional_identity}:conflict:{flag_index}"))

    entries.extend(_weakness_scan_entries(out_dir))

    entries.extend(_parse_audit_entries(out_dir))
    # A producer may repeat the same stable signal in multiple sidecars. Keep the first occurrence so
    # reviewers never see duplicate actions for the same versioned clarification subject.
    deduped: dict[str, dict[str, Any]] = {}
    for entry in entries:
        deduped.setdefault(str(entry["clarification_id"]), entry)
    return list(deduped.values())


def _weakness_scan_entries(out_dir: Path, functional_items: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    """WS4 弱词/可测性扫描（Cap3）：功能需求级复跑 extract_guards.vague_acceptance + 弱词词典。

    弱词命中 →「模糊」（important）；可测性不足（vague_acceptance 返回不可测验收）→「缺失」（blocking）。
    汇入既有四分类与就绪判定，不新建报告；LLM 标的 ambiguity 布尔保留原位并列展示，不合并。
    手工需求（manual_requirements.jsonl）走完全相同的下游扫描。
    """
    from extract_guards import vague_acceptance
    from requirement_schema import requirement_identity

    items: list[dict[str, Any]] = []
    if isinstance(functional_items, list):
        items.extend(item for item in functional_items if isinstance(item, dict))
    else:
        # 自加载 functional_requirements.json（避免依赖 collect_questions 的条件局部变量）
        synth_path = out_dir / "functional_requirements.json"
        if synth_path.exists():
            try:
                payload = json.loads(synth_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                payload = {}
            for item in (payload.get("items") or []) if isinstance(payload, dict) else []:
                if isinstance(item, dict):
                    items.append(item)
    try:
        from review_state import read_manual_requirements
        items.extend(read_manual_requirements(out_dir))
    except Exception:
        pass

    weak_words = _load_weak_words()
    entries: list[dict[str, Any]] = []
    for item in items:
        identity = str(requirement_identity(item) or "")
        title = str(item.get("title") or item.get("functional_key") or identity)
        section = str(item.get("source_section") or "")
        quote = str(item.get("source_quote") or "")
        surface = " ".join(str(item.get(field) or "") for field in ("title", "objective", "description")) \
            + " " + " ".join(str(value) for value in (item.get("behaviors") or []))
        lowered = surface.casefold()
        hits = sorted({word for word in weak_words if word and word.casefold() in lowered})
        if hits:
            entries.append(_entry(
                CAT_AMBIGUOUS,
                f"功能「{title[:40]}」含弱词/模糊表述（{', '.join(hits[:6])}），请细化为可测、可判定的需求",
                section=section, quote=quote, source_id=identity,
                signal="weakness:vague_word", tier=TIER_HARD, audience=AUDIENCE_INTERNAL,
                blocker_level=BLOCKER_IMPORTANT,
                evidence={"weak_words": hits[:10], "title": title[:80], "surface": surface[:200]},
                subject_key=f"{identity}:weak-word"))
        for vague_text in vague_acceptance(item):
            entries.append(_entry(
                CAT_MISSING,
                f"功能「{title[:40]}」的验收含不可测表述：{str(vague_text)[:100]}",
                section=section, quote=quote, source_id=identity,
                signal="weakness:untestable", tier=TIER_HARD, audience=AUDIENCE_INTERNAL,
                blocker_level=BLOCKER_BLOCKING,
                evidence={"vague_acceptance": str(vague_text)[:200], "title": title[:80]},
                subject_key=identity + ":untestable:" + _hash_payload({"vague": str(vague_text)})[:10]))
    return entries


def _parse_audit_entries(out_dir: Path) -> list[dict[str, Any]]:
    """解析层守恒审计信号（2026-07-08 审计 H5）：此前 READY 门的全部信号源都产自
    过滤器**之后**——噪声误标/区域误标把内容静默丢掉时，零疑问信号、照判 READY。
    这里从 blocks.jsonl 直接算收支，异常比例升为内部核对必答项。"""
    blocks_path = out_dir / "blocks.jsonl"
    if not blocks_path.exists():
        return []
    blocks = read_jsonl(blocks_path)
    if not blocks:
        return []
    entries: list[dict[str, Any]] = []
    total_chars = sum(len(str(b.get("text") or "")) for b in blocks)
    noise_chars = sum(len(str(b.get("text") or "")) for b in blocks if b.get("noise"))
    body_blocks = sum(1 for b in blocks if str(b.get("doc_region") or "body") == "body")
    if total_chars and noise_chars / total_chars > 0.2:
        entries.append(_entry(
            CAT_CONFLICT,
            f"解析层将 {noise_chars / total_chars:.0%} 的文字标为页眉页脚噪声（正常约 5-10%），"
            "请抽查 blocks.jsonl 中 noise=true 的块是否误伤正文",
            signal="parse_audit:noise_char_ratio", audience=AUDIENCE_INTERNAL))
    if len(blocks) >= 20 and body_blocks / len(blocks) < 0.5:
        entries.append(_entry(
            CAT_CONFLICT,
            f"仅 {body_blocks / len(blocks):.0%} 的块被判为正文区（其余为前言/目录等），"
            "文档前半可能被 Scope 标题误判整体标成前言——请核对 doc_region 分布",
            signal="parse_audit:body_ratio", audience=AUDIENCE_INTERNAL))
    return entries


def readiness_verdict(
    out_dir: Path,
    questions: int = 0,
    *,
    unresolved_blocking: int | None = None,
    unresolved_important: int | None = None,
    unresolved_internal: int = 0,
    resolved_internal: int = 0,
    resolved: int = 0,
) -> dict[str, Any]:
    """Evaluate extraction failures, coverage, blockers, and ordinary issues independently.

    ``questions`` remains the backward-compatible ordinary-issue input for callers that have not
    adopted blocker levels. New callers pass the two explicit unresolved counts.
    """
    reasons: list[str] = []
    quality: dict[str, Any] = {}
    qp = out_dir / "ai_extract_quality.json"
    if qp.exists():
        try:
            quality = json.loads(qp.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            quality = {}
        if not isinstance(quality, dict):
            raise ValueError("ai_extract_quality.json must contain a JSON object")
    blocks_path = out_dir / "blocks.jsonl"
    requirements_path = out_dir / "ai_requirements.jsonl"
    if blocks_path.exists() and requirements_path.exists():
        from ai_extract import _coverage_quality_fields, _current_non_requirement_ids
        from extract_units import body_blocks

        live_coverage = _coverage_quality_fields(
            read_jsonl(requirements_path),
            body_blocks(read_jsonl(blocks_path)),
            expert_excluded_block_ids=_current_non_requirement_ids(out_dir),
        )
        quality.update(live_coverage)
    failed = int(quality.get("failed_sections") or 0)
    core_coverage = quality.get("core_coverage_pct")
    if isinstance(core_coverage, (int, float)):
        coverage = core_coverage
        coverage_basis = "core"
    else:
        coverage = quality.get("coverage_pct")
        coverage_basis = "legacy" if isinstance(coverage, (int, float)) else "unavailable"
    if failed > 0:
        reasons.append(f"抽取失败单元 {failed} 个")
    if isinstance(coverage, (int, float)) and coverage < READY_MIN_COVERAGE:
        basis_label = "（旧版未分层口径）" if coverage_basis == "legacy" else ""
        reasons.append(f"覆盖率{basis_label} {coverage}% < {READY_MIN_COVERAGE:g}%")
    if unresolved_blocking is None:
        unresolved_blocking = 0
    if unresolved_important is None:
        unresolved_important = int(questions)
    if unresolved_blocking > 0:
        reasons.append(f"未解决阻塞项 {unresolved_blocking} 条")
    if unresolved_internal > 0:
        reasons.append(f"未解决内部核对 {unresolved_internal} 条")
    if unresolved_important > READY_MAX_QUESTIONS:
        reasons.append(f"普通待澄清 {unresolved_important} 条 > {READY_MAX_QUESTIONS}")
    unresolved_total = int(unresolved_blocking) + int(unresolved_important)
    return {"verdict": "NEEDS WORK" if reasons else "READY",
            "reasons": reasons, "questions": unresolved_total,
            "unresolved_blocking": int(unresolved_blocking),
            "unresolved_important": int(unresolved_important),
            "unresolved_internal": int(unresolved_internal),
            "resolved_internal": int(resolved_internal),
            "unresolved": unresolved_total,
            "resolved": int(resolved),
            "ordinary_problem_limit": READY_MAX_QUESTIONS,
            "coverage_pct": coverage,
            "coverage_basis": coverage_basis,
            "legacy_coverage": coverage_basis == "legacy",
            "failed_sections": failed}


def render_markdown(entries: list[dict[str, Any]], readiness: dict[str, Any]) -> str:
    hard = [e for e in entries if e.get("tier", TIER_HARD) == TIER_HARD]
    soft = [e for e in entries if e.get("tier") == TIER_SOFT]
    gap = [e for e in entries if e.get("tier") == TIER_GAP]
    coverage_basis_label = {
        "core": "core 分层口径",
        "legacy": "legacy 旧版未分层口径",
        "unavailable": "不可用",
    }.get(str(readiness.get("coverage_basis") or ""), "未标注")
    lines = ["# 需求澄清问题清单", "",
             f"**就绪判定：{readiness['verdict']}**"
             + (f"（{'；'.join(readiness['reasons'])}）" if readiness["reasons"] else ""),
             f"覆盖率口径：{coverage_basis_label}",
             f"未解决必答 {readiness.get('questions', len(hard))} 条"
             f"（阻塞 {readiness.get('unresolved_blocking', 0)} / 普通 {readiness.get('unresolved_important', len(hard))}）"
             f" · 已确认内部核对 {readiness.get('resolved_internal', 0)} 条"
             f" · 参考 {len(soft)} 条（模型自报的假设/开放问题，留档备查）"
             + (f" · 遗漏候选 {len(gap)} 条（覆盖缺口机器检出，请核对）" if gap else ""), ""]

    def emit_items(group_entries: list[dict[str, Any]]) -> None:
        order = [CAT_CONFLICT, CAT_MISSING, CAT_AMBIGUOUS, CAT_ASSUMPTION]
        for cat in order:
            group = [e for e in group_entries if e["category"] == cat]
            if not group:
                continue
            lines.append(f"#### {cat}（{len(group)}）")
            for i, e in enumerate(group, 1):
                lines.append(f"{i}. {e['question']}")
                meta = [x for x in (f"§{e['section']}" if e["section"] else "",
                                    f"来源 {e['source_id'][:12]}" if e["source_id"] else "",
                                    e["signal"], e.get("clarification_id", "")) if x]
                if e.get("audience") == AUDIENCE_INTERNAL and e.get("check_action"):
                    labels = {"verified_ok": "确认无误", "issue_confirmed": "确认有问题", "deferred": "暂缓"}
                    current = "当前" if e.get("check_state_current") else "证据已变化，需复核"
                    meta.append(f"{labels.get(e['check_action'], e['check_action'])}·{current}")
                if e["quote"]:
                    lines.append(f"   > {e['quote']}")
                if meta:
                    lines.append(f"   *{' · '.join(meta)}*")
            lines.append("")

    for blocker, blocker_label in (
        (BLOCKER_BLOCKING, "阻塞级"),
        (BLOCKER_IMPORTANT, "重要级"),
    ):
        level_entries = [e for e in hard if e.get("blocker_level") == blocker]
        if not level_entries:
            continue
        lines.append(f"# {blocker_label}（{len(level_entries)}）")
        for audience, audience_label in (
            (AUDIENCE_CUSTOMER, "必答·问客户"),
            (AUDIENCE_INTERNAL, "必答·内部核对"),
        ):
            audience_entries = [e for e in level_entries if e.get("audience") == audience]
            if not audience_entries:
                continue
            lines.append(f"## {audience_label}（{len(audience_entries)}）")
            modules = sorted({str(e.get("module") or MODULE_UNASSIGNED) for e in audience_entries})
            for module in modules:
                module_entries = [e for e in audience_entries
                                  if str(e.get("module") or MODULE_UNASSIGNED) == module]
                lines.append(f"### 模块：{module}（{len(module_entries)}）")
                emit_items(module_entries)

    def emit_aux(group_entries: list[dict[str, Any]], heading: str) -> None:
        lines.append(f"# {heading}")
        modules = sorted({str(e.get("module") or MODULE_UNASSIGNED) for e in group_entries})
        for module in modules:
            module_entries = [e for e in group_entries
                              if str(e.get("module") or MODULE_UNASSIGNED) == module]
            lines.append(f"### 模块：{module}（{len(module_entries)}）")
            emit_items(module_entries)

    if gap:
        emit_aux(gap, f"遗漏候选（{len(gap)}）——覆盖缺口机器检出，核对是漏抽还是非需求文本")
    if soft:
        emit_aux(soft, f"参考（{len(soft)}）——模型自报，抽查即可")
    if not entries:
        lines.append("（无待澄清问题）")
    return "\n".join(lines) + "\n"


def write_xlsx(entries: list[dict[str, Any]], readiness: dict[str, Any], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    blocker_order = {BLOCKER_BLOCKING: 0, BLOCKER_IMPORTANT: 1}

    def sort_key(entry: dict[str, Any]) -> tuple[Any, ...]:
        return (
            blocker_order.get(str(entry.get("blocker_level") or ""), 9),
            str(entry.get("module") or MODULE_UNASSIGNED),
            str(entry.get("category") or ""),
            str(entry.get("clarification_id") or ""),
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "必答-问客户"
    customer_header = [
        "序号", "分类", "问题", "出处章节", "原文引用", "来源需求", "信号", "答复", "采纳(是/否)",
        "澄清ID", "阻塞级", "模块", "证据指纹",
    ]
    ws.append(customer_header)
    hard = [e for e in entries if e.get("tier", TIER_HARD) == TIER_HARD]
    soft = [e for e in entries if e.get("tier") == TIER_SOFT]
    gap = [e for e in entries if e.get("tier") == TIER_GAP]
    customer = sorted([e for e in hard if e.get("audience") != AUDIENCE_INTERNAL], key=sort_key)
    internal = sorted([e for e in hard if e.get("audience") == AUDIENCE_INTERNAL], key=sort_key)
    for i, e in enumerate(customer, 1):
        ws.append([i, e["category"], formula_safe(e["question"]), formula_safe(e["section"]),
                   formula_safe(e["quote"]), formula_safe(e["source_id"]), formula_safe(e["signal"]),
                   "", "", formula_safe(e["clarification_id"]), e["blocker_level"],
                   formula_safe(e["module"]), formula_safe(e["evidence_fingerprint"])])

    ws_int = wb.create_sheet("必答-内部核对")
    internal_header = [
        "序号", "分类", "问题", "出处章节", "原文引用", "来源需求", "信号",
        "澄清ID", "阻塞级", "模块", "证据指纹",
        "当前处置", "当前核对人", "当前时间", "当前备注",
        "新处置(确认无误/确认有问题/暂缓)", "核对人", "备注",
    ]
    ws_int.append(internal_header)
    action_labels = {"verified_ok": "确认无误", "issue_confirmed": "确认有问题", "deferred": "暂缓"}
    for i, e in enumerate(internal, 1):
        ws_int.append([i, e["category"], formula_safe(e["question"]), formula_safe(e["section"]),
                       formula_safe(e["quote"]), formula_safe(e["source_id"]), formula_safe(e["signal"]),
                       formula_safe(e["clarification_id"]), e["blocker_level"], formula_safe(e["module"]),
                       formula_safe(e["evidence_fingerprint"]),
                       action_labels.get(str(e.get("check_action") or ""), ""),
                       formula_safe(str(e.get("check_actor") or "")), str(e.get("check_timestamp") or ""),
                       formula_safe(str(e.get("check_note") or "")), "", "", ""])
    action_validation = DataValidation(
        type="list", formula1='"确认无误,确认有问题,暂缓"', allow_blank=True
    )
    ws_int.add_data_validation(action_validation)
    if internal:
        action_validation.add(f"P2:P{len(internal) + 1}")

    aux_header = ["序号", "分类", "问题", "出处章节", "原文引用", "来源需求", "信号",
                  "澄清ID", "阻塞级", "模块", "证据指纹"]
    ws_gap = wb.create_sheet("遗漏候选(内部核对)")
    ws_gap.append(aux_header)
    for i, e in enumerate(sorted(gap, key=sort_key), 1):
        ws_gap.append([i, e["category"], formula_safe(e["question"]), formula_safe(e["section"]),
                       formula_safe(e["quote"]), formula_safe(e["source_id"]), formula_safe(e["signal"]),
                       formula_safe(e["clarification_id"]), e["blocker_level"], formula_safe(e["module"]),
                       formula_safe(e["evidence_fingerprint"])])
    ws_soft = wb.create_sheet("参考(模型自报)")
    ws_soft.append(aux_header)
    for i, e in enumerate(sorted(soft, key=sort_key), 1):
        ws_soft.append([i, e["category"], formula_safe(e["question"]), formula_safe(e["section"]),
                        formula_safe(e["quote"]), formula_safe(e["source_id"]), formula_safe(e["signal"]),
                        formula_safe(e["clarification_id"]), e["blocker_level"], formula_safe(e["module"]),
                        formula_safe(e["evidence_fingerprint"])])
    ws2 = wb.create_sheet("就绪判定")
    ws2.append(["判定", readiness["verdict"]])
    for r in readiness["reasons"]:
        ws2.append(["原因", formula_safe(r)])
    ws2.append(["未解决待澄清", readiness["questions"]])
    ws2.append(["未解决阻塞项", readiness.get("unresolved_blocking")])
    ws2.append(["未解决普通问题", readiness.get("unresolved_important")])
    ws2.append(["未解决内部核对", readiness.get("unresolved_internal")])
    ws2.append(["已确认内部核对", readiness.get("resolved_internal")])
    ws2.append(["覆盖率%", readiness.get("coverage_pct")])
    ws2.append(["失败单元", readiness.get("failed_sections")])
    from xlsx_io import safe_save_workbook
    safe_save_workbook(wb, path)


def load_answers(out_dir: Path) -> dict[tuple[str, str], dict[str, Any]]:
    """已回灌答复：(来源需求id, 问题) → 答复条目。容错读，坏行跳过。"""
    root = Path(out_dir).expanduser().resolve()
    with clarification_answers_lock(root):
        return _load_answers_unlocked(
            governed_artifact_path(root, ANSWERS_FILE, category="state")
        )


def _load_answers_unlocked(path: Path) -> dict[tuple[str, str], dict[str, Any]]:
    answers: dict[tuple[str, str], dict[str, Any]] = {}
    if not path.exists():
        return answers
    for line in path.read_text(encoding="utf-8").splitlines():
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        key = (str(row.get("source_id") or ""), str(row.get("question") or ""))
        answers[key] = row
    return answers


@contextmanager
def clarification_answers_lock(
    out_dir: Path,
    *,
    timeout_s: float = _ANSWER_LOCK_TIMEOUT_S,
    stale_after_s: float = _ANSWER_LOCK_STALE_AFTER_S,
) -> Iterator[None]:
    """Serialize customer-answer readers and writers across processes."""
    root = Path(out_dir).expanduser().resolve()
    lock_path = governed_artifact_path(root, ANSWERS_LOCK, category="state")
    with _answer_process_lock_for(lock_path.parent):
        deadline = time.monotonic() + timeout_s
        fd: int | None = None
        while fd is None:
            try:
                fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError:
                if _remove_stale_answer_lock(lock_path, stale_after_s):
                    continue
                if time.monotonic() >= deadline:
                    raise TimeoutError(f"timed out waiting for clarification answer lock: {lock_path}")
                time.sleep(0.01)
        try:
            os.write(fd, str(os.getpid()).encode("ascii"))
            yield
        finally:
            os.close(fd)
            try:
                lock_path.unlink()
            except FileNotFoundError:
                pass


def _answer_process_lock_for(out_dir: Path) -> RLock:
    with _ANSWER_LOCKS_GUARD:
        return _ANSWER_LOCKS.setdefault(out_dir, RLock())


def _remove_stale_answer_lock(lock_path: Path, stale_after_s: float) -> bool:
    if stale_after_s < 0:
        return False
    try:
        age_s = time.time() - lock_path.stat().st_mtime
    except FileNotFoundError:
        return True
    if age_s < stale_after_s:
        return False
    try:
        lock_path.unlink()
    except FileNotFoundError:
        return True
    return True


def _atomic_write_answers(path: Path, rows: list[dict[str, Any]]) -> None:
    tmp_path = path.with_name(f"{path.name}.{os.getpid()}.tmp")
    try:
        with tmp_path.open("w", encoding="utf-8", newline="\n") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")
            handle.flush()
            os.fsync(handle.fileno())
        for attempt in range(_ANSWER_REPLACE_ATTEMPTS):
            try:
                os.replace(tmp_path, path)
                break
            except PermissionError:
                if attempt + 1 >= _ANSWER_REPLACE_ATTEMPTS:
                    raise
                time.sleep(_ANSWER_REPLACE_RETRY_DELAY_S)
    finally:
        try:
            tmp_path.unlink()
        except FileNotFoundError:
            pass


def load_current_answers(out_dir: Path) -> dict[tuple[str, str], dict[str, Any]]:
    """Return adopted customer answers whose evidence still matches the current source.

    Customer answers become authoritative analysis input, so a stable clarification id alone is
    insufficient: an answer captured for an older quote must remain in the audit trail without
    being injected into a later document generation.
    """
    answers = load_answers(out_dir)
    if not answers:
        return {}
    current_entries = {
        str(entry.get("clarification_id") or ""): entry
        for entry in collect_questions(Path(out_dir))
        if entry.get("audience") != AUDIENCE_INTERNAL
        and str(entry.get("clarification_id") or "")
    }
    current: dict[tuple[str, str], dict[str, Any]] = {}
    for key, answer in answers.items():
        clarification_id = str(answer.get("clarification_id") or "")
        entry = current_entries.get(clarification_id)
        if not entry or not answer.get("adopted", True):
            continue
        if str(answer.get("evidence_fingerprint") or "") != str(
            entry.get("evidence_fingerprint") or ""
        ):
            continue
        current[key] = answer
    return current


def import_answers(out_dir: Path, xlsx_path: Path) -> dict[str, Any]:
    """从填好的 clarification_questions.xlsx「必答」sheet 读回答复列 → 落 ANSWERS_FILE。

    闭环的另一半：评审会带走清单，答复填回同一文件，导入后 analyze 把答复当权威客户输入
    （注入富化 prompt + 数值有据基线），澄清报告把已采纳答复的问题消解出必答区。
    """
    from openpyxl import load_workbook
    out_dir = Path(out_dir).expanduser().resolve()
    wb = load_workbook(Path(xlsx_path).expanduser(), data_only=True, read_only=True)
    try:
        sheet_name = next((n for n in ("必答-问客户", "必答") if n in wb.sheetnames), None)
        if sheet_name is None:
            raise ValueError("工作簿缺少「必答-问客户」sheet——请用本工具导出的澄清清单填写答复")
        ws = wb[sheet_name]
        headers = [str(cell.value or "").strip() for cell in next(
            ws.iter_rows(min_row=1, max_row=1)
        )]
        header_index = {name: index for index, name in enumerate(headers)}
        imported_entries: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 8:
                continue
            answer = str(row[7] or "").strip()
            if not answer:
                continue
            entry = {
                "source_id": str(row[5] or "").strip(),
                "category": str(row[1] or "").strip(),
                "question": str(row[2] or "").strip(),
                "answer": answer,
                "adopted": str(row[8] if len(row) > 8 else "").strip() not in ("否", "no", "N", "n"),
                "clarification_id": str(row[9] if len(row) > 9 else "").strip(),
                "evidence_fingerprint": str(
                    row[header_index["证据指纹"]]
                    if "证据指纹" in header_index and len(row) > header_index["证据指纹"]
                    else ""
                ).strip(),
            }
            imported_entries.append(entry)
    finally:
        wb.close()
    with clarification_answers_lock(out_dir):
        answers_path = governed_artifact_path(out_dir, ANSWERS_FILE, category="state")
        merged = _load_answers_unlocked(answers_path)
        for entry in imported_entries:
            merged[(entry["source_id"], entry["question"])] = entry
        _atomic_write_answers(answers_path, list(merged.values()))
    return {"imported": len(imported_entries), "total_answers": len(merged),
            "written": [ANSWERS_FILE]}


def import_internal_checks(
    out_dir: Path,
    xlsx_path: Path,
    *,
    actor: str | None = None,
) -> dict[str, Any]:
    """Import explicit acknowledge actions from the internal-check worksheet.

    Unlike customer clarifications, these rows do not require an answer. The reviewer records one
    of three audited dispositions: verified_ok, issue_confirmed, or deferred.
    """
    from openpyxl import load_workbook

    aliases = {
        "确认无误": "verified_ok",
        "verified_ok": "verified_ok",
        "确认有问题": "issue_confirmed",
        "issue_confirmed": "issue_confirmed",
        "暂缓": "deferred",
        "deferred": "deferred",
    }
    out_dir = Path(out_dir).expanduser().resolve()
    wb = load_workbook(Path(xlsx_path).expanduser(), data_only=True, read_only=True)
    try:
        if "必答-内部核对" not in wb.sheetnames:
            raise ValueError("工作簿缺少「必答-内部核对」sheet")
        ws = wb["必答-内部核对"]
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        def column(name: str, *, prefix: bool = False) -> int:
            for index, value in enumerate(headers):
                if value == name or (prefix and value.startswith(name)):
                    return index
            raise ValueError(f"「必答-内部核对」sheet 缺少“{name}”列，请重新生成澄清清单")

        id_col = column("澄清ID")
        evidence_col = column("证据指纹")
        blocker_col = column("阻塞级")
        module_col = column("模块")
        signal_col = column("信号")
        source_col = column("来源需求")
        action_col = column("新处置", prefix=True)
        actor_col = column("核对人")
        note_col = column("备注")
        submitted: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_action = str(row[action_col] if len(row) > action_col else "").strip()
            if not raw_action:
                continue
            action = aliases.get(raw_action)
            if action not in VALID_CHECK_ACTIONS:
                raise ValueError(f"无效内部核对处置：{raw_action}")
            clarification_id = str(row[id_col] if len(row) > id_col else "").strip()
            evidence_fingerprint = str(row[evidence_col] if len(row) > evidence_col else "").strip()
            row_actor = str(row[actor_col] if len(row) > actor_col else "").strip()
            note = str(row[note_col] if len(row) > note_col else "").strip()
            submitted.append({
                "clarification_id": clarification_id,
                "action": action,
                "evidence_fingerprint": evidence_fingerprint,
                # Workbook metadata is retained only for diagnostics. The stored event takes
                # blocker/module/signal/source from the current report generation.
                "workbook_blocker_level": str(
                    row[blocker_col] if len(row) > blocker_col else ""
                ).strip(),
                "workbook_module": str(row[module_col] if len(row) > module_col else "").strip(),
                "workbook_signal": str(row[signal_col] if len(row) > signal_col else "").strip(),
                "workbook_source_id": str(row[source_col] if len(row) > source_col else "").strip(),
                "actor": row_actor or actor,
                "note": note,
            })
    finally:
        wb.close()

    from omission_actions import extraction_operation_lock

    with extraction_operation_lock(out_dir, operation="clarification-check-import"):
        accepted, missing, stale, ineligible, duplicate = _prepare_internal_check_rows(
            out_dir, submitted
        )
        events = apply_clarification_check_actions_batch(out_dir, accepted)
    return {
        "imported": len(events),
        "requested": len(submitted),
        "stale": stale,
        "missing": missing,
        "ineligible": ineligible,
        "duplicates": duplicate,
        "total_states": len(read_clarification_check_states(out_dir)),
        "written": ["clarification_check_states.jsonl"] if events else [],
    }


def batch_apply_internal_checks(
    out_dir: Path,
    checks: list[dict[str, Any]],
    *,
    action: str = "verified_ok",
    actor: str | None = None,
    note: str = "",
) -> dict[str, Any]:
    """Apply current internal checks in one audited batch.

    Each submitted row must match the current clarification id and evidence fingerprint. Stale,
    missing, duplicate, or customer-facing rows are reported and never written.
    """
    if action not in VALID_CHECK_ACTIONS:
        raise ValueError(f"invalid clarification check action: {action}")
    root = Path(out_dir).expanduser().resolve()
    from omission_actions import extraction_operation_lock

    # Keep the evidence generation fixed between validation and the single atomic state write.
    # Without this lease a concurrent full/targeted extraction could turn an accepted row stale
    # after validation but before persistence.
    with extraction_operation_lock(root, operation="clarification-check-batch"):
        accepted, missing, stale, ineligible, duplicate = _prepare_internal_check_batch(
            root,
            checks,
            action=action,
            actor=actor,
            note=note,
        )
        events = apply_clarification_check_actions_batch(root, accepted)

    def grouped(field: str) -> dict[str, int]:
        counts: dict[str, int] = {}
        for event in events:
            key = str(event.get(field) or "未分类")
            counts[key] = counts.get(key, 0) + 1
        return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))

    report = run_report(root) if events else None
    return {
        "requested": len(checks),
        "applied": len(events),
        "stale": stale,
        "missing": missing,
        "ineligible": ineligible,
        "duplicates": duplicate,
        "by_signal": grouped("signal"),
        "by_module": grouped("module"),
        "readiness": (report or {}).get("readiness"),
        "written": ["clarification_check_states.jsonl"] if events else [],
    }


def _prepare_internal_check_batch(
    root: Path,
    checks: list[dict[str, Any]],
    *,
    action: str,
    actor: str | None,
    note: str,
) -> tuple[list[dict[str, Any]], list[str], list[str], list[str], list[str]]:
    submitted = [
        {
            **row,
            "action": action,
            "actor": actor,
            "note": note,
        }
        for row in checks
    ]
    return _prepare_internal_check_rows(root, submitted)


def _prepare_internal_check_rows(
    root: Path,
    checks: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str], list[str], list[str], list[str]]:
    current_entries = collect_questions(root)
    all_current = {
        str(entry.get("clarification_id") or ""): entry
        for entry in current_entries
    }
    current = {
        str(entry.get("clarification_id") or ""): entry
        for entry in current_entries
        if entry.get("tier", TIER_HARD) == TIER_HARD
        and entry.get("audience") == AUDIENCE_INTERNAL
    }
    accepted: list[dict[str, Any]] = []
    missing: list[str] = []
    stale: list[str] = []
    ineligible: list[str] = []
    duplicate: list[str] = []
    seen: set[str] = set()
    for row in checks:
        clarification_id = str(row.get("clarification_id") or "").strip()
        if not clarification_id:
            ineligible.append(clarification_id)
            continue
        if clarification_id in seen:
            duplicate.append(clarification_id)
            continue
        seen.add(clarification_id)
        entry = all_current.get(clarification_id)
        if entry is None:
            missing.append(clarification_id)
            continue
        if clarification_id not in current:
            ineligible.append(clarification_id)
            continue
        action = str(row.get("action") or "").strip()
        if action not in VALID_CHECK_ACTIONS:
            ineligible.append(clarification_id)
            continue
        submitted_fingerprint = str(row.get("evidence_fingerprint") or "").strip()
        if not submitted_fingerprint or submitted_fingerprint != str(
            entry.get("evidence_fingerprint") or ""
        ):
            stale.append(clarification_id)
            continue
        accepted.append({
            "clarification_id": clarification_id,
            "action": action,
            "evidence_fingerprint": submitted_fingerprint,
            "blocker_level": entry.get("blocker_level") or BLOCKER_IMPORTANT,
            "module": entry.get("module") or MODULE_UNASSIGNED,
            "signal": entry.get("signal") or "",
            "source_id": entry.get("source_id") or "",
            "actor": row.get("actor"),
            "note": str(row.get("note") or ""),
        })
    return accepted, missing, stale, ineligible, duplicate


def current_internal_checks(out_dir: Path) -> dict[str, Any]:
    """Return current hard internal checks with effective, fingerprint-validated states."""
    root = Path(out_dir).expanduser().resolve()
    entries = [
        entry for entry in collect_questions(root)
        if entry.get("tier", TIER_HARD) == TIER_HARD
        and entry.get("audience") == AUDIENCE_INTERNAL
    ]
    _attach_internal_check_states(entries, root)
    unresolved = [entry for entry in entries if not _internal_check_resolved(entry)]
    groups: dict[str, dict[str, Any]] = {}
    for entry in unresolved:
        signal = str(entry.get("signal") or "未分类")
        group = groups.setdefault(signal, {
            "signal": signal,
            "count": 0,
            "blocking": 0,
            "modules": {},
        })
        group["count"] += 1
        if entry.get("blocker_level") == BLOCKER_BLOCKING:
            group["blocking"] += 1
        module = str(entry.get("module") or MODULE_UNASSIGNED)
        group["modules"][module] = group["modules"].get(module, 0) + 1
    return {
        "schema": "clarification-internal-checks/v1",
        "total": len(entries),
        "unresolved": len(unresolved),
        "entries": unresolved,
        "groups": sorted(groups.values(), key=lambda row: (-row["blocking"], -row["count"], row["signal"])),
    }


def _attach_internal_check_states(entries: list[dict[str, Any]], out_dir: Path) -> None:
    states = read_clarification_check_states(out_dir)
    for entry in entries:
        if (entry.get("tier", TIER_HARD) != TIER_HARD
                or entry.get("audience") != AUDIENCE_INTERNAL):
            continue
        state = states.get(str(entry.get("clarification_id") or ""))
        entry["check_action"] = ""
        entry["check_state_current"] = None
        entry["check_actor"] = ""
        entry["check_timestamp"] = ""
        entry["check_note"] = ""
        entry["state"] = ""
        entry["actor"] = ""
        entry["timestamp"] = ""
        entry["note"] = ""
        if not state:
            continue
        state_fingerprint = str(state.get("evidence_fingerprint") or "")
        current = bool(state_fingerprint) and state_fingerprint == str(entry.get("evidence_fingerprint") or "")
        effective_state = str(state.get("state") or state.get("action") or "")
        entry["check_action"] = effective_state
        entry["check_state_current"] = current
        entry["check_actor"] = str(state.get("actor") or "")
        entry["check_timestamp"] = str(state.get("timestamp") or "")
        entry["check_note"] = str(state.get("note") or "")
        entry["state"] = effective_state
        entry["actor"] = entry["check_actor"]
        entry["timestamp"] = entry["check_timestamp"]
        entry["note"] = entry["check_note"]


def _internal_check_resolved(entry: dict[str, Any]) -> bool:
    return (
        entry.get("audience") == AUDIENCE_INTERNAL
        and entry.get("check_action") == "verified_ok"
        and entry.get("check_state_current") is True
    )


def _organize_entries(entries: list[dict[str, Any]]) -> dict[str, Any]:
    """Nest entries by blocker level, audience, and effective module for JSON consumers."""
    organized: dict[str, Any] = {BLOCKER_BLOCKING: {}, BLOCKER_IMPORTANT: {}}
    for entry in entries:
        blocker = str(entry.get("blocker_level") or BLOCKER_IMPORTANT)
        audience = str(entry.get("audience") or AUDIENCE_CUSTOMER)
        module = str(entry.get("module") or MODULE_UNASSIGNED)
        organized.setdefault(blocker, {}).setdefault(audience, {}).setdefault(module, []).append(entry)
    return organized


def _find_answer(entry: dict[str, Any], answers: dict, answers_by_id: dict) -> dict | None:
    hit = answers_by_id.get(str(entry.get("clarification_id") or ""))
    if hit is None:
        hit = answers.get((entry.get("source_id") or "", entry.get("question") or ""))
    return hit


def _answer_is_current(entry: dict[str, Any], answer: dict | None) -> bool:
    return bool(
        answer
        and str(answer.get("evidence_fingerprint") or "")
        == str(entry.get("evidence_fingerprint") or "")
    )


def _customer_answer_resolved(entry: dict[str, Any], answers: dict, answers_by_id: dict) -> bool:
    """客户问题已消解的唯一判定：有答复 + 采纳 + 证据指纹当前（agent_state 与报告共用）。"""
    if entry.get("audience") == AUDIENCE_INTERNAL:
        return False
    hit = _find_answer(entry, answers, answers_by_id)
    return bool(hit and hit.get("adopted", True) and _answer_is_current(entry, hit))


def unresolved_hard_questions(out_dir: Path) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """「必答未解决」判定的唯一实现（Phase 1.5 口径收敛：agent_state 不再自维护一份）。

    返回 (未解决必答条目, 计数)。计数键：blocking / important / internal /
    resolved_internal / resolved。"""
    out_dir = Path(out_dir).expanduser().resolve()
    entries = [
        e for e in collect_questions(out_dir)
        if e.get("tier", TIER_HARD) == TIER_HARD
    ]
    _attach_internal_check_states(entries, out_dir)
    answers = load_answers(out_dir)
    answers_by_id = {
        str(row.get("clarification_id") or ""): row
        for row in answers.values()
        if str(row.get("clarification_id") or "")
    }
    unresolved: list[dict[str, Any]] = []
    resolved = 0
    resolved_internal = 0
    for e in entries:
        if e.get("audience") == AUDIENCE_INTERNAL:
            if _internal_check_resolved(e):
                resolved_internal += 1
                continue
            unresolved.append(e)
            continue
        if _customer_answer_resolved(e, answers, answers_by_id):
            resolved += 1
            continue
        unresolved.append(e)
    blocking = sum(1 for e in unresolved if e.get("blocker_level") == BLOCKER_BLOCKING)
    internal = sum(1 for e in unresolved if e.get("audience") == AUDIENCE_INTERNAL)
    return unresolved, {
        "blocking": blocking,
        "important": len(unresolved) - blocking,
        "internal": internal,
        "resolved_internal": resolved_internal,
        "resolved": resolved + resolved_internal,
    }


def run_report(out_dir: Path) -> dict[str, Any]:
    out_dir = Path(out_dir).expanduser().resolve()
    if not (out_dir / "ai_requirements.jsonl").exists():
        # 缺输入响亮失败（仓库纪律：不产"0 问题"的假清单掩盖打错目录）
        raise FileNotFoundError(
            f"ai_requirements.jsonl not found in {out_dir} — 先跑「AI 抽取」再生成澄清清单")
    entries = collect_questions(out_dir)
    _attach_internal_check_states(entries, out_dir)
    answers = load_answers(out_dir)
    answers_by_id = {
        str(row.get("clarification_id") or ""): row
        for row in answers.values()
        if str(row.get("clarification_id") or "")
    }
    resolved = 0
    if answers:
        kept: list[dict[str, Any]] = []
        for e in entries:
            if _customer_answer_resolved(e, answers, answers_by_id):
                resolved += 1            # 已答复采纳 → 消解，不再出现在清单
                continue
            hit = _find_answer(e, answers, answers_by_id)
            if hit and e.get("audience") != AUDIENCE_INTERNAL:
                e["answer_state_current"] = _answer_is_current(e, hit)
            kept.append(e)
        entries = kept
    hard_entries = [e for e in entries if e.get("tier", TIER_HARD) == TIER_HARD]
    resolved_internal = [e for e in hard_entries if _internal_check_resolved(e)]
    unresolved_hard = [e for e in hard_entries if not _internal_check_resolved(e)]
    unresolved_blocking = sum(
        1 for e in unresolved_hard if e.get("blocker_level") == BLOCKER_BLOCKING
    )
    unresolved_important = sum(
        1 for e in unresolved_hard if e.get("blocker_level") != BLOCKER_BLOCKING
    )
    unresolved_internal = sum(
        1 for e in unresolved_hard if e.get("audience") == AUDIENCE_INTERNAL
    )
    hard_count = len(unresolved_hard)
    readiness = readiness_verdict(
        out_dir,
        hard_count,
        unresolved_blocking=unresolved_blocking,
        unresolved_important=unresolved_important,
        unresolved_internal=unresolved_internal,
        resolved_internal=len(resolved_internal),
        resolved=resolved + len(resolved_internal),
    )
    (out_dir / REPORT_MD).write_text(render_markdown(entries, readiness), encoding="utf-8")
    write_xlsx(entries, readiness, out_dir / REPORT_XLSX)
    by_cat: dict[str, int] = {}
    for e in entries:
        by_cat[e["category"]] = by_cat.get(e["category"], 0) + 1
    from requirement_record import check_provenance, provenance
    warnings: list[str] = []
    ana_path = out_dir / "engineering_analysis.json"
    if ana_path.exists():
        try:
            from requirements_analysis import ANALYZE_PROMPT_VERSION
            warn = check_provenance(json.loads(ana_path.read_text(encoding="utf-8")),
                                    expect_producer="requirements_analysis",
                                    current_version=ANALYZE_PROMPT_VERSION)
            if warn:
                warnings.append(warn)
        except Exception:  # 血统校验永不阻断出报告
            pass
    claim_ledger_summary: dict[str, Any]
    try:
        from claim_views import build_claim_clarification_views

        claim_views = build_claim_clarification_views(out_dir, uncertain_limit=50)
        claim_metrics = claim_views["metrics"]
        uncertain_claims = claim_views["uncertain_catalog"]
        effective_metrics = dict(claim_metrics.get("effective_metrics") or {})
        claim_ledger_summary = {
            "available": bool(claim_metrics.get("available")),
            "phase": claim_metrics.get("phase"),
            "document_effective_revision": claim_metrics.get(
                "document_effective_revision"
            ),
            "document_ready": claim_metrics.get("document_ready"),
            "effective_fresh": bool(claim_metrics.get("effective_fresh")),
            "freshness_reasons": list(
                claim_metrics.get("freshness_reasons") or []
            ),
            "metrics": {
                key: effective_metrics.get(key)
                for key in (
                    "verified_coverage_ratio",
                    "verified_semantic_exclusion_ratio",
                    "verified_exclusion_ratio",
                    "structural_exclusion_ratio",
                )
            },
            "open_claim_count": int(
                effective_metrics.get("uncertain_count") or 0
            ),
            "uncertain_claims": [
                {
                    "claim_id": row.get("claim_id"),
                    "locator": row.get("locator"),
                    "text": str(row.get("text") or "")[:120],
                }
                for row in uncertain_claims.get("rows") or []
            ],
        }
        if not claim_ledger_summary["available"]:
            claim_ledger_summary["reason"] = (
                claim_metrics.get("reason") or "claim ledger unavailable"
            )
    except Exception as exc:
        claim_ledger_summary = {
            "available": False,
            "effective_fresh": False,
            "error": f"{type(exc).__name__}: {exc}",
            "metrics": {},
            "open_claim_count": None,
            "uncertain_claims": [],
        }
    report = {"questions": hard_count,        # 未解决必答数（就绪门口径；GUI 消息同源）
              "provenance": provenance("clarification_report", CLARIFICATION_REPORT_VERSION),
              "upstream_warnings": warnings,
              "questions_total": len(entries),
              "soft_questions": sum(1 for e in entries if e.get("tier") == TIER_SOFT),
              "coverage_candidates": sum(1 for e in entries if e.get("tier") == TIER_GAP),
              "compliance_gaps": sum(
                  1 for e in entries
                  if str(e.get("signal") or "").startswith("consistency:compliance_uncovered")
              ),
              "customer_questions": sum(1 for e in entries
                                         if e.get("tier", TIER_HARD) == TIER_HARD
                                         and e.get("audience") != AUDIENCE_INTERNAL),
              "internal_checks": sum(1 for e in entries
                                     if e.get("tier", TIER_HARD) == TIER_HARD
                                     and e.get("audience") == AUDIENCE_INTERNAL),
              "resolved_by_answers": resolved,
              "resolved_by_checks": len(resolved_internal),
              "unresolved_blocking": unresolved_blocking,
              "unresolved_important": unresolved_important,
              "unresolved_internal": unresolved_internal,
              "resolved_internal": len(resolved_internal),
              "unresolved": hard_count,
              "resolved": resolved + len(resolved_internal),
              "by_category": by_cat,
              "by_blocker_level": {
                  BLOCKER_BLOCKING: sum(1 for e in entries
                                        if e.get("blocker_level") == BLOCKER_BLOCKING),
                  BLOCKER_IMPORTANT: sum(1 for e in entries
                                         if e.get("blocker_level") != BLOCKER_BLOCKING),
              },
              "organized_entries": _organize_entries(entries),
              "claim_ledger": claim_ledger_summary,
              "readiness": readiness, "entries": entries,
              "written": [REPORT_MD, REPORT_XLSX, REPORT_JSON]}
    from input_completeness import attach_input_completeness

    attach_input_completeness(report, out_dir)
    (out_dir / REPORT_JSON).write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Aggregate clarification questions + readiness verdict.")
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args(argv)
    try:
        report = run_report(args.out)
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps({k: report[k] for k in ("questions", "by_category", "readiness", "written")},
                     ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
