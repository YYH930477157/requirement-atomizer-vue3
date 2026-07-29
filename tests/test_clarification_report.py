"""澄清问题清单 + 就绪判定回归（确定性零 LLM）。"""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import clarification_report as cr
import claim_review_actions
from tests.test_claim_artifacts import _catalog, _publish, _requirement


def seed(tmp: Path, *, reqs=None, analysis=None, consistency=None, quality=None) -> None:
    with (tmp / "ai_requirements.jsonl").open("w", encoding="utf-8") as f:
        for r in (reqs if reqs is not None else []):
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    if analysis is not None:
        (tmp / "engineering_analysis.json").write_text(
            json.dumps(analysis, ensure_ascii=False), encoding="utf-8")
    if consistency is not None:
        (tmp / "consistency_report.json").write_text(
            json.dumps(consistency, ensure_ascii=False), encoding="utf-8")
    if quality is not None:
        (tmp / "ai_extract_quality.json").write_text(
            json.dumps(quality, ensure_ascii=False), encoding="utf-8")


class CollectQuestionsTests(unittest.TestCase):
    def test_aggregates_all_signal_sources_with_categories(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp,
                 reqs=[{"title": "泄漏率", "source_section": "7.13", "source_quote": "q1",
                        "suspicion_reasons": ["原文数值未带全", "验收不可测"]}],
                 analysis={"items": [{"source_section": "5.3", "source_quote": "q2",
                                      "source_requirement_ids": ["AIR-1"],
                                      "open_questions": ["升级包大小上限是多少？"],
                                      "assumptions": ["假定采用双区存储"]}]},
                 consistency={"obis_coreference": [{"code": "1-0:1.8.0", "values_differ": True}],
                              "duplicate_groups": [{"source_quote": "dup quote here"}]})
            entries = cr.collect_questions(tmp)

        cats = sorted(e["category"] for e in entries)
        self.assertEqual(cats.count(cr.CAT_MISSING), 2)       # 漏值 + open_question
        self.assertEqual(cats.count(cr.CAT_AMBIGUOUS), 1)     # 验收不可测
        self.assertEqual(cats.count(cr.CAT_ASSUMPTION), 1)
        self.assertEqual(cats.count(cr.CAT_CONFLICT), 2)      # OBIS 发散 + 重复组
        assumption = next(e for e in entries if e["category"] == cr.CAT_ASSUMPTION)
        self.assertIn("双区存储", assumption["question"])
        self.assertIn("请确认", assumption["question"])

    def test_absent_sources_tolerated(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp, reqs=[{"title": "干净需求", "source_quote": "q"}])
            self.assertEqual(cr.collect_questions(tmp), [])


class ReadinessTests(unittest.TestCase):
    def test_committed_claim_summary_is_informational_and_preserves_readiness(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            catalog = _catalog()
            seed(
                out,
                reqs=[_requirement(catalog)],
                consistency={"coverage": {
                    "scope": "core",
                    "requirement_like": 1,
                    "covered": 1,
                    "uncovered_count": 0,
                    "uncovered_samples": [],
                    "excluded": {"count": 0, "block_ids": [], "samples": []},
                }},
                quality={
                    "failed_sections": 0,
                    "coverage_pct": 100.0,
                    "core_coverage_pct": 100.0,
                },
            )
            before = cr.run_report(out)
            _publish(out, catalog)
            claim_review_actions.fold_effective_ledger(
                out,
                actor_trigger="clarification-integration-test",
            )

            after = cr.run_report(out)

        self.assertEqual(after["readiness"], before["readiness"])
        self.assertEqual(after["questions"], before["questions"])
        self.assertEqual(after["soft_questions"], before["soft_questions"])
        claim_summary = after["claim_ledger"]
        self.assertTrue(claim_summary["available"])
        self.assertTrue(claim_summary["effective_fresh"])
        self.assertEqual(claim_summary["uncertain_claims"], [])
        self.assertEqual(
            claim_summary["metrics"]["verified_coverage_ratio"]["numerator"],
            1,
        )

    def test_claim_summary_derives_metrics_and_uncertain_rows_from_one_context(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            catalog = _catalog()
            seed(
                out,
                reqs=[_requirement(catalog)],
                quality={"failed_sections": 0, "coverage_pct": 100.0},
            )
            _publish(out, catalog)
            claim_review_actions.fold_effective_ledger(
                out,
                actor_trigger="clarification-single-context-seed",
            )
            import claim_views

            original_context = claim_views._context
            with patch("claim_views._context", wraps=original_context) as context:
                report = cr.run_report(out)

        self.assertEqual(context.call_count, 1)
        self.assertTrue(report["claim_ledger"]["available"])

    def test_ready_when_clean(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp, reqs=[], quality={"failed_sections": 0, "coverage_pct": 80.0})
            v = cr.readiness_verdict(tmp, questions=5)
        self.assertEqual(v["verdict"], "READY")
        self.assertEqual(v["reasons"], [])
        self.assertEqual(v["coverage_basis"], "legacy")
        self.assertTrue(v["legacy_coverage"])

    def test_needs_work_reasons_accumulate(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp, reqs=[], quality={"failed_sections": 2, "coverage_pct": 50.0})
            v = cr.readiness_verdict(tmp, questions=99)
        self.assertEqual(v["verdict"], "NEEDS WORK")
        self.assertEqual(len(v["reasons"]), 3)                 # 失败单元 + 覆盖率 + 待澄清

    def test_readiness_prefers_core_coverage(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp, reqs=[], quality={
                "failed_sections": 0,
                "coverage_pct": 10.0,
                "core_coverage_pct": 80.0,
            })
            verdict = cr.readiness_verdict(tmp)

        self.assertEqual(verdict["verdict"], "READY")
        self.assertEqual(verdict["coverage_pct"], 80.0)
        self.assertEqual(verdict["coverage_basis"], "core")
        self.assertFalse(verdict["legacy_coverage"])

    def test_structurally_invalid_quality_json_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            (out / "ai_extract_quality.json").write_text("[]", encoding="utf-8")

            with self.assertRaises(ValueError):
                cr.readiness_verdict(out)

    def test_structurally_invalid_consistency_json_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            (out / "consistency_report.json").write_text("[]", encoding="utf-8")

            with self.assertRaises(ValueError):
                cr.current_internal_checks(out)

    def test_unresolved_compliance_gap_is_a_blocker_even_when_core_is_ready(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            seed(
                out,
                reqs=[],
                consistency={"coverage": {
                    "scope": "core",
                    "requirement_like": 1,
                    "covered": 1,
                    "uncovered_count": 0,
                    "uncovered_samples": [],
                    "compliance": {
                        "requirement_like": 1,
                        "covered": 0,
                        "uncovered_count": 1,
                        "uncovered_block_ids": ["B-COMP"],
                        "uncovered_samples": [{
                            "block_id": "B-COMP",
                            "section": "2.1",
                            "text": "A valid type certificate shall be supplied.",
                        }],
                    },
                    "excluded": {"count": 0, "block_ids": [], "samples": []},
                }},
                quality={"failed_sections": 0, "core_coverage_pct": 100.0},
            )
            report = cr.run_report(out)

        entry = next(row for row in report["entries"] if row["signal"] == "consistency:compliance_uncovered")
        self.assertEqual(entry["tier"], cr.TIER_HARD)
        self.assertEqual(entry["blocker_level"], cr.BLOCKER_BLOCKING)
        self.assertEqual(report["readiness"]["verdict"], "NEEDS WORK")
        self.assertIn("claim_ledger", report)
        self.assertFalse(report["claim_ledger"]["available"])


class TierTests(unittest.TestCase):
    """信号分级：硬信号（确定性检出）必答进就绪门；软信号（模型自报）留档不计门限。
    真实教训：v10 数据 303 条假设 + 277 条 open_questions 把清单膨胀到 612 条不可用。"""

    def test_soft_signals_do_not_trip_readiness_gate(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp,
                 reqs=[{"title": "干净需求", "source_quote": "q"}],
                 analysis={"items": [{"source_section": "5", "source_requirement_ids": ["A"],
                                      "source_quote": "q",
                                      "open_questions": [f"软问题{i}" for i in range(20)],
                                      "assumptions": [f"假设{i}" for i in range(20)]}]},
                 quality={"failed_sections": 0, "coverage_pct": 80.0})
            report = cr.run_report(tmp)

        self.assertEqual(report["questions"], 0)                 # 必答 0
        self.assertEqual(report["soft_questions"], 40)
        self.assertEqual(report["questions_total"], 40)
        self.assertEqual(report["readiness"]["verdict"], "READY")   # 软信号不触发 NEEDS WORK

    def test_hard_signals_trip_gate(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp,
                 reqs=[{"title": f"T{i}", "source_section": "4", "source_quote": "q",
                        "suspicion_reasons": ["原文数值未带全"]} for i in range(31)],
                 quality={"failed_sections": 0, "coverage_pct": 80.0})
            report = cr.run_report(tmp)
        self.assertEqual(report["questions"], 31)
        self.assertEqual(report["readiness"]["verdict"], "NEEDS WORK")

    def test_markdown_separates_tiers(self) -> None:
        entries = [cr._entry(cr.CAT_MISSING, "硬问题"),
                   cr._entry(cr.CAT_ASSUMPTION, "软问题", tier=cr.TIER_SOFT)]
        md = cr.render_markdown(entries, {"verdict": "READY", "reasons": []})
        self.assertIn("必答·问客户（1）", md)
        self.assertIn("参考（1）", md)
        self.assertLess(md.find("硬问题"), md.find("软问题"))   # 必答在前


class AnswersRoundtripTests(unittest.TestCase):
    """答复回灌闭环：xlsx 填答复 → import → 报告消解 + 富化可见（analyze 侧另有注入测试）。"""

    def test_import_and_resolution(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp,
                 reqs=[{"title": "泄漏率", "source_section": "7.13", "source_quote": "q1",
                        "suspicion_reasons": ["原文数值未带全"]}],
                 quality={"failed_sections": 0, "coverage_pct": 80.0})
            first = cr.run_report(tmp)
            self.assertEqual(first["questions"], 1)

            # 模拟评审会：打开 xlsx 在「必答」sheet 填答复
            from openpyxl import load_workbook
            wb = load_workbook(tmp / cr.REPORT_XLSX)
            ws = wb["必答-问客户"]
            self.assertEqual(ws.cell(row=1, column=8).value, "答复")
            ws.cell(row=2, column=8, value="限值为 25 cm3/h")
            ws.cell(row=2, column=9, value="是")
            filled = tmp / "filled.xlsx"
            wb.save(filled)

            result = cr.import_answers(tmp, filled)
            self.assertEqual(result["imported"], 1)
            answers = cr.load_answers(tmp)
            self.assertEqual(len(answers), 1)
            entry = next(iter(answers.values()))
            self.assertEqual(entry["answer"], "限值为 25 cm3/h")
            self.assertTrue(entry["adopted"])

            second = cr.run_report(tmp)
            self.assertEqual(second["questions"], 0)               # 已答复采纳 → 消解
            self.assertEqual(second["resolved_by_answers"], 1)
            self.assertEqual(second["readiness"]["verdict"], "READY")

    def test_not_adopted_answer_keeps_question(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp,
                 reqs=[{"title": "T", "source_section": "4", "source_quote": "q",
                        "suspicion_reasons": ["原文数值未带全"]}],
                 quality={"failed_sections": 0, "coverage_pct": 80.0})
            first = cr.run_report(tmp)
            entry = first["entries"][0]
            with (tmp / cr.ANSWERS_FILE).open("w", encoding="utf-8") as f:
                import json as _json
                f.write(_json.dumps({"source_id": entry["source_id"],
                                     "question": entry["question"],
                                     "answer": "待定", "adopted": False},
                                    ensure_ascii=False) + chr(10))
            second = cr.run_report(tmp)
            self.assertEqual(second["questions"], 1)               # 未采纳不消解

    def test_customer_answer_is_stale_after_evidence_changes(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            requirement = {
                "ai_req_id": "AIR-1",
                "title": "Limit",
                "source_section": "4",
                "source_quote": "The configured limit shall be exposed.",
                "suspicion_reasons": ["原文数值未带全"],
            }
            seed(out, reqs=[requirement], quality={"failed_sections": 0, "coverage_pct": 80.0})
            entry = cr.run_report(out)["entries"][0]
            (out / cr.ANSWERS_FILE).write_text(json.dumps({
                "source_id": entry["source_id"],
                "question": entry["question"],
                "answer": "限值为 25。",
                "adopted": True,
                "clarification_id": entry["clarification_id"],
                "evidence_fingerprint": entry["evidence_fingerprint"],
            }, ensure_ascii=False) + "\n", encoding="utf-8")

            seed(out, reqs=[dict(requirement, source_quote="The revised limit shall be exposed.")],
                 quality={"failed_sections": 0, "coverage_pct": 80.0})
            report = cr.run_report(out)
            current_answers = cr.load_current_answers(out)

        self.assertEqual(report["questions"], 1)
        self.assertFalse(report["entries"][0]["answer_state_current"])
        self.assertEqual(current_answers, {})

    def test_closed_omission_is_removed_from_the_next_report(self) -> None:
        from omission_actions import apply_omission_action

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            seed(
                out,
                reqs=[],
                consistency={"coverage": {
                    "uncovered_count": 1,
                    "uncovered_samples": [{
                        "block_id": "B3", "section": "4",
                        "text": "The device shall expose a diagnostic flag.",
                    }],
                }},
                quality={"failed_sections": 0, "coverage_pct": 80.0},
            )
            (out / "blocks.jsonl").write_text(json.dumps({
                "block_id": "B3", "text": "The device shall expose a diagnostic flag.",
            }) + "\n", encoding="utf-8")
            first = cr.run_report(out)
            apply_omission_action(out, block_id="B3", status="non_requirement", actor="reviewer")
            second = cr.run_report(out)

        self.assertEqual(first["coverage_candidates"], 1)
        self.assertEqual(second["coverage_candidates"], 0)

    def test_closed_block_outside_current_candidates_does_not_reduce_count(self) -> None:
        from omission_actions import apply_omission_action

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            seed(
                out,
                reqs=[],
                consistency={"coverage": {
                    "uncovered_count": 1,
                    "uncovered_block_ids": ["B3"],
                    "uncovered_samples": [{
                        "block_id": "B3", "section": "4",
                        "text": "The device shall expose a diagnostic flag.",
                    }],
                }},
                quality={"failed_sections": 0, "coverage_pct": 80.0},
            )
            (out / "blocks.jsonl").write_text(
                "".join(json.dumps(row) + "\n" for row in [
                    {"block_id": "B3", "text": "The device shall expose a diagnostic flag."},
                    {"block_id": "B9", "text": "An unrelated old candidate."},
                ]),
                encoding="utf-8",
            )
            apply_omission_action(out, block_id="B9", status="non_requirement", actor="reviewer")

            report = cr.run_report(out)

        self.assertEqual(report["coverage_candidates"], 1)


class AudienceSplitTests(unittest.TestCase):
    def test_internal_checks_separated_from_customer_questions(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp,
                 reqs=[{"title": "A", "source_section": "4", "source_quote": "q",
                        "suspicion_reasons": ["原文数值未带全", "引用非逐字"]}],
                 quality={"failed_sections": 0, "coverage_pct": 80.0})
            report = cr.run_report(tmp)
        self.assertEqual(report["customer_questions"], 1)      # 漏值 → 问客户
        self.assertEqual(report["internal_checks"], 1)         # 引用非逐字 → 内部核对
        self.assertEqual(report["questions"], 2)               # 就绪门仍数全部硬信号


class RunReportTests(unittest.TestCase):
    def test_end_to_end_writes_three_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            seed(tmp,
                 reqs=[{"title": "T", "source_section": "4.1", "source_quote": "=EVIL()",
                        "suspicion_reasons": ["原文数值未带全"]}],
                 quality={"failed_sections": 0, "coverage_pct": 80.0})
            report = cr.run_report(tmp)

            self.assertEqual(report["questions"], 1)
            self.assertEqual(report["provenance"]["producer_version"],
                             "clarification/v7-claim-ledger-info")
            self.assertEqual(report["readiness"]["verdict"], "NEEDS WORK")
            self.assertEqual(report["readiness"]["unresolved_blocking"], 1)
            md = (tmp / cr.REPORT_MD).read_text(encoding="utf-8")
            self.assertIn("缺失", md)
            from openpyxl import load_workbook
            wb = load_workbook(tmp / cr.REPORT_XLSX)
            ws = wb["必答-问客户"]
            self.assertNotEqual(ws.cell(row=2, column=5).data_type, "f")   # 公式已中和
            self.assertIn("必答-内部核对", wb.sheetnames)
            self.assertIn("参考(模型自报)", wb.sheetnames)
            self.assertIn("就绪判定", wb.sheetnames)

    def test_missing_input_raises(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            with self.assertRaises(FileNotFoundError):
                cr.run_report(Path(td))


class BlockerAndIdentityTests(unittest.TestCase):
    def test_high_risk_signals_are_blocking_and_ids_are_stable(self) -> None:
        reasons = [
            "数字漂移",
            "数值配对待核",
            "表文数值不一致",
            "二遍复核:方向或上下限反转",
        ]
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            seed(out, reqs=[{
                "ai_req_id": "AIR-1",
                "title": "限值",
                "module": "计量",
                "source_section": "4",
                "source_quote": "The limit is 5 l/h.",
                "description": "限值为 6 l/h",
                "suspicion_reasons": reasons,
            }])
            first = cr.collect_questions(out)
            second = cr.collect_questions(out)

        self.assertEqual(len(first), 4)
        self.assertTrue(all(e["blocker_level"] == cr.BLOCKER_BLOCKING for e in first))
        self.assertTrue(all(e["audience"] == cr.AUDIENCE_INTERNAL for e in first))
        self.assertEqual(
            [e["clarification_id"] for e in first],
            [e["clarification_id"] for e in second],
        )
        self.assertTrue(all(e["evidence_fingerprint"] for e in first))

    def test_evidence_change_preserves_identity_but_requires_reconfirmation(self) -> None:
        from clarification_check_states import apply_clarification_check_action

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            req = {"ai_req_id": "AIR-1", "title": "T", "source_section": "4",
                   "source_quote": "original", "suspicion_reasons": ["引用非逐字"]}
            seed(out, reqs=[req])
            before = cr.run_report(out)["entries"][0]
            apply_clarification_check_action(
                out, before["clarification_id"], "verified_ok",
                evidence_fingerprint=before["evidence_fingerprint"], actor="reviewer",
            )
            seed(out, reqs=[dict(req, source_quote="changed")])
            report = cr.run_report(out)
            after = report["entries"][0]
        self.assertNotEqual(before["evidence_fingerprint"], after["evidence_fingerprint"])
        self.assertEqual(before["clarification_id"], after["clarification_id"])
        self.assertFalse(after["check_state_current"])
        self.assertEqual(report["questions"], 1)

    def test_distinct_second_pass_findings_keep_distinct_identities(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            seed(out, reqs=[{
                "ai_req_id": "AIR-1",
                "title": "Limit",
                "source_section": "4",
                "source_quote": "The limit shall remain below the maximum.",
                "source_block_ids": ["B1"],
                "suspicion_reasons": [
                    "二遍复核:方向或上下限反转",
                    "二遍复核:主语或对象错配",
                ],
            }])
            entries = cr.collect_questions(out)

        self.assertEqual(len(entries), 2)
        self.assertEqual(len({entry["clarification_id"] for entry in entries}), 2)

    def test_module_override_is_reported_as_effective_module(self) -> None:
        from ai_review_actions import apply_ai_review_action

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            seed(out, reqs=[{"ai_req_id": "AIR-1", "title": "T", "module": "时钟",
                             "source_quote": "q", "suspicion_reasons": ["数字漂移"]}])
            apply_ai_review_action(out, "AIR-1", "needs_discussion", module_override="计量精度")
            entry = cr.collect_questions(out)[0]
        self.assertEqual(entry["module"], "计量精度")

    def test_readiness_separates_blockers_from_ordinary_threshold(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            seed(out, reqs=[], quality={"failed_sections": 0, "coverage_pct": 80.0})
            blocked = cr.readiness_verdict(
                out, unresolved_blocking=1, unresolved_important=0
            )
            ordinary_ok = cr.readiness_verdict(
                out, unresolved_blocking=0, unresolved_important=30
            )
            ordinary_many = cr.readiness_verdict(
                out, unresolved_blocking=0, unresolved_important=31
            )
        self.assertEqual(blocked["verdict"], "NEEDS WORK")
        self.assertEqual(ordinary_ok["verdict"], "READY")
        self.assertEqual(ordinary_many["verdict"], "NEEDS WORK")


class InternalCheckClosureTests(unittest.TestCase):
    def _seed_internal(self, out: Path) -> None:
        seed(out, reqs=[{
            "ai_req_id": "AIR-1",
            "title": "T",
            "module": "计量",
            "source_section": "4",
            "source_quote": "source text",
            "suspicion_reasons": ["引用非逐字"],
        }], quality={"failed_sections": 0, "coverage_pct": 80.0})

    def test_verified_ok_resolves_only_matching_evidence(self) -> None:
        from clarification_check_states import apply_clarification_check_action

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed_internal(out)
            entry = cr.run_report(out)["entries"][0]
            apply_clarification_check_action(
                out,
                entry["clarification_id"],
                "verified_ok",
                evidence_fingerprint="stale-fingerprint",
                actor="reviewer",
            )
            stale = cr.run_report(out)
            self.assertEqual(stale["questions"], 1)
            self.assertFalse(stale["entries"][0]["check_state_current"])
            self.assertEqual(stale["readiness"]["verdict"], "NEEDS WORK")

            apply_clarification_check_action(
                out,
                entry["clarification_id"],
                "verified_ok",
                evidence_fingerprint=entry["evidence_fingerprint"],
                actor="reviewer",
                note="逐字核对完成",
            )
            resolved = cr.run_report(out)

        self.assertEqual(resolved["questions"], 0)
        self.assertEqual(resolved["resolved_by_checks"], 1)
        self.assertEqual(resolved["readiness"]["resolved_internal"], 1)
        self.assertEqual(resolved["entries"][0]["check_note"], "逐字核对完成")
        self.assertEqual(resolved["entries"][0]["state"], "verified_ok")
        self.assertEqual(resolved["entries"][0]["actor"], "reviewer")
        self.assertTrue(resolved["entries"][0]["timestamp"])
        self.assertEqual(resolved["entries"][0]["note"], "逐字核对完成")
        self.assertEqual(resolved["readiness"]["unresolved"], 0)
        self.assertEqual(resolved["readiness"]["resolved"], 1)

    def test_issue_confirmed_and_deferred_remain_unresolved(self) -> None:
        from clarification_check_states import apply_clarification_check_action

        for action in ("issue_confirmed", "deferred"):
            with self.subTest(action=action), tempfile.TemporaryDirectory() as td:
                out = Path(td)
                self._seed_internal(out)
                entry = cr.run_report(out)["entries"][0]
                apply_clarification_check_action(
                    out, entry["clarification_id"], action,
                    evidence_fingerprint=entry["evidence_fingerprint"], actor="reviewer",
                )
                report = cr.run_report(out)
                self.assertEqual(report["questions"], 1)
                self.assertEqual(report["unresolved_internal"], 1)
                self.assertEqual(report["readiness"]["verdict"], "NEEDS WORK")

    def test_internal_action_roundtrip_from_xlsx(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed_internal(out)
            first = cr.run_report(out)
            wb = load_workbook(out / cr.REPORT_XLSX)
            ws = wb["必答-内部核对"]
            headers = {str(cell.value): cell.column for cell in ws[1]}
            ws.cell(2, headers["新处置(确认无误/确认有问题/暂缓)"], "确认无误")
            ws.cell(2, headers["核对人"], "张工")
            ws.cell(2, headers["备注"], "已对照原文")
            filled = out / "filled-internal.xlsx"
            wb.save(filled)
            wb.close()

            imported = cr.import_internal_checks(out, filled)
            second = cr.run_report(out)
            from clarification_check_states import read_clarification_check_states
            stored = read_clarification_check_states(out)[first["entries"][0]["clarification_id"]]

        self.assertEqual(first["questions"], 1)
        self.assertEqual(imported["imported"], 1)
        self.assertEqual(second["questions"], 0)
        self.assertEqual(second["entries"][0]["check_actor"], "张工")
        self.assertTrue(second["entries"][0]["check_timestamp"])
        state = second["entries"][0]
        self.assertEqual(state["blocker_level"], cr.BLOCKER_IMPORTANT)
        self.assertEqual(state["module"], "计量")
        self.assertEqual(stored["blocker_level"], cr.BLOCKER_IMPORTANT)
        self.assertEqual(stored["module"], "计量")

    def test_internal_xlsx_import_rejects_stale_evidence_without_writing(self) -> None:
        from openpyxl import load_workbook
        from clarification_check_states import read_clarification_check_states

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed_internal(out)
            first = cr.run_report(out)
            entry = first["entries"][0]
            wb = load_workbook(out / cr.REPORT_XLSX)
            ws = wb["必答-内部核对"]
            headers = {str(cell.value): cell.column for cell in ws[1]}
            ws.cell(2, headers["证据指纹"], "stale-fingerprint")
            ws.cell(2, headers["新处置(确认无误/确认有问题/暂缓)"], "确认无误")
            filled = out / "filled-stale-internal.xlsx"
            wb.save(filled)
            wb.close()

            imported = cr.import_internal_checks(out, filled)
            states = read_clarification_check_states(out)

        self.assertEqual(imported["imported"], 0)
        self.assertEqual(imported["stale"], [entry["clarification_id"]])
        self.assertEqual(states, {})

    def test_json_is_organized_by_blocker_audience_and_module(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed_internal(out)
            report = cr.run_report(out)
        grouped = report["organized_entries"][cr.BLOCKER_IMPORTANT]
        self.assertIn("计量", grouped[cr.AUDIENCE_INTERNAL])


class BatchInternalCheckTests(unittest.TestCase):
    def test_batch_applies_only_current_evidence_and_reports_categories(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            seed(out, reqs=[
                {
                    "ai_req_id": "AI-1", "title": "A", "module": "计量",
                    "source_section": "4", "source_quote": "The meter shall do A.",
                    "suspicion_reasons": ["引用非逐字"],
                },
                {
                    "ai_req_id": "AI-2", "title": "B", "module": "通信协议",
                    "source_section": "5", "source_quote": "The meter shall do B.",
                    "suspicion_reasons": ["引用非逐字"],
                },
            ], quality={"failed_sections": 0, "core_coverage_pct": 100.0})
            entries = [
                entry for entry in cr.collect_questions(out)
                if entry["audience"] == cr.AUDIENCE_INTERNAL
            ]
            result = cr.batch_apply_internal_checks(out, [
                {
                    "clarification_id": entries[0]["clarification_id"],
                    "evidence_fingerprint": entries[0]["evidence_fingerprint"],
                },
                {
                    "clarification_id": entries[1]["clarification_id"],
                    "evidence_fingerprint": "stale-fingerprint",
                },
            ], actor="reviewer", note="同类证据已逐项核对")

            states = cr.read_clarification_check_states(out)

        self.assertEqual(result["applied"], 1)
        self.assertEqual(result["stale"], [entries[1]["clarification_id"]])
        self.assertEqual(result["by_signal"], {"suspicion:引用": 1})
        self.assertEqual(len(states), 1)
        self.assertEqual(next(iter(states.values()))["note"], "同类证据已逐项核对")


if __name__ == "__main__":
    unittest.main()
