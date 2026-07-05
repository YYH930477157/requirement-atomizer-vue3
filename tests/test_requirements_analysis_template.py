"""requirements_analysis_template 回归（unittest 风格——pytest 未装，模块级函数不会被 discover 收集）。"""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from requirements_analysis_template import (
    extract_template_vocabulary,
    fallback_template_vocabulary,
)


class TemplateVocabularyTests(unittest.TestCase):
    def test_extracts_sheet_modules_and_submodules(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            path = tmp_path / "template.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "时钟需求"
            ws.append(["关闭", "序号", "子模块", "描述", "需求"])
            ws.append(["", 1, "时钟", "夏令时：", "支持"])
            ws.append(["", 2, "时钟同步", "时区：", "东八区"])
            ws.append(["", 3, "时钟", "重复：", "忽略"])
            ws2 = wb.create_sheet("协议栈需求")
            ws2.append(["关闭", "序号", "子模块", "描述", "需求"])
            ws2.append(["", 1, "通信口1", "通信方式：", "Optical"])
            wb.save(path)

            vocab = extract_template_vocabulary(path)

            assert "时钟需求" in vocab["modules"]
            assert vocab["submodules_by_module"]["时钟需求"] == ["时钟", "时钟同步"]
            assert vocab["submodules_by_module"]["协议栈需求"] == ["通信口1"]

    def test_fallback_vocabulary_contains_core_modules(self) -> None:
        vocab = fallback_template_vocabulary()

        assert "系统需求" in vocab["modules"]
        assert "协议栈需求" in vocab["modules"]


class TemplateKnowledgeTests(unittest.TestCase):
    """模板知识（公司标准做法）抽取与检索：说明/示例列进富化 prompt 的地基。"""

    def _make_knowledge_template(self, path: Path) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "时钟需求"
        ws.append(["关闭", "序号", "子模块", "描述", "需求模版", "需求", "说明、示例、注意事项"])
        ws.append(["", 1, "时钟", "历法：", "公历", "公历", "① 公历 ② 波斯历法"])
        ws.append(["", 2, "时钟", "夏令时：", "支持", "支持", ""])
        ws.append(["", 3, "时钟", "时钟精度：", "±0.5s/天", "±0.5s/天", "对应程序中 RTC_CAL 宏定义"])
        wb.create_sheet("计量列表").append(["占位"])   # "列表" sheet 跳过
        wb.save(path)

    def test_extract_knowledge_reads_question_default_notes(self) -> None:
        from requirements_analysis_template import extract_template_knowledge
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "t.xlsx"
            self._make_knowledge_template(path)
            knowledge = extract_template_knowledge(path)
        self.assertEqual(list(knowledge), ["时钟需求"])        # "列表" sheet 不进知识
        self.assertEqual(knowledge["时钟需求"][0], {
            "submodule": "时钟", "question": "历法：", "default": "公历",
            "notes": "① 公历 ② 波斯历法"})
        self.assertEqual(knowledge["时钟需求"][2]["notes"], "对应程序中 RTC_CAL 宏定义")

    def test_extract_knowledge_missing_path_returns_empty(self) -> None:
        from requirements_analysis_template import extract_template_knowledge
        self.assertEqual(extract_template_knowledge(None), {})
        self.assertEqual(extract_template_knowledge(Path("/no/such/file.xlsx")), {})

    def test_select_references_picks_relevant_rows_only(self) -> None:
        from requirements_analysis_template import (
            extract_template_knowledge, select_template_references)
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "t.xlsx"
            self._make_knowledge_template(path)
            knowledge = extract_template_knowledge(path)
        req = {"module": "时钟", "title": "时钟精度要求",
               "description": "时钟精度须优于每天 5 秒", "source_quote": "clock accuracy"}
        refs = select_template_references(knowledge, req)
        self.assertTrue(refs)
        self.assertEqual(refs[0]["question"], "时钟精度：")       # 相关行排最前
        # 模块无 sheet / 零词面重叠 → 不注入（宁漏勿错）
        self.assertEqual(select_template_references(knowledge, {"module": "安全", "title": "x"}), [])
        self.assertEqual(select_template_references(
            knowledge, {"module": "时钟", "title": "wholly unrelated words"}), [])

    def test_select_references_respects_char_cap(self) -> None:
        from requirements_analysis_template import select_template_references
        knowledge = {"时钟需求": [
            {"submodule": "时钟", "question": f"时钟精度项{i}：", "default": "", "notes": "长" * 900}
            for i in range(4)]}
        req = {"module": "时钟", "title": "时钟精度", "description": "", "source_quote": ""}
        refs = select_template_references(knowledge, req, k=4, max_chars=1500)
        self.assertEqual(len(refs), 1)                            # 第二条起超帽被截

    def test_render_references_format(self) -> None:
        from requirements_analysis_template import render_template_references
        text = render_template_references([
            {"submodule": "时钟", "question": "历法：", "default": "公历", "notes": "①公历②波斯历法"}])
        self.assertEqual(text, "时钟 | 历法： | 默认:公历 | 说明:①公历②波斯历法")


if __name__ == "__main__":
    unittest.main()
