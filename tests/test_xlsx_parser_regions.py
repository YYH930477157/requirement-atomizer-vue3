"""End-to-end tests for xlsx_parser region-detection wiring (WS1 wk7).

Uses openpyxl to synthesize real .xlsx workbooks (no real LLM, no real
external dependency beyond openpyxl which is already installed). Verifies:
  * multi-sheet OBIS key_missing => honest parse_incomplete on the keyed
    table blocks (never a silent merge);
  * cross-sheet shared keys => linked, no audit signal;
  * single sheet without OBIS => legacy behaviour unchanged (no new audit).
"""
from __future__ import annotations

import random
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from parsers.xlsx_parser import _RowColumnIntervals, extract_xlsx


def _write_workbook(path: Path, sheets: dict[str, list[list[str]]]) -> None:
    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(title=name)
        first = False
        ws.title = name
        for row_index, row in enumerate(rows, start=1):
            for column_index, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=column_index, value=value)
    wb.save(path)


class MultiSheetObisLinkTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = Path(tempfile.mkdtemp())
        self.addCleanup(_cleanup, self.tmp)

    def _table_blocks(self, blocks):
        return [b for b in blocks if b.get("type") == "table"]

    def test_key_missing_blocks_silently_marked(self) -> None:
        # Two keyed tables across two sheets with ZERO shared keys.
        path = self.tmp / "no_shared.xlsx"
        _write_workbook(path, {
            "s1": [["OBIS", "Desc"], ["0-0:96.1.0", "clock"]],
            "s2": [["OBIS", "Desc"], ["1-1:32.0.0", "assoc"]],
        })
        blocks, _items, _cells = extract_xlsx(path)
        tables = self._table_blocks(blocks)
        self.assertEqual(len(tables), 2)
        # Both keyed tables must honestly carry the link-block audit.
        for block in tables:
            self.assertTrue(block["parse_incomplete"], block["table_id"])
            reason = block["parse_incomplete_reason"]
            self.assertIsNotNone(reason)
            # Either as the main reason or as an additional_reasons entry.
            payloads = []
            if isinstance(reason, dict) and reason.get("code") == "xlsx_multi_sheet_link_blocked":
                payloads.append(reason)
            for extra in (reason.get("additional_reasons") or []) if isinstance(reason, dict) else []:
                if isinstance(extra, dict) and extra.get("code") == "xlsx_multi_sheet_link_blocked":
                    payloads.append(extra)
            self.assertTrue(payloads, f"no link-block audit on {block['table_id']}")
            self.assertEqual(payloads[0]["status"], "key_missing")

    def test_shared_keys_linked_no_audit(self) -> None:
        path = self.tmp / "shared.xlsx"
        _write_workbook(path, {
            "s1": [["OBIS", "Desc"], ["0-0:96.1.0", "clock"]],
            "s2": [["OBIS", "Ref"], ["0-0:96.1.0", "see s1"]],
        })
        blocks, _items, _cells = extract_xlsx(path)
        tables = self._table_blocks(blocks)
        for block in tables:
            # Linked => no link-block audit: parse_incomplete stays at whatever
            # the sheet-level parse produced (False for this clean workbook).
            reason = block.get("parse_incomplete_reason")
            codes = []
            if isinstance(reason, dict):
                codes.append(reason.get("code"))
                codes.extend(
                    e.get("code") for e in (reason.get("additional_reasons") or [])
                    if isinstance(e, dict)
                )
            self.assertNotIn("xlsx_multi_sheet_link_blocked", codes)

    def test_single_sheet_legacy_behaviour_unchanged(self) -> None:
        path = self.tmp / "single.xlsx"
        _write_workbook(path, {"only": [["Name", "Value"], ["foo", "1"], ["bar", "2"]]})
        blocks, items, cells = extract_xlsx(path)
        tables = self._table_blocks(blocks)
        self.assertEqual(len(tables), 1)
        self.assertFalse(tables[0]["parse_incomplete"])
        # parse_incomplete_reason absent or None — legacy default.
        self.assertFalse(tables[0].get("parse_incomplete_reason"))


class RowColumnIntervalsDifferentialTests(unittest.TestCase):
    """_RowColumnIntervals（bisect stabbing 计数）的差分等价测试。

    区间索引替代守恒计数/表内跳过的逐格枚举时，等价性证据曾只是一次性的
    reverse-patch 对比（不可在仓库内复现）。这里用固定种子的随机矩形
    （含退化：反向行段/列段、单格、重叠、相邻、空区域集）把等价性永久钉住：
    每个格子的 count/contains 必须与朴素逐格参考实现完全一致。"""

    @staticmethod
    def _naive_count(
        regions: list[tuple[int, int, int, int]], row_index: int, column_index: int
    ) -> int:
        return sum(
            1
            for min_row, min_col, max_row, max_col in regions
            if min_row <= row_index <= max_row and min_col <= column_index <= max_col
        )

    @staticmethod
    def _random_rectangles(
        rng: random.Random, max_row: int, max_column: int, count: int
    ) -> list[tuple[int, int, int, int]]:
        regions: list[tuple[int, int, int, int]] = []
        for _ in range(count):
            row_a, row_b = rng.randint(1, max_row), rng.randint(1, max_row)
            col_a, col_b = rng.randint(1, max_column), rng.randint(1, max_column)
            min_row, max_row = min(row_a, row_b), max(row_a, row_b)
            min_col, max_col = min(col_a, col_b), max(col_a, col_b)
            # 概率注入畸形矩形：反向行段/列段（空跨度；朴素实现里天然覆盖 0 格）
            if rng.random() < 0.25:
                min_row, max_row = max_row + 1, min_row
            if rng.random() < 0.25:
                min_col, max_col = max_col + 1, min_col
            regions.append((min_row, min_col, max_row, max_col))
        return regions

    def test_randomized_regions_match_naive_reference(self) -> None:
        rng = random.Random(20260815)
        for sheet_index in range(40):
            max_row = rng.randint(1, 10)
            max_column = rng.randint(1, 10)
            regions = self._random_rectangles(
                rng, max_row, max_column, rng.randint(0, 8)  # 0 = 空 sheet
            )
            index = _RowColumnIntervals(regions)
            # 探测窗比网格大一圈：区域边界外/索引 0 也必须与参考一致
            for row_index in range(0, max_row + 2):
                for column_index in range(0, max_column + 2):
                    expected = self._naive_count(regions, row_index, column_index)
                    self.assertEqual(
                        index.count(row_index, column_index),
                        expected,
                        f"sheet {sheet_index} regions={regions} "
                        f"cell=({row_index},{column_index})",
                    )
                    self.assertEqual(
                        index.contains(row_index, column_index),
                        expected > 0,
                        f"contains mismatch sheet {sheet_index} regions={regions} "
                        f"cell=({row_index},{column_index})",
                    )

    def test_degenerate_rectangles_never_yield_negative_counts(self) -> None:
        """畸形矩形守卫：反向区间被跳过，计数永不为负（守恒门不得被静默穿过）。

        无守卫时反向列区间 (1,5,2,3) 的 stabbing 计数在中段会得到 -1，而守恒门
        只看 ==0 / >1，负数会静默放行。守卫后：受影响格子一律计 0（表现为
        dropped，fail-closed 大声报警），单格等正常矩形不受影响。"""
        inverted_columns = _RowColumnIntervals([(1, 5, 2, 3)])
        for row_index in range(0, 4):
            for column_index in range(0, 8):
                count = inverted_columns.count(row_index, column_index)
                self.assertGreaterEqual(
                    count, 0, f"negative count at ({row_index},{column_index})"
                )
                self.assertFalse(inverted_columns.contains(row_index, column_index))
        inverted_rows = _RowColumnIntervals([(3, 1, 2, 4)])
        self.assertEqual(inverted_rows.count(3, 1), 0)
        self.assertFalse(inverted_rows.contains(3, 1))
        # 正常矩形（含单格）行为不变
        single_cell = _RowColumnIntervals([(2, 2, 2, 2)])
        self.assertEqual(single_cell.count(2, 2), 1)
        self.assertTrue(single_cell.contains(2, 2))
        self.assertEqual(single_cell.count(2, 3), 0)
        block = _RowColumnIntervals([(1, 1, 2, 2), (2, 2, 3, 3)])
        self.assertEqual(block.count(2, 2), 2)  # 重叠处叠加计数
        self.assertEqual(block.count(1, 1), 1)
        self.assertEqual(block.count(4, 4), 0)


def _cleanup(path: Path) -> None:
    import shutil
    shutil.rmtree(path, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
