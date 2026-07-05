"""模板映射器（架构转向的第二步：理解在前、映射在后）。

方向裁定（用户拍板）：**不是**拿模板问题去审讯文档（带着答案找问题，先入为主且跨地区失灵）；
而是先用抽取管线按文档自身逻辑读懂需求（ai_requirements.jsonl），再把**已理解的需求**对齐到
公司标准化需求列表（V2.3.x）的槽位——匹配上的用文档答案填槽 + 章节溯源，匹配不上的留作
新增行候选。模板 = 交付格式 + 完整性保底，不是理解工具。

防幻觉：seq 只认清单内编号（程序校验）；answer 的数字/编码必须 ⊆ 来源需求文本（drift 护栏，
编造即弃该指派并留痕）；unmatched 由程序核算（全集减已指派），不信 LLM 自报。

用法：python -m template_mapper --template <V2.3.x.xlsx> --out <atomizer 输出目录>
      --sheet 时钟需求 [--route openai_compatible]
产物：<out>/软件需求列表-映射.xlsx（模板副本，目标 sheet 填好）+ <out>/mapping_report.json
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path
from typing import Any, Callable

from cosem_behavior_spec import extract_codes, extract_ints
from io_utils import read_jsonl
from text_normalize import formula_safe

LOGGER = logging.getLogger("requirement_atomizer")

ChatFn = Callable[[str, str], dict[str, Any]]

MAPPER_PROMPT_VERSION = "template-map-v1"
MAPPED_WORKBOOK = "软件需求列表-映射.xlsx"
MAPPING_REPORT = "mapping_report.json"

# 模板统一表头（V2.3.x 实测 19 个需求 sheet 全一致）
_COL_SEQ = 2          # 序号
_COL_SUBMODULE = 3    # 子模块
_COL_QUESTION = 4     # 描述
_COL_DEFAULT = 5      # 需求模版
_COL_ANSWER = 6       # 需求
_COL_IS_CUSTOMER = 8  # 是否客户需求
_COL_SECTION = 9      # 客户需求章节

SYSTEM_PROMPT = (
    "你是电表软件需求分析工程师。给你两份材料：【标准问题清单】（公司标准化需求列表某模块的"
    "问题行：seq 编号、子模块、问题、默认值）和【已理解的客户需求】（从客户标准文档抽取，"
    "含 id、标题、描述、原文引用、章节）。任务：判断哪些需求回答了哪些问题。"
    "输出 JSON：{\"assignments\": [{\"seq\": <清单内编号>, \"answer\": \"<该问题的客户答案，"
    "值/参数必须来自需求内容原文，不得编造>\", \"source_req_ids\": [\"<需求id>\"]}], "
    "\"unmatched_req_ids\": [\"<清单里没有对应问题的需求id>\"]}。"
    "规则：seq 只能取清单里出现过的编号；answer 用简洁中文陈述客户的具体要求（数值/编码原样照抄）；"
    "一个问题可由多条需求共同回答，一条需求也可回答多个问题；**没有把握就不指派（宁缺勿滥）**；"
    "客户需求与清单问题无关时归入 unmatched_req_ids。只输出 JSON。"
)


def load_question_bank(template_path: Path, sheet: str) -> list[dict[str, Any]]:
    """读模板某模块 sheet 的标准问题行（描述列非空）。保留 xlsx 真实行号供回写。"""
    from openpyxl import load_workbook
    wb = load_workbook(template_path, data_only=True, read_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"模板中不存在 sheet：{sheet}（可用：{', '.join(wb.sheetnames[:8])}…）")
        ws = wb[sheet]
        questions: list[dict[str, Any]] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < _COL_QUESTION or not row[_COL_QUESTION - 1]:
                continue
            questions.append({
                "xlsx_row": row_idx,
                "seq": str(row[_COL_SEQ - 1] or "").strip() or str(row_idx),
                "submodule": str(row[_COL_SUBMODULE - 1] or "").strip(),
                "question": str(row[_COL_QUESTION - 1] or "").strip(),
                "default": str(row[_COL_DEFAULT - 1] or "").strip(),
            })
        return questions
    finally:
        wb.close()  # read_only 句柄必须关（Windows 下不关会锁住模板文件）


def sheet_module(sheet: str) -> str:
    """sheet 名 → 抽取轨模块名（"时钟需求"→"时钟"；"负荷曲线"→"曲线"）。"""
    special = {"负荷曲线": "曲线", "P1需求": "通信协议", "MBUS需求": "通信协议",
               "报警窃电需求": "窃电", "push需求": "Push"}
    if sheet in special:
        return special[sheet]
    return sheet[:-2] if sheet.endswith("需求") else sheet


def select_requirements(requirements: list[dict[str, Any]], sheet: str) -> list[dict[str, Any]]:
    """按模块归属预筛给映射器的需求（省 token；跨模块遗漏由未答问题的查漏网兜——第三期）。"""
    module = sheet_module(sheet)
    picked = []
    for req in requirements:
        labels = [str(x) for x in (req.get("labels") or [])]
        if str(req.get("module") or "").strip() == module or module in labels:
            picked.append(req)
    return picked


def build_user_prompt(questions: list[dict[str, Any]], requirements: list[dict[str, Any]]) -> str:
    bank = [{"seq": q["seq"], "submodule": q["submodule"], "question": q["question"],
             "default": q["default"]} for q in questions]
    reqs = [{"id": str(r.get("ai_req_id") or ""), "title": r.get("title"),
             "description": r.get("description"), "source_quote": r.get("source_quote"),
             "source_section": r.get("source_section")} for r in requirements]
    return ("【标准问题清单】\n" + json.dumps(bank, ensure_ascii=False, indent=1)
            + "\n\n【已理解的客户需求】\n" + json.dumps(reqs, ensure_ascii=False, indent=1))


def _answer_drift(answer: str, sources: list[dict[str, Any]]) -> list[str]:
    """answer 的数字/编码必须有据（⊆ 来源需求文本合集）。返回漂移项；空=通过。"""
    basis = " ".join(
        " ".join([str(r.get("title") or ""), str(r.get("description") or ""),
                  str(r.get("source_quote") or ""), json.dumps(r.get("threshold_table") or {})])
        for r in sources)
    drift = sorted((extract_codes(answer) - extract_codes(basis))
                   | (extract_ints(answer) - extract_ints(basis)))
    return drift


def map_requirements(questions: list[dict[str, Any]], requirements: list[dict[str, Any]],
                     chat: ChatFn) -> dict[str, Any]:
    """LLM 指派 + 程序护栏。返回 {assignments, unmatched_requirements, rejected}。"""
    by_id = {str(r.get("ai_req_id") or ""): r for r in requirements}
    valid_seqs = {q["seq"] for q in questions}
    payload = chat(SYSTEM_PROMPT, build_user_prompt(questions, requirements))

    assignments: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    assigned_req_ids: set[str] = set()
    for item in (payload.get("assignments") or []):
        if not isinstance(item, dict):
            continue
        seq = str(item.get("seq") or "").strip()
        answer = str(item.get("answer") or "").strip()
        src_ids = [str(x) for x in (item.get("source_req_ids") or []) if str(x) in by_id]
        if seq not in valid_seqs:
            rejected.append({"seq": seq, "reason": "seq 不在清单内（编造槽位）"})
            continue
        if not answer or not src_ids:
            rejected.append({"seq": seq, "reason": "缺 answer 或有效来源需求"})
            continue
        sources = [by_id[i] for i in src_ids]
        drift = _answer_drift(answer, sources)
        if drift:
            rejected.append({"seq": seq, "reason": f"答案数值/编码无据：{', '.join(drift[:5])}"})
            continue
        sections = sorted({str(s.get("source_section") or "").strip()
                           for s in sources if s.get("source_section")})
        assignments.append({"seq": seq, "answer": answer, "source_req_ids": src_ids,
                            "customer_sections": sections})
        assigned_req_ids.update(src_ids)

    # unmatched 由程序核算（全集 − 已指派），不信 LLM 自报
    unmatched = [r for r in requirements if str(r.get("ai_req_id") or "") not in assigned_req_ids]
    return {"assignments": assignments, "rejected": rejected,
            "unmatched_requirements": [
                {"ai_req_id": r.get("ai_req_id"), "title": r.get("title"),
                 "source_section": r.get("source_section")} for r in unmatched]}


def apply_to_workbook(template_path: Path, sheet: str, questions: list[dict[str, Any]],
                      assignments: list[dict[str, Any]], out_path: Path) -> Path:
    """模板副本回写：目标 sheet 的 需求/是否客户需求/客户需求章节 三列，其余分毫不动。"""
    from openpyxl import load_workbook
    wb = load_workbook(template_path)   # 保留格式/其余 sheet
    ws = wb[sheet]
    row_by_seq = {q["seq"]: q["xlsx_row"] for q in questions}
    for a in assignments:
        row = row_by_seq.get(a["seq"])
        if not row:
            continue
        ws.cell(row=row, column=_COL_ANSWER, value=formula_safe(a["answer"]))
        ws.cell(row=row, column=_COL_IS_CUSTOMER, value="是")
        if a.get("customer_sections"):
            ws.cell(row=row, column=_COL_SECTION,
                    value=formula_safe("、".join(a["customer_sections"])))
    wb.save(out_path)
    return out_path


def run_mapping(out_dir: Path, template_path: Path, sheet: str,
                chat: ChatFn | None = None, route: str | None = None) -> dict[str, Any]:
    out_dir = Path(out_dir).expanduser().resolve()
    reqs_path = out_dir / "ai_requirements.jsonl"
    if not reqs_path.exists():
        raise FileNotFoundError(f"ai_requirements.jsonl not found in {out_dir} — 先跑「AI 抽取」（理解在前）")
    requirements = read_jsonl(reqs_path)
    # 原始行没有 ai_req_id 字段（内容稳定 id 由视图层计算）——映射器同样用唯一主键函数补齐，
    # 与批注/裁决同一 id 体系（将来映射结果可回链批注）
    from ai_review_actions import source_ai_requirement_id
    for req in requirements:
        req["ai_req_id"] = source_ai_requirement_id(req)
    questions = load_question_bank(template_path, sheet)
    picked = select_requirements(requirements, sheet)

    if chat is None:
        from ai_extract import DEFAULT_PIPELINE_PATH, config_for_route
        from llm_client import chat_json
        config = config_for_route(route or "openai_compatible", DEFAULT_PIPELINE_PATH)
        if config is None:
            raise RuntimeError("openai_compatible 端点未配置（映射是 LLM 分类任务，无 stub 路径）")
        chat = lambda s, u: chat_json(config, s, u)  # noqa: E731

    if picked:
        result = map_requirements(questions, picked, chat)
    else:
        result = {"assignments": [], "rejected": [], "unmatched_requirements": []}

    answered = {a["seq"] for a in result["assignments"]}
    report = {
        "prompt_version": MAPPER_PROMPT_VERSION,
        "sheet": sheet,
        "questions": len(questions),
        "candidate_requirements": len(picked),
        "answered": len(answered),
        "unanswered_seqs": [q["seq"] for q in questions if q["seq"] not in answered],
        **result,
    }
    out_xlsx = out_dir / MAPPED_WORKBOOK
    apply_to_workbook(template_path, sheet, questions, result["assignments"], out_xlsx)
    (out_dir / MAPPING_REPORT).write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    report["written"] = [MAPPED_WORKBOOK, MAPPING_REPORT]
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Map understood requirements onto the company template sheet.")
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--route", default="openai_compatible")
    args = parser.parse_args(argv)
    try:
        report = run_mapping(args.out, args.template, args.sheet, route=args.route)
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps({k: report[k] for k in
                      ("sheet", "questions", "candidate_requirements", "answered", "written")},
                     ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
