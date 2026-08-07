"""Tests for unextracted_registry.py (A7)."""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from unextracted_registry import (
    KIND_FRONT_MATTER_BLOCK,
    KIND_HEADER_CHANNEL,
    KIND_HIDDEN_SHEET,
    KIND_NOISE_BLOCK,
    KIND_SKIPPED_SHEET,
    KIND_TEXTBOX_CHANNEL,
    build_unextracted_registry,
    collect_unextracted_clarification_entries,
    load_unextracted_registry,
    summarize_unextracted_counts,
    write_unextracted_registry,
)


class BuildRegistryTests(unittest.TestCase):
    def test_noise_blocks_are_registered(self):
        blocks = [
            {"block_id": "BLK-000001", "text": "Scope", "section_path": ["Scope"], "noise": False},
            {"block_id": "BLK-000002", "text": "Page 12", "section_path": [], "noise": True},
        ]
        registry = build_unextracted_registry(Path("/tmp/fake.docx"), blocks)
        self.assertEqual(registry["total"], 1)
        entry = registry["entries"][0]
        self.assertEqual(entry["kind"], KIND_NOISE_BLOCK)
        self.assertEqual(entry["source_id"], "BLK-000002")
        self.assertEqual(entry["text_preview"], "Page 12")
        self.assertIn("噪声", entry["reason"])

    def test_front_matter_blocks_are_registered(self):
        blocks = [
            {"block_id": "BLK-000001", "text": "Cover title", "section_path": [], "noise": False, "doc_region": "front_matter"},
            {"block_id": "BLK-000002", "text": "Scope", "section_path": ["Scope"], "noise": False, "doc_region": "body"},
        ]
        registry = build_unextracted_registry(Path("/tmp/fake.docx"), blocks)
        kinds = [e["kind"] for e in registry["entries"]]
        self.assertIn(KIND_FRONT_MATTER_BLOCK, kinds)
        self.assertEqual(registry["by_kind"][KIND_FRONT_MATTER_BLOCK], 1)

    def test_textbox_channel_blocks_are_registered(self):
        blocks = [
            {"block_id": "BLK-000001", "text": "Body paragraph", "section_path": [], "content_channel": "body"},
            {"block_id": "BLK-000002", "text": "Textbox note", "section_path": [], "content_channel": "textbox"},
            {"block_id": "BLK-000003", "text": "Header note", "section_path": [], "content_channel": "header"},
        ]
        registry = build_unextracted_registry(Path("/tmp/fake.docx"), blocks)
        kinds = {e["kind"] for e in registry["entries"]}
        self.assertEqual(kinds, {KIND_TEXTBOX_CHANNEL, KIND_HEADER_CHANNEL})

    def test_xlsx_hidden_sheets_are_registered(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "hidden.xlsx"
            wb = Workbook()
            wb.active.title = "Visible"
            wb.create_sheet("Hidden")
            wb["Hidden"].sheet_state = "hidden"
            wb.save(path)
            wb.close()

            registry = build_unextracted_registry(path, [])
            self.assertEqual(registry["total"], 1)
            entry = registry["entries"][0]
            self.assertEqual(entry["kind"], KIND_HIDDEN_SHEET)
            self.assertEqual(entry["source_id"], "Hidden")
            self.assertEqual(entry["evidence"]["sheet_state"], "hidden")

    def test_xlsx_skipped_sheets_are_registered(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "skipped.xlsx"
            wb = Workbook()
            wb.active.title = "需求清单"
            wb.create_sheet("需求模版Release notes")
            wb.save(path)
            wb.close()

            registry = build_unextracted_registry(path, [])
            kinds = {e["kind"] for e in registry["entries"]}
            self.assertIn(KIND_SKIPPED_SHEET, kinds)
            skipped = [e for e in registry["entries"] if e["kind"] == KIND_SKIPPED_SHEET]
            self.assertEqual(skipped[0]["source_id"], "需求模版Release notes")


class WriteLoadTests(unittest.TestCase):
    def test_round_trip(self):
        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            registry = build_unextracted_registry(Path("/tmp/fake.docx"), [
                {"block_id": "BLK-000001", "text": "Page 12", "section_path": [], "noise": True},
            ])
            path = write_unextracted_registry(out_dir, registry, use_governed_path=False)
            self.assertTrue(path.exists())
            loaded = load_unextracted_registry(out_dir)
            self.assertIsNotNone(loaded)
            self.assertEqual(loaded["schema"], registry["schema"])
            self.assertEqual(loaded["total"], 1)

    def test_summarize_returns_counts(self):
        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            registry = build_unextracted_registry(Path("/tmp/fake.docx"), [
                {"block_id": "BLK-000001", "text": "Page 12", "section_path": [], "noise": True},
                {"block_id": "BLK-000002", "text": "Cover", "section_path": [], "noise": False, "doc_region": "front_matter"},
            ])
            write_unextracted_registry(out_dir, registry, use_governed_path=False)
            summary = summarize_unextracted_counts(out_dir)
            self.assertTrue(summary["available"])
            self.assertEqual(summary["total"], 2)
            self.assertEqual(summary["by_kind"][KIND_NOISE_BLOCK], 1)
            self.assertEqual(summary["by_kind"][KIND_FRONT_MATTER_BLOCK], 1)

    def test_clarification_entries_filter(self):
        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            registry = build_unextracted_registry(Path("/tmp/fake.docx"), [
                {"block_id": "BLK-000001", "text": "Page 12", "section_path": [], "noise": True},
                {"block_id": "BLK-000002", "text": "HiddenSheet", "section_path": [], "noise": False, "doc_region": "front_matter"},
                {"block_id": "BLK-000003", "text": "Textbox note", "section_path": [], "content_channel": "textbox"},
            ])
            write_unextracted_registry(out_dir, registry, use_governed_path=False)
            entries = collect_unextracted_clarification_entries(out_dir)
            kinds = {e["kind"] for e in entries}
            self.assertIn(KIND_NOISE_BLOCK, kinds)
            self.assertIn(KIND_FRONT_MATTER_BLOCK, kinds)
            self.assertIn(KIND_TEXTBOX_CHANNEL, kinds)
            self.assertNotIn(KIND_HIDDEN_SHEET, kinds)


if __name__ == "__main__":
    unittest.main()
