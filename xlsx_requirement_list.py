"""Excel 需求清单型分流（A6）。

识别「需求清单」型 xlsx（变体表头 + 复用 requirements_analysis_template 词表机制），
将行映射为需求条目；校验失败时退回既有 table 处理。默认关。
同时把历史 xlsx 聚合为基本需求库候选（base_library_candidates.jsonl）。
"""
from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from requirements_analysis_template import _normalize_cell_value, _SKIPPED_SHEET_TITLES


XLSX_REQUIREMENT_LIST_SWITCH = "RATOMIZER_XLSX_REQUIREMENT_LIST"
XLSX_REQUIREMENT_LIST_VERSION = "xlsx-requirement-list-v1"
BASE_LIBRARY_CANDIDATES_FILE = "base_library_candidates.jsonl"

# 变体表头词典：需求清单常见列名（中英文、缩写、同义词）
_HEADER_ALIASES: dict[str, set[str]] = {
    "title": {"标题", "需求标题", "title", "requirement title", "需求名", "name"},
    "description": {"描述", "需求描述", "description", "requirement description", "正文", "需求正文", "question"},
    "module": {"模块", "module", "所属模块", "功能域", "domain"},
    "submodule": {"子模块", "submodule", "子系统", "subsystem"},
    "source_quote": {"原文引用", "source quote", "quote", "引用", "来源"},
    "acceptance": {"验收标准", "acceptance", "验收", "criteria"},
}


def _detect_header_columns(row: tuple[Any, ...]) -> dict[str, int]:
    """根据变体表头词典识别列位置；返回 {field: column_index}。"""
    mapping: dict[str, int] = {}
    for column_index, value in enumerate(row):
        normalized = _normalize_cell_value(value).lower()
        if not normalized:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if normalized in aliases:
                mapping[field] = column_index
                break
    return mapping


def _is_requirement_list_sheet(worksheet: Worksheet) -> bool:
    """启发式判断 sheet 是否为需求清单：标题不在跳过清单，且前 5 行能识别出描述列。"""
    title = str(worksheet.title or "").strip()
    if not title or title in _SKIPPED_SHEET_TITLES or title.endswith("列表"):
        return False
    for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
        mapping = _detect_header_columns(row)
        if "description" in mapping or "title" in mapping:
            return True
    return False


def _row_to_requirement(
    row: tuple[Any, ...],
    mapping: dict[str, int],
    *,
    module: str = "",
    submodule: str = "",
) -> dict[str, Any] | None:
    """把一行映射为需求候选；无标题/描述时返回 None。"""
    title_col = mapping.get("title")
    desc_col = mapping.get("description")
    title = _normalize_cell_value(row[title_col]) if title_col is not None and title_col < len(row) else ""
    description = _normalize_cell_value(row[desc_col]) if desc_col is not None and desc_col < len(row) else ""
    if not title and not description:
        return None
    if not title:
        title = description[:60]
    if not description:
        description = title
    return {
        "title": title,
        "description": description,
        "module": module or (_normalize_cell_value(row[mapping["module"]]) if mapping.get("module") is not None and mapping["module"] < len(row) else ""),
        "submodule": submodule or (_normalize_cell_value(row[mapping["submodule"]]) if mapping.get("submodule") is not None and mapping["submodule"] < len(row) else ""),
        "source_quote": _normalize_cell_value(row[mapping["source_quote"]]) if mapping.get("source_quote") is not None and mapping["source_quote"] < len(row) else "",
        "acceptance_criteria": ([_normalize_cell_value(row[mapping["acceptance"]])] if mapping.get("acceptance") is not None and mapping["acceptance"] < len(row) and _normalize_cell_value(row[mapping["acceptance"]]) else []),
        "candidate_version": XLSX_REQUIREMENT_LIST_VERSION,
    }


def extract_requirement_list_candidates(input_path: Path) -> list[dict[str, Any]]:
    """从 xlsx 抽取需求清单候选；无法识别时返回空列表（调用方退回 table 路径）。"""
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    candidates: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            if not _is_requirement_list_sheet(worksheet):
                continue
            module = str(worksheet.title or "").strip()
            header_mapping: dict[str, int] | None = None
            for row in worksheet.iter_rows(values_only=True):
                if header_mapping is None:
                    mapping = _detect_header_columns(row)
                    if "description" in mapping or "title" in mapping:
                        header_mapping = mapping
                    continue
                req = _row_to_requirement(row, header_mapping, module=module)
                if req:
                    candidates.append(req)
    finally:
        workbook.close()
    return candidates


def requirement_list_enabled() -> bool:
    return os.environ.get(XLSX_REQUIREMENT_LIST_SWITCH, "").strip().lower() in {"1", "true", "yes", "on"}


def write_base_library_candidates(out_dir: Path, candidates: list[dict[str, Any]]) -> Path:
    """把历史 xlsx 聚合候选写入 base_library_candidates.jsonl（专家筛选是后续人工动作）。"""
    out_dir = Path(out_dir).expanduser().resolve()
    path = out_dir / BASE_LIBRARY_CANDIDATES_FILE
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for candidate in candidates:
            handle.write(json.dumps(candidate, ensure_ascii=False) + "\n")
    return path
