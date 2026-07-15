"""批次一（0714 整体 review 落地）回归：

- E2 冻结归属注入：富化 prompt 携带上游冻结 ownership,模型只写不判;归属变化令缓存 key 变化。
- E1a 部分降级上屏：enrich_degraded>0 时 note 必须出现（此前只有全灭才提示）。
（S1 合批 / S2 自适应限流 / E1b 遗漏候选 的回归在各自实现后追加于此文件。）
"""
from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from requirements_analysis_agent import build_analysis_prompt, slim_vocabulary


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + "\n",
                    encoding="utf-8")


class OwnershipInjectionPromptTests(unittest.TestCase):
    def test_prompt_instructs_frozen_ownership(self) -> None:
        prompt = build_analysis_prompt(
            [{"ai_req_id": "AI-1", "module": "时钟", "ownership": "software"}],
            slim_vocabulary({"modules": ["时钟"]}, "时钟"))
        user = prompt["user"]
        self.assertIn("冻结", user)
        self.assertIn("绝不改判", user)
        self.assertIn("给定归属", user)          # ownership_reason 字段说明改为解释给定归属
        self.assertIn('"ownership": "software"', user)   # 冻结值随需求 JSON 注入


class OwnershipInjectionEnrichTests(unittest.TestCase):
    SOURCE = {"ai_req_id": "AI-1", "module": "计量",
              "title": "数据存储",
              "description": "The meter shall store data.",
              "source_quote": "The meter shall store data."}
    VOCAB = {"modules": ["计量"], "submodules_by_module": {"计量": []}}

    def _item(self, ownership: str) -> dict:
        return {"analysis_id": "SRA-001", "ownership": ownership,
                "ownership_reason": "Matched software rule term: store",
                "ownership_source": "rule"}

    def test_prompt_req_carries_frozen_ownership(self) -> None:
        from requirements_analysis import _llm_enrich_item
        captured: list[str] = []

        def chat(system: str, user: str) -> dict:
            captured.append(user)
            return {"items": [{"software_requirement_text": "存储数据的软件逻辑。",
                               "ownership": "software"}]}

        ok, _ = _llm_enrich_item(self._item("software"), self.SOURCE, self.VOCAB,
                                 chat, {}, "m")
        self.assertTrue(ok)
        self.assertIn('"ownership": "software"', captured[0])

    def test_ownership_change_invalidates_cache_key(self) -> None:
        from requirements_analysis import _llm_enrich_item
        calls: list[str] = []
        cache: dict = {}

        def chat(system: str, user: str) -> dict:
            calls.append(user)
            return {"items": [{"software_requirement_text": "存储数据的软件逻辑。"}]}

        _llm_enrich_item(self._item("software"), self.SOURCE, self.VOCAB, chat, cache, "m")
        _llm_enrich_item(self._item("co_design"), self.SOURCE, self.VOCAB, chat, cache, "m")
        self.assertEqual(len(calls), 2)          # 归属不同 → key 不同 → 各自真调
        _llm_enrich_item(self._item("software"), self.SOURCE, self.VOCAB, chat, cache, "m")
        self.assertEqual(len(calls), 2)          # 归属相同 → 命中缓存零调用

    def test_echoed_frozen_ownership_reason_adopted(self) -> None:
        from requirements_analysis import _llm_enrich_item

        def chat(system: str, user: str) -> dict:
            return {"items": [{"software_requirement_text": "存储数据的软件逻辑。",
                               "ownership": "software",
                               "ownership_reason": "存储与协议逻辑属软件职责"}]}

        item = self._item("software")
        ok, _ = _llm_enrich_item(item, self.SOURCE, self.VOCAB, chat, {}, "m")
        self.assertTrue(ok)
        self.assertEqual(item["ownership_reason"], "存储与协议逻辑属软件职责")
        self.assertEqual(item["ownership_reason_source"], "llm")


class PartialDegradeNoteTests(unittest.TestCase):
    def _seed(self, out: Path) -> None:
        _write_jsonl(out / "ai_requirements.jsonl", [
            {"ai_req_id": "AI-1", "title": "数据存储", "module": "数据存储",
             "description": "The meter shall store data.",
             "source_quote": "The meter shall store data."},
            {"ai_req_id": "AI-2", "title": "时钟同步", "module": "时钟",
             "description": "The meter shall sync clock.",
             "source_quote": "The meter shall sync clock."},
        ])

    def test_partial_degrade_sets_note(self) -> None:
        from requirements_analysis import run_requirements_analysis

        def chat(system: str, user: str) -> dict:
            if "store data" in user:
                return {"items": [{"software_requirement_text": "存储数据的软件逻辑。"}]}
            raise RuntimeError("boom")   # 第二条调用失败 → 单条降级

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out)
            result = run_requirements_analysis(out, route="openai_compatible", chat=chat)
            self.assertEqual(result["enriched"], 1)
            self.assertEqual(result["enrich_degraded"], 1)
            self.assertIn("部分降级", result.get("note", ""))
            self.assertIn("1/2", result["note"])

    def test_full_success_has_no_note(self) -> None:
        from requirements_analysis import run_requirements_analysis

        def chat(system: str, user: str) -> dict:
            return {"items": [{"software_requirement_text": "对应软件逻辑成文。"}]}

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out)
            result = run_requirements_analysis(out, route="openai_compatible", chat=chat)
            self.assertEqual(result["enrich_degraded"], 0)
            self.assertNotIn("note", result)


class CoverageGapClarificationTests(unittest.TestCase):
    """E1b：覆盖缺口的遗漏候选进澄清清单（独立档,不进就绪门,不混入模型自报）。"""

    def _seed(self, out: Path, consistency: dict) -> None:
        _write_jsonl(out / "ai_requirements.jsonl", [
            {"ai_req_id": "AI-1", "title": "需求", "source_quote": "q"}])
        (out / "consistency_report.json").write_text(
            json.dumps(consistency, ensure_ascii=False), encoding="utf-8")

    def test_uncovered_samples_become_gap_entries(self) -> None:
        import clarification_report as cr
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out, {"coverage": {
                "measured": True, "uncovered_count": 2,
                "uncovered_samples": [
                    {"block_id": "BLK-1", "section": "4.5", "text": "The XDEV shall close the valve."},
                    {"block_id": "BLK-2", "section": "", "text": "All meters shall meet this."},
                ]}})
            entries = cr.collect_questions(out)
        gap = [e for e in entries if e.get("tier") == cr.TIER_GAP]
        self.assertEqual(len(gap), 2)
        self.assertEqual(gap[0]["source_id"], "BLK-1")     # 溯源可回链批注视图
        self.assertEqual(gap[0]["section"], "4.5")
        self.assertIn("XDEV shall close", gap[0]["quote"])
        self.assertTrue(all(e["audience"] == cr.AUDIENCE_INTERNAL for e in gap))

    def test_legacy_plain_string_samples_tolerated(self) -> None:
        import clarification_report as cr
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out, {"coverage": {"measured": True, "uncovered_count": 1,
                                          "uncovered_samples": ["Legacy uncovered text."]}})
            entries = cr.collect_questions(out)
        gap = [e for e in entries if e.get("tier") == cr.TIER_GAP]
        self.assertEqual(len(gap), 1)
        self.assertEqual(gap[0]["quote"], "Legacy uncovered text.")

    def test_gap_entries_do_not_trip_readiness_gate(self) -> None:
        import clarification_report as cr
        samples = [{"block_id": f"B{i}", "section": "", "text": f"Uncovered requirement {i}."}
                   for i in range(40)]   # 超过 READY_MAX_QUESTIONS=30
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out, {"coverage": {"measured": True, "uncovered_count": 40,
                                          "uncovered_samples": samples}})
            report = cr.run_report(out)
        self.assertEqual(report["questions"], 0)                 # 必答口径不含遗漏候选
        self.assertEqual(report["coverage_candidates"], 40)
        self.assertEqual(report["readiness"]["verdict"], "READY")

    def test_sample_cap_overflow_leaves_trace(self) -> None:
        import clarification_report as cr
        samples = [{"block_id": "B1", "section": "", "text": "Uncovered one."}]
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out, {"coverage": {"measured": True, "uncovered_count": 113,
                                          "uncovered_samples": samples}})
            entries = cr.collect_questions(out)
        overflow = [e for e in entries if e["signal"] == "consistency:uncovered_overflow"]
        self.assertEqual(len(overflow), 1)
        self.assertIn("112", overflow[0]["question"])            # 无声截断禁令：超上限必留痕

    def test_gap_sheet_written_and_hard_sheets_unpolluted(self) -> None:
        import clarification_report as cr
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out, {"coverage": {"measured": True, "uncovered_count": 1,
                                          "uncovered_samples": [
                                              {"block_id": "B1", "section": "4.5",
                                               "text": "The XDEV shall close the valve."}]}})
            cr.run_report(out)
            wb = load_workbook(out / cr.REPORT_XLSX, read_only=True)
            try:
                self.assertIn("遗漏候选(内部核对)", wb.sheetnames)
                gap_rows = list(wb["遗漏候选(内部核对)"].iter_rows(min_row=2, values_only=True))
                self.assertEqual(len(gap_rows), 1)
                self.assertIn("XDEV shall close", str(gap_rows[0][4]))
                self.assertEqual(list(wb["必答-问客户"].iter_rows(min_row=2, values_only=True)), [])
            finally:
                wb.close()


class BatchMappingTests(unittest.TestCase):
    """S1：合批响应槽位映射——宁缺勿错（缺槽走单条回退,绝不张冠李戴）。"""

    def test_slot_echo_mapping(self) -> None:
        from requirements_analysis import _map_batch_items
        mapped = _map_batch_items({"items": [
            {"enrich_slot": 1, "software_requirement_text": "乙"},
            {"enrich_slot": 0, "software_requirement_text": "甲"},
        ]}, 2)
        self.assertEqual(mapped[0]["software_requirement_text"], "甲")
        self.assertEqual(mapped[1]["software_requirement_text"], "乙")

    def test_index_alignment_only_when_counts_match_and_no_slots(self) -> None:
        from requirements_analysis import _map_batch_items
        mapped = _map_batch_items({"items": [{"a": 1}, {"b": 2}]}, 2)
        self.assertEqual(len(mapped), 2)                     # 全缺槽+数量吻合 → 按序
        self.assertEqual(_map_batch_items({"items": [{"a": 1}]}, 2), {})   # 数量不吻合 → 空
        self.assertEqual(_map_batch_items({"items": "bad"}, 2), {})

    def test_partial_slots_do_not_index_align(self) -> None:
        from requirements_analysis import _map_batch_items
        mapped = _map_batch_items({"items": [
            {"enrich_slot": 0, "text": "有槽"}, {"text": "缺槽"},
        ]}, 2)
        self.assertEqual(list(mapped.keys()), [0])           # 缺槽条目不冒充槽 1


class BatchEnrichmentTests(unittest.TestCase):
    """S1：同模块合批富化——调用数下降,护栏逐条不放宽,失败回退单条。"""

    def _seed(self, out: Path, n: int, module: str = "事件记录") -> None:
        _write_jsonl(out / "ai_requirements.jsonl", [
            {"ai_req_id": f"AI-{i}", "title": f"任务{i}",
             "description": f"The meter shall do task {i}.",
             "source_quote": f"shall do task {i}",
             "source_block_ids": [f"B-{i}"], "module": module}
            for i in range(n)
        ])

    def test_same_module_jobs_are_batched(self) -> None:
        from requirements_analysis import run_requirements_analysis
        calls: list[str] = []

        def chat(system: str, user: str) -> dict:
            calls.append(user)
            data = json.loads(user.split("需求 JSON:")[-1].strip())
            return {"items": [
                {"enrich_slot": entry.get("enrich_slot", 0),
                 "source_requirement_ids": [entry.get("ai_req_id")],
                 "software_requirement_text":
                     f"实现任务 {entry['description'].split('task ')[1].rstrip('.')} 的处理逻辑。"}
                for entry in data
            ]}

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out, 5)
            result = run_requirements_analysis(out, route="openai_compatible", chat=chat)
            self.assertEqual(result["enriched"], 5)
            self.assertEqual(len(calls), 2)                  # 4 条一批 + 1 条单发
            self.assertIn("enrich_slot", calls[0] + calls[1])
            payload = json.loads((out / "engineering_analysis.json").read_text(encoding="utf-8"))
            for item in payload["items"]:
                idx = item["source_requirement_ids"][0].split("-")[1]
                self.assertEqual(item["software_requirement_text"],
                                 f"实现任务 {idx} 的处理逻辑。")    # 槽位映射不串条

    def test_batch_env_one_restores_per_item_calls(self) -> None:
        import os
        from unittest.mock import patch

        from requirements_analysis import run_requirements_analysis
        calls: list[str] = []

        def chat(system: str, user: str) -> dict:
            calls.append(user)
            return {"items": [{"software_requirement_text": "对应处理逻辑成文。"}]}

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out, 3)
            with patch.dict(os.environ, {"RATOMIZER_ANALYZE_BATCH": "1"}):
                result = run_requirements_analysis(out, route="openai_compatible", chat=chat)
            self.assertEqual(result["enriched"], 3)
            self.assertEqual(len(calls), 3)                  # 逐条模式恢复
            self.assertNotIn("enrich_slot", "".join(calls))

    def test_batch_failure_falls_back_to_single_calls(self) -> None:
        from requirements_analysis import run_requirements_analysis
        calls: list[str] = []

        def chat(system: str, user: str) -> dict:
            calls.append(user)
            if "enrich_slot" in user:
                raise RuntimeError("batch boom")             # 整批失败
            return {"items": [{"software_requirement_text": "单条回退成文。"}]}

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out, 4)
            result = run_requirements_analysis(out, route="openai_compatible", chat=chat)
            self.assertEqual(result["enriched"], 4)          # 回退后全部成功
            self.assertEqual(result["enrich_degraded"], 0)
            self.assertEqual(len(calls), 5)                  # 1 批失败 + 4 单条

    def test_cross_item_code_borrowing_rejected_per_item(self) -> None:
        """合批不放宽护栏：B 条借用 A 条条款原文里的 OBIS → 仅 B 被硬拒。"""
        from requirements_analysis import _llm_enrich_batch

        ctx_a = {"template_refs": "", "exemplars": "", "answers": "",
                 "doc_context": "", "siblings": "",
                 "section_context": "Clause A: use 0-0:96.1.0.255 here."}
        ctx_b = dict(ctx_a, section_context="Clause B: no codes at all.")
        req_a = {"ai_req_id": "AI-A", "module": "安全",
                 "description": "Requirement alpha.", "source_quote": "Requirement alpha."}
        req_b = {"ai_req_id": "AI-B", "module": "安全",
                 "description": "Requirement beta.", "source_quote": "Requirement beta."}
        item_a = {"analysis_id": "SRA-001", "ownership": "software",
                  "ownership_reason": "rule", "ownership_source": "rule"}
        item_b = {"analysis_id": "SRA-002", "ownership": "software",
                  "ownership_reason": "rule", "ownership_source": "rule"}

        def chat(system: str, user: str) -> dict:
            return {"items": [
                {"enrich_slot": 0, "software_requirement_text": "读取对象 0-0:96.1.0.255。"},
                {"enrich_slot": 1, "software_requirement_text": "读取对象 0-0:96.1.0.255。"},
            ]}

        results = _llm_enrich_batch(
            [(item_a, req_a, ctx_a, "software"), (item_b, req_b, ctx_b, "software")],
            {"modules": ["安全"]}, chat, {}, "m")
        by_id = {r[0]["analysis_id"]: r for r in results}
        self.assertTrue(by_id["SRA-001"][1])                 # A 有据 → 采纳
        self.assertFalse(by_id["SRA-002"][1])                # B 跨条借码 → 硬拒降级
        self.assertTrue(any("fabricated" in i or "编造" in i for i in by_id["SRA-002"][2]))

    def test_hardware_jobs_batched_with_slot_echo(self) -> None:
        from requirements_analysis import run_requirements_analysis
        calls: list[str] = []

        def chat(system: str, user: str) -> dict:
            calls.append(user)
            data = json.loads(user.split("需求 JSON:")[-1].strip())
            return {"items": [
                {"enrich_slot": entry.get("enrich_slot", 0),
                 "hardware_translation": f"硬件条目{'甲乙丙'[int(entry.get('enrich_slot', 0))]}的中文说明",
                 "ownership_reason": "外壳与阀门为机械部件"}
                for entry in data
            ]}

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            _write_jsonl(out / "ai_requirements.jsonl", [
                {"ai_req_id": f"AI-{i}", "title": f"外壳{i}",
                 "description": f"The meter case variant {chr(65 + i)} shall be sealed metal enclosure.",
                 "source_quote": "sealed metal enclosure",
                 "ownership_override": "hardware", "module": "机械结构"}
                for i in range(3)
            ])
            result = run_requirements_analysis(out, route="openai_compatible", chat=chat)
            self.assertEqual(result["enriched"], 3)
            self.assertEqual(len(calls), 1)                  # 3 条硬件一批
            payload = json.loads((out / "engineering_analysis.json").read_text(encoding="utf-8"))
            translations = {item["hardware_translation"] for item in payload["items"]}
            self.assertEqual(len(translations), 3)           # 槽位映射不串条


class SpecEnrichBatchTests(unittest.TestCase):
    """S1b：装配描述富化合批（无蓝皮书条目成批,带条款条目恒单发,护栏逐条）。"""

    def _reqs(self, n: int) -> list[dict]:
        return [{"id": f"REQ-{i}", "title": f"行为 {i}",
                 "description": f"template desc {i}",
                 "source_quote": f"The meter shall behave {i}.", "labels": ["事件记录"]}
                for i in range(n)]

    def test_plain_items_batched_after_fast_fail_sample(self) -> None:
        import spec_enrich
        from unittest.mock import patch
        from llm_client import LLMClientConfig

        calls: list[str] = []

        def fake_chat(config, system, user):
            calls.append(system)
            if system == spec_enrich.SYSTEM_PROMPT_BATCH:
                entries = json.loads(user)
                return {"items": [{"enrich_slot": e["enrich_slot"],
                                   "description": f"改写后的行为描述（{e['title']}）"}
                                  for e in entries]}
            return {"description": "改写后的单条描述"}

        with tempfile.TemporaryDirectory() as td, \
                patch("spec_enrich.chat_json", side_effect=fake_chat):
            reqs = self._reqs(10)
            enriched, rejected, failed = spec_enrich.enrich_descriptions(
                reqs, config=LLMClientConfig(base_url="http://x", model="m"),
                cache_path=Path(td) / "c.jsonl", concurrency=2)
        self.assertEqual((enriched, rejected, failed), (10, 0, 0))
        batch_calls = [s for s in calls if s == spec_enrich.SYSTEM_PROMPT_BATCH]
        single_calls = [s for s in calls if s == spec_enrich.SYSTEM_PROMPT]
        self.assertEqual(len(single_calls), 5)           # 快速失败探测样本保持单发
        self.assertEqual(len(batch_calls), 1)            # 其余 5 条一批（batch=6 装得下）
        for req in reqs[5:]:
            self.assertIn("改写后的行为描述", req["description"])   # 槽位映射逐条落对

    def test_batch_drift_rejected_per_item(self) -> None:
        import spec_enrich
        from unittest.mock import patch
        from llm_client import LLMClientConfig

        def fake_chat(config, system, user):
            if system == spec_enrich.SYSTEM_PROMPT_BATCH:
                entries = json.loads(user)
                items = []
                for e in entries:
                    desc = (f"含无据数字 99999 的描述" if e["title"].endswith("5")
                            else f"改写后的行为描述（{e['title']}）")
                    items.append({"enrich_slot": e["enrich_slot"], "description": desc})
                return {"items": items}
            return {"description": "改写后的单条描述"}

        with tempfile.TemporaryDirectory() as td, \
                patch("spec_enrich.chat_json", side_effect=fake_chat):
            reqs = self._reqs(8)
            enriched, rejected, failed = spec_enrich.enrich_descriptions(
                reqs, config=LLMClientConfig(base_url="http://x", model="m"),
                cache_path=Path(td) / "c.jsonl", concurrency=1)
        self.assertEqual(failed, 0)
        self.assertEqual(rejected, 1)                    # 只有借数那条被拒
        self.assertEqual(enriched, 7)
        self.assertEqual(reqs[5]["description"], "template desc 5")   # 被拒保留模板
        self.assertIn("漂移", str(reqs[5].get("notes")))

    def test_missing_slot_falls_back_to_single_call(self) -> None:
        import spec_enrich
        from unittest.mock import patch
        from llm_client import LLMClientConfig

        calls: list[str] = []

        def fake_chat(config, system, user):
            calls.append(system)
            if system == spec_enrich.SYSTEM_PROMPT_BATCH:
                entries = json.loads(user)
                return {"items": [{"enrich_slot": e["enrich_slot"],
                                   "description": f"改写后的行为描述（{e['title']}）"}
                                  for e in entries[:-1]]}   # 漏最后一槽
            return {"description": "单条回退的改写描述"}

        with tempfile.TemporaryDirectory() as td, \
                patch("spec_enrich.chat_json", side_effect=fake_chat):
            reqs = self._reqs(8)
            enriched, rejected, failed = spec_enrich.enrich_descriptions(
                reqs, config=LLMClientConfig(base_url="http://x", model="m"),
                cache_path=Path(td) / "c.jsonl", concurrency=1)
        self.assertEqual((enriched, rejected, failed), (8, 0, 0))
        self.assertEqual(reqs[7]["description"], "单条回退的改写描述")
        self.assertEqual(len([s for s in calls if s == spec_enrich.SYSTEM_PROMPT]), 6)  # 5 样本+1 回退

    def test_batch_env_one_restores_per_item(self) -> None:
        import os
        import spec_enrich
        from unittest.mock import patch
        from llm_client import LLMClientConfig

        calls: list[str] = []

        def fake_chat(config, system, user):
            calls.append(system)
            return {"description": "改写后的单条描述"}

        with tempfile.TemporaryDirectory() as td, \
                patch("spec_enrich.chat_json", side_effect=fake_chat), \
                patch.dict(os.environ, {"RATOMIZER_ENRICH_BATCH": "1"}):
            reqs = self._reqs(8)
            enriched, rejected, failed = spec_enrich.enrich_descriptions(
                reqs, config=LLMClientConfig(base_url="http://x", model="m"),
                cache_path=Path(td) / "c.jsonl", concurrency=1)
        self.assertEqual((enriched, rejected, failed), (8, 0, 0))
        self.assertEqual(calls.count(spec_enrich.SYSTEM_PROMPT), 8)   # 全部单发
        self.assertNotIn(spec_enrich.SYSTEM_PROMPT_BATCH, calls)


class CoverageGapMarkdownTests(unittest.TestCase):
    def test_markdown_renders_gap_section(self) -> None:
        import clarification_report as cr
        entries = [
            cr._entry(cr.CAT_MISSING, "该段疑似含需求但未被覆盖", quote="The XDEV shall act.",
                      source_id="B1", signal="consistency:uncovered",
                      tier=cr.TIER_GAP, audience=cr.AUDIENCE_INTERNAL),
        ]
        md = cr.render_markdown(entries, {"verdict": "READY", "reasons": [], "questions": 0})
        self.assertIn("遗漏候选（1）", md)
        self.assertIn("遗漏候选 1 条", md)
        self.assertIn("The XDEV shall act.", md)


if __name__ == "__main__":
    unittest.main()
