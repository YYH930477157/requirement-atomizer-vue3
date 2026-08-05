from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from requirement_schema import (
    MANUAL_MODULE_DEFAULT,
    PROVENANCE_MANUAL,
    requirement_identity,
    verification_excel_columns,
)
from requirements_analysis_rules import classify_ownership
from requirements_analysis_schema import (
    OWNERSHIP_CO_DESIGN,
    OWNERSHIP_HARDWARE,
    OWNERSHIP_SOFTWARE,
)
from review_state import read_manual_requirements, read_verification_states
from text_normalize import formula_safe

# WP2 兜底渲染（2026-07-23 用户裁定）：被标"待澄清"的字段若留有原始候选
# （clarify_fallback），透出时带上完整标注——读者一眼知道这不是可实施依据。
CLARIFY_MARK = "待澄清"
_CLARIFY_FALLBACK_LABEL = "原始候选（未经依据校验，仅供参考，不得作为实现依据）"


def clarify_display_text(item: dict[str, Any], field: str) -> str:
    """正文类字段的透出文本：标了待澄清且有原始候选 → 待澄清标注 + 兜底候选；
    否则原样返回字段值（非待澄清字段不受影响）。"""
    value = str(item.get(field) or "")
    if value != CLARIFY_MARK:
        return value
    fallback = (item.get("clarify_fallback") or {}).get(field)
    if not fallback:
        return value
    return f"{CLARIFY_MARK}（未经依据校验，需专家核补）\n{_CLARIFY_FALLBACK_LABEL}：{fallback}"


def _fallback_lines(item: dict[str, Any], field: str, prefix: str = "") -> list[str]:
    """列表字段的兜底行：整列被标待澄清且有原始候选 → 标注行 + 逐条带标注的候选。"""
    value = item.get(field)
    if value != [CLARIFY_MARK]:
        return []
    fallback = (item.get("clarify_fallback") or {}).get(field)
    if not isinstance(fallback, list) or not fallback:
        return []
    lines = [f"{prefix}{CLARIFY_MARK}（未经依据校验，需专家核补）"]
    lines.extend(f"{prefix}{_CLARIFY_FALLBACK_LABEL}：{entry}" for entry in fallback)
    return lines


HEADERS = [
    "关闭",
    "序号",
    "子模块",
    "描述",
    "需求模版",
    "需求",
    "说明、示例、注意事项",
    "是否客户需求",
    "客户需求章节",
    "驱动/硬件相关",
    "项目负责人确认",
    "测试负责人确认",
    "研发测试确认",
    "功能是否实现",
    "测试用例号",
    "测试是否完成",
]

_SOFTWARE_OWNERSHIPS = {OWNERSHIP_SOFTWARE, OWNERSHIP_CO_DESIGN}
_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")
# openpyxl 会对控制字符抛 IllegalCharacterError（PDF 文本层常带 \x0b 等）——写入前剥离
_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _safe_cell(value: Any) -> Any:
    """单元格防护：控制字符剥离 + 公式注入中和（formula_safe，与 spec_excel 同一纪律）。

    没有这层，'=HYPERLINK(...)' 开头的抽取文本会以活公式落进交付给研发的工作簿——
    这是 spec-data-integrity 阶段修掉的同类漏洞，此处必须同样设防。
    """
    if isinstance(value, str):
        value = _ILLEGAL_XLSX_CHARS.sub("", value)
    return formula_safe(value)


def _shape_manual_item(record: dict[str, Any]) -> dict[str, Any]:
    """把手工需求记录塑形为分析项形态（与 _excel_row/_notes_text 字段对齐）。

    归属走确定性 classify_ownership（按其 objective/behaviors 文本）；显式 ownership_override
    走 reviewer_override 通道覆盖。source_quote/source_section 留空——追溯列以空明示无文档来源。
    """
    objective = str(record.get("objective") or record.get("title") or "").strip()
    behaviors = [str(value).strip() for value in (record.get("behaviors") or []) if str(value).strip()]
    classified = classify_ownership({
        "title": record.get("title") or objective,
        "description": record.get("description") or objective,
        "requirement": objective,
        "module": record.get("module") or MANUAL_MODULE_DEFAULT,
        "source_quote": "",
        "labels": record.get("labels") or [],
    })
    item: dict[str, Any] = dict(classified)
    ownership_override = str(record.get("ownership_override") or "").strip()
    if ownership_override in (OWNERSHIP_SOFTWARE, OWNERSHIP_CO_DESIGN, OWNERSHIP_HARDWARE):
        item["ownership"] = ownership_override
        item["ownership_source"] = "reviewer_override"
    item.update({
        "analysis_id": str(record.get("functional_requirement_id") or ""),
        "functional_requirement_id": str(record.get("functional_requirement_id") or ""),
        "module": str(record.get("module") or MANUAL_MODULE_DEFAULT),
        "submodule": str(record.get("module") or MANUAL_MODULE_DEFAULT),
        "description": objective or str(record.get("title") or ""),
        "requirement": objective,
        "software_requirement_text": objective or str(record.get("title") or ""),
        "source_section": "",
        "source_requirement_ids": [],
        "source_quote": "",
        "objective": objective,
        "behaviors": behaviors,
        "labels": list(record.get("labels") or []),
        "notes": [],
        "manual_actor": str(record.get("manual_actor") or ""),
        "_manual": True,
    })
    return item


def _enrich_items_for_export(items: list[dict[str, Any]], out_dir: Path) -> list[dict[str, Any]]:
    """合并 verification 覆盖 + 手工需求，返回新列表（不修改调用方 items，避免污染 engineering_analysis.json）。

    每条带 _trace_id（需求追溯ID，回灌按行定位用）与 verification 子对象。
    """
    root = Path(out_dir).expanduser().resolve()
    verification_overlay = read_verification_states(root)
    enriched: list[dict[str, Any]] = []
    for item in items:
        copy = dict(item)
        trace_id = requirement_identity(item)
        copy["_trace_id"] = trace_id
        record = verification_overlay.get(trace_id)
        copy["verification"] = record.get("verification") if isinstance(record, dict) else None
        enriched.append(copy)
    # 手工需求走完全相同下游：归属/澄清/导出/状态机
    for manual in read_manual_requirements(root):
        enriched.append(_shape_manual_item(manual))
    return enriched


def write_software_requirements_xlsx(items: list[dict[str, Any]], output_path: Path) -> Path:
    output_path = Path(output_path)
    wb = Workbook()
    wb.remove(wb.active)

    enriched = _enrich_items_for_export(items, output_path.parent)
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in enriched:
        if item.get("ownership") not in _SOFTWARE_OWNERSHIPS:
            continue
        module = str(item.get("module") or "unmapped")
        grouped.setdefault(module, []).append(item)

    if not grouped:
        grouped["软件需求"] = []

    used_titles: set[str] = set()
    for module, rows in grouped.items():
        ws = wb.create_sheet(_unique_sheet_title(module, used_titles))
        ws.append(HEADERS)
        _style_header(ws)
        ws.freeze_panes = "A2"
        for index, item in enumerate(rows, start=1):
            ws.append([_safe_cell(value) for value in _excel_row(index, item)])

    from xlsx_io import safe_save_workbook
    return safe_save_workbook(wb, output_path)
    return output_path


def _excel_row(index: int, item: dict[str, Any]) -> list[Any]:
    source_chapter = str(item.get("source_section") or "").strip()
    if not source_chapter:
        source_chapter = ",".join(str(value) for value in item.get("source_requirement_ids") or [])

    # 六列数据源从空字符串改读 verification 子对象（与导出 xlsx 六列一一对应、列位/样式不变）
    pm, tl, dt, implemented, test_case_ids, test_completed = verification_excel_columns(
        item.get("verification"))

    return [
        "",
        index,
        item.get("submodule") or "",
        item.get("description") or "",
        "",
        clarify_display_text(item, "software_requirement_text") or item.get("requirement") or "",
        _notes_text(item),
        "" if item.get("_manual") else "是",
        source_chapter,
        "是" if item.get("ownership") == OWNERSHIP_CO_DESIGN else "",
        pm,
        tl,
        dt,
        implemented,
        test_case_ids,
        test_completed,
    ]


def _notes_text(item: dict[str, Any]) -> str:
    notes: list[str] = []
    objective = str(item.get("objective") or "").strip()
    if objective:
        notes.append(f"功能目标：{objective}")
    notes.extend(f"功能行为：{value}" for value in item.get("behaviors") or [])
    role_labels = {
        "configure": "配置", "detect": "检测/触发", "execute": "执行/控制",
        "store": "存储/归档", "query": "查询/读取", "report": "上报/传输",
        "access": "权限/访问", "recover": "恢复/重试", "behavior": "其它行为",
    }
    for entry in item.get("lifecycle_behaviors") or []:
        if isinstance(entry, dict):
            role = role_labels.get(str(entry.get("role") or "behavior"), str(entry.get("role") or "其它行为"))
            behavior = str(entry.get("behavior") or "").strip()
            if behavior:
                notes.append(f"生命周期-{role}：{behavior}")
    source_modules = [str(value).strip() for value in item.get("source_modules") or [] if str(value).strip()]
    if len(source_modules) > 1:
        notes.append("跨模块来源：" + "、".join(source_modules))
    notes.extend(f"前置条件：{value}" for value in item.get("preconditions") or [])
    notes.extend(f"数据约束：{value}" for value in item.get("data_constraints") or [])
    for variant in item.get("variants") or []:
        if isinstance(variant, dict):
            name = str(variant.get("name") or "变体").strip()
            behavior = str(variant.get("behavior") or "").strip()
            notes.append(f"功能变体 {name}：{behavior}" if behavior else f"功能变体：{name}")
    notes.extend(f"异常处理：{value}" for value in item.get("exceptions") or [])
    related = [str(value).strip() for value in item.get("related_dlms_objects") or [] if str(value).strip()]
    if related:
        notes.append("关联 DLMS 对象：" + "、".join(related))
    # 硬件依赖落交付列（审计 P1-b：WP2 只写字段不透出，研发看不到依赖内容及其待澄清
    # 状态）——clarify_display_text 处理：待澄清时自动带"未经依据校验+原始候选"标注
    hardware_dependency = clarify_display_text(item, "hardware_dependency").strip()
    if hardware_dependency:
        notes.append(f"硬件依赖：{hardware_dependency}")
    notes.extend(f"待澄清冲突：{value}" for value in item.get("conflict_flags") or [])
    # 富化软标随交付物同行：编造数字/遗漏漂移必须在研发看的列里可见（2026-07-08 审计 B1）
    notes.extend(f"⚠ 富化待核：{value}" for value in item.get("enrichment_warnings") or [])
    # 参数表优先（数值是研发的命根子：粒径/成分/限值清单必须出现在交付物里，不能只留在中间产物）
    table = item.get("threshold_table")
    if isinstance(table, dict) and table.get("rows"):
        columns = [str(c) for c in table.get("columns") or []]
        if columns:
            notes.append("参数表：" + " | ".join(columns))
        for row in table.get("rows") or []:
            notes.append("  " + " | ".join(str(cell) for cell in (row if isinstance(row, list) else [row])))
    guidance_fallback = _fallback_lines(item, "developer_guidance")
    if guidance_fallback:
        notes.extend(guidance_fallback)
    else:
        notes.extend(str(value) for value in item.get("developer_guidance") or [])
    options_fallback = _fallback_lines(item, "design_options", "设计候选（非规范约束）：")
    if options_fallback:
        notes.extend(options_fallback)
    else:
        notes.extend(f"设计候选（非规范约束）：{value}" for value in item.get("design_options") or [])
    # 验收建议进交付列（此前 acceptance_criteria 不落 xlsx 任何列——富化白算）
    criteria_fallback = _fallback_lines(item, "acceptance_criteria", "验收建议：")
    if criteria_fallback:
        notes.extend(criteria_fallback)
    else:
        notes.extend(f"验收建议：{value}" for value in item.get("acceptance_criteria") or [])
    notes.extend(f"假设：{value}" for value in item.get("assumptions") or [])
    # 归属判定随行（真实反馈 2026-07-12：软件件全链路无"为什么"）
    ownership = str(item.get("ownership") or "").strip()
    reason = str(item.get("ownership_reason") or "").strip()
    if ownership and reason:
        labels = {"software": "软件", "hardware": "硬件", "co_design": "软硬件协同"}
        suffix = "，LLM 判定" if item.get("ownership_reason_source") == "llm" else ""
        if str(item.get("ownership_source") or "") == "reviewer_override":
            suffix += "；已人工覆盖"
        notes.append(f"归属判定：{labels.get(ownership, ownership)}（依据：{reason}{suffix}）")
    source_quote = str(item.get("source_quote") or "").strip()
    if source_quote:
        notes.append(f"原文：{source_quote}")
    notes.extend(f"待确认：{value}" for value in item.get("open_questions") or [])
    # WS4 追溯：需求追溯ID（回灌按行定位）+ 手工来源明示
    trace_id = str(item.get("_trace_id") or item.get("functional_requirement_id")
                   or item.get("analysis_id") or "").strip()
    if trace_id:
        notes.append(f"需求追溯ID：{trace_id}")
    if item.get("_manual"):
        notes.append("来源：手工录入（无文档来源）")
    return "\n".join(notes)


def _style_header(ws: Any) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _safe_sheet_title(value: str) -> str:
    title = _INVALID_SHEET_TITLE_CHARS.sub("", value).strip()[:31]
    return title or "软件需求"


def _unique_sheet_title(value: str, used_titles: set[str]) -> str:
    base = _safe_sheet_title(value)
    title_key = base.casefold()
    if title_key not in used_titles:
        used_titles.add(title_key)
        return base

    index = 2
    while True:
        suffix = f"~{index}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        title_key = title.casefold()
        if title_key not in used_titles:
            used_titles.add(title_key)
            return title
        index += 1
