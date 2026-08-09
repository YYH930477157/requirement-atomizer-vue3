from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import logging
import os
import re
import sys
import time
import uuid
from contextlib import contextmanager
from functools import wraps
from pathlib import Path
from threading import RLock
from typing import Any, Iterator, Sequence

import ai_extract
import claim_ledger
from agent_policy import AGENT_POLICY_VERSION
from assemble_spec import assemble
from atomize import run_atomizer_pipeline
from claim_artifacts import ClaimArtifactError
from engineering_composer import compose_engineering_requirements, write_engineering_requirements
from functional_synthesis import FUNCTIONAL_REQUIREMENTS, FUNCTIONAL_SYNTHESIS_VERSION, run_functional_synthesis
from export_requirements import export_requirements
from llm_pipeline import DEFAULT_DOMAIN_PACK_PATH, DEFAULT_PIPELINE_PATH, read_jsonl, run_review_pipeline
from requirement_kb.cli import default_kb_paths, package_root
from requirements_analysis import requirements_analysis_enrichment_enabled, run_requirements_analysis
from requirements_analysis_schema import normalize_ownership
from result_package import (
    ResultPackageCorrupt,
    ResultPackageError,
    ResultPackagePartialError,
    commit_analysis_completion,
    detect_result_layout,
    initialize_result_package,
    governed_artifact_path,
    load_result_package,
    package_artifact_path,
    package_root_for_analysis_root,
    publish_registered_deliverables,
    record_analysis_partial,
    record_analysis_failure,
    record_package_warning,
    resolve_analysis_root,
)
from spec_export import export_spec


LOGGER = logging.getLogger("requirement_atomizer")
ASSEMBLED_JSON = "dlms_cosem_spec_requirements.json"
PROGRESS_PREFIX = "__RATOMIZER_PROGRESS__"
BLUE_BOOK_INDEX_ENV = "RATOMIZER_BLUE_BOOK_INDEX"
REQUIREMENTS_ANALYSIS_OUTPUTS = [
    "engineering_analysis.json",
    "hardware_items.md",
    "co_design_items.md",
    "compliance_items.json",
    "compliance_items.md",
    "software_requirements.xlsx",
]


def _leased_pipeline_stage(stage: str):
    """Keep targeted extraction out while a downstream stage consumes AI requirements."""
    def decorate(func):
        @wraps(func)
        def wrapped(out_dir: Path, *args, **kwargs):
            from omission_actions import extraction_operation_lock

            root = Path(out_dir).expanduser().resolve()
            with extraction_operation_lock(root, operation=f"stage:{stage}"):
                if "ai_requirements.jsonl" in STAGE_INPUTS.get(stage, []):
                    from api_server import final_ai_requirements_are_stale

                    if final_ai_requirements_are_stale(root):
                        raise RuntimeError(
                            "AI extraction belongs to an older parsed document; rerun ai-extract first"
                        )
                before = stage_input_files_fingerprint(root, stage)
                payload = func(root, *args, **kwargs)
                after = stage_input_files_fingerprint(root, stage)
                if before != after:
                    raise RuntimeError(f"{stage} inputs changed while the stage was running")
                if isinstance(payload, dict):
                    payload["_input_files_fingerprint"] = after
                return payload
        return wrapped
    return decorate


def resolve_bundled_path(path: Path | None) -> Path | None:
    """把前端预设送来的相对资源路径解析成可用的绝对路径。

    前端预设送相对名（KB 用 "knowledge_bases/…json"、domain pack 用 "domain_packs/dlms_cosem"），
    dev 下 cwd=仓库根能命中，但打包后后端 cwd=resources/backend 命中不到。这里：绝对路径或 cwd 下
    已存在的相对路径原样保留（兼容 CLI 习惯）；否则按 package_root() 解析（dev=仓库根 /
    Electron=resources/，捆绑资源真正所在）。
    """
    if path is None:
        return None
    path = Path(path)
    if path.is_absolute() or path.exists():
        return path
    return package_root() / path


def resolve_kb_paths(kb_paths: list[Path] | None) -> list[Path]:
    """把显式 --kb 路径列表逐个解析（见 resolve_bundled_path）；None 时用 default_kb_paths()。"""
    if kb_paths is None:
        return default_kb_paths()
    return [resolve_bundled_path(path) for path in kb_paths]


def _budget_document_id(input_path: Path | None, out_dir: Path) -> str:
    """文档预算单稳定 document_id：输入文件名优先，其次输出目录名。"""
    if input_path is not None:
        stem = Path(input_path).stem.strip()
        if stem:
            return stem
    return (out_dir.name or "document").strip() or "document"


def _attach_budget_ledger_for_run(out_dir: Path, input_path: Path | None = None):
    """S1-1：``RATOMIZER_LLM_BUDGET=1`` 时创建文档预算单并 ``attach()`` 到 llm_client 钩子。

    开关未开返回 None（既有行为逐字节不变）。开启后所有 LLM HTTP 调用经钩子扣减，超额在
    ``intercept`` 事前拦截（调用方既有 stub catch 接管），落盘 ``llm_budget.json`` 供 cost-report。
    """
    try:
        from llm_budget import LLMBudgetLedger, budget_enabled
    except Exception:  # noqa: BLE001 — 预算模块不可用不得阻断主流程
        return None
    if not budget_enabled():
        return None
    try:
        ledger = LLMBudgetLedger.for_document(
            _budget_document_id(input_path, out_dir), out_dir=out_dir
        )
        ledger.attach()
        return ledger
    except Exception as exc:  # noqa: BLE001 — 预算单创建/挂载失败不得阻断主流程
        logging.getLogger("requirement_atomizer").warning("文档预算单 attach 失败：%s", exc)
        return None


def _detach_budget_ledger(ledger) -> None:
    """S1-1：退出时 save（落盘 cost-report 数据源）+ detach（llm_client 回到无预算行为）。"""
    if ledger is None:
        return
    try:
        ledger.save()
    except Exception as exc:  # noqa: BLE001
        logging.getLogger("requirement_atomizer").warning("文档预算单 save 失败：%s", exc)
    try:
        ledger.detach()
    except Exception:  # noqa: BLE001
        pass


@contextmanager
def _budget_stage(ledger, stage: str):
    """把一个流水线环节包进预算单的 stage 上下文（无活动预算单则空操作）。"""
    if ledger is None:
        yield
        return
    with ledger.enter_stage(stage):
        yield


def _attach_dual_track_proposer(route: str | None) -> bool:
    """S1-4：``RATOMIZER_TABLE_DUAL_TRACK=1`` 且有 openai_compatible route 时，挂双轨提议器。

    atomize 自身只做几何校验 + 假设派生 + 落盘（零 LLM）；提议器闭包捕获 config 调
    ``llm_table_understanding.propose_table_structure``。无 route（stub）/ 开关关 /
    config 解析失败 → 不挂，atomize 走确定性 ``analyze_table``（OFF 字节不变）。
    """
    try:
        from table_structure import dual_track_enabled
    except Exception:  # noqa: BLE001
        return False
    if not dual_track_enabled() or not route or route == "stub":
        return False
    try:
        from ai_extract import DEFAULT_PIPELINE_PATH, config_for_route
        from atomize import set_table_dual_track_proposer
        from llm_table_understanding import propose_table_structure

        config = config_for_route(route, DEFAULT_PIPELINE_PATH)
    except Exception:  # noqa: BLE001
        return False
    if config is None:
        return False

    def proposer(parsed_table, *, table_id="", block_id="", section_path=None):
        return propose_table_structure(parsed_table, config=config)

    try:
        set_table_dual_track_proposer(proposer)
        return True
    except Exception:  # noqa: BLE001
        return False


def _detach_dual_track_proposer() -> None:
    try:
        from atomize import clear_table_dual_track_proposer

        clear_table_dual_track_proposer()
    except Exception:  # noqa: BLE001
        pass


# 链内阶段名 → 预算单环节（仅 LLM 承载环节映射，确定性环节不进账本分摊）
_CHAIN_BUDGET_STAGES = {
    "ai-extract": "functional_extract",
    "assemble": "spec_enrich",
    "requirements-analysis": "analyze_enrich",
}


def run_pipeline_task(
    input_path: Path,
    out_dir: Path,
    *,
    skip_review: bool = False,
    llm_route: str | None = None,
    review_scope: str | None = None,
    llm_review_limit: int = 0,
    chunk_chars: int = 3500,
    kb_paths: list[Path] | None = None,
    domain_pack_dir: Path | None = None,
) -> dict[str, Any]:
    input_path = input_path.expanduser().resolve()
    out_dir = out_dir.expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    # S1-1：开启 RATOMIZER_LLM_BUDGET 时挂文档预算单（attach 后所有 LLM 调用经钩子扣减 +
    # 超额事前拦截；save 落盘 cost-report 数据源）。开关未开返回 None，行为逐字节不变。
    budget = _attach_budget_ledger_for_run(out_dir, input_path)
    _attach_dual_track_proposer(llm_route)  # S1-4：双轨开且有 route 时挂提议器（atomize 用）
    try:
        atomize_config = {
            "chunk_chars": chunk_chars,
            "kb_paths": [str(path) for path in resolve_kb_paths(kb_paths)],
            "domain_pack_dir": str(resolve_bundled_path(domain_pack_dir) or ""),
        }
        atomize_reused = stage_is_reusable(
            out_dir, "atomize", input_path=input_path, config=atomize_config)
        if atomize_reused:
            manifest = json.loads((out_dir / "manifest.json").read_text(encoding="utf-8"))
            manifest["resume_action"] = "skipped"
            update_run_manifest(out_dir, "atomize", "ok", outputs=STAGE_REQUIRED_OUTPUTS["atomize"],
                                action="skipped", input_path=input_path, config=atomize_config)
            emit_progress({"stage": "pipeline_stage", "step": "atomize", "status": "skipped", "percent": 100})
        else:
            update_run_manifest(out_dir, "atomize", "running")
            emit_progress({"stage": "pipeline_stage", "step": "atomize", "status": "running", "percent": 10})
            try:
                # 预算单环节：structure_hypothesis（atomize 的 LLM 仅在 WS1 双轨开时出现）
                with _budget_stage(budget, "structure_hypothesis"):
                    manifest = run_atomizer_pipeline(
                        input_path,
                        out_dir,
                        chunk_chars=chunk_chars,
                        kb_paths=resolve_kb_paths(kb_paths),
                        domain_pack_dir=resolve_bundled_path(domain_pack_dir),
                    )
            except Exception as exc:
                update_run_manifest(out_dir, "atomize", "failed", error=str(exc))
                raise
            update_run_manifest(out_dir, "atomize", "ok", outputs=STAGE_REQUIRED_OUTPUTS["atomize"],
                                action="ran", input_path=input_path, config=atomize_config)
            emit_progress({"stage": "pipeline_stage", "step": "atomize", "status": "ok", "percent": 100})

        # 审计 P1-d：用户 --kb 必须贯通到审查阶段——此前 atomize 按客户 KB 匹配、审查工具
        # 却落回默认 KB 复核（KB 双轨错配）。None（未显式传 kb）保持旧默认行为；显式传入时
        # 同一份解析结果进 run_review_pipeline 与阶段指纹，两侧文件集合严格一致。
        review_kb_paths = resolve_kb_paths(kb_paths) if kb_paths is not None else None
        review_config: dict[str, Any] = {
            # 审计 R2-H2：scope/limit 改变审查覆盖面，必须进阶段指纹——否则先 targeted 后
            # all、或先限量后全量时指纹不变，阶段被整体跳过。
            "review_scope": review_scope,
            "llm_review_limit": llm_review_limit,
        }
        if review_kb_paths is not None:
            review_config["kb_paths"] = [str(path) for path in review_kb_paths]
        # 审计 R2-H2/H3：--domain-pack 同时喂审查（review_policy 合并）与阶段指纹，与
        # atomize 同轨取解析后的包目录；未传时保持 review 默认捆绑包行为不变（显式 None
        # 会关掉 merge_review_policy 的默认包合并）。
        review_domain_pack_path: Path | None = None
        resolved_domain_pack_dir = resolve_bundled_path(domain_pack_dir)
        if resolved_domain_pack_dir is not None:
            review_config["domain_pack_dir"] = str(resolved_domain_pack_dir)
            review_domain_pack_path = resolved_domain_pack_dir / "pack.yaml"
        if skip_review:
            review = None
        elif atomize_reused and stage_is_reusable(out_dir, "llm-review", route=llm_route, config=review_config):
            review = skipped_stage_payload(out_dir, "llm-review")
            update_run_manifest(out_dir, "llm-review", "ok", route=llm_route, outputs=STAGE_REQUIRED_OUTPUTS["llm-review"], action="skipped", config=review_config)
            emit_progress({"stage": "pipeline_stage", "step": "llm-review", "status": "skipped", "percent": 100})
        else:
            update_run_manifest(out_dir, "llm-review", "running", config=review_config)
            emit_progress({"stage": "pipeline_stage", "step": "llm-review", "status": "running", "percent": 0})
            try:
                with _budget_stage(budget, "llm_review"):  # 预算单环节：llm_review
                    review = run_review_pipeline(
                        out_dir,
                        route=llm_route,
                        scope=review_scope,
                        llm_review_limit=llm_review_limit,
                        progress_callback=emit_progress,
                        kb_paths=review_kb_paths,
                        domain_pack_path=review_domain_pack_path or DEFAULT_DOMAIN_PACK_PATH,
                    )
            except Exception as exc:
                update_run_manifest(out_dir, "llm-review", "failed", error=str(exc), config=review_config)
                raise
            update_run_manifest(out_dir, "llm-review", "ok", route=llm_route, outputs=STAGE_REQUIRED_OUTPUTS["llm-review"], action="ran", config=review_config)
        return {
            "kind": "pipeline",
            "out_dir": str(out_dir),
            "input": str(input_path),
            "manifest": manifest,
            "review": review,
            "summary": _stage_summary(out_dir),
        }
    finally:
        _detach_budget_ledger(budget)
        _detach_dual_track_proposer()


def export_task(out_dir: Path, formats: list[str]) -> dict[str, Any]:
    out_dir = out_dir.expanduser().resolve()
    written = export_requirements(out_dir, formats=formats)
    return {
        "kind": "export",
        "out_dir": str(out_dir),
        "written": written,
        "summary": _stage_summary(out_dir),
    }


def resolve_blue_book_index(explicit: Path | None, out_dir: Path) -> Path | None:
    """蓝皮书索引路径：显式参数 > 环境变量 > 约定位置自动探测；都没有 → None（与无索引一致）。

    桌面「运行」链没有索引输入口（GUI 面板本期不做）——自动探测让桌面用户零配置享受 P2 行为
    富化：把编译好的 blue_book_index.json 放在输出目录（或 dev 仓库 out/bluebook/）即可。
    """
    if explicit is not None:
        return explicit
    env_value = os.environ.get(BLUE_BOOK_INDEX_ENV, "").strip()
    if env_value:
        return Path(env_value)
    candidates = (
        out_dir / "blue_book_index.json",
        out_dir / "bluebook" / "blue_book_index.json",
        package_root() / "out" / "bluebook" / "blue_book_index.json",  # dev 仓库编译位置
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


@_leased_pipeline_stage("assemble")
def assemble_task(
    out_dir: Path,
    *,
    formats: list[str] | None = None,
    enrich_route: str | None = None,
    blue_book_index_path: Path | None = None,
) -> dict[str, Any]:
    out_dir = out_dir.expanduser().resolve()
    reviews = out_dir / "llm_review_results.jsonl"
    reviews_path = reviews if reviews.exists() else None
    blue_book_index_path = resolve_blue_book_index(blue_book_index_path, out_dir)
    doc, breakdown = assemble(
        out_dir,
        reviews_path,
        source=out_dir.name,
        extracted_at=datetime.datetime.now().isoformat(timespec="seconds"),
        enrich_route=enrich_route,
        blue_book_index_path=blue_book_index_path,
    )
    target = out_dir / ASSEMBLED_JSON
    target.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    written = [str(target)]
    if formats:
        written.extend(str(out_dir / name) for name in export_spec(out_dir, formats=formats, reviews_path=reviews_path))
    return {
        "kind": "assemble",
        "out_dir": str(out_dir),
        "count": len(doc.get("requirements", [])),
        "analysis": doc.get("analysis", {}),
        "breakdown": breakdown,
        # 出处追溯：本次装配用了哪个蓝皮书索引（None=未注入，行为与 P2 之前一致）
        "blue_book_index": str(blue_book_index_path) if blue_book_index_path else None,
        "written": written,
        "summary": _stage_summary(out_dir),
    }


@_leased_pipeline_stage("functional-synthesis")
def functional_synthesis_task(out_dir: Path, *, route: str = "stub") -> dict[str, Any]:
    out_dir = out_dir.expanduser().resolve()
    return run_functional_synthesis(out_dir, route=route)


@_leased_pipeline_stage("compose")
def compose_task(out_dir: Path) -> dict[str, Any]:
    out_dir = out_dir.expanduser().resolve()
    model = compose_engineering_requirements(out_dir)
    written = write_engineering_requirements(out_dir, model)
    analysis = model.get("analysis", {})
    return {
        "kind": "compose",
        "out_dir": str(out_dir),
        "count": int(analysis.get("requirement_functions") or len(model.get("requirement_functions", []))),
        "analysis": analysis,
        "written": written,
        "summary": _stage_summary(out_dir),
    }


@_leased_pipeline_stage("requirements-analysis")
def requirements_analysis_task(
    out_dir: Path,
    *,
    route: str = "stub",
    template_path: Path | None = None,
) -> dict[str, Any]:
    out_dir = out_dir.expanduser().resolve()
    analysis = run_requirements_analysis(
        out_dir,
        route=route,
        template_path=resolve_template_path(template_path),
        progress_callback=emit_progress,  # 富化逐条上报（GUI n/total，并发度走 RATOMIZER_LLM_CONCURRENCY）
    )
    # 只上报真实存在的产物（此前无条件列出 4 个文件，失败时载荷撒谎）
    names = analysis.get("written") or REQUIREMENTS_ANALYSIS_OUTPUTS
    written = [str(out_dir / name) for name in names if (out_dir / name).exists()]
    return {
        "kind": "requirements_analysis",
        "out_dir": str(out_dir),
        "analysis": analysis,
        "written": written,
        "summary": _stage_summary(out_dir),
    }


def resolve_template_path(template_path: Path | None) -> Path | None:
    if template_path is None:
        return None
    path = Path(template_path).expanduser()
    if not path.is_absolute() and not path.exists():
        path = resolve_bundled_path(path) or path
    if not path.exists():
        raise FileNotFoundError(f"Template file does not exist: {path}")
    return path.resolve()


@_leased_pipeline_stage("clarification-report")
def clarification_report_task(out_dir: Path) -> dict[str, Any]:
    """澄清问题清单 + 就绪判定（确定性零 LLM）：全链疑问信号聚合成评审会可用的问客户清单。"""
    from clarification_report import run_report

    out_dir = out_dir.expanduser().resolve()
    report = run_report(out_dir)
    return {
        "kind": "clarification_report",
        "out_dir": str(out_dir),
        "questions": report["questions"],
        "by_category": report["by_category"],
        "readiness": report["readiness"],
        "written": [str(out_dir / name) for name in report.get("written") or []
                    if (out_dir / name).exists()],
        "summary": _stage_summary(out_dir),
    }


def import_clarification_workbook_task(
    out_dir: Path,
    workbook_path: Path,
    *,
    actor: str = "desktop-import",
) -> dict[str, Any]:
    """Import customer answers and internal acknowledgements from one report workbook."""
    from openpyxl import load_workbook
    from clarification_report import import_answers, import_internal_checks, run_report

    out_dir = out_dir.expanduser().resolve()
    workbook_path = workbook_path.expanduser().resolve()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        has_internal_sheet = "必答-内部核对" in workbook.sheetnames
        internal_headers = (
            [str(cell.value or "").strip() for cell in workbook["必答-内部核对"][1]]
            if has_internal_sheet else []
        )
    finally:
        workbook.close()
    if not has_internal_sheet:
        raise ValueError("工作簿缺少「必答-内部核对」sheet，请重新生成澄清清单")
    required = {"澄清ID", "证据指纹", "阻塞级", "模块", "信号", "来源需求", "核对人", "备注"}
    missing = sorted(required.difference(internal_headers))
    if not any(value.startswith("新处置") for value in internal_headers):
        missing.append("新处置(确认无误/确认有问题/暂缓)")
    if missing:
        raise ValueError(f"「必答-内部核对」sheet 缺少列：{', '.join(missing)}")
    answers = import_answers(out_dir, workbook_path)
    checks = import_internal_checks(out_dir, workbook_path, actor=actor)
    report = run_report(out_dir)
    return {
        "kind": "clarification_answers",
        "out_dir": str(out_dir),
        "imported": int(answers.get("imported") or 0),
        "internal_imported": int(checks.get("imported") or 0),
        "readiness": report.get("readiness") or {},
        "questions": int(report.get("questions") or 0),
        "written": list(dict.fromkeys([
            *(str(value) for value in (answers.get("written") or [])),
            *(str(value) for value in (checks.get("written") or [])),
            *(str(value) for value in (report.get("written") or [])),
        ])),
    }


# ---------------------------------------------------------------------------
# WS4 能力补齐：verification 回写-回灌、手工入口、状态机回退、需求库、依赖推荐
# 全程零 LLM 调用；共享状态文件写走锁 + 原子替换（review_state）。
# ---------------------------------------------------------------------------
_WS4_TRACE_ID_RE = re.compile(r"需求追溯ID[：:]\s*([^\n\r]+)")


def import_verification_workbook_task(
    out_dir: Path,
    workbook_path: Path,
    *,
    actor: str = "desktop-verification",
) -> dict[str, Any]:
    """回灌线下改过的 software_requirements.xlsx 六列 → verification_states.jsonl。

    复用 import-clarification-answers 的解析模式：按需求追溯ID（notes 列）定位行，
    读六列单元格 → verification 子对象，CAS 指纹失配（结构字段漂移）拒绝自动合入转人工。

    T3-2 CAS 分桶：回灌闸只比对**结构列**（子模块 + 客户需求章节）——``description``（叙述）
    不再进闸，叙述措辞变化不再误拒回灌。结构漂移行进 ``rejected``；叙述漂移（结构匹配、
    描述变化）进 ``narrative_review`` 清单——状态仍回灌（不吊销），仅提示专家复核措辞。
    S1-10b：``rejected`` 每条含 requirement_id + xlsx 物理行号 + sheet + 原因。
    """
    from openpyxl import load_workbook
    from requirement_schema import parse_verification_columns, structural_fingerprint_from_cells
    from requirements_analysis_rules import apply_verification_override, load_requirement_index

    root = out_dir.expanduser().resolve()
    index = load_requirement_index(root)
    workbook_path = workbook_path.expanduser().resolve()
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    # (rid, six, structural_cells, description_cell, row_number, sheet_title)
    harvested: list[tuple[str, list[Any], tuple[Any, ...], Any, int, str]] = []
    try:
        for sheet in wb.worksheets:
            header = [str(cell.value or "").strip() for cell in next(
                sheet.iter_rows(min_row=1, max_row=1), [])]
            if "项目负责人确认" not in header:
                continue
            col = {name: idx for idx, name in enumerate(header)}
            # row_number = xlsx 物理行号（min_row=2 → 首条数据行是第 2 行）
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue
                notes_idx = col.get("说明、示例、注意事项", 6)
                notes = str(row[notes_idx] if len(row) > notes_idx else "")
                match = _WS4_TRACE_ID_RE.search(notes)
                if not match:
                    continue
                rid = match.group(1).strip()
                six = [
                    row[col.get(name, idx)] if len(row) > col.get(name, idx) else ""
                    for name, idx in (
                        ("项目负责人确认", 10), ("测试负责人确认", 11), ("研发测试确认", 12),
                        ("功能是否实现", 13), ("测试用例号", 14), ("测试是否完成", 15),
                    )
                ]
                # T3-2：结构闸只用子模块 + 客户需求章节（描述=叙述，降级为复核提示）
                structural_cells = (
                    row[col.get("子模块", 2)] if len(row) > col.get("子模块", 2) else "",
                    row[col.get("客户需求章节", 8)] if len(row) > col.get("客户需求章节", 8) else "",
                )
                description_cell = row[col.get("描述", 3)] if len(row) > col.get("描述", 3) else ""
                harvested.append((rid, six, structural_cells, description_cell, row_number, sheet.title))
    finally:
        wb.close()

    imported = stale = missing = 0
    rejected: list[dict[str, Any]] = []
    narrative_review: list[dict[str, Any]] = []
    for rid, six, structural_cells, description_cell, row_number, sheet_title in harvested:
        entry = index.get(rid)
        if not entry:
            missing += 1
            rejected.append({
                "requirement_id": rid, "row": row_number, "sheet": sheet_title,
                "reason": "需求追溯ID不在当前索引（需求可能已删除或尚未生成）",
            })
            continue
        # CAS 结构闸：子模块 + 客户需求章节 指纹必须与当前需求一致（结构漂移=转人工）
        if structural_fingerprint_from_cells(*structural_cells) != entry["cell_fingerprint"]:
            stale += 1
            rejected.append({
                "requirement_id": rid, "row": row_number, "sheet": sheet_title,
                "reason": "结构字段失配（子模块/客户需求章节已变化，请人工核对后再回灌）",
            })
            continue
        # 叙述复核提示：结构匹配但描述变化 → 不吊销，进 narrative_review 提示专家复核
        item_description = str((entry.get("item") or {}).get("description") or "")
        if str(description_cell or "").strip() and item_description and \
                _backfill_description_drifted(description_cell, item_description):
            narrative_review.append({
                "requirement_id": rid, "row": row_number, "sheet": sheet_title,
                "reason": "描述（叙述）变化——状态已回灌，请复核措辞是否仍准确",
            })
        verification = parse_verification_columns(six, actor_fallback=actor)
        # 仅当六列至少有一项非默认值才写（避免空行覆盖既有状态）
        if verification == default_verification_for_check():
            continue
        apply_verification_override(root, rid, verification, actor=actor,
                                    evidence_fingerprint=entry["fingerprint"])
        imported += 1
    return {
        "kind": "verification_import",
        "out_dir": str(root),
        "imported": imported,
        "stale": stale,
        "missing": missing,
        # S1-10b：拒绝清单精确到 requirement_id + xlsx 行号 + sheet + 原因
        "rejected": rejected,
        # T3-2：叙述复核清单（状态已回灌不吊销，仅提示）
        "narrative_review": narrative_review,
        "written": ["verification_states.jsonl"] if imported else [],
    }


def _backfill_description_drifted(cell_description: Any, item_description: str) -> bool:
    """回灌描述（叙述）漂移判定：折叠空白 + 去控制字符后逐字不等即视为叙述变化。

    仅用于 ``narrative_review`` 复核提示，不参与 CAS 吊销；宽松归一避免无害空白差异误报。
    """
    import re as _re
    norm_cell = _re.sub(r"\s+", " ", str(cell_description or "")).strip()
    norm_item = _re.sub(r"\s+", " ", str(item_description or "")).strip()
    return bool(norm_cell) and bool(norm_item) and norm_cell != norm_item


def default_verification_for_check() -> Any:
    """空 verification（用于回灌跳过全空行）。延迟导入避免顶层依赖。"""
    from requirement_schema import default_verification
    return default_verification()


def set_verification_task(
    out_dir: Path,
    requirement_id: str,
    *,
    implemented: str | None = None,
    test_completed: bool | None = None,
    test_case_ids: str | None = None,
    confirm_pm: bool | None = None,
    confirm_tl: bool | None = None,
    confirm_dt: bool | None = None,
    actor: str = "desktop-verification",
) -> dict[str, Any]:
    """直接写入一条 verification 覆盖（CLI 数据入口；六列字段分散为 flag）。"""
    from requirement_schema import IMPLEMENTED_VALUES, default_verification, normalize_verification
    from requirements_analysis_rules import apply_verification_override

    root = out_dir.expanduser().resolve()
    patch = default_verification()
    if implemented is not None:
        if implemented not in IMPLEMENTED_VALUES:
            raise ValueError(f"非法 implemented 值：{implemented}（可选 {IMPLEMENTED_VALUES}）")
        patch["implemented"] = implemented
    if test_completed is not None:
        patch["test_completed"] = bool(test_completed)
    if test_case_ids is not None:
        patch["test_case_ids"] = [item.strip() for item in re.split(r"[;\n,、 ]+", test_case_ids) if item.strip()]
    if confirm_pm is not None:
        patch["project_manager_confirm"] = {"confirmed": bool(confirm_pm), "by": actor, "at": ""}
    if confirm_tl is not None:
        patch["test_lead_confirm"] = {"confirmed": bool(confirm_tl), "by": actor, "at": ""}
    if confirm_dt is not None:
        patch["dev_test_confirm"] = {"confirmed": bool(confirm_dt), "by": actor, "at": ""}
    record = apply_verification_override(root, requirement_id, normalize_verification(patch), actor=actor)
    return {
        "kind": "verification_set",
        "out_dir": str(root),
        "requirement_id": requirement_id,
        "verification": record.get("verification"),
        "lifecycle_state": record.get("lifecycle_state"),
        "written": ["verification_states.jsonl"],
    }


def rollback_requirement_task(
    out_dir: Path,
    requirement_id: str,
    target_state: str,
    *,
    actor: str,
    reason: str,
) -> dict[str, Any]:
    """人工回退需求生命周期（唯一使状态下落的路径；回退事件 append-only 留痕）。"""
    from requirements_analysis_rules import rollback_requirement_lifecycle

    root = out_dir.expanduser().resolve()
    record = rollback_requirement_lifecycle(root, requirement_id, target_state,
                                            actor=actor, reason=reason)
    return {
        "kind": "requirement_rollback",
        "out_dir": str(root),
        "requirement_id": requirement_id,
        "lifecycle_state": record.get("lifecycle_state"),
        "written": ["verification_states.jsonl", "requirement_lifecycle_events.jsonl"],
    }


def add_manual_requirement_task(
    out_dir: Path,
    *,
    objective: str,
    behaviors: str | None = None,
    module: str = "",
    ownership: str = "",
    priority: str = "P1",
    notes: str = "",
    actor: str = "desktop-manual",
) -> dict[str, Any]:
    """手工建需求入口（provenance=manual，追溯列留空不伪引）。走完全相同下游。"""
    from requirements_analysis_rules import record_manual_requirement

    root = out_dir.expanduser().resolve()
    if isinstance(behaviors, (list, tuple)):
        behavior_list = [str(item).strip() for item in behaviors if str(item).strip()]
    elif behaviors:
        behavior_list = [item.strip() for item in str(behaviors).split(",") if item.strip()]
    else:
        behavior_list = []
    record = record_manual_requirement(
        root, objective=objective, behaviors=behavior_list, module=module,
        ownership=ownership, priority=priority, notes=notes, actor=actor,
    )
    return {
        "kind": "manual_requirement",
        "out_dir": str(root),
        "functional_requirement_id": record.get("functional_requirement_id"),
        "written": ["manual_requirements.jsonl"],
    }


def build_requirement_library_task(
    project_dirs: list[Path],
    output_path: Path,
    *,
    include_unconfirmed: bool = False,
) -> dict[str, Any]:
    """汇总各项目 functional_requirements 为带项目元数据的 JSONL 检索库（不引入数据库）。

    S1-10d 需求库入库质量门：默认仅收录 lifecycle>=confirmed 的条目（draft 视为未确认，不入库）；
    ``include_unconfirmed=True`` 显式收录 draft 条目。每条 entry 携带 ``lifecycle_state``，供
    采纳 UI 默认隐藏未确认条目。functional_requirements 经 governed 双路径探测读取
    （package_v1 下在 .ratomizer/pipeline/，裸根拼会落空——B1 类寻址失守）。
    """
    import json as _json
    from requirement_schema import (
        LIFECYCLE_CONFIRMED, library_entry_from_requirement, lifecycle_rank, requirement_identity,
    )
    from requirements_analysis_rules import _read_functional_requirements_payload
    from review_state import read_verification_states

    output_path = output_path.expanduser().resolve()
    entries: list[dict[str, Any]] = []
    sources: list[dict[str, Any]] = []
    total_skipped_unconfirmed = 0
    confirmed_floor = lifecycle_rank(LIFECYCLE_CONFIRMED)
    for project_dir in project_dirs:
        root = Path(project_dir).expanduser().resolve()
        payload = _read_functional_requirements_payload(root)
        items = payload.get("items") if isinstance(payload, dict) else None
        if not isinstance(items, list) or not items:
            sources.append({"project_dir": str(root), "imported": 0,
                            "reason": "functional_requirements.json 缺失"})
            continue
        # 生命周期来自 verification_states.jsonl（governed state 路径，for_write=False 不建目录）
        lifecycle_by_rid = {
            str(row.get("requirement_id") or "").strip(): str(row.get("lifecycle_state") or "draft")
            for row in read_verification_states(root).values()
        }
        project_name = str(payload.get("source") or root.name) if isinstance(payload, dict) else root.name
        created_at = ""
        prov = payload.get("provenance") if isinstance(payload, dict) else None
        if isinstance(prov, dict):
            created_at = str(prov.get("generated_at") or "")
        count = 0
        skipped_unconfirmed = 0
        for item in items or []:
            if not isinstance(item, dict):
                continue
            rid = requirement_identity(item)
            lifecycle_state = lifecycle_by_rid.get(rid) or "draft"
            # 质量门：默认仅收录 lifecycle>=confirmed；draft 不入库（视为未确认）
            if not include_unconfirmed and lifecycle_rank(lifecycle_state) < confirmed_floor:
                skipped_unconfirmed += 1
                continue
            entry = library_entry_from_requirement(
                item, project=project_name, doc_source=str(root), created_at=created_at)
            entry["lifecycle_state"] = lifecycle_state
            entries.append(entry)
            count += 1
        total_skipped_unconfirmed += skipped_unconfirmed
        sources.append({"project_dir": str(root), "project": project_name, "imported": count,
                        "skipped_unconfirmed": skipped_unconfirmed})
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = output_path.with_suffix(output_path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8", newline="\n") as handle:
        for entry in entries:
            handle.write(_json.dumps(entry, ensure_ascii=False) + "\n")
    os.replace(tmp, output_path)
    return {
        "kind": "requirement_library",
        "library": str(output_path),
        "entries": len(entries),
        "skipped_unconfirmed": total_skipped_unconfirmed,
        "include_unconfirmed": bool(include_unconfirmed),
        "sources": sources,
    }


def search_requirements_task(
    library_path: Path,
    query: str,
    *,
    limit: int = 20,
    retriever: Any = None,
) -> dict[str, Any]:
    """召回历史相似需求（T3-4：经 ``RequirementRetriever`` 插件点，默认词面，可注入）。

    ``retriever`` 注入优先（测试/外部向量插件）；否则按需求库建词面默认检索器。明确不引入
    向量依赖——``RATOMIZER_REQUIREMENT_RETRIEVER=vector`` 当前如实回退词面并标 ``retriever_kind``。
    任何检索器产出仍是同一 entry 形态，下游确定性校验不放松。
    """
    import json as _json
    from requirement_schema import build_requirement_retriever

    library_path = library_path.expanduser().resolve()
    if not library_path.exists():
        raise FileNotFoundError(f"需求库不存在：{library_path}（先 build-requirement-library）")
    library: list[dict[str, Any]] = []
    with library_path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                library.append(_json.loads(line))
            except _json.JSONDecodeError:
                continue
    retriever_obj = build_requirement_retriever(library, retriever=retriever)
    results = retriever_obj.search(query, limit=limit)
    return {
        "kind": "requirement_search",
        "library": str(library_path),
        "query": query,
        "matches": len(results),
        "results": results,
        "retriever_kind": getattr(retriever_obj, "retriever_kind", "literal"),
    }


def unified_search_requirements_task(
    query: str,
    *,
    limit: int = 20,
    retriever: Any = None,
) -> dict[str, Any]:
    """WS-C3：跨三库统一检索（requirement / base / solution）。

    任一库配置存在即启用；未配置库如实跳过。默认词面，可注入外部 retriever。
    """
    from unified_requirement_retriever import build_unified_retriever, default_library_paths

    paths = default_library_paths()
    configured = {k: v for k, v in paths.items() if v is not None}
    if not configured:
        return {
            "kind": "unified_requirement_search",
            "query": query,
            "matches": 0,
            "results": [],
            "retriever_kind": "unavailable",
            "note": "未配置任何需求库（RATOMIZER_REQUIREMENT_LIBRARY / BASE_LIBRARY / SOLUTION_LIBRARY）",
        }
    retriever_obj = build_unified_retriever(library_paths=paths, retriever=retriever)
    results = retriever_obj.search(query, limit=limit)
    return {
        "kind": "unified_requirement_search",
        "query": query,
        "matches": len(results),
        "results": results,
        "retriever_kind": getattr(retriever_obj, "retriever_kind", "unknown"),
        "source_counts": {
            source: sum(1 for r in results if r.get("library_source") == source)
            for source in configured
        },
    }


def recommend_dependencies_task(out_dir: Path) -> dict[str, Any]:
    """对当前项目跑确定性依赖/父子候选推荐（只生产值，不动 schema）。

    T3-1：候选状态反映 RTM 边事件流的最新裁决——accept→``accepted``、reject→``rejected``、
    未裁决→``pending``。物化库（``dependency_decisions.jsonl``）只含 accept；事件流含全部，
    故以事件流回放为准（最后决策胜出）。
    """
    from requirements_analysis_rules import dependency_candidates_for_project
    from review_state import read_rtm_edge_events, replay_rtm_edges

    root = out_dir.expanduser().resolve()
    candidates = dependency_candidates_for_project(root)
    replay = replay_rtm_edges(read_rtm_edge_events(root))
    edges = replay.get("edges") or {}
    decided_by_tuple = {
        (e.get("from"), e.get("to"), e.get("kind")): e["decision"]
        for e in edges.values()
    }
    rendered = []
    for candidate in candidates:
        key = (candidate.get("from"), candidate.get("to"), candidate.get("kind"))
        candidate = dict(candidate)
        decision = decided_by_tuple.get(key)
        candidate["status"] = {"accept": "accepted", "reject": "rejected"}.get(decision, "pending")
        rendered.append(candidate)
    return {
        "kind": "dependency_candidates",
        "out_dir": str(root),
        "candidates": rendered,
        "pending": sum(1 for c in rendered if c["status"] == "pending"),
        "accepted": sum(1 for c in rendered if c["status"] == "accepted"),
        "rejected": sum(1 for c in rendered if c["status"] == "rejected"),
    }


def decide_dependency_task(
    out_dir: Path,
    *,
    frm: str,
    to: str,
    kind: str,
    accepted: bool,
    actor: str = "desktop-dependency",
    reason: str = "",
) -> dict[str, Any]:
    """依赖候选裁决：接受才写库；拒绝不落库。"""
    from requirement_schema import DEPENDENCY_KINDS
    from requirements_analysis_rules import apply_dependency_decision

    if kind not in DEPENDENCY_KINDS:
        raise ValueError(f"非法依赖类型：{kind}（可选 {DEPENDENCY_KINDS}）")
    root = out_dir.expanduser().resolve()
    result = apply_dependency_decision(
        root, {"from": frm, "to": to, "kind": kind}, accepted=accepted, actor=actor, reason=reason,
    )
    # T3-1：accept 与 reject 都追加 RTM 边事件流；accept 另落物化库
    written = ["requirement_rtm_edges.jsonl"]
    if result.get("written"):
        written.append("dependency_decisions.jsonl")
    return {
        "kind": "dependency_decision",
        "out_dir": str(root),
        **result,
        "written": written,
    }


def adjudicate_task(
    out_dir: Path,
    *,
    route: str | None = None,
    actor: str = "desktop-adjudicator",
) -> dict[str, Any]:
    """WS-B：运行功能需求级 AI 裁决（默认全关；LLM 不可用时全部进 review）。"""
    from adjudicate import adjudicate_all

    root = out_dir.expanduser().resolve()
    summary = adjudicate_all(root, route=route, actor=actor)
    return {
        "kind": "adjudicate",
        "out_dir": str(root),
        **summary,
    }


def overturn_adjudication_task(
    out_dir: Path,
    *,
    functional_requirement_id: str,
    new_decision: str,
    actor: str,
    reason: str,
) -> dict[str, Any]:
    """WS-B：人工推翻自动裁决结果，写回裁决流（actor/reason 必填）。"""
    from adjudicate import overturn_adjudication

    root = out_dir.expanduser().resolve()
    record = overturn_adjudication(
        root, functional_requirement_id,
        new_decision=new_decision, actor=actor, reason=reason,
    )
    return {
        "kind": "adjudication_overturn",
        "out_dir": str(root),
        "record": record,
        "written": ["adjudication_results.jsonl"],
    }


@_leased_pipeline_stage("template-write")
def template_write_task(out_dir: Path, template_path: Path) -> dict[str, Any]:
    """成文：analyze 结果按公司标准化需求列表 V2.3.x 格式追加进对应模块 sheet（确定性零 LLM）。"""
    from template_writer import run_writer

    out_dir = out_dir.expanduser().resolve()
    report = run_writer(out_dir, template_path.expanduser().resolve())
    return {
        "kind": "template_write",
        "out_dir": str(out_dir),
        "report": report,
        "written": [str(out_dir / name) for name in report.get("written") or []
                    if (out_dir / name).exists()],
        "summary": _stage_summary(out_dir),
    }


def ai_extract_task(out_dir: Path, *, route: str | None, limit_sections: int | None = None,
                    sample_ratio: float | None = None) -> dict[str, Any]:
    """AI 主抽 + 双引擎合并：AI 行为需求 + 确定性结构需求 → merged_spec.xlsx/json。

    试抽模式（「测试运行」用）：sample_ratio 按比例抽样（0.2=全文 1/5，随文档自适应）；
    limit_sections 固定 N 章（CLI 用）。"""
    out_dir = out_dir.expanduser().resolve()
    result = ai_extract.run_ai_extract(out_dir, route=route, merge_deterministic=True,
                                       progress_callback=emit_progress,
                                       limit_sections=limit_sections,
                                       sample_ratio=sample_ratio)
    return {
        "kind": "ai_extract",
        "out_dir": str(out_dir),
        "route": result.get("route"),
        "count": result.get("requirements", 0),
        "merged": result.get("merged", {}),
        "sections": result.get("sections", 0),
        "failed_sections": result.get("failed_sections", 0),
        "code_drift_flagged": result.get("code_drift_flagged", 0),
        "int_drift_flagged": result.get("int_drift_flagged", 0),
        "note": result.get("note", ""),
        "quality": result.get("quality", {}),
        "sampled": result.get("sampled"),
        "consistency": result.get("consistency", {}),
        "claim_shadow": result.get("claim_shadow", {}),
        "written": [str(out_dir / name) for name in result.get("written", [])],
        "summary": _stage_summary(out_dir),
    }


RUN_MANIFEST = "run_manifest.json"
STAGES_DIR = "_stages"
_MANIFEST_LOCKS: dict[Path, RLock] = {}
_MANIFEST_LOCKS_GUARD = RLock()
_MANIFEST_LOCK_TIMEOUT_S = 10.0
_MANIFEST_LOCK_STALE_AFTER_S = 300.0
_REPLACE_ATTEMPTS = 5
_REPLACE_RETRY_DELAY_S = 0.02

# 阶段名 == 子命令名（manifest 键与 CLI 一致，GUI 单步按钮与 chain 写同一本账）
CHAIN_ORDER = ["ai-extract", "functional-synthesis", "assemble", "requirements-analysis", "template-write",
               "clarification-report", "compose", "export-annotation-html"]

# 结果包发布纪律（2026-08-03 审查 I1/I3，spec §8.2/§15）：
# - 只有这些会改动已注册根交付物的写命令才允许触发发布；
#   只读命令（summary 等）永不发布——发布会做恢复写 + 全量复制 + marker 重写。
# - 活动 attempt 期间即使白名单命令也不发布：根交付物保持上一完成代，
#   由 result-package-complete 在全部阶段验证后一次性发布（失败重跑不污染旧完成结果）。
PUBLISHING_COMMANDS = frozenset({
    "run", "chain", "ai-extract", "clarification-report", "requirements-analysis",
    "template-write", "assemble", "compose", "export",
    "export-annotation-html", "import-ai-decisions", "import-clarification-answers",
})
_RESULT_PACKAGE_COMMANDS = frozenset({
    "result-package-start", "result-package-complete",
    "result-package-fail", "result-package-status",
})
STAGE_INPUTS: dict[str, list[str]] = {
    "atomize": [],
    "llm-review": ["atomic_requirements.jsonl", "llm_tasks.jsonl"],
    "ai-extract": ["blocks.jsonl", "table_items.jsonl", "table_cell_items.jsonl",
                   "table_cell_dispositions.jsonl",
                   "llm_review_results.jsonl", "review_states.jsonl",
                   "ai_supplements.jsonl"],
    "assemble": ["table_items.jsonl", "table_cell_items.jsonl", "table_cell_dispositions.jsonl", "atomic_requirements.jsonl",
                 "llm_review_results.jsonl",
                 "ai_supplements.jsonl"],
    "functional-synthesis": ["ai_requirements.jsonl", "ai_requirements.meta.json", "blocks.jsonl",
                             "ai_review_states.jsonl", "ai_supplements.jsonl"],
    "requirements-analysis": [FUNCTIONAL_REQUIREMENTS, "ai_requirements.jsonl", "ai_review_states.jsonl",
                              "clarification_answers.jsonl", "blocks.jsonl", "term_map.json",
                              "ai_requirements.meta.json", "ai_supplements.jsonl"],
    "template-write": ["engineering_analysis.json", "ai_requirements.meta.json",
                       "ai_supplements.jsonl"],
    "clarification-report": [FUNCTIONAL_REQUIREMENTS, "ai_requirements.jsonl", "engineering_analysis.json",
                             "consistency_report.json", "blocks.jsonl", "ai_review_states.jsonl",
                             "clarification_answers.jsonl", "clarification_check_states.jsonl",
                             "omission_states.jsonl", "ai_requirements.meta.json", "ai_supplements.jsonl",
                             "claim_effective_ledger.jsonl", "claim_effective.meta.json",
                             "claim_queue_proposals.jsonl", "claim_effective_health.json"],
    "compose": ["atomic_requirements.jsonl", "table_items.jsonl", "table_cell_items.jsonl", "table_cell_dispositions.jsonl",
                "ai_requirements.meta.json",
                "ai_supplements.jsonl"],
    "export-annotation-html": ["blocks.jsonl", "table_items.jsonl", "table_cell_items.jsonl",
                               "table_cell_dispositions.jsonl",
                               "ai_requirements.jsonl",
                               "engineering_analysis.json", "ai_review_states.jsonl",
                               "annotation_translations.json", "ai_requirements.meta.json",
                               "ai_supplements.jsonl", "claim_catalog.jsonl",
                               "claim_catalog.meta.json", "claim_coverage_groups.jsonl",
                               "claim_ledger.jsonl", "claim_shadow_metrics.json",
                               "claim_verifier_attempts.jsonl", "claim_generation.meta.json",
                               "claim_effective_ledger.jsonl", "claim_effective.meta.json",
                               "claim_queue_proposals.jsonl", "claim_structural_overrides.jsonl",
                               ".claim_publication.journal.json",
                               ".claim_effective_publication.journal.json"],
}


STAGE_REQUIRED_OUTPUTS: dict[str, list[str]] = {
    "atomize": [
        "manifest.json",
        "blocks.jsonl",
        "chunks.jsonl",
        "table_items.jsonl",
        "table_cell_items.jsonl",
        "table_cell_dispositions.jsonl",
        "atomic_requirements.jsonl",
        "llm_tasks.jsonl",
        "quality_report.json",
        "summary.md",
    ],
    "llm-review": ["llm_review_results.jsonl", "review_states.jsonl"],
    "ai-extract": [
        "ai_requirements.jsonl",
        "ai_requirements.meta.json",
        "compliance_requirements.json",
        "merged_spec_requirements.json",
        "claim_catalog.jsonl",
        "claim_catalog.meta.json",
        "claim_coverage_groups.jsonl",
        "claim_ledger.jsonl",
        "claim_shadow_metrics.json",
        "claim_verifier_attempts.jsonl",
        "claim_generation.meta.json",
    ],
    "functional-synthesis": [FUNCTIONAL_REQUIREMENTS],
    "assemble": [ASSEMBLED_JSON],
    "requirements-analysis": REQUIREMENTS_ANALYSIS_OUTPUTS,
    "template-write": ["软件需求列表-成文.xlsx", "template_writer_report.json"],
    "clarification-report": ["clarification_questions.md", "clarification_questions.xlsx", "clarification_report.json"],
    "compose": ["engineering_requirements/engineering_requirements.json"],
    "export-annotation-html": ["document_annotation.html"],
}


STAGE_IMPLEMENTATION_REVISIONS = {
    # v12：table-structure-v6 候选闭环（未类型化冒号规格与拒收矩阵 marker
    # 进入可操作 cell 审核面）——leaf plan/catalog 输入变化，全部输入须重解析
    # v11：table-structure-v5 证据契约（受控矩阵轴、当前 PDF 无 merge 几何不再
    # 启用旧 group-header 启发式、类型化 colon_spec 资格）——结构与 claims 变化，
    # 全部输入须重解析
    # v10：table-structure-v4 复审闭环（正向 matrix_dimension_evidence 取代黑名单、
    # merge 证据 [] 与 None 显式区分、单格无结构证据行转 ambiguous_structure_cells
    # 候选、other 表资格按强义务信号授权）——blocks/table_items/cells 变化，
    # 全部输入须重解析
    # v9：table-structure-v3 审核闭环（结构角色/内容资格解耦为 structural_role ×
    # obligation_signal × 内容守恒三维、同行义务句携前置标识格 Header=Value 上下文、
    # 合成表头/处置列矩阵闸、同格重复句去重、xlsx sheet 守恒计数器 + 无缓存公式
    # fail-closed、合并格被覆盖文本逐字校验、PDF 几何冲突显式状态）——blocks/
    # table_items/cells 变化，全部输入须重解析
    # v8：table-structure-v2（cell 级闭环 + table_cell_items.jsonl + 标题/表头确定性识别
    # + merge 证据）——blocks/table_items 结构面变化，全部输入须重解析
    # v7：render_table_text 取消 20 行截断（大参数表 21 行起内容进不了管线,STO 实证）
    # v6：表格块扁平文本取消 [:5000] 截断（初始提交遗留）——大参数表 88% 内容此前
    # 进不了抽取管线（STO/俄标实证）；blocks 内容变化,docx 输入须重解析
    # v5：PDF 清单段合并（名词式清单项并整段，微块可锚定）——块结构变化，PDF 输入须重解析
    "atomize": "v13",
    # v6：cell 级 assemble 输入（cells source_block）+ 权威 row/cell ID 去重键 + 发布断言
    # ——section/unit 文本与 source_blocks 结构变,旧 ai-extract 缓存失效重抽
    # v5：表格行级化(封堵一 chunk 表头注入 + 行级 source_blocks/rows + 封堵二去重 + 封堵三澄清聚合)
    "ai-extract": "v7",
    "assemble": "v2",
    # v4: consumes is_compliance_requirement; compliance-rules v2 invalidates old caches.
    "functional-synthesis": "v4",
    # v6：hardware_dependency 落交付物渲染（xlsx 说明列/co_design_items.md）——
    # 纯渲染变更不动 analyze 缓存版本，靠 impl 戳让阶段重跑重渲染（审计 P1-b）
    "requirements-analysis": "v6",
    # v5：完整性元数据进入阶段输入，旧缓存不得缺 incomplete_inputs。
    "template-write": "v5",
    "clarification-report": "v6",
    # v1.5：compose 首次绑定完整性元数据，显式升级阶段实现戳。
    "compose": "v2",
}

_STAGE_BASE_PRODUCERS = {
    "atomize": "atomize",
    "assemble": "assemble_spec/v1",
    "template-write": "template_writer/v1",
    "clarification-report": "clarification/v8-param-row-aggregate",
    "compose": "engineering_composer/v1",
    # v16-cell-claim-projection：P0-2 cell claim 落公共 records（claims_json/claim_zones
    # 不再丢失生产 table_cell claim）+ P1-3 静态 HTML 按物理 R×C/merge anchor 在
    # <th>/<td> 内渲染 cell claim 入口
    # + v15-cell-merge-span-context：cell_context 增 row_span/column_span/
    # covered_coordinates（UI 标题/表头 cell 按钮 + 合并跨度 DOM 渲染）
    # + table_cell claim 几何热区真实落页（消费 cell bbox 记录）+ 不可用负载
    # 也携 cell_context（非影印模式可取上下文）
    # + v14-claim-distribution-claim-focus：cell 级闭环（cell claim 记录/几何/卡片）
    # + v13 三线联合戳——v13=行区占比切片互斥（远端 rowcell）
    # + claim-distribution（Phase 1 块级角标）+ claim-focus（Phase 1.5 claim span/row 级定位，
    # 经 claim_focus 确定性映射）——三侧缓存产物一并失效；v12=表格行级热区、v11=几何回填
    "export-annotation-html": "doc_annotation_export/v16-cell-claim-projection",
    "run": "pipeline/v1",
    "llm-review": "review/v1",
}


def stage_producer(stage: str, *, out_dir: Path | None = None,
                   kb_paths: list[Any] | None = None) -> str:
    """阶段 → 生产者版本戳（产物血统：今天拿 v9 数据当新结果看的事故，靠它绝迹）。

    llm-review 额外拼入审查代码版本（prompt/cache/tools，审计 R2-H2）与工具证据
    内容指纹（KB/blocks/原子需求/蓝皮书索引，审计 P1-d）——审查代码升级或实际
    读取的证据变了，旧审查产物不得继续复用；out_dir 缺席时保持基础戳+代码版本
    （无目录语境兼容）。"""
    producer = _STAGE_BASE_PRODUCERS.get(stage, stage)
    # Phase 0 reserves policy lineage for future agent stages without invalidating any
    # current pipeline cache. There is intentionally no agent stage in CHAIN_ORDER yet.
    if stage.startswith("agent-"):
        producer = f"{producer}+{AGENT_POLICY_VERSION}"
    try:
        if stage == "ai-extract":
            # 版本戳必须覆盖全部影响产物的代码层；否则 guards/verify/compliance 升级后
            # chain 续跑仍可能复用旧结果。阶段戳采用完整 producer lineage；付费 section
            # cache 只钉实际改变缓存行的子集，避免纯发布后处理升级触发无必要重抽。
            from ai_extract import (
                AI_REQUIREMENTS_PRODUCER_LINEAGE_VERSION,
                producer_lineage_versions,
            )
            producer = "+".join((
                AI_REQUIREMENTS_PRODUCER_LINEAGE_VERSION,
                *producer_lineage_versions().values(),
            ))
        elif stage == "atomize":
            # PDF text repair changes blocks consumed by every downstream stage. Include both
            # the algorithm version and the bundled vocabulary content in the producer so a
            # repaired parser cannot silently reuse an old atomize run. Source alignment is a
            # separate parser-output contract used by the claim catalog conservation audit.
            from parsers.pdf_parser import (
                PDF_TEXT_REPAIR_VERSION,
                pdf_layout_switch_fingerprint,
                text_repair_vocabulary_fingerprint,
            )
            from source_spans import (
                SOURCE_ALIGNMENT_VERSION,
                SOURCE_TRANSFORMATION_POLICY_VERSION,
                SOURCE_TRANSFORMATION_RULESET_VERSION,
            )
            from table_structure import TABLE_STRUCTURE_VERSION
            producer = (
                f"{producer}+{PDF_TEXT_REPAIR_VERSION}"
                f"+repair-vocab-{text_repair_vocabulary_fingerprint()}"
                # W8：D1/D2/D3 版式修复开关状态必须进戳——ON/OFF 不得共用缓存；
                # fingerprint 由 pdf_parser 单点生成，两处默认值不漂移
                f"+{pdf_layout_switch_fingerprint()}"
                f"+{SOURCE_ALIGNMENT_VERSION}"
                f"+{SOURCE_TRANSFORMATION_POLICY_VERSION}"
                f"+{SOURCE_TRANSFORMATION_RULESET_VERSION}"
                f"+{TABLE_STRUCTURE_VERSION}"
            )
            # A9：tender 适配开关仅在开启时进戳，默认关时保持 producer 不变
            if os.environ.get("RATOMIZER_TENDER_TABLE_FILTER", "0").strip().lower() not in {"0", "false", "off"}:
                from tender_table_filter import TENDER_TABLE_FILTER_VERSION
                producer = f"{producer}+{TENDER_TABLE_FILTER_VERSION}"
            if os.environ.get("RATOMIZER_TENDER_REGION_FILTER", "0").strip().lower() not in {"0", "false", "off"}:
                from tender_regions import TENDER_REGION_FILTER_VERSION
                producer = f"{producer}+{TENDER_REGION_FILTER_VERSION}"
            if os.environ.get("RATOMIZER_TENDER_FIGURE_PAGE_FILTER", "0").strip().lower() not in {"0", "false", "off"}:
                from unextracted_registry import UNEXTRACTED_REGISTRY_VERSION
                producer = f"{producer}+figure-page-v1"
        elif stage == "llm-review":
            # 代码版本必须进戳（审计 R2-H2）：prompt/cache/tools 任一 bump 后旧阶段
            # 产物不得复用——此前 llm-review 是唯一不拼代码版本的阶段。
            from llm_pipeline import LLM_REVIEW_CACHE_VERSION, PROMPT_VERSION
            from review_tools import REVIEW_TOOLS_VERSION
            producer = (f"{producer}+{PROMPT_VERSION}+{LLM_REVIEW_CACHE_VERSION}"
                        f"+{REVIEW_TOOLS_VERSION}")
            if out_dir is not None:
                from review_tools import evidence_fingerprint
                producer = f"{producer}+evidence-{evidence_fingerprint(out_dir, kb_paths)}"
        elif stage == "assemble":
            # assemble 会运行 spec_enrich；富化 prompt/护栏升级必须让阶段续跑失效，
            # 否则旧 run_manifest 会在富化缓存检查之前跳过整个阶段。
            from spec_enrich import ENRICH_GUARDS_VERSION, ENRICH_PROMPT_VERSION
            producer = (f"{producer}+{ENRICH_PROMPT_VERSION}"
                        f"+{ENRICH_GUARDS_VERSION}")
        elif stage == "requirements-analysis":
            from requirements_analysis import ANALYZE_PROMPT_VERSION, UNFOUNDED_RULE_VERSION
            from requirements_analysis_rules import ANALYZE_RULES_VERSION
            producer = (
                f"{ANALYZE_PROMPT_VERSION}+{UNFOUNDED_RULE_VERSION}"
                f"+{ANALYZE_RULES_VERSION}"
            )
        elif stage == "functional-synthesis":
            producer = FUNCTIONAL_SYNTHESIS_VERSION
        elif stage == "export-annotation-html":
            from claim_focus import CLAIM_FOCUS_ADAPTER_VERSION
            from doc_annotation_export import (ANNOTATION_TRANSLATION_GUARDS_VERSION,
                                                ANNOTATION_TRANSLATION_STRATEGY_VERSION,
                                                CLAIM_ANNOTATION_VERSION)
            from doc_facsimile import DOC_FACSIMILE_VERSION
            # docx/xlsx 影印支路（WP-A）进戳：转换层版本变化 → 旧影印产物不得复用
            producer = (f"{producer}+{CLAIM_ANNOTATION_VERSION}"
                        f"+{CLAIM_FOCUS_ADAPTER_VERSION}"
                        f"+{ANNOTATION_TRANSLATION_STRATEGY_VERSION}"
                        f"+{ANNOTATION_TRANSLATION_GUARDS_VERSION}+{DOC_FACSIMILE_VERSION}")
        if stage in {
            "ai-extract", "functional-synthesis", "assemble", "requirements-analysis", "template-write",
            "clarification-report", "compose", "export-annotation-html",
        }:
            from omission_actions import AI_SUPPLEMENT_VERSION
            producer = f"{producer}+{AI_SUPPLEMENT_VERSION}"
    except Exception:  # pragma: no cover - 版本戳失败不阻断任务
        pass
    revision = STAGE_IMPLEMENTATION_REVISIONS.get(stage)
    return f"{producer}+impl-{revision}" if revision else producer


def read_run_manifest(out_dir: Path) -> dict[str, Any]:
    root = Path(out_dir).expanduser().resolve()
    package_root = package_root_for_analysis_root(root)
    path = (package_artifact_path(package_root, "run_manifest")
            if package_root is not None else root / RUN_MANIFEST)
    try:
        data = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}
    except (json.JSONDecodeError, OSError):
        data = {}
    return data if isinstance(data, dict) else {}


def _relative_outputs(out_dir: Path, outputs: list[str | Path]) -> list[str]:
    root = Path(out_dir).expanduser().resolve()
    result: list[str] = []
    for item in outputs:
        path = Path(item)
        if path.is_absolute():
            try:
                path = path.resolve().relative_to(root)
            except ValueError:
                pass
        result.append(path.as_posix())
    return result


def _stage_outputs(stage: str, entry: dict[str, Any] | None = None) -> list[str]:
    outputs = (entry or {}).get("outputs") if isinstance(entry, dict) else None
    if isinstance(outputs, list) and outputs:
        return [str(item) for item in outputs]
    return list(STAGE_REQUIRED_OUTPUTS.get(stage, []))


def _outputs_exist(out_dir: Path, outputs: list[str]) -> bool:
    if not outputs:
        return False
    root = Path(out_dir).expanduser().resolve()
    for name in outputs:
        path = governed_artifact_path(root, name, for_write=False)
        if not path.exists() or path.is_dir():
            return False
        try:
            if path.stat().st_size <= 0 and name not in {
                "ai_requirements.jsonl",
                "claim_catalog.jsonl",
                "claim_coverage_groups.jsonl",
                "claim_ledger.jsonl",
                "claim_effective_ledger.jsonl",
            }:
                return False
        except OSError:
            return False
    return True


def _hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


_STAGE_MUTABLE_INPUTS = {
    "export-annotation-html": {"annotation_translations.json"},
}

_CLAIM_STAGE_OUTPUTS = {
    "claim_catalog.jsonl",
    "claim_catalog.meta.json",
    "claim_coverage_groups.jsonl",
    "claim_ledger.jsonl",
    "claim_effective_ledger.jsonl",
    "claim_shadow_metrics.json",
    "claim_verifier_attempts.jsonl",
    "claim_generation.meta.json",
    "claim_effective.meta.json",
}

_CLAIM_EFFECTIVE_RUNTIME_OUTPUTS = {
    "claim_effective_ledger.jsonl",
    "claim_effective.meta.json",
    "claim_review_events.jsonl",
    "claim_queue_proposals.jsonl",
    "claim_effective_health.json",
}


def stage_input_files_fingerprint(out_dir: Path, stage: str) -> str:
    """Hash immutable stage inputs for the duration of a downstream read lease."""
    root = Path(out_dir).expanduser().resolve()
    ignored = _STAGE_MUTABLE_INPUTS.get(stage, set())
    payload = []
    for name in STAGE_INPUTS.get(stage, []):
        if name in ignored:
            continue
        path = governed_artifact_path(root, name, for_write=False)
        payload.append({
            "path": name,
            "sha256": _hash_file(path) if path.is_file() else None,
        })
    return hashlib.sha256(json.dumps(
        payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"),
    ).encode("utf-8")).hexdigest()


def _resource_content_fingerprint(path: Path | None) -> str | None:
    """Hash a KB file or a deterministic directory tree for atomize reuse checks."""
    if path is None:
        return None
    path = Path(path).expanduser().resolve()
    if path.is_file():
        return _hash_file(path)
    if not path.is_dir():
        return None
    entries = []
    for child in sorted(item for item in path.rglob("*") if item.is_file()):
        entries.append({
            "path": str(child.relative_to(path)).replace("\\", "/"),
            "sha256": _hash_file(child),
        })
    return hashlib.sha256(json.dumps(
        entries, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")).hexdigest()


def stage_input_fingerprint(out_dir: Path, stage: str, *, route: str | None = None,
                            template_path: Path | None = None,
                            input_path: Path | None = None,
                            config: dict[str, Any] | None = None) -> str:
    root = Path(out_dir).expanduser().resolve()
    inputs: list[dict[str, Any]] = []
    for name in STAGE_INPUTS.get(stage, []):
        path = governed_artifact_path(root, name, for_write=False)
        inputs.append({
            "path": name,
            "sha256": _hash_file(path) if path.exists() and path.is_file() else None,
        })
    template = Path(template_path).expanduser().resolve() if template_path else None
    payload = {
        "stage": stage,
        "producer": stage_producer(
            stage,
            out_dir=root,
            kb_paths=(config or {}).get("kb_paths") if stage == "llm-review" else None,
        ),
        "route": route or "",
        "inputs": inputs,
        "atomize_resources": (
            {
                "knowledge_bases": [
                    {
                        "path": str(Path(path).expanduser().resolve()),
                        "sha256": _resource_content_fingerprint(Path(path)),
                    }
                    for path in (config or {}).get("kb_paths") or []
                ],
                "domain_pack": (
                    {
                        "path": str(Path((config or {}).get("domain_pack_dir")).expanduser().resolve()),
                        "sha256": _resource_content_fingerprint(
                            Path((config or {}).get("domain_pack_dir"))
                        ),
                    }
                    if (config or {}).get("domain_pack_dir") else None
                ),
            }
            if stage == "atomize" else None
        ),
        # 审计 R2-H2：llm-review 合并 domain-pack 的 review_policy，包内容变必须令阶段
        # 失效（仅当实际传入 domain pack 时；atomize 的包内容走上方 atomize_resources）。
        "domain_pack": (
            {
                "path": str(Path((config or {}).get("domain_pack_dir")).expanduser().resolve()),
                "sha256": _resource_content_fingerprint(
                    Path((config or {}).get("domain_pack_dir"))
                ),
            }
            if stage == "llm-review" and (config or {}).get("domain_pack_dir") else None
        ),
        "template": ({"path": str(template), "sha256": _hash_file(template)}
                     if template and template.exists() and template.is_file() else None),
        "source_input": ({"path": str(Path(input_path).expanduser().resolve()),
                          "sha256": _hash_file(Path(input_path).expanduser().resolve())}
                         if input_path and Path(input_path).expanduser().resolve().is_file() else None),
        "config": config or {},
        "llm_pipeline": ({"path": str(DEFAULT_PIPELINE_PATH), "sha256": _hash_file(DEFAULT_PIPELINE_PATH)}
                         if stage in {"llm-review", "ai-extract", "functional-synthesis", "assemble", "requirements-analysis"}
                         and DEFAULT_PIPELINE_PATH.exists() else None),
        "llm": {key: os.environ.get(key, "") for key in (
            "RATOMIZER_LLM_BASE_URL", "RATOMIZER_LLM_MODEL", "RATOMIZER_LLM_MAX_TOKENS",
            "RATOMIZER_LLM_TEMPERATURE", "RATOMIZER_AI_UNIT_MODE", "RATOMIZER_AI_SELFCHECK",
            "RATOMIZER_AI_SELFCHECK_ROUNDS",
            # 二遍复核开关/轮数改变产物 → 指纹必须失效（专家审核 0715:缺席使复核
            # 变更后 chain 续跑直接跳过 ai-extract,新设置静默零生效）
            "RATOMIZER_AI_VERIFY", "RATOMIZER_AI_VERIFY_ROUNDS",
            # 合批条数改变 prompt 形状 → 产物可能不同 → 指纹必须失效（0714 批次二）
            "RATOMIZER_ANALYZE_BATCH", "RATOMIZER_ENRICH_BATCH",
        )},
        "requirements_analysis_enrich": (
            requirements_analysis_enrichment_enabled()
            if stage == "requirements-analysis" else None
        ),
    }
    encoded = json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _stage_manifest_path(out_dir: Path, stage: str) -> Path:
    root = Path(out_dir).expanduser().resolve()
    package_root = package_root_for_analysis_root(root)
    stages_root = (package_artifact_path(package_root, "stage_state")
                   if package_root is not None else root / STAGES_DIR)
    return stages_root / stage / "stage_manifest.json"


@contextmanager
def _run_manifest_lock(
    out_dir: Path,
    *,
    timeout_s: float = _MANIFEST_LOCK_TIMEOUT_S,
    stale_after_s: float = _MANIFEST_LOCK_STALE_AFTER_S,
) -> Iterator[None]:
    root = Path(out_dir).expanduser().resolve()
    root.mkdir(parents=True, exist_ok=True)
    with _manifest_process_lock_for(root):
        package_root = package_root_for_analysis_root(root)
        lock_path = (package_artifact_path(package_root, "run_manifest_lock", for_write=True)
                     if package_root is not None else root / "run_manifest.lock")
        deadline = time.monotonic() + timeout_s
        fd: int | None = None
        while fd is None:
            try:
                fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            except FileExistsError:
                if _remove_stale_manifest_lock(lock_path, stale_after_s):
                    continue
                if time.monotonic() >= deadline:
                    raise TimeoutError(f"timed out waiting for run manifest lock: {lock_path}")
                time.sleep(0.01)
        try:
            os.write(fd, str(os.getpid()).encode("ascii"))
            yield
        finally:
            os.close(fd)
            try:
                lock_path.unlink()
            except FileNotFoundError:
                pass


def _manifest_process_lock_for(out_dir: Path) -> RLock:
    with _MANIFEST_LOCKS_GUARD:
        return _MANIFEST_LOCKS.setdefault(out_dir, RLock())


def _remove_stale_manifest_lock(lock_path: Path, stale_after_s: float) -> bool:
    if stale_after_s < 0:
        return False
    try:
        age_s = time.time() - lock_path.stat().st_mtime
    except FileNotFoundError:
        return True
    if age_s < stale_after_s:
        return False
    try:
        lock_path.unlink()
    except FileNotFoundError:
        return True
    return True


def _replace_with_retry(source: Path, target: Path) -> None:
    for attempt in range(_REPLACE_ATTEMPTS):
        try:
            os.replace(source, target)
            return
        except PermissionError:
            if attempt + 1 >= _REPLACE_ATTEMPTS:
                raise
            time.sleep(_REPLACE_RETRY_DELAY_S)


def _atomic_write_json(path: Path, text: str) -> None:
    """原子写：临时文件 + os.replace。崩溃中途不会留下半截 JSON（旧 write_text 会，
    read_run_manifest 吞 JSONDecodeError 后静默丢弃所有阶段记录 → 强制全量重跑）。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
    try:
        with tmp.open("w", encoding="utf-8", newline="\n") as handle:
            handle.write(text)
            handle.flush()
            os.fsync(handle.fileno())
        _replace_with_retry(tmp, path)
    finally:
        try:
            tmp.unlink()
        except FileNotFoundError:
            pass


def _write_stage_manifest(out_dir: Path, stage: str, entry: dict[str, Any]) -> None:
    path = _stage_manifest_path(out_dir, stage)
    _atomic_write_json(path, json.dumps({"stage": stage, **entry}, ensure_ascii=False, indent=2) + "\n")


def _stage_is_reusable(out_dir: Path, stage: str, *,
                       route: str | None = None,
                       input_path: Path | None = None,
                       template_path: Path | None = None,
                       config: dict[str, Any] | None = None,
                       require_claim_generation: bool = True) -> bool:
    data = read_run_manifest(out_dir)
    stages = data.get("stages") if isinstance(data.get("stages"), dict) else {}
    entry = stages.get(stage) if isinstance(stages.get(stage), dict) else None
    # 出处方向性守卫（评审修正 2026-07-09）：请求真 LLM 时，复用 stub/来历不明的产物有害
    # （空行为需求被标"已完成"）——必须有带 route 的台账条目佐证；请求 stub/确定性阶段时，
    # 复用任何现成产物无害（最坏反而复用了更好的），保留遗留目录续跑价值（文件存在即可）。
    if not entry:
        return False
    if route == "openai_compatible" and entry.get("route") != route:
        return False
    if entry:
        if entry.get("status") != "ok":
            return False
        if entry.get("producer") and entry.get("producer") != stage_producer(
                stage, out_dir=out_dir, kb_paths=(config or {}).get("kb_paths")):
            return False
        recorded_fingerprint = str(entry.get("input_fingerprint") or "")
        recorded_route = str(entry.get("route") or "")
        fingerprint_route = recorded_route if route == "stub" and recorded_route == "openai_compatible" else route
        current_fingerprint = stage_input_fingerprint(
            out_dir, stage, route=fingerprint_route, template_path=template_path,
            input_path=input_path, config=config)
        if recorded_fingerprint != current_fingerprint:
            return False
        recorded_files = str(entry.get("input_files_fingerprint") or "")
        if recorded_files and recorded_files != stage_input_files_fingerprint(out_dir, stage):
            return False
    outputs = _stage_outputs(stage, entry)
    if stage == "ai-extract":
        # Runtime effective sidecars are independently recoverable. Legacy
        # manifests may list them, but their absence/staleness never reruns extraction.
        outputs = [
            name
            for name in outputs
            if Path(name).name not in _CLAIM_EFFECTIVE_RUNTIME_OUTPUTS
        ]
    if stage == "ai-extract" and not require_claim_generation:
        outputs = [name for name in outputs if Path(name).name not in _CLAIM_STAGE_OUTPUTS]
    if not _outputs_exist(out_dir, outputs):
        return False
    if stage == "ai-extract" and require_claim_generation:
        try:
            from claim_artifacts import (
                committed_base_versions_are_current,
                load_committed_claim_base,
            )

            snapshot = load_committed_claim_base(out_dir)
            if not committed_base_versions_are_current(snapshot):
                return False
            from ai_extract import (
                resolve_claim_shadow_verify,
                resolve_claim_shadow_verify_max_calls,
                resolve_claim_shadow_verify_max_total_tokens,
            )

            shadow_meta = dict(snapshot.get("generation_meta", {}).get("shadow_meta") or {})
            expected_verifier = bool(
                route not in {None, "stub"}
                and resolve_claim_shadow_verify()
                and resolve_claim_shadow_verify_max_calls() > 0
                and resolve_claim_shadow_verify_max_total_tokens() > 0
            )
            if shadow_meta.get("semantic_verifier_enabled") is not expected_verifier:
                return False
        except ClaimArtifactError:
            return False
    if stage == "atomize" and input_path is not None:
        try:
            manifest = json.loads((Path(out_dir) / "manifest.json").read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return False
        if Path(str(manifest.get("input") or "")).expanduser().resolve() != Path(input_path).expanduser().resolve():
            return False
    return True


def stage_is_reusable(out_dir: Path, stage: str, *,
                      route: str | None = None,
                      input_path: Path | None = None,
                      template_path: Path | None = None,
                      config: dict[str, Any] | None = None) -> bool:
    return _stage_is_reusable(
        out_dir,
        stage,
        route=route,
        input_path=input_path,
        template_path=template_path,
        config=config,
        require_claim_generation=True,
    )


# ---------------------------------------------------------------------------
# WS3 章节级增量重跑（条款候选粒度，默认关闭）
# ---------------------------------------------------------------------------
# 把 ``stage_is_reusable`` 的"整批指纹变化即全量重跑"思想细化到条款候选粒度：条款候选
# 自带单元格坐标（block_ids）与内容哈希，重解析后逐候选比对哈希，仅哈希变化者及其映射
# 的功能需求进重跑队列。默认 ``RATOMIZER_INCREMENTAL_RERUN=0``（关），``stage_is_reusable``
# 的 bool 契约与全量重跑行为不变；本判定是独立函数，由调用方在增量模式开时使用。与全量
# 重跑共用同一 ``claim_artifacts.hash_json`` 幂等键空间——重试 / 续跑不二次扣费。
INCREMENTAL_RERUN_ENV = "RATOMIZER_INCREMENTAL_RERUN"
_CLAUSE_CANDIDATE_HASH_DOMAIN = "clause-candidate-hash/v1"


def incremental_rerun_enabled(value: str | None = None) -> bool:
    """``RATOMIZER_INCREMENTAL_RERUN`` 是否开启（默认关）。"""
    raw = os.environ.get(INCREMENTAL_RERUN_ENV) if value is None else value
    return str(raw or "").strip().lower() in {"1", "true", "yes", "on"}


def _clause_candidate_key(chunk: dict[str, Any]) -> tuple[str, ...]:
    ids = tuple(str(b) for b in (chunk.get("block_ids") or []) if str(b))
    if ids:
        return ids
    # 无 block_ids 的候选回退到 section + heading 作为稳定身份键。
    return (str(chunk.get("section_id") or ""), str(chunk.get("heading") or ""))


def clause_candidate_fingerprint(chunk: dict[str, Any]) -> str:
    """单条款候选的内容哈希（section_path + heading + text + block_ids）。

    复用 ``claim_artifacts.hash_json``——与全量重跑的整批指纹、claim / llm_budget 幂等键
    同一命名空间。WS1 坐标证据（block_ids）天然就是增量检测键。
    """
    from claim_artifacts import hash_json

    payload = {
        "section_path": [str(s) for s in (chunk.get("section_path") or [])],
        "heading": str(chunk.get("heading") or ""),
        "text": str(chunk.get("text") or ""),
        "block_ids": [str(b) for b in (chunk.get("block_ids") or [])],
    }
    return hash_json(_CLAUSE_CANDIDATE_HASH_DOMAIN, payload)


def diff_clause_candidates(
    old_chunks: Sequence[dict[str, Any]],
    new_chunks: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """逐候选比对内容哈希，返回变化的候选身份键（block_ids 元组）。

    返回 ``{changed, added, removed, unchanged_count, old_count, new_count}``。
    ``changed ∪ added`` 即需重跑的候选范围。
    """
    old_by_key: dict[tuple[str, ...], str] = {}
    for chunk in old_chunks:
        old_by_key.setdefault(_clause_candidate_key(chunk), clause_candidate_fingerprint(chunk))
    new_by_key: dict[tuple[str, ...], str] = {}
    for chunk in new_chunks:
        new_by_key.setdefault(_clause_candidate_key(chunk), clause_candidate_fingerprint(chunk))

    changed: list[tuple[str, ...]] = []
    added: list[tuple[str, ...]] = []
    removed: list[tuple[str, ...]] = []
    unchanged = 0
    for key, fp in new_by_key.items():
        if key not in old_by_key:
            added.append(key)
        elif old_by_key[key] != fp:
            changed.append(key)
        else:
            unchanged += 1
    for key in old_by_key:
        if key not in new_by_key:
            removed.append(key)
    return {
        "changed": sorted(changed),
        "added": sorted(added),
        "removed": sorted(removed),
        "unchanged_count": unchanged,
        "old_count": len(old_by_key),
        "new_count": len(new_by_key),
    }


def incremental_rerun_plan(
    old_chunks: Sequence[dict[str, Any]],
    new_chunks: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """变化候选 → 重跑队列（block_ids 并集）+ 内容寻址幂等键。

    重跑队列 = ``(changed ∪ added)`` 候选的 block_ids 并集。``rerun_idempotency_key`` 绑定
    变化候选集合的内容哈希——与全量重跑共用同一 ``hash_json`` 幂等键空间。其映射的功能需求
    （经 ``source_block_ids`` 交集）进重跑队列由调用方在增量模式开时执行（本切片只交付判定
    机制，不改 functional_extract 等消费者）。
    """
    from claim_artifacts import hash_json

    diff = diff_clause_candidates(old_chunks, new_chunks)
    rerun_block_ids: list[str] = []
    for key in [*diff["changed"], *diff["added"]]:
        for block_id in key:
            if block_id and block_id not in rerun_block_ids:
                rerun_block_ids.append(block_id)
    total_new_blocks = sum(len(_clause_candidate_key(c)) for c in new_chunks if _clause_candidate_key(c))
    plan_key = hash_json("incremental-rerun-plan/v1", {
        "changed": diff["changed"],
        "added": diff["added"],
        "removed": diff["removed"],
    })
    return {
        "schema": "incremental-rerun-plan/v1",
        "enabled": incremental_rerun_enabled(),
        "rerun_block_ids": rerun_block_ids,
        "rerun_block_count": len(rerun_block_ids),
        "rerun_ratio": round(len(rerun_block_ids) / max(1, total_new_blocks), 4),
        "diff": diff,
        "rerun_idempotency_key": plan_key,
    }


# ---------------------------------------------------------------------------
# WS-E E1 首次全量闭合门
# ---------------------------------------------------------------------------
# 在 claim full 模式（三档已建）之上叠加「全部条目确认完成 → READY」门控：
# 未全确认 / 有 blocking 澄清 / 守恒未闭合 / claim 模式不是 full → 不 READY，并列出缺口清单。
# 复用既有 readiness 判定与 claim_ledger mode，不新造语义。
FULL_CLOSURE_SCHEMA = "full-closure/v1"


def evaluate_full_closure(out_dir: Path) -> dict[str, Any]:
    """Evaluate the WS-E E1 first full-closure gate.

    Returns ``{schema, ready, claim_mode, gaps}`` where ``gaps`` lists every
    blocking condition. The gate is strict: any gap means ``ready=False``.
    """
    root = Path(out_dir).expanduser().resolve()
    gaps: list[dict[str, Any]] = []

    # 1. claim ledger 模式必须是 full（sampling/baseline_gate 不保证全量闭合）
    claim_mode = claim_ledger.resolve_claim_ledger_mode()
    if claim_mode != "full":
        gaps.append({
            "kind": "claim_mode_not_full",
            "current": claim_mode,
            "required": "full",
            "message": f"Claim ledger mode is '{claim_mode}'; full closure requires 'full'",
        })

    # 2. claim 侧 document_ready + 无 uncertain claim + 无 pending 结构复核
    try:
        from claim_views import build_claim_view

        metrics = build_claim_view(root, "metrics")
        document_ready = metrics.get("document_ready")
        effective_metrics = dict(metrics.get("effective_metrics") or {})
        uncertain_count = int(effective_metrics.get("uncertain_count") or 0)
        pending_structural = int(metrics.get("structural_review_pending_count") or 0)
        if not document_ready:
            gaps.append({
                "kind": "claim_document_not_ready",
                "document_ready": document_ready,
                "uncertain_count": uncertain_count,
                "pending_structural_reviews": pending_structural,
            })
        if uncertain_count > 0:
            gaps.append({
                "kind": "uncertain_claims",
                "count": uncertain_count,
                "message": f"{uncertain_count} eligible claim(s) remain uncertain",
            })
        if pending_structural > 0:
            gaps.append({
                "kind": "pending_structural_reviews",
                "count": pending_structural,
                "message": f"{pending_structural} structural review(s) pending",
            })
    except Exception as exc:  # pragma: no cover - defensive
        gaps.append({
            "kind": "claim_metrics_unavailable",
            "error": str(exc)[:200],
        })

    # 3. 分析侧 readiness（blocking 澄清、覆盖率、失败章节）
    try:
        from agent_state import load_analysis_state

        state = load_analysis_state(root)
        readiness = state.readiness
        if readiness.get("verdict") != "READY":
            gaps.append({
                "kind": "readiness_not_ready",
                "verdict": readiness.get("verdict"),
                "reasons": list(readiness.get("reasons") or []),
            })
        # 全部 B-track 条目必须有非 draft/candidate 的裁决
        unreviewed_statuses = {"draft", "candidate"}
        unreviewed = [
            req for req in state.requirements
            if str(req.get("status") or "draft").lower() in unreviewed_statuses
        ]
        if unreviewed:
            gaps.append({
                "kind": "unreviewed_requirements",
                "count": len(unreviewed),
                "sample_ids": [
                    str(req.get("ai_req_id") or req.get("requirement_id") or "")
                    for req in unreviewed[:10]
                ],
            })
    except Exception as exc:  # pragma: no cover - defensive
        gaps.append({
            "kind": "analysis_state_unavailable",
            "error": str(exc)[:200],
        })

    # 4. functional_extract 守恒门（A-5，2026-08-07）：注释承诺"守恒未闭合 → 不 READY"落地。
    # 原实现只检查 claim 模式 / claim ready / 分析 readiness / 全部已裁决，从未直接消费
    # functional_extract 守恒状态——守恒未闭合时仍可能判 ready=True（缺口清单也不含
    # conservation_open）。此处显式接入：守恒计算过且 ok=False → ready=False 且缺口含
    # conservation_open。守恒未计算（functional_requirements.json 缺席）不阻塞，与
    # orchestration_gaps._conservation_gaps / adjudicate.hard_basis_check 同口径（不伪造
    # 未计算的信号）。
    try:
        from requirements_analysis_rules import _read_functional_requirements_payload

        fr_payload = _read_functional_requirements_payload(root) or {}
        conservation = fr_payload.get("conservation")
        if isinstance(conservation, dict) and not conservation.get("ok", True):
            missing = [str(b) for b in (conservation.get("missing_block_ids") or []) if str(b)]
            extra = [str(b) for b in (conservation.get("extra_block_ids") or []) if str(b)]
            duplicate = [str(b) for b in (conservation.get("duplicate_assignments") or []) if str(b)]
            mismatches = conservation.get("evidence_mismatches") or []
            mismatch_count = len(mismatches) if isinstance(mismatches, list) else 0
            gaps.append({
                "kind": "conservation_open",
                "missing_block_ids": missing,
                "extra_block_ids": extra,
                "duplicate_assignments": duplicate,
                "evidence_mismatches_count": mismatch_count,
                "message": (
                    f"functional_extract 守恒未闭合：missing={len(missing)} "
                    f"extra={len(extra)} duplicate={len(duplicate)} "
                    f"evidence_mismatches={mismatch_count}"
                ),
            })
    except Exception as exc:  # pragma: no cover - defensive
        gaps.append({
            "kind": "conservation_check_unavailable",
            "error": str(exc)[:200],
        })

    return {
        "schema": FULL_CLOSURE_SCHEMA,
        "ready": len(gaps) == 0,
        "claim_mode": claim_mode,
        "gaps": gaps,
    }


# ---------------------------------------------------------------------------
# WS-E E2 增量闭合：文档 diff 驱动变更集
# ---------------------------------------------------------------------------
# 复用 WS3 incremental_rerun_plan 的条款候选哈希，把变化 block_ids 映射到需求，
# 产出新增/失效/沿用三分类报告。
CHANGESET_SCHEMA = "requirement-changeset/v1"


def build_requirement_changeset(
    old_requirements: Sequence[dict[str, Any]],
    new_requirements: Sequence[dict[str, Any]],
    old_chunks: Sequence[dict[str, Any]],
    new_chunks: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """Build an added/obsolete/retained changeset from clause-candidate diff.

    ``old_chunks`` / ``new_chunks`` are clause-candidate descriptors as consumed
    by ``incremental_rerun_plan``. A new requirement is *added* if its id did not
    exist in the old set; *obsolete* if it disappeared; *retained* if it stayed.
    Retained requirements whose ``source_block_ids`` intersect the changed block
    set are flagged with ``reason: source_changed`` so the caller can rerun them.
    """
    plan = incremental_rerun_plan(old_chunks, new_chunks)
    changed_block_ids = set(plan["rerun_block_ids"])

    def req_id(req: dict[str, Any]) -> str:
        return str(req.get("ai_req_id") or req.get("requirement_id") or "")

    def source_blocks(req: dict[str, Any]) -> set[str]:
        return {str(b) for b in (req.get("source_block_ids") or []) if str(b)}

    old_by_id = {req_id(r): r for r in old_requirements}
    new_by_id = {req_id(r): r for r in new_requirements}

    added: list[dict[str, Any]] = []
    obsolete: list[dict[str, Any]] = []
    retained: list[dict[str, Any]] = []

    for req_id_key, req in new_by_id.items():
        blocks = source_blocks(req)
        changed = blocks & changed_block_ids
        if req_id_key not in old_by_id:
            added.append({
                "id": req_id_key,
                "reason": "new_requirement",
                "source_blocks": sorted(blocks),
                "changed_source_blocks": sorted(changed),
            })
        else:
            retained.append({
                "id": req_id_key,
                "reason": "source_changed" if changed else "unchanged",
                "source_blocks": sorted(blocks),
                "changed_source_blocks": sorted(changed),
            })

    for req_id_key, req in old_by_id.items():
        if req_id_key not in new_by_id:
            obsolete.append({
                "id": req_id_key,
                "reason": "removed_requirement",
                "source_blocks": sorted(source_blocks(req)),
            })

    return {
        "schema": CHANGESET_SCHEMA,
        "incremental_rerun_plan": plan,
        "added": added,
        "obsolete": obsolete,
        "retained": retained,
        "counts": {
            "added": len(added),
            "obsolete": len(obsolete),
            "retained": len(retained),
        },
    }


def ai_requirements_are_reusable(
    out_dir: Path,
    *,
    route: str | None,
    config: dict[str, Any] | None = None,
) -> bool:
    """Validate the extraction layer while deliberately ignoring claim artifacts."""
    if not _stage_is_reusable(
        out_dir,
        "ai-extract",
        route=route,
        config=config,
        require_claim_generation=False,
    ):
        return False
    root = Path(out_dir).expanduser().resolve()
    requirements_path = root / "ai_requirements.jsonl"
    metadata_path = root / "ai_requirements.meta.json"
    if not requirements_path.is_file() or not metadata_path.is_file():
        return False
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        from ai_extract import extraction_input_fingerprint

        return (
            isinstance(metadata, dict)
            and metadata.get("schema") == "ai-requirements-final/v1"
            and str(metadata.get("input_fingerprint") or "")
            == extraction_input_fingerprint(root)
            and int(metadata.get("failed_sections") or 0) == 0
        )
    except (OSError, ValueError, json.JSONDecodeError):
        return False


# 链内跳过各阶段 summary（0714 批次二 S7）：8 个阶段各算一遍 build_output_summary
# （三份 jsonl 全量读+遍历）然后被 chain 逐个 pop 丢弃,链尾统一算一份即可。
# chain 单线程顺序跑,module 级布尔够用（desktop 任务本就每命令独立进程）。
_CHAIN_ACTIVE = False


def _stage_summary(out_dir: Path) -> dict[str, Any]:
    if _CHAIN_ACTIVE:
        return {}
    return build_output_summary(out_dir)


def _claim_effective_summary(out_dir: Path) -> dict[str, Any]:
    try:
        from claim_views import build_claim_view

        view = build_claim_view(out_dir, "metrics")
        effective_metrics = dict(view.get("effective_metrics") or {})
        return {
            "document_ready": view.get("document_ready"),
            "effective_fresh": bool(view.get("effective_fresh")),
            "open_claim_count": effective_metrics.get("uncertain_count"),
        }
    except Exception as exc:
        return {
            "document_ready": None,
            "effective_fresh": False,
            "open_claim_count": None,
            "effective_error": str(exc)[:300],
        }


def _claim_component_manifest(out_dir: Path) -> dict[str, Any]:
    try:
        from claim_artifacts import load_committed_effective_snapshot_readonly

        snapshot = load_committed_effective_snapshot_readonly(out_dir)
        generation = dict(snapshot["generation_meta"])
        shadow_meta = dict(generation.get("shadow_meta") or {})
        effective = dict(snapshot["effective_meta"])
        return {
            "base_generation_id": effective.get("base_generation_id"),
            "document_effective_revision": effective.get(
                "document_effective_revision"
            ),
            "event_prefix_sha256": effective.get("event_prefix_sha256"),
            "last_event_seq": effective.get("last_event_seq"),
            "catalog": dict(snapshot.get("catalog_meta") or {}).get(
                "catalog_version"
            ),
            "coverage": dict(shadow_meta.get("versions") or {}).get(
                "coverage_validator"
            ),
            "effective": effective.get("effective_ledger_schema"),
            "bridge": effective.get("bridge_version"),
            "reducer": effective.get("reducer_version"),
            "queue": effective.get("queue_version"),
            "versions": dict(effective.get("versions") or {}),
        }
    except Exception as exc:
        return {"status": "unavailable", "error": str(exc)[:300]}


def skipped_stage_payload(out_dir: Path, stage: str) -> dict[str, Any]:
    data = read_run_manifest(out_dir)
    stages = data.get("stages") if isinstance(data.get("stages"), dict) else {}
    entry = stages.get(stage) if isinstance(stages.get(stage), dict) else {}
    outputs = _stage_outputs(stage, entry)
    root = Path(out_dir).expanduser().resolve()
    payload = {
        "kind": stage.replace("-", "_"),
        "out_dir": str(root),
        "resume_action": "skipped",
        "written": [str(root / name) for name in outputs],
    }
    if stage == "ai-extract":
        try:
            from claim_artifacts import load_committed_shadow

            snapshot = load_committed_shadow(root)
            shadow_meta = dict(snapshot.get("generation_meta", {}).get("shadow_meta") or {})
            payload["claim_shadow"] = {
                "status": "published",
                "accounting_status": shadow_meta.get("accounting_status"),
                "resolution_status": shadow_meta.get("resolution_status"),
                "termination_reason": shadow_meta.get("termination_reason"),
                "metrics": snapshot.get("metrics") or {},
                **_claim_effective_summary(root),
            }
        except Exception as exc:
            payload["claim_shadow"] = {"status": "stale", "error": str(exc)[:300]}
    return payload


def update_run_manifest(out_dir: Path, stage: str, status: str, *,
                        error: str | None = None, route: str | None = None,
                        outputs: list[str | Path] | None = None,
                        action: str | None = None,
                        input_fingerprint: str | None = None,
                        input_files_fingerprint: str | None = None,
                        template_path: Path | None = None,
                        input_path: Path | None = None,
                        config: dict[str, Any] | None = None) -> None:
    """run_manifest.json：out_dir 的显式状态账本（阶段/状态/版本/时间）。写失败不阻断任务。"""
    import datetime as _dt
    root = Path(out_dir).expanduser().resolve()
    try:
        package_root = package_root_for_analysis_root(root)
        path = (package_artifact_path(package_root, "run_manifest", for_write=True)
                if package_root is not None else root / RUN_MANIFEST)
    except ResultPackageError as exc:
        # marker/journal 损坏不阻断阶段记账（"写失败不阻断"契约，2026-08-03 审查 I2）
        LOGGER.warning("run_manifest 记账：结果包布局不可读，退回分析根直写：%s", exc)
        package_root = None
        path = root / RUN_MANIFEST
    attempt_run_id: str | None = None
    if package_root is not None:
        try:
            package = load_result_package(package_root)
        except ResultPackageError as exc:
            LOGGER.warning("run_manifest 记账：结果包 marker 不可读，跳过 attempt 绑定：%s", exc)
        else:
            active_attempt = package.get("active_attempt")
            if (
                isinstance(active_attempt, dict)
                and stage in (active_attempt.get("requested_stages") or [])
            ):
                attempt_run_id = str(active_attempt.get("run_id") or "") or None
    now = _dt.datetime.now().isoformat(timespec="seconds")
    try:
        with _run_manifest_lock(root):
            data = read_run_manifest(root)
            stages = data.get("stages")
            if not isinstance(stages, dict):
                stages = {}
            entry = stages.get(stage) if isinstance(stages.get(stage), dict) else {}
            producer = stage_producer(stage, out_dir=root, kb_paths=(config or {}).get("kb_paths"))
            if status == "running":
                entry = {"status": "running", "started": now, "producer": producer}
            else:
                entry.update({"status": status, "finished": now, "producer": producer})
                if route:
                    entry["route"] = route   # stub 降级 ≠ 真 LLM：账本必须可区分（2026-07-08 审计）
                if outputs is not None:
                    entry["outputs"] = _relative_outputs(root, outputs)
                elif status == "ok" and "outputs" not in entry and stage in STAGE_REQUIRED_OUTPUTS:
                    entry["outputs"] = list(STAGE_REQUIRED_OUTPUTS[stage])
                if status == "ok":
                    entry["input_fingerprint"] = input_fingerprint or stage_input_fingerprint(
                        root, stage, route=route, template_path=template_path,
                        input_path=input_path, config=config)
                    if input_files_fingerprint:
                        entry["input_files_fingerprint"] = input_files_fingerprint
                if action:
                    entry["last_action"] = action
                if error:
                    entry["error"] = str(error)[:300]
                else:
                    entry.pop("error", None)
                if stage == "ai-extract" and status in {"ok", "partial"}:
                    entry["claim_components"] = _claim_component_manifest(root)
            if attempt_run_id is not None:
                entry["attempt_run_id"] = attempt_run_id
            stages[stage] = entry
            data.update({"manifest_version": 2, "stages": stages, "updated": now})
            _atomic_write_json(path, json.dumps(data, ensure_ascii=False, indent=2) + "\n")
            _write_stage_manifest(root, stage, entry)
    except OSError:  # pragma: no cover - manifest 写失败不阻断任务本体
        LOGGER.warning("run_manifest 写入失败（忽略）：%s", path)


def _stage_completion_status(stage: str, payload: Any) -> str:
    """Return a non-reusable terminal status when a stage produced incomplete output."""
    if stage == "ai-extract" and isinstance(payload, dict):
        quality = payload.get("quality") if isinstance(payload.get("quality"), dict) else {}
        failed_sections = int(payload.get("failed_sections") or quality.get("failed_sections") or 0)
        if failed_sections > 0:
            return "partial"
    if stage == "export-annotation-html" and isinstance(payload, dict):
        translations = payload.get("translations")
        if isinstance(translations, dict):
            unresolved = int(translations.get("unresolved") or 0)
            failed_calls = int(translations.get("failed_calls") or 0)
            if unresolved > 0 or failed_calls > 0:
                return "partial"
    return "ok"


def chain_task(out_dir: Path, *, stages: list[str], route: str = "stub",
               template_path: Path | None = None,
               sample_ratio: float | None = None,
               limit_sections: int | None = None,
               annotation_layout_mode: str = "pdf_original") -> dict[str, Any]:
    """交付物链的后端单命令编排（F1：编排从 App.vue 搬回后端——headless/批量/CI 的地基）。

    阶段按 CHAIN_ORDER 归一排序去重；template-write 无模板路径时前置报错（不跑一半才死）；
    逐阶段发进度事件并记 run_manifest；任一阶段失败 → 记账后整链响亮失败。
    """
    out_dir = out_dir.expanduser().resolve()
    unknown = [s for s in stages if s not in CHAIN_ORDER]
    if unknown:
        raise ValueError(f"未知阶段：{', '.join(unknown)}（可用：{', '.join(CHAIN_ORDER)}）")
    ordered = [s for s in CHAIN_ORDER if s in set(stages)]
    if not ordered:
        raise ValueError("阶段清单为空")
    if "template-write" in ordered and template_path is None:
        raise ValueError("template-write 阶段需要 --template（公司需求列表模板路径）")

    # 归一化与单命令落账一致（R8，0710 评审）：显式 0/0.0 与 None 指纹必须同形
    sample_ratio = sample_ratio or None
    limit_sections = limit_sections or None
    ai_config = {"sample_ratio": sample_ratio, "limit_sections": limit_sections}
    ai_dependent = {"functional-synthesis", "requirements-analysis", "template-write",
                    "clarification-report"}
    if route == "stub" and ai_dependent.intersection(ordered) and not stage_is_reusable(
            out_dir, "ai-extract", route="stub", config=ai_config):
        raise ValueError(
            "当前链包含功能重组/需求分析等 AI 依赖阶段，但没有可复用的 AI 抽取产物；"
            "请使用 openai_compatible 完成 AI 抽取后再继续。"
        )

    runners: dict[str, Any] = {
        "ai-extract": lambda: ai_extract_task(out_dir, route=route,
                                              limit_sections=limit_sections,
                                              sample_ratio=sample_ratio),
        "functional-synthesis": lambda: functional_synthesis_task(out_dir, route=route),
        "assemble": lambda: assemble_task(out_dir, enrich_route=route if route != "stub" else None),
        "requirements-analysis": lambda: requirements_analysis_task(
            out_dir, route=route, template_path=template_path),
        "template-write": lambda: template_write_task(out_dir, template_path),
        "clarification-report": lambda: clarification_report_task(out_dir),
        "compose": lambda: compose_task(out_dir),
        "export-annotation-html": lambda: export_annotation_html_task(
            out_dir, route=route, layout_mode=annotation_layout_mode),
    }

    results: dict[str, Any] = {}
    skipped_stages: list[str] = []
    payload: dict[str, Any] = {"kind": "chain", "out_dir": str(out_dir), "stages": ordered}
    global _CHAIN_ACTIVE
    _CHAIN_ACTIVE = True   # 链内各阶段跳过 summary;finally 复位,失败路径不污染后续任务
    chain_budget = _attach_budget_ledger_for_run(out_dir)  # S1-1：开启预算单时挂账本
    try:
        llm_stages = {"ai-extract", "functional-synthesis", "assemble", "requirements-analysis",
                      "export-annotation-html"}
        for index, stage in enumerate(ordered, start=1):
            emit_progress({"stage": "chain", "step": stage, "completed": index - 1,
                           "total": len(ordered), "percent": int((index - 1) * 100 / len(ordered))})
            stage_route = route if stage in llm_stages else None
            stage_template = template_path if stage in {"requirements-analysis", "template-write"} else None
            if stage == "ai-extract":
                stage_config = {"sample_ratio": sample_ratio, "limit_sections": limit_sections}
            elif stage == "export-annotation-html":
                stage_config = {"layout_mode": annotation_layout_mode}
            else:
                stage_config = None
            reusable = stage_is_reusable(
                out_dir,
                stage,
                route=stage_route,
                template_path=stage_template,
                config=stage_config,
            )
            if (stage == "ai-extract" and not reusable
                    and ai_requirements_are_reusable(
                        out_dir,
                        route=stage_route,
                        config=stage_config,
                    )):
                scope = "sample" if sample_ratio or limit_sections else "full"
                try:
                    refresh = ai_extract.refresh_claim_shadow(
                        out_dir,
                        route=stage_route,
                        scope=scope,
                    )
                except Exception as exc:
                    error = f"ai-extract claim shadow refresh failed: {exc}"
                    update_run_manifest(
                        out_dir,
                        stage,
                        "failed",
                        route=stage_route,
                        error=error,
                        action="ledger_refresh_failed",
                        template_path=stage_template,
                        config=stage_config,
                    )
                    raise RuntimeError(error) from exc
                existing_stages = read_run_manifest(out_dir).get("stages", {})
                existing_entry = (
                    existing_stages.get(stage, {}) if isinstance(existing_stages, dict) else {}
                )
                outputs = list(dict.fromkeys([
                    *_stage_outputs(stage, existing_entry),
                    *(refresh.get("written") or []),
                ]))
                refresh["written"] = [str(out_dir / name) for name in outputs]
                refresh["resume_action"] = "ledger_refreshed"
                update_run_manifest(
                    out_dir,
                    stage,
                    "ok",
                    route=str(existing_entry.get("route") or stage_route or "") or None,
                    outputs=outputs,
                    action="ledger_refreshed",
                    template_path=stage_template,
                    config=stage_config,
                )
                results[stage] = refresh
                emit_progress({
                    "stage": "chain",
                    "step": stage,
                    "status": "ledger_refreshed",
                    "completed": index,
                    "total": len(ordered),
                    "percent": int(index * 100 / len(ordered)),
                })
                continue
            if reusable:
                stage_payload = skipped_stage_payload(out_dir, stage)
                skipped_stages.append(stage)
                existing_stages = read_run_manifest(out_dir).get("stages", {})
                existing_entry = existing_stages.get(stage, {}) if isinstance(existing_stages, dict) else {}
                preserved_route = str(existing_entry.get("route") or stage_route or "") or None
                update_run_manifest(out_dir, stage, "ok", route=preserved_route,
                                    outputs=stage_payload.get("written") or _stage_outputs(stage), action="skipped",
                                    input_fingerprint=str(existing_entry.get("input_fingerprint") or "") or None,
                                    template_path=stage_template, config=stage_config)
                emit_progress({"stage": "chain", "step": stage, "status": "skipped",
                               "completed": index, "total": len(ordered),
                               "percent": int(index * 100 / len(ordered))})
            else:
                update_run_manifest(out_dir, stage, "running")
                try:
                    # S1-1：LLM 承载环节进预算单 stage 上下文（按 _CHAIN_BUDGET_STAGES 映射；
                    # 无活动预算单时空操作，确定性环节不进账本分摊）
                    with _budget_stage(chain_budget, _CHAIN_BUDGET_STAGES.get(stage, "default")):
                        stage_payload = runners[stage]()
                except Exception as exc:
                    update_run_manifest(out_dir, stage, "failed", error=str(exc))
                    raise RuntimeError(f"{stage} 阶段失败：{exc}") from exc
                stage_outputs = stage_payload.get("written") if isinstance(stage_payload, dict) else None
                actual_route = (str(stage_payload.get("route") or "").strip()
                                if isinstance(stage_payload, dict) else "") or stage_route
                leased_input_files = (str(stage_payload.get("_input_files_fingerprint") or "")
                                      if isinstance(stage_payload, dict) else "") or None
                completion_status = _stage_completion_status(stage, stage_payload)
                update_run_manifest(out_dir, stage, completion_status, route=actual_route,
                                    outputs=stage_outputs or _stage_outputs(stage), action="ran",
                                    template_path=stage_template, config=stage_config,
                                    input_files_fingerprint=leased_input_files)
            stage_payload = dict(stage_payload or {})
            stage_payload.pop("_input_files_fingerprint", None)
            stage_payload.pop("summary", None)   # 各阶段的 summary 体积大且重复，链尾统一给一份
            results[stage] = stage_payload
            # 顶层聚合：GUI 消息只看这几个键，不必翻 results
            if stage == "ai-extract":
                for key in ("consistency", "sampled", "quality", "count", "claim_shadow"):
                    if stage_payload.get(key) is not None:
                        payload[key] = stage_payload[key]
            elif stage == "requirements-analysis":
                payload["analysis"] = stage_payload.get("analysis")
            elif stage == "template-write":
                payload["template"] = stage_payload.get("report")
            elif stage == "clarification-report":
                payload["readiness"] = stage_payload.get("readiness")
                payload["questions"] = stage_payload.get("questions")
            # 阶段降级/告警上提（2026-07-08 审计 2-C）：此前 note 埋在 results 里，
            # stub 降级/部分章节失败时 GUI 一律显示「运行完成」全绿
            note = stage_payload.get("note")
            analysis = stage_payload.get("analysis")
            if not note and isinstance(analysis, dict):
                note = analysis.get("note")
            if note:
                payload.setdefault("stage_notes", []).append(f"{stage}: {note}")
            if stage not in skipped_stages:
                emit_progress({"stage": "chain", "step": stage, "completed": index,
                               "total": len(ordered), "percent": int(index * 100 / len(ordered))})

        from adjudication_bank import resolve_bank_path, update_bank
        bank_path = resolve_bank_path()
        if bank_path and "requirements-analysis" in ordered:
            try:   # 收割失败不影响链结果
                payload["adjudication_bank"] = update_bank(bank_path, out_dir)
            except Exception as exc:  # pragma: no cover
                LOGGER.warning("裁决样本库收割失败（忽略）：%s", exc)
        # WS-H：成文导出后自动 harvest（默认关，env RATOMIZER_HARVEST=1 启用）
        if "template-write" in ordered and os.environ.get("RATOMIZER_HARVEST", "").strip().lower() in {"1", "true", "yes", "on"}:
            try:
                from harvest import harvest_assets
                payload["harvest"] = harvest_assets(out_dir, actor="chain-harvest")
            except Exception as exc:  # pragma: no cover
                LOGGER.warning("WS-H harvest 失败（忽略）：%s", exc)
        payload["results"] = results
        payload["skipped_stages"] = skipped_stages
        # V3 WS-A A3：整篇对账 sidecar（RATOMIZER_RECONCILE=1 时链尾自动跑一次；
        # 默认关=行为面零变化；sidecar 失败不阻断链结果）
        try:
            from reconcile import reconcile_enabled
            if reconcile_enabled():
                payload["reconcile"] = reconcile_task(out_dir, route=route)
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning("整篇对账 sidecar 失败（不阻断链）：%s", exc)
    finally:
        _CHAIN_ACTIVE = False
        _detach_budget_ledger(chain_budget)  # S1-1：落盘 cost-report 数据源 + 卸载钩子
    payload["summary"] = build_output_summary(out_dir)
    return payload


@_leased_pipeline_stage("export-annotation-html")
def export_annotation_html_task(out_dir: Path, route: str | None = None,
                                layout_mode: str = "pdf_original") -> dict[str, Any]:
    """生成可分享的文档批注 HTML bundle（内含 localStorage 裁决 + 导出 JSON）。

    route=openai_compatible 时补齐块级"说明"标记的原文中文翻译（内容哈希缓存
    annotation_translations.json，重导出零调用）；渲染本体保持确定性。"""
    import doc_annotation_export
    out_dir = out_dir.expanduser().resolve()
    path, translations = doc_annotation_export.export_annotation_bundle(
        out_dir, route=route, layout_mode=layout_mode)
    written = [str(path)]
    if translations.get("source_pdf"):
        written.append(str(translations["source_pdf"]))
    written.extend(str(item) for item in (translations.get("page_files") or []))
    payload = {"kind": "annotation_html", "out_dir": str(out_dir),
               "path": str(path), "route": str(translations.get("route") or "stub"),
               "layout_mode_requested": translations.get("layout_mode_requested"),
               "layout_mode": translations.get("layout_mode"),
               "source_pdf": translations.get("source_pdf"),
               "annotation_overlay": bool(translations.get("annotation_overlay")),
               "translations": translations, "written": written}
    notes: list[str] = []
    if route and route != "stub" and payload["route"] != "openai_compatible":
        notes.append("LLM 不可用，块级说明未翻译（原文照排，开启后重导出自动补齐）")
    if translations.get("rejected"):
        notes.append(f"{translations['rejected']} 条翻译含无据编码/数字被拒（保留原文）")
    if translations.get("failed_calls"):
        notes.append(f"{translations['failed_calls']} 批翻译调用失败（重新导出自动补齐）")
    if translations.get("pdf_render_error"):
        notes.append("PDF 批注覆盖层生成失败，已回退浏览器原版查看器")
    if str(translations.get("layout_mode") or "") == "pdf_original":
        # 数据处置提醒（0714 评审）：原版影印 bundle 内含完整原始 PDF + 整页影印图,
        # 数据面与优化模式（仅抽取片段）完全不同——分享文件夹=分享整份客户文档
        notes.append("原版影印导出为文件夹（含完整原始 PDF 与整页影印图），对外分享前请确认可提供整份文档")
    if notes:
        payload["note"] = "；".join(notes)
    return payload


def import_ai_decisions_task(out_dir: Path, decisions_file: Path) -> dict[str, Any]:
    """把 HTML 导出的裁决 JSON 回灌到 ai_review_states.jsonl（合进交付物）。"""
    import ai_review_actions
    out_dir = out_dir.expanduser().resolve()
    data = json.loads(Path(decisions_file).expanduser().read_text(encoding="utf-8"))
    decisions = data.get("decisions") if isinstance(data, dict) else data
    applied = 0
    skipped = 0
    needs_reconfirmation = 0
    conflicts = 0
    ownership_skipped = 0
    for d in (decisions or []):
        rid = str((d or {}).get("ai_req_id") or "").strip()
        status = str((d or {}).get("status") or "").strip()
        if not rid or not status:
            skipped += 1
            continue
        submitted_source = str(d.get("source_fingerprint") or "").strip()
        submitted_subject = str(d.get("review_subject_fingerprint") or "").strip()
        expected_target_fingerprint = str(
            d.get("expected_target_fingerprint") or ""
        ).strip()
        expected_target_publication_revision = str(
            d.get("expected_target_publication_revision") or ""
        ).strip()
        expected_target_authority_write_revision = str(
            d.get("expected_target_authority_write_revision") or ""
        ).strip()
        if not all((
            submitted_source,
            submitted_subject,
            expected_target_fingerprint,
            expected_target_publication_revision,
            expected_target_authority_write_revision,
        )):
            skipped += 1
            needs_reconfirmation += 1
            continue
        # 归属值单独校验：仅归属非法时丢归属、保留整行裁决（status/模块/意见不陪葬）
        ownership = str(d.get("ownership_override") or "").strip() or None
        if ownership:
            try:
                normalize_ownership(ownership)
            except ValueError:
                ownership = None
                ownership_skipped += 1
        try:
            from api_server import find_current_ai_requirement
            from omission_actions import extraction_operation_lock

            with extraction_operation_lock(out_dir, operation="import-ai-decision"):
                current = find_current_ai_requirement(out_dir, rid)
                if current is None:
                    raise ai_review_actions.AIReviewAuthorityConflict(
                        "AI requirement is not present in the current run",
                        current_revision="",
                    )
                current_cas = (
                    str(current.get("source_fingerprint") or ""),
                    str(current.get("review_subject_fingerprint") or ""),
                    str(current.get("target_fingerprint") or ""),
                    str(current.get("target_publication_revision") or ""),
                    str(current.get("target_authority_write_revision") or ""),
                )
                submitted_cas = (
                    submitted_source,
                    submitted_subject,
                    expected_target_fingerprint,
                    expected_target_publication_revision,
                    expected_target_authority_write_revision,
                )
                if submitted_cas != current_cas:
                    raise ai_review_actions.AIReviewAuthorityConflict(
                        "AI requirement or review authority changed",
                        current_revision=current_cas[-1],
                    )
                ai_review_actions.apply_ai_review_action(
                    out_dir,
                    rid,
                    status,
                    module_override=(d.get("module_override") or None),
                    ownership_override=ownership,
                    reason=(d.get("reason") or ""),
                    actor="html-import",
                    source_fingerprint_value=current_cas[0],
                    review_subject_fingerprint_value=current_cas[1],
                    review_anchor_fingerprint_value=(
                        ai_review_actions.review_anchor_fingerprint(current)
                    ),
                    expected_target_authority_write_revision=current_cas[-1],
                )
            applied += 1
        except ai_review_actions.AIReviewAuthorityConflict:
            skipped += 1
            conflicts += 1
        except ValueError:
            skipped += 1
    payload: dict[str, Any] = {"kind": "ai_decisions_import", "out_dir": str(out_dir),
                               "applied": applied, "skipped": skipped}
    if needs_reconfirmation:
        payload["needs_reconfirmation"] = needs_reconfirmation
    if conflicts:
        payload["conflicts"] = conflicts
    if ownership_skipped:
        payload["ownership_skipped"] = ownership_skipped
    # 裁决回流交付物：导入后立即重建 merged_spec（免 LLM）
    if applied and (out_dir / "ai_requirements.jsonl").exists():
        rebuilt = ai_extract.rebuild_merged_spec(out_dir)
        payload["rebuilt"] = rebuilt
        try:
            from adjudication_bank import resolve_bank_path, update_bank

            bank_path = resolve_bank_path()
            if bank_path is not None:
                payload["adjudication_bank"] = update_bank(bank_path, out_dir)
        except Exception as exc:  # 导入/重建已成功；学习资产失败只留告警
            LOGGER.warning("HTML 裁决导入后样本库收割失败（忽略）：%s", exc)
    return payload


def build_output_summary(out_dir: Path) -> dict[str, Any]:
    out_dir = out_dir.expanduser().resolve()
    requirements = read_jsonl(out_dir / "atomic_requirements.jsonl")
    reviews = read_jsonl(out_dir / "llm_review_results.jsonl")
    states = read_jsonl(
        governed_artifact_path(out_dir, "review_states.jsonl", category="state")
    )
    status_counts: dict[str, int] = {}
    type_counts: dict[str, int] = {}
    confidence_counts = {"high": 0, "medium": 0, "low": 0}
    for row in requirements:
        requirement_type = str(row.get("requirement_type") or row.get("type") or "unknown")
        type_counts[requirement_type] = type_counts.get(requirement_type, 0) + 1
        confidence = row.get("confidence")
        if isinstance(confidence, (int, float)):
            if confidence >= 0.9:
                confidence_counts["high"] += 1
            elif confidence >= 0.7:
                confidence_counts["medium"] += 1
            else:
                confidence_counts["low"] += 1
    for state in states:
        status = str(state.get("status") or "unknown")
        status_counts[status] = status_counts.get(status, 0) + 1
    return {
        "counts": {
            "requirements": len(requirements),
            "reviews": len(reviews),
            "review_states": len(states),
        },
        "status_counts": status_counts,
        "type_counts": type_counts,
        "confidence_counts": confidence_counts,
        "run_manifest": read_run_manifest(out_dir),
    }


def orchestrate_task(
    out_dir: Path,
    *,
    max_rounds: int | None = None,
    allow_llm: bool = False,
    actor: str = "orchestration-loop",
) -> dict[str, Any]:
    """T2 编排环入口（CHAIN_ORDER 之外的 sidecar，同 agent-loop 纪律）。

    读四类缺口 → 经既有 allow_llm 授权通道发起 spot_extract/targeted_reextract → 写
    orchestration_trace.jsonl，直到收敛或达上限。裁决仍在专家面板；编排环只决定"该看哪里"。
    默认 allow_llm=False（只读缺口 + extract 缺口转人工），显式授权才发起 LLM 补抽。
    """
    from orchestration_loop import (
        resolve_allow_llm,
        resolve_max_rounds,
        run_orchestration_loop,
    )

    root = out_dir.expanduser().resolve()
    # CLI 显式值优先；缺省时 ENV（RATOMIZER_ORCHESTRATION_MAX_ROUNDS）覆盖，再缺省回 8。
    rounds = resolve_max_rounds(max_rounds)
    authorized = resolve_allow_llm(bool(allow_llm))
    summary = run_orchestration_loop(
        root, max_rounds=rounds, allow_llm=authorized, actor=actor
    )
    return {
        "kind": "orchestrate",
        "out_dir": str(root),
        "summary": summary,
        "written": [
            summary.get("trace_file") or "orchestration_trace.jsonl",
            "orchestration_summary.json",
        ],
    }


def reconcile_task(out_dir: Path, *, route: str = "stub") -> dict[str, Any]:
    """V3 WS-A A3 整篇对账入口（CHAIN_ORDER 之外的 sidecar，同 orchestrate 纪律）。

    规则筛疑 + LLM 裁定两段；LLM 不可用（stub/无 key/预算耗尽）如实 rules_only。
    产物 reconcile_report.json（governed pipeline）+ 摘要并入根 quality_report.json。
    """
    from reconcile import run_reconcile

    root = out_dir.expanduser().resolve()
    return run_reconcile(root, route=route)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run Requirement Atomizer desktop tasks.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    run_parser = subparsers.add_parser("run")
    run_parser.add_argument("--input", type=Path, required=True)
    run_parser.add_argument("--out", type=Path, required=True)
    run_parser.add_argument("--skip-review", action="store_true")
    run_parser.add_argument("--llm-route", choices=["stub", "openai_compatible"], default=None)
    run_parser.add_argument("--review-scope", choices=["targeted", "all"], default=None)
    run_parser.add_argument("--llm-review-limit", type=int, default=0)
    run_parser.add_argument("--chunk-chars", type=int, default=3500)
    run_parser.add_argument("--kb", type=Path, action="append", default=[])
    run_parser.add_argument("--domain-pack", type=Path, default=None)

    export_parser = subparsers.add_parser("export")
    export_parser.add_argument("--out", type=Path, required=True)
    export_parser.add_argument("--formats", default="csv,md")

    assemble_parser = subparsers.add_parser("assemble")
    assemble_parser.add_argument("--out", type=Path, required=True)
    assemble_parser.add_argument("--formats", default="xlsx,docx,md")
    assemble_parser.add_argument("--enrich-route", default="")
    assemble_parser.add_argument("--blue-book-index", type=Path, default=None)

    synthesis_parser = subparsers.add_parser("functional-synthesis")
    synthesis_parser.add_argument("--out", type=Path, required=True)
    synthesis_parser.add_argument("--llm-route", choices=["stub", "openai_compatible"], default="stub")

    compose_parser = subparsers.add_parser("compose")
    compose_parser.add_argument("--out", type=Path, required=True)

    bank_parser = subparsers.add_parser("adjudication-bank")
    bank_parser.add_argument("--out", type=Path, required=True)
    bank_parser.add_argument("--bank", type=Path, required=True)

    answers_parser = subparsers.add_parser("import-clarification-answers")
    answers_parser.add_argument("--out", type=Path, required=True)
    answers_parser.add_argument("--file", type=Path, required=True)

    chain_parser = subparsers.add_parser("chain")
    chain_parser.add_argument("--out", type=Path, required=True)
    chain_parser.add_argument("--stages", required=True, help="逗号分隔的阶段清单（按依赖自动排序）")
    # 默认与独立子命令一致（openai_compatible）：headless 忘带路由时"没配 key 响亮失败"
    # 优于"静默 stub 产空行为需求且 manifest 全 ok"（2026-07-08 审计 A4）
    chain_parser.add_argument("--llm-route", default="openai_compatible", choices=["stub", "openai_compatible"])
    chain_parser.add_argument("--template", type=Path, default=None)
    chain_parser.add_argument("--sample-ratio", type=float, default=None)
    chain_parser.add_argument("--limit-sections", type=int, default=None)
    chain_parser.add_argument("--annotation-layout-mode", choices=["optimized", "pdf_original"],
                              default="pdf_original")

    clarification_parser = subparsers.add_parser("clarification-report")
    clarification_parser.add_argument("--out", type=Path, required=True)

    template_write_parser = subparsers.add_parser("template-write")
    template_write_parser.add_argument("--out", type=Path, required=True)
    template_write_parser.add_argument("--template", type=Path, required=True)

    requirements_analysis_parser = subparsers.add_parser("requirements-analysis")
    requirements_analysis_parser.add_argument("--out", type=Path, required=True)
    requirements_analysis_parser.add_argument("--template", type=Path, default=None)
    requirements_analysis_parser.add_argument("--llm-route", choices=["stub", "openai_compatible"], default="stub")

    ai_extract_parser = subparsers.add_parser("ai-extract")
    ai_extract_parser.add_argument("--out", type=Path, required=True)
    ai_extract_parser.add_argument("--llm-route", choices=["stub", "openai_compatible"], default="openai_compatible")
    ai_extract_parser.add_argument("--limit-sections", type=int, default=0)
    ai_extract_parser.add_argument("--sample-ratio", type=float, default=0.0)

    summary_parser = subparsers.add_parser("summary")
    summary_parser.add_argument("--out", type=Path, required=True)

    anno_parser = subparsers.add_parser("export-annotation-html")
    anno_parser.add_argument("--out", type=Path, required=True)
    anno_parser.add_argument("--route", choices=["stub", "openai_compatible"], default=None,
                             help="openai_compatible 时补齐块级说明标记的中文翻译（缓存复用）")
    anno_parser.add_argument("--layout-mode", choices=["optimized", "pdf_original"],
                             default="pdf_original")

    import_parser = subparsers.add_parser("import-ai-decisions")
    import_parser.add_argument("--out", type=Path, required=True)
    import_parser.add_argument("--file", type=Path, required=True, help="HTML 导出的裁决 JSON")

    # --- WS4 能力补齐子命令（全程零 LLM 调用）---
    verify_import_parser = subparsers.add_parser(
        "import-verification",
        help="回灌线下改过的 software_requirements.xlsx 六列 → verification_states.jsonl")
    verify_import_parser.add_argument("--out", type=Path, required=True)
    verify_import_parser.add_argument("--file", type=Path, required=True)
    verify_import_parser.add_argument("--actor", default="desktop-verification")

    verify_set_parser = subparsers.add_parser(
        "set-verification", help="直接写入一条需求 verification 覆盖（状态机前进迁移）")
    verify_set_parser.add_argument("--out", type=Path, required=True)
    verify_set_parser.add_argument("--requirement-id", required=True)
    verify_set_parser.add_argument("--implemented", default=None,
                                   choices=["not_started", "in_progress", "done"])
    verify_set_parser.add_argument("--test-completed", default=None, choices=["true", "false"])
    verify_set_parser.add_argument("--test-case-ids", default=None, help="分号/逗号分隔的测试用例号")
    verify_set_parser.add_argument("--confirm-pm", default=None, choices=["true", "false"])
    verify_set_parser.add_argument("--confirm-tl", default=None, choices=["true", "false"])
    verify_set_parser.add_argument("--confirm-dt", default=None, choices=["true", "false"])
    verify_set_parser.add_argument("--actor", default="desktop-verification")

    rollback_parser = subparsers.add_parser(
        "rollback-requirement", help="人工回退需求生命周期（回退事件 append-only 留痕）")
    rollback_parser.add_argument("--out", type=Path, required=True)
    rollback_parser.add_argument("--requirement-id", required=True)
    rollback_parser.add_argument("--target", required=True,
                                 choices=["draft", "confirmed", "implemented", "verified"])
    rollback_parser.add_argument("--actor", required=True)
    rollback_parser.add_argument("--reason", required=True)

    manual_parser = subparsers.add_parser(
        "add-manual-requirement", help="手工建需求（provenance=manual，追溯列留空不伪引）")
    manual_parser.add_argument("--out", type=Path, required=True)
    manual_parser.add_argument("--objective", required=True)
    manual_parser.add_argument("--behaviors", default=None, help="逗号分隔的行为列表")
    manual_parser.add_argument("--module", default="")
    manual_parser.add_argument("--ownership", default="", choices=["", "software", "hardware", "co_design"])
    manual_parser.add_argument("--priority", default="P1")
    manual_parser.add_argument("--notes", default="")
    manual_parser.add_argument("--actor", default="desktop-manual")

    lib_build_parser = subparsers.add_parser(
        "build-requirement-library", help="汇总各项目 functional_requirements 为 JSONL 检索库")
    lib_build_parser.add_argument("--projects", type=Path, nargs="+", required=True)
    lib_build_parser.add_argument("--library", type=Path, required=True, help="输出的检索库 JSONL 路径")
    lib_build_parser.add_argument(
        "--include-unconfirmed", action="store_true",
        help="默认仅收录 lifecycle>=confirmed 的条目；此开关显式收录 draft（未确认）条目")

    lib_search_parser = subparsers.add_parser(
        "search-requirements", help="词面集合重叠度召回历史相似需求")
    lib_search_parser.add_argument("--library", type=Path, required=True)
    lib_search_parser.add_argument("--query", required=True)
    lib_search_parser.add_argument("--limit", type=int, default=20)

    dep_rec_parser = subparsers.add_parser(
        "recommend-dependencies", help="确定性依赖/父子候选推荐（只生产值，不动 schema）")
    dep_rec_parser.add_argument("--out", type=Path, required=True)

    dep_dec_parser = subparsers.add_parser(
        "decide-dependency", help="依赖候选裁决（接受才写库，拒绝不落库）")
    dep_dec_parser.add_argument("--out", type=Path, required=True)
    dep_dec_parser.add_argument("--from", dest="from_id", required=True)
    dep_dec_parser.add_argument("--to", required=True)
    dep_dec_parser.add_argument("--kind", required=True, choices=["depend", "exclude", "refine"])
    dep_dec_parser.add_argument("--accept", choices=["true", "false"], default="true")
    dep_dec_parser.add_argument("--actor", default="desktop-dependency")
    dep_dec_parser.add_argument("--reason", default="")

    claim_acceptance_parser = subparsers.add_parser("claim-shadow-acceptance")
    claim_acceptance_parser.add_argument("--input", type=Path, required=True)
    claim_acceptance_parser.add_argument("--output", type=Path)

    claim_packet_parser = subparsers.add_parser("claim-shadow-review-packet")
    claim_packet_parser.add_argument("--input", type=Path, required=True)
    claim_packet_parser.add_argument("--output-dir", type=Path, required=True)

    claim_import_parser = subparsers.add_parser("claim-shadow-review-import")
    claim_import_parser.add_argument("--input", type=Path, required=True)
    claim_import_parser.add_argument("--decisions", type=Path, required=True)
    claim_import_parser.add_argument("--output", type=Path, required=True)
    claim_import_parser.add_argument("--golden-manifest", type=Path, required=True)

    claim_fold_parser = subparsers.add_parser("claim-ledger-fold")
    claim_fold_parser.add_argument(
        "--out-dir",
        "--out",
        dest="out",
        type=Path,
        required=True,
    )

    # T2 编排环（agent_loop 升格）：缺口驱动的再规划，裁决仍在专家面板。
    orchestrate_parser = subparsers.add_parser(
        "orchestrate",
        help="编排环：读缺口→授权补抽→写 trace，直到收敛或达上限（NEEDS WORK 交人）")
    orchestrate_parser.add_argument("--out-dir", "--out", dest="out", type=Path, required=True)
    orchestrate_parser.add_argument(
        "--max-rounds", type=int, default=None,
        help=f"每文档最大编排轮次（默认 8，上限 50；env RATOMIZER_ORCHESTRATION_MAX_ROUNDS）")
    orchestrate_parser.add_argument(
        "--allow-llm", action="store_true",
        help="授权编排环发起 spot_extract/targeted_reextract（默认关闭=只读缺口转人工；"
             "env RATOMIZER_ORCHESTRATION_ALLOW_LLM=1 等效）")
    orchestrate_parser.add_argument("--actor", default="orchestration-loop")

    # V3 WS-A A3：整篇对账 sidecar（CHAIN_ORDER 之外，同 orchestrate 纪律）。
    reconcile_parser = subparsers.add_parser(
        "reconcile",
        help="整篇对账：规则筛疑+LLM 裁定两段，写 reconcile_report.json 并并入 quality_report")
    reconcile_parser.add_argument("--out-dir", "--out", dest="out", type=Path, required=True)
    reconcile_parser.add_argument(
        "--llm-route", choices=["stub", "openai_compatible"], default="stub",
        help="裁定投票路由（默认 stub=仅规则筛疑 rules_only）")

    # WS-H：知识沉淀闭环（成文导出后自动/手动 harvest）
    harvest_parser = subparsers.add_parser(
        "harvest",
        help="执行 WS-H 知识沉淀闭环：收割裁决样本、confirmed 需求、方案、领域知识、语言资产、校准资产",
    )
    harvest_parser.add_argument("--out", type=Path, required=True)
    harvest_parser.add_argument("--actor", default="desktop-harvest")

    cost_report_parser = subparsers.add_parser("cost-report")
    cost_report_parser.add_argument(
        "--out-dir", "--out", dest="out", type=Path, required=True,
        help="结果目录（读 governed state/llm_budget.json）",
    )

    package_start_parser = subparsers.add_parser("result-package-start")
    package_start_parser.add_argument("--out", type=Path, required=True)
    package_start_parser.add_argument("--input", type=Path, required=True)
    package_start_parser.add_argument("--stages", required=True)

    package_complete_parser = subparsers.add_parser("result-package-complete")
    package_complete_parser.add_argument("--out", type=Path, required=True)
    package_complete_parser.add_argument("--run-id", required=True)
    package_complete_parser.add_argument("--completed-stages", required=True)

    package_fail_parser = subparsers.add_parser("result-package-fail")
    package_fail_parser.add_argument("--out", type=Path, required=True)
    package_fail_parser.add_argument("--run-id", required=True)
    package_fail_parser.add_argument("--error", required=True)

    package_status_parser = subparsers.add_parser("result-package-status")
    package_status_parser.add_argument("--out", type=Path, required=True)
    package_status_parser.add_argument(
        "--verify",
        action="store_true",
        # S5：显式完整校验（「打开已有结果」）——重算交付物/完成证据 SHA
        help="recompute deliverable and completion-evidence hashes (fail on mismatch)",
    )
    return parser.parse_args(argv)


def _manifest_context_from_args(args: argparse.Namespace) -> dict[str, Any]:
    command = str(getattr(args, "command", "") or "")
    context: dict[str, Any] = {}
    if command in {"llm-review", "ai-extract", "requirements-analysis"}:
        context["route"] = getattr(args, "llm_route", None)
    elif command == "assemble":
        context["route"] = getattr(args, "enrich_route", None) or None
    elif command == "export-annotation-html":
        context["route"] = getattr(args, "route", None) or None
        context["config"] = {"layout_mode": getattr(args, "layout_mode", "pdf_original")}
    if command in {"requirements-analysis", "template-write"}:
        context["template_path"] = getattr(args, "template", None)
    if command == "ai-extract":
        context["config"] = {
            "sample_ratio": getattr(args, "sample_ratio", 0.0) or None,
            "limit_sections": getattr(args, "limit_sections", 0) or None,
        }
    return context


def setup_run_logging(out_dir: Path | None, *, allow_root_files: bool = True) -> None:
    """后端任务日志：stderr（Electron 收集持久化）+ <输出目录>/run.log（跟着交付物走）。

    此前 GUI 路径全链路零日志：LOGGER.info（LLM 调用时长/自检轮次/富化被拒原因/降级）被
    Python 兜底 handler 丢弃，Electron 只在任务失败时把 stderr 拼进弹窗。排查"为什么慢/
    为什么产物长这样"无从下手——本函数让每次运行在输出目录留下完整可追溯日志。幂等可重入。

    allow_root_files=False 时（只读探测，如 summary 预览空目录）不在非 package 目录的
    根创建 run.log/llm_trace.jsonl：根目录偶发文件只是垃圾，不产生任何价值。
    """
    logger = logging.getLogger("requirement_atomizer")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
    have = {getattr(h, "_ratomizer_tag", None) for h in logger.handlers}
    if "stderr" not in have:
        handler = logging.StreamHandler(sys.stderr)
        handler.setFormatter(fmt)
        handler._ratomizer_tag = "stderr"  # type: ignore[attr-defined]
        logger.addHandler(handler)
    if out_dir is not None and "runlog" not in have:
        try:
            out_dir = Path(out_dir).expanduser().resolve()
            package_root = package_root_for_analysis_root(out_dir)
            if package_root is None and detect_result_layout(out_dir) == "package_v1":
                package_root = out_dir
            log_path: Path | None = None
            if package_root is not None:
                log_path = package_artifact_path(package_root, "run_log", for_write=True)
            elif allow_root_files:
                log_path = out_dir / "run.log"
            if log_path is not None:
                log_path.parent.mkdir(parents=True, exist_ok=True)
                file_handler = logging.FileHandler(log_path, encoding="utf-8")
                file_handler.setFormatter(fmt)
                file_handler._ratomizer_tag = "runlog"  # type: ignore[attr-defined]
                logger.addHandler(file_handler)
        except OSError:  # 日志落盘失败不阻断任务
            pass
    # LLM 消息级追踪：完整收发落 <out>/llm_trace.jsonl（含 prompt/响应全文/token 用量）。
    # 默认开（体量 ~几 MB/次运行，与 blocks.jsonl 同量级）；设 RATOMIZER_LLM_TRACE=0 可关。
    if out_dir is not None and os.environ.get("RATOMIZER_LLM_TRACE", "").strip() != "0":
        import llm_client
        resolved = Path(out_dir).expanduser().resolve()
        package_root = package_root_for_analysis_root(resolved)
        if package_root is None and detect_result_layout(resolved) == "package_v1":
            package_root = resolved
        trace_path: Path | None = None
        if package_root is not None:
            trace_path = package_artifact_path(package_root, "llm_trace", for_write=True)
        elif allow_root_files:
            trace_path = resolved / "llm_trace.jsonl"
        if trace_path is not None:
            llm_client.set_trace_path(trace_path)


def teardown_run_logging() -> None:
    """关闭并摘除 run.log 的 FileHandler（任务进程一次性；不关会锁住输出目录文件句柄）。"""
    logger = logging.getLogger("requirement_atomizer")
    for handler in list(logger.handlers):
        if getattr(handler, "_ratomizer_tag", None) == "runlog":
            handler.close()
            logger.removeHandler(handler)
    import llm_client
    llm_client.set_trace_path(None)


def _fail_with_envelope(kind: str, error_type: str, exc: BaseException, code: int) -> int:
    """结构化失败面：stdout 落 JSON envelope（CLI 契约），stderr 落同一 JSON 行
    （Electron runDesktopTaskProcess 非零退出时以 stderr 为错误消息）。"""
    envelope = {
        "kind": kind,
        "ok": False,
        "error": {"type": error_type, "message": str(exc)},
    }
    print_json_payload(envelope)
    print(json.dumps(envelope, ensure_ascii=True), file=sys.stderr)
    return code


def _result_package_main(args: argparse.Namespace) -> int:
    """result-package-* 子命令：成功/失败都输出单个 JSON envelope，exit code 结构化——
    0 成功；2 输入/布局拒绝（如 legacy 目录需显式迁移）；3 marker/journal 损坏等校验失败；
    1 未分类异常（traceback 落 stderr）。2026-08-03 审查 S1/I2：此前异常直接裸 traceback。
    """
    kind = args.command.replace("-", "_")
    try:
        package_root = args.out.expanduser().resolve()
        if args.command == "result-package-start":
            package = initialize_result_package(
                package_root,
                input_path=args.input,
                requested_stages=split_formats(args.stages),
            )
        elif args.command == "result-package-complete":
            package = commit_analysis_completion(
                package_root,
                run_id=args.run_id,
                completed_stages=split_formats(args.completed_stages),
            )
        elif args.command == "result-package-fail":
            package = record_analysis_failure(
                package_root,
                run_id=args.run_id,
                error=args.error,
            )
        else:  # result-package-status
            layout = detect_result_layout(package_root)
            package = (
                load_result_package(package_root, verify=bool(getattr(args, "verify", False)))
                if layout == "package_v1"
                else None
            )
            print_json_payload({
                "kind": kind,
                "ok": True,
                "out_dir": str(package_root),
                "analysis_root": str(resolve_analysis_root(package_root)),
                "layout": layout,
                "package": package,
            })
            return 0
        print_json_payload({
            "kind": kind,
            "ok": True,
            "out_dir": str(package_root),
            "analysis_root": str(resolve_analysis_root(package_root)),
            "package": package,
        })
        return 0
    except ResultPackageCorrupt as exc:
        # S5：verify 发现交付物/完成证据哈希不一致是独立稳定错误面——
        # 桌面端据此显示"结果文件已被修改"
        if "changed" in str(exc):
            return _fail_with_envelope(
                kind, "result_package_modified",
                ResultPackageCorrupt(f"结果文件已被修改：{exc}"), 3,
            )
        return _fail_with_envelope(kind, "result_package_corrupt", exc, 3)
    except ResultPackagePartialError as exc:
        # I6：部分阶段降级是稳定错误码——桌面端据此显示"分析未完成（部分阶段
        # 降级）"而非"运行失败"；语义仍 fail-closed（exit 2，不冒充完成）
        # R2：拒绝即锁内持久化终止 attempt（marker 不得停留 running/running，
        # 否则重开结果误显"运行中"）；持久化失败不掩盖原错误码
        if args.command == "result-package-complete":
            try:
                record_analysis_partial(
                    package_root, run_id=args.run_id, error=str(exc),
                )
            except Exception:
                LOGGER.warning("partial attempt 状态持久化失败", exc_info=True)
        return _fail_with_envelope(kind, "requested_stage_partial", exc, 2)
    except ResultPackageError as exc:
        # 含 "legacy flat output requires explicit migration" 等布局拒绝
        return _fail_with_envelope(kind, "input_error", exc, 2)
    except Exception as exc:
        logging.getLogger("requirement_atomizer").exception("%s 失败", args.command)
        return _fail_with_envelope(kind, "internal_error", exc, 1)


def _cost_report_main(args: argparse.Namespace) -> int:
    """WS3 成本看板：读文档级预算单记账流水，输出分环节消耗 / 缓存命中 / 路由分布。

    数据全部来自预算单记账流水（无新增埋点）。无预算单（未启用 ``RATOMIZER_LLM_BUDGET``）
    时如实返回 ``available=false``，绝不伪造数据。
    """
    from llm_budget import LLMBudgetLedger, cost_report

    out_dir = Path(args.out).expanduser().resolve()
    try:
        ledger = LLMBudgetLedger.load(out_dir)
    except (OSError, ValueError) as exc:
        return _fail_with_envelope("cost-report", "input_error", exc, 2)
    if ledger is None:
        payload = {
            "ok": True,
            "kind": "cost-report",
            "available": False,
            "reason": "no document budget ledger (RATOMIZER_LLM_BUDGET disabled)",
        }
        print(json.dumps(payload, ensure_ascii=False))
        return 0
    report = cost_report(ledger)
    print(json.dumps({
        "ok": True, "kind": "cost-report", "available": True, "report": report,
    }, ensure_ascii=False))
    return 0


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if args.command in _RESULT_PACKAGE_COMMANDS:
        return _result_package_main(args)
    if args.command == "claim-shadow-acceptance":
        from claim_acceptance import main as claim_acceptance_main
        forwarded = ["--input", str(args.input)]
        if args.output is not None:
            forwarded.extend(["--output", str(args.output)])
        return claim_acceptance_main(forwarded)
    if args.command == "claim-shadow-review-packet":
        from claim_review_packet import main as claim_review_packet_main
        return claim_review_packet_main([
            "--input", str(args.input),
            "--output-dir", str(args.output_dir),
        ])
    if args.command == "claim-shadow-review-import":
        from claim_review_import import main as claim_review_import_main
        return claim_review_import_main([
            "--input", str(args.input),
            "--decisions", str(args.decisions),
            "--output", str(args.output),
            "--golden-manifest", str(args.golden_manifest),
        ])
    if args.command == "cost-report":
        return _cost_report_main(args)
    package_root: Path | None = None
    original_out = getattr(args, "out", None)
    out_layout: str | None = None
    if original_out is not None:
        original_out = Path(original_out).expanduser().resolve()
        try:
            out_layout = detect_result_layout(original_out)
        except ResultPackageError as exc:
            # 残留发布 journal / 损坏 marker：结构化失败，不再对所有桌面命令裸 traceback
            # （2026-08-03 审查 I2）。
            if isinstance(exc, ResultPackageCorrupt):
                return _fail_with_envelope(args.command, "result_package_corrupt", exc, 3)
            return _fail_with_envelope(args.command, "input_error", exc, 2)
        if out_layout == "package_v1":
            package_root = original_out
            args.out = resolve_analysis_root(original_out)
    setup_run_logging(
        getattr(args, "out", None),
        # 只读 summary 预览空目录时不在其根留 run.log/llm_trace.jsonl——那是纯垃圾文件
        # （哨兵判定已不再被它们干扰，但也没必要往用户新目录里写东西）。
        allow_root_files=not (args.command == "summary" and out_layout == "empty"),
    )
    logging.getLogger("requirement_atomizer").info("desktop task 开始：%s", args.command)
    try:
        if args.command == "run":
            payload = run_pipeline_task(
                args.input,
                args.out,
                skip_review=args.skip_review,
                llm_route=args.llm_route,
                review_scope=args.review_scope,
                llm_review_limit=args.llm_review_limit,
                chunk_chars=args.chunk_chars,
                kb_paths=args.kb,
                domain_pack_dir=args.domain_pack,
            )
        elif args.command == "export":
            payload = export_task(args.out, split_formats(args.formats))
        elif args.command == "assemble":
            payload = assemble_task(
                args.out,
                formats=split_formats(args.formats),
                enrich_route=args.enrich_route or None,
                blue_book_index_path=args.blue_book_index,
            )
        elif args.command == "functional-synthesis":
            payload = functional_synthesis_task(args.out, route=args.llm_route)
        elif args.command == "compose":
            payload = compose_task(args.out)
        elif args.command == "adjudication-bank":
            from adjudication_bank import update_bank
            payload = {"kind": "adjudication_bank", **update_bank(args.bank, args.out)}
        elif args.command == "import-clarification-answers":
            payload = import_clarification_workbook_task(args.out, args.file)
        elif args.command == "chain":
            payload = chain_task(args.out, stages=[x.strip() for x in args.stages.split(",") if x.strip()],
                                 route=args.llm_route, template_path=args.template,
                                 sample_ratio=args.sample_ratio or None,
                                 limit_sections=args.limit_sections or None,
                                 annotation_layout_mode=args.annotation_layout_mode)
        elif args.command == "clarification-report":
            payload = clarification_report_task(args.out)
        elif args.command == "template-write":
            payload = template_write_task(args.out, args.template)
        elif args.command == "requirements-analysis":
            payload = requirements_analysis_task(args.out, route=args.llm_route, template_path=args.template)
        elif args.command == "ai-extract":
            payload = ai_extract_task(args.out, route=args.llm_route,
                                      limit_sections=args.limit_sections or None,
                                      sample_ratio=args.sample_ratio or None)
        elif args.command == "export-annotation-html":
            payload = export_annotation_html_task(
                args.out,
                route=getattr(args, "route", None),
                layout_mode=args.layout_mode,
            )
        elif args.command == "import-ai-decisions":
            payload = import_ai_decisions_task(args.out, args.file)
        elif args.command == "import-verification":
            payload = import_verification_workbook_task(args.out, args.file, actor=args.actor)
        elif args.command == "set-verification":
            payload = set_verification_task(
                args.out, args.requirement_id,
                implemented=args.implemented,
                test_completed=(None if args.test_completed is None else args.test_completed == "true"),
                test_case_ids=args.test_case_ids,
                confirm_pm=(None if args.confirm_pm is None else args.confirm_pm == "true"),
                confirm_tl=(None if args.confirm_tl is None else args.confirm_tl == "true"),
                confirm_dt=(None if args.confirm_dt is None else args.confirm_dt == "true"),
                actor=args.actor,
            )
        elif args.command == "rollback-requirement":
            payload = rollback_requirement_task(
                args.out, args.requirement_id, args.target, actor=args.actor, reason=args.reason)
        elif args.command == "add-manual-requirement":
            payload = add_manual_requirement_task(
                args.out, objective=args.objective, behaviors=args.behaviors, module=args.module,
                ownership=args.ownership, priority=args.priority, notes=args.notes, actor=args.actor)
        elif args.command == "build-requirement-library":
            payload = build_requirement_library_task(
                args.projects, args.library, include_unconfirmed=args.include_unconfirmed)
        elif args.command == "search-requirements":
            payload = search_requirements_task(args.library, args.query, limit=args.limit)
        elif args.command == "recommend-dependencies":
            payload = recommend_dependencies_task(args.out)
        elif args.command == "decide-dependency":
            payload = decide_dependency_task(
                args.out, frm=args.from_id, to=args.to, kind=args.kind,
                accepted=args.accept == "true", actor=args.actor, reason=args.reason)
        elif args.command == "claim-ledger-fold":
            from claim_review_actions import fold_effective_ledger

            payload = {
                "kind": "claim_ledger_fold",
                **fold_effective_ledger(
                    args.out,
                    actor_trigger="desktop-claim-ledger-fold",
                ),
            }
        elif args.command == "orchestrate":
            payload = orchestrate_task(
                args.out,
                max_rounds=args.max_rounds,
                allow_llm=args.allow_llm,
                actor=args.actor,
            )
        elif args.command == "reconcile":
            payload = reconcile_task(args.out, route=args.llm_route)
        elif args.command == "harvest":
            from harvest import harvest_assets
            payload = {"kind": "harvest", **harvest_assets(args.out, actor=args.actor)}
        else:
            payload = {"kind": "summary", "out_dir": str(args.out.expanduser().resolve()), "summary": build_output_summary(args.out)}
    except Exception as exc:
        logging.getLogger("requirement_atomizer").exception("desktop task 失败：%s", args.command)
        # 单步命令也记账（chain 内部已逐阶段记，不重复）
        if args.command in CHAIN_ORDER and getattr(args, "out", None):
            update_run_manifest(args.out, args.command, "failed", error=str(exc),
                                **_manifest_context_from_args(args))
        print(json.dumps({"error": str(exc)}, ensure_ascii=True), file=sys.stderr)
        return 1
    finally:
        teardown_run_logging()
    if args.command in CHAIN_ORDER and getattr(args, "out", None):
        manifest_context = _manifest_context_from_args(args)
        actual_route = str(payload.get("route") or "").strip() if isinstance(payload, dict) else ""
        if actual_route:
            manifest_context["route"] = actual_route
        if args.command == "export-annotation-html" and isinstance(payload, dict):
            manifest_context["outputs"] = payload.get("written") or None
        leased_input_files = (str(payload.pop("_input_files_fingerprint", "") or "")
                              if isinstance(payload, dict) else "") or None
        if leased_input_files:
            manifest_context["input_files_fingerprint"] = leased_input_files
        update_run_manifest(
            args.out,
            args.command,
            _stage_completion_status(args.command, payload),
            **manifest_context,
        )
    if package_root is not None and isinstance(payload, dict):
        # out_dir 永远还原为 package root，不把 .ratomizer/pipeline 泄漏给 Electron 最近会话。
        payload["out_dir"] = str(package_root)
        published = _maybe_publish_after_command(args.command, package_root, payload)
        if published:
            published_by_id = {item["artifact_id"]: item for item in published}
            normalized_written: list[str] = []
            for value in payload.get("written") or []:
                name = Path(value).name
                match = next(
                    (
                        item for item in published_by_id.values()
                        if Path(item["path"]).name == name
                    ),
                    None,
                )
                normalized_written.append(
                    str(package_root / match["path"]) if match is not None else str(value)
                )
            if "written" in payload:
                payload["written"] = normalized_written
    print_json_payload(payload)
    return 0


def _maybe_publish_after_command(
    command: str,
    package_root: Path,
    payload: dict[str, Any],
) -> list[dict[str, Any]] | None:
    """已完成结果上的写命令结束后发布根交付物；返回发布清单（未发布为 None）。

    失败降级（2026-08-03 审查 I2，spec §11）：发布异常（锁超时/磁盘/journal 损坏）
    不掩盖已成功的阶段——记 run.log + marker warnings[] + payload warnings，
    run_manifest 里的阶段结果保持原样。
    """
    if command not in PUBLISHING_COMMANDS:
        return None
    try:
        package = load_result_package(package_root)
    except ResultPackageError as exc:
        LOGGER.warning("结果包不可读，跳过交付物发布：%s", exc)
        payload.setdefault("warnings", []).append(
            f"deliverable publication skipped: {exc}"
        )
        return None
    if isinstance(package.get("active_attempt"), dict):
        # 活动 attempt：根交付物保持上一完成代，由 result-package-complete 一次性发布。
        return None
    try:
        return publish_registered_deliverables(package_root)
    except Exception as exc:  # 锁超时/磁盘错误/journal 损坏等
        LOGGER.exception("交付物发布失败（阶段本身已成功）：%s", exc)
        warning = f"deliverable publication failed: {exc}"
        payload.setdefault("warnings", []).append(warning)
        try:
            record_package_warning(package_root, warning)
        except Exception:  # marker 本身不可写时只留日志
            LOGGER.warning("marker warnings 记录失败", exc_info=True)
        return None


def split_formats(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def emit_progress(event: dict[str, Any]) -> None:
    print(f"{PROGRESS_PREFIX}{json.dumps(event, ensure_ascii=True)}", flush=True)


def print_json_payload(payload: dict[str, Any]) -> None:
    """Write IPC JSON safely even when packaged Windows stdout is GBK."""
    print(json.dumps(payload, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    raise SystemExit(main())
