from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_MODULES = (
    "系统需求",
    "计量需求",
    "时钟需求",
    "费率需求",
    "显示需求",
    "需量需求",
    "结算需求",
    "负荷曲线",
    "报警窃电需求",
    "电网质量需求",
    "升级需求",
    "负控需求",
    "状态字需求",
    "事件需求",
    "协议栈需求",
    "push需求",
    "P1需求",
    "MBUS需求",
    "预付费需求",
)

_SKIPPED_SHEET_TITLES = {
    "需求模版Release notes",
    "原始需求对应表",
    "需求变更管理",
}


def fallback_template_vocabulary() -> dict[str, Any]:
    modules = list(DEFAULT_MODULES)
    return {
        "modules": modules,
        "submodules_by_module": {module: [] for module in modules},
    }


def extract_template_vocabulary(path: Path | None) -> dict[str, Any]:
    if path is None or not path.exists():
        return fallback_template_vocabulary()

    workbook = load_workbook(path, data_only=True, read_only=True)
    modules: list[str] = []
    submodules_by_module: dict[str, list[str]] = {}

    try:
        for worksheet in workbook.worksheets:
            module = worksheet.title.strip()
            if not module or module.endswith("列表") or module in _SKIPPED_SHEET_TITLES:
                continue

            modules.append(module)
            submodules_by_module[module] = _extract_submodules(worksheet)
    finally:
        workbook.close()

    return {
        "modules": modules,
        "submodules_by_module": submodules_by_module,
    }


def _extract_submodules(worksheet: Worksheet) -> list[str]:
    header_column: int | None = None
    header_row: int | None = None

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=5, values_only=True),
        start=1,
    ):
        for column_index, cell_value in enumerate(row):
            if _normalize_cell_value(cell_value) == "子模块":
                header_column = column_index
                header_row = row_number
                break
        if header_column is not None:
            break

    if header_column is None or header_row is None:
        return []

    submodules: list[str] = []
    seen: set[str] = set()
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if header_column >= len(row):
            continue
        value = _normalize_cell_value(row[header_column])
        if not value or value in seen:
            continue
        seen.add(value)
        submodules.append(value)

    return submodules


def _normalize_cell_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


# --- 模板知识（公司标准做法）抽取与检索 -----------------------------------
# 模板的「描述/需求模版/说明、示例、注意事项」列是公司工程记忆（选项枚举、条件性做法、
# 固件宏名），注入 analyze 富化 prompt 供对齐粒度/术语/通用做法。运行时从用户模板路径
# 现读（模板=公司知识不进仓，也不落编译索引——避免陈旧+保密扩散）。检索是确定性词面
# 匹配（宁漏勿错，与蓝皮书 RAG 同哲学）：零重叠 → 不注入。

_KNOWLEDGE_COL_SUBMODULE = 3   # 子模块
_KNOWLEDGE_COL_QUESTION = 4    # 描述
_KNOWLEDGE_COL_DEFAULT = 5     # 需求模版
_KNOWLEDGE_COL_NOTES = 7       # 说明、示例、注意事项
DEFAULT_REFERENCE_TOP_K = 8
DEFAULT_REFERENCE_MAX_CHARS = 1500


def extract_template_knowledge(path: Path | None) -> dict[str, list[dict[str, str]]]:
    """按 sheet 读模板需求行（描述列非空），供富化检索。缺文件 → {}（与词表降级一致）。"""
    if path is None or not Path(path).exists():
        return {}
    workbook = load_workbook(path, data_only=True, read_only=True)
    knowledge: dict[str, list[dict[str, str]]] = {}
    try:
        for worksheet in workbook.worksheets:
            sheet = worksheet.title.strip()
            if not sheet or sheet.endswith("列表") or sheet in _SKIPPED_SHEET_TITLES:
                continue
            rows: list[dict[str, str]] = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < _KNOWLEDGE_COL_QUESTION:
                    continue
                question = _normalize_cell_value(row[_KNOWLEDGE_COL_QUESTION - 1])
                if not question:
                    continue
                rows.append({
                    "submodule": _normalize_cell_value(row[_KNOWLEDGE_COL_SUBMODULE - 1]),
                    "question": question,
                    "default": _normalize_cell_value(row[_KNOWLEDGE_COL_DEFAULT - 1])
                    if len(row) >= _KNOWLEDGE_COL_DEFAULT else "",
                    "notes": _normalize_cell_value(row[_KNOWLEDGE_COL_NOTES - 1])
                    if len(row) >= _KNOWLEDGE_COL_NOTES else "",
                })
            if rows:
                knowledge[sheet] = rows
    finally:
        workbook.close()
    return knowledge


def _match_tokens(text: str) -> set[str]:
    """打分词元：ASCII 词（≥2 字符，casefold）+ 中文 2-gram。确定性、零依赖。"""
    import re
    tokens: set[str] = set()
    for word in re.findall(r"[A-Za-z0-9_.\-]{2,}", text):
        tokens.add(word.casefold())
    han = re.findall(r"[一-鿿]", text)
    for i in range(len(han) - 1):
        tokens.add(han[i] + han[i + 1])
    return tokens


def select_template_references(
    knowledge: dict[str, list[dict[str, str]]],
    source_req: dict[str, Any],
    *,
    k: int = DEFAULT_REFERENCE_TOP_K,
    max_chars: int = DEFAULT_REFERENCE_MAX_CHARS,
) -> list[dict[str, str]]:
    """给单条需求选公司标准做法参考行：模块定 sheet → 词面重叠打分 top-k → 字数帽。"""
    from template_writer import MODULE_TO_SHEET
    sheet = MODULE_TO_SHEET.get(str(source_req.get("module") or "").strip(), "")
    rows = knowledge.get(sheet) or []
    if not rows:
        return []
    req_tokens = _match_tokens(" ".join(
        str(source_req.get(field) or "")
        for field in ("title", "description", "source_quote")))
    if not req_tokens:
        return []
    scored: list[tuple[int, int, dict[str, str]]] = []
    for order, row in enumerate(rows):
        row_tokens = _match_tokens(" ".join((row["submodule"], row["question"], row["notes"])))
        score = len(req_tokens & row_tokens)
        if score > 0:
            scored.append((score, order, row))
    scored.sort(key=lambda x: (-x[0], x[1]))   # 分高优先，同分按模板行序（稳定可复现）
    picked: list[dict[str, str]] = []
    used = 0
    for _score, _order, row in scored[:k]:
        size = len(row["question"]) + len(row["default"]) + len(row["notes"])
        if picked and used + size > max_chars:
            break
        picked.append(row)
        used += size
    return picked


def render_template_references(refs: list[dict[str, str]]) -> str:
    """参考行渲染为 prompt 文本（也作缓存指纹基底——同一渲染两用，保证指纹与实际注入一致）。"""
    lines = []
    for row in refs:
        parts = [row["submodule"] or "-", row["question"]]
        if row["default"]:
            parts.append(f"默认:{row['default']}")
        if row["notes"]:
            parts.append(f"说明:{row['notes']}")
        lines.append(" | ".join(parts))
    return "\n".join(lines)
