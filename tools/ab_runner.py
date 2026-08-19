"""去原子化方案 §3.2 + §五(M3)：只切 ``RATOMIZER_FUNCTIONAL_EXTRACT`` 的完整 B 轨 A/B runner。

与旧 ``tools/shadow_run.py`` 的本质区别（旧工具不可作为翻转依据的四个缺陷逐条修掉）：

1. **完整 B 轨**：A 路 ``ai-extract → functional-synthesis → requirements-analysis →
   template-write``；B 路 ``functional-extract → requirements-analysis → template-write``
   （B 路传旧阶段名，chain 在 ``RATOMIZER_FUNCTIONAL_EXTRACT=1`` 时单点替换）。
2. **控制变量**：相同解析产物、相同模板、相同 route/model/温度/prompt 版本——**唯一
   差异是 ``RATOMIZER_FUNCTIONAL_EXTRACT``**（运行 env 全量快照进报告供审计）。
3. **失败即 FAIL**：直抽异常、产物缺失、execution_status != ok（stub 降级/mixed）、
   守恒未闭合一律判 FAIL；stub route 直接拒绝（旧 shadow 对 stub 产物显式放行是缺陷）。
4. **对比对象是最终 xlsx**：不只比中间 JSON；逐份文档判定，不用跨文档平均值掩盖
   单份问题（多文档时任一份 FAIL 即整体 FAIL，逐份明细保留）。

M3（去原子化修复方案 §五，2026-08-15）最终 XLSX 质量门：

- **真值集 §5.1**：``--truth`` JSONL 每行 ``truth_id/document_id/section_id/expected_text``
  + ``conditions/exceptions/negations/numbers/units/codes``（schema
  ``schemas/functional_truth.schema.json``）。加载即校验必需键，坏行响亮报错；
  客户原文不进仓——仓库只保存 schema 与合成 fixture，真实真值集走机器本地路径。
  ``document_id`` 按文档键匹配（parsed-dir 名 / ``sha256:<blocks.jsonl sha256>``），
  ``*`` 为文档无关通配。
- **一对一匹配 §5.2**：truth × 最终 XLSX 行的确定性最大权重一对一匹配（自实现
  Hungarian/Kuhn-Munkres，O(n^3)，零第三方依赖；同一 produced 行不得匹配多条 truth，
  结果对真值输入顺序稳定），输出 TP/FN/FP/precision/recall/F1。优化目标分层全部编码
  进整数权重一次求解：先最大化匹配数（TP），再最大化总匹配分，同分平局按
  (truth_id, XLSX 行序) 稳定决胜。分数 = 0.40*正文 token 覆盖 + 0.20*section 一致 +
  0.20*条件/例外/否定一致 + 0.20*数值单位编码不冲突；覆盖 <0.5、冲突、双方 section
  均非空且不相等（跨条款借位）的候选对不可匹配（不进图）。
- **最终 XLSX 读取 §5.3**：openpyxl 读 ``软件需求列表-成文.xlsx``，按模板表头别名
  （中英文，见 ``XLSX_COLUMN_ALIASES``）定位需求正文/条件/验收标准/模块等列；签名
  sheet（表头含 序号+子模块）body 别名不命中时按写入器列契约兜底
  （``template_writer.WRITER_COLUMN_CONTRACT``——计量需求 sheet 的「需求」列被电表
  类型列拆分，读的正是写入器写的列位）；**模板校准**：``--template`` 的行界
  （``_load_template_extents``，与 ``template_writer._next_seq`` 同一追加权威）把
  annex 清单 sheet（计量列表/费率列表/事件列表等）与需求 sheet 样例行剥离为模板
  自带内容（``template_rows_skipped``/``template_only_sheets``）——produced 行 =
  行界之后的真实追加行，无需空模板；缺必需列、空正文行、不可读单元格 → 该文档
  FAIL；precision/recall 必须基于最终 XLSX 行（functional JSON 指标仅诊断）；对
  **全部真值行**计算条件/例外/否定/数值/单位/编码保存率——分母 = 全部真值行的期望
  条目，未匹配真值（FN）的期望条目计为未保存（整份文档的真值信息保存率，与
  recall 同口径的诚实分母；报告以 ``preservation_denominator_scope =
  "all_truth_rows"`` 说明口径；数值/单位/编码被替换 = 冲突 = 不保存）。
- **强制阈值 §5.4**：``REQUIRED_THRESHOLD_KEYS`` 全部 14 项缺一不可；**缺真值集、缺
  阈值文件、或缺任一必需阈值 → NO_GATE**（不 PASS）。阈值按文档独立配置：thresholds
  JSON 可含 ``{"documents": {"<parsed-dir 名或文档 sha256>": {...覆盖...}}}``，文档层
  覆盖默认层（顶层键）；无默认层而文档层缺键同样 NO_GATE。绝不跨文档平均。

判定语义：链失败/产物缺失/执行不完整/守恒未闭合/stub 降级/最终交付物缺席或缺列空行
→ FAIL；门不完备（缺真值/缺阈值）→ **NO_GATE**（本报告不能作为 Go/No-Go 依据，不给
PASS）；两者齐备且全部阈值达标才 PASS。退出码：0=PASS、1=NO_GATE、2=FAIL。

用法（真实语料，需已 parse 的输出目录与公司模板）::

    python tools/ab_runner.py \
        --parsed-dir <含 blocks.jsonl/chunks.jsonl 的 parse 产物目录> \
        --template <公司需求列表模板 xlsx（空模板）> \
        --route openai_compatible \
        --truth <功能级真值集 jsonl（schemas/functional_truth.schema.json）> \
        --thresholds <门槛 json（14 必需键 + 可选 documents 覆盖层）> \
        [--out ab_report.json]

单测纪律：测试注入 ``chain_runner`` 回调，禁止真实 LLM 调用。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sys
import tempfile
import traceback
from collections import Counter
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

SWITCH_ENV = "RATOMIZER_FUNCTIONAL_EXTRACT"

PARSED_ARTIFACTS = (
    "blocks.jsonl", "chunks.jsonl", "table_items.jsonl", "table_cell_items.jsonl",
    "table_cell_dispositions.jsonl", "doc_map.json", "unextracted_registry.json",
    "parse_audit.json",
)

# v3（2026-08-17）：最终 XLSX 读取器模板校准——produced 行 = 模板行界之后的追加行
# （模板样例行/annex 清单 sheet 剥离进 template_rows_skipped/template_only_sheets），
# 签名 sheet body 别名不命中时按写入器列契约兜底（contract_body_sheets）。
REPORT_SCHEMA = "ab-runner-report/v3"

FINAL_XLSX = "软件需求列表-成文.xlsx"

# ---------------------------------------------------------------------------
# §5.1 真值集格式
# ---------------------------------------------------------------------------
# 完整 schema 见 schemas/functional_truth.schema.json；此处为加载期必需键校验。
TRUTH_REQUIRED_KEYS = ("truth_id", "document_id", "section_id", "expected_text")
TRUTH_LIST_KEYS = ("conditions", "exceptions", "negations", "numbers", "units", "codes")
TRUTH_GLOBAL_DOCUMENT = "*"  # 文档无关通配（仅限合成 fixture/单测方便，真值集应写实名）

# ---------------------------------------------------------------------------
# §5.4 强制阈值（14 项全部必需，缺任一 → NO_GATE）
# ---------------------------------------------------------------------------
REQUIRED_THRESHOLD_KEYS = (
    "min_truth_precision",
    "min_truth_recall",
    "min_truth_f1",
    "min_condition_preservation",
    "min_exception_preservation",
    "min_negation_preservation",
    "min_number_preservation",
    "min_unit_preservation",
    "min_code_preservation",
    "max_duplicate_rate",
    "max_oversplit_rate",
    "max_undersplit_rate",
    "max_manual_action_estimate",
    "max_final_row_growth_ratio",
)
# 旧诊断键：可提供，评估但不强制。
OPTIONAL_THRESHOLD_KEYS = ("max_preservation_blocking_losses",)
THRESHOLD_KEYS = tuple(REQUIRED_THRESHOLD_KEYS) + OPTIONAL_THRESHOLD_KEYS

# ---------------------------------------------------------------------------
# §5.3 最终 XLSX 模板列别名表（中英文常见表头；归一化 = 去空白 + casefold）
# ---------------------------------------------------------------------------
XLSX_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    # 需求正文（必需列）：模板 V2.3.x 的「需求」列
    "body": ("需求", "需求正文", "需求内容", "需求文本", "需求描述", "软件需求",
             "requirement", "requirement text", "requirement description",
             "software requirement", "software requirement text"),
    # 源条款（section 匹配/拆分率用）
    "section": ("客户需求章节", "需求章节", "源文章节", "源章节", "章节", "条款",
                "section", "source section", "clause", "chapter"),
    # 模块/子模块（上下文）
    "module": ("模块", "子模块", "功能模块", "module", "submodule", "sub-module",
               "function module"),
    # 条件类列（保存率上下文）
    "condition": ("条件", "前置条件", "前提条件", "适用条件", "前提",
                  "condition", "conditions", "precondition", "pre-condition"),
    # 验收标准列（保存率上下文）
    "acceptance": ("验收标准", "验收准则", "验证标准", "验收",
                   "acceptance", "acceptance criteria", "acceptance criterion",
                   "verification criteria", "verify criteria"),
    # 说明列（保存率上下文）：模板 V2.3.x 的「说明、示例、注意事项」
    "notes": ("说明、示例、注意事项", "说明", "说明示例注意事项", "说明与示例",
              "示例、注意事项", "备注", "notes", "note", "remarks", "remark"),
    # 描述列（保存率上下文）
    "description": ("描述", "问题描述", "description", "problem description"),
}
REQUIRED_XLSX_COLUMNS = ("body",)
# 参与「保存率上下文」拼接的逻辑列（正文 + 条件/验收/说明/描述）
CONTEXT_COLUMNS = ("body", "condition", "acceptance", "notes", "description")

_XLSX_ERROR_LITERALS = frozenset({
    "#div/0!", "#n/a", "#name?", "#null!", "#num!", "#ref!", "#value!",
    "#spill!", "#calc!", "#getting_data",
})

# 数值/单位/编码抽取（§5.2/§5.3 保存与冲突判定，确定性正则）
_NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")
# 点分数字链（OBIS 风格编码，≥3 组）：先于数值抽取剥离，避免 0.0.1.0.0.255 被拆成数字
_DOTTED_CODE_RE = re.compile(r"(?<![\d.])(\d+(?:\.\d+){2,})(?![\d.])")
# 已识别单位 token 集（冲突判定：期望单位丢失而行内另有单位 → 冲突）
_UNIT_WORDS = (
    "kvarh", "kwh", "kvar", "kva", "kw", "varh", "wh", "var", "va", "mah", "ah",
    "ma", "kv", "mv", "khz", "hz", "ms", "min", "hours", "hour", "hrs", "hr",
    "days", "day", "weeks", "week", "bytes", "byte", "kb", "mb", "gb", "tb",
    "db", "ppm", "°c", "k", "s", "h", "a", "v", "w", "%",
)
_UNIT_TOKEN_RE = re.compile(
    "(?<![A-Za-z])(" + "|".join(sorted(map(re.escape, _UNIT_WORDS), key=len, reverse=True))
    + ")(?![A-Za-z])",
    re.IGNORECASE,
)


@dataclass
class PathResult:
    label: str
    ok: bool
    error: str = ""
    chain_payload: dict[str, Any] = field(default_factory=dict)
    execution_status: str = ""
    conservation_ok: bool | None = None
    functional_count: int | None = None
    xlsx_rows: int | None = None
    xlsx_struct: dict[str, Any] = field(default_factory=dict)
    env_switch: str = ""


@contextmanager
def _switch_env(value: str):
    """唯一允许的差异变量：RATOMIZER_FUNCTIONAL_EXTRACT；退出时如实恢复。"""
    original = os.environ.get(SWITCH_ENV)
    os.environ[SWITCH_ENV] = value
    try:
        yield
    finally:
        if original is None:
            os.environ.pop(SWITCH_ENV, None)
        else:
            os.environ[SWITCH_ENV] = original


def _copy_parsed_artifacts(parsed_dir: Path, target: Path) -> None:
    for name in PARSED_ARTIFACTS:
        source = parsed_dir / name
        if source.is_file():
            shutil.copy2(source, target / name)


def _read_functional_product(out_dir: Path) -> dict[str, Any]:
    from requirements_analysis_rules import _read_functional_requirements_payload

    payload = _read_functional_requirements_payload(out_dir)
    return payload if isinstance(payload, dict) else {}


# ---------------------------------------------------------------------------
# §5.1 真值集加载/校验/按文档选择
# ---------------------------------------------------------------------------

def _validate_truth_row(row: Any, *, source: str) -> dict[str, Any]:
    """校验并归一化一条真值行；坏行响亮报错（ValueError，带来源定位）。"""
    if not isinstance(row, dict):
        raise ValueError(f"{source}: truth row must be a JSON object, got {type(row).__name__}")
    for key in TRUTH_REQUIRED_KEYS:
        value = row.get(key)
        if not isinstance(value, str) or not value.strip():
            raise ValueError(
                f"{source}: truth row missing required key {key!r}: {json.dumps(row, ensure_ascii=False)[:300]}")
    normalized = dict(row)
    for key in TRUTH_LIST_KEYS:
        value = row.get(key)
        if value is None:
            value = []
        if not isinstance(value, list) or any(
                not isinstance(entry, str) or not entry.strip() for entry in value):
            raise ValueError(
                f"{source}: truth row {row['truth_id']!r} field {key!r} must be a list "
                f"of non-empty strings: {value!r}")
        normalized[key] = list(value)
    return normalized


def _load_truth(path: Path | None) -> list[dict[str, Any]] | None:
    """加载真值集 JSONL：逐行校验必需键/列表类型 + truth_id 去重，坏行响亮报错。"""
    if path is None:
        return None
    rows: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for lineno, line in enumerate(Path(path).read_text(encoding="utf-8").splitlines(), 1):
        if not line.strip():
            continue
        try:
            payload = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{path}:{lineno}: truth row is not valid JSON: {exc}") from exc
        row = _validate_truth_row(payload, source=f"{path}:{lineno}")
        if row["truth_id"] in seen_ids:
            raise ValueError(f"{path}:{lineno}: duplicate truth_id {row['truth_id']!r}")
        seen_ids.add(row["truth_id"])
        rows.append(row)
    return rows


def _document_keys(parsed_dir: Path) -> dict[str, str]:
    """文档键候选（thresholds documents 层与 truth document_id 匹配用）。

    name = parsed-dir 名；sha256 = blocks.jsonl 的 sha256 前缀键；path = 解析绝对路径。
    """
    resolved = Path(parsed_dir).resolve()
    blocks = resolved / "blocks.jsonl"
    digest = hashlib.sha256(blocks.read_bytes()).hexdigest() if blocks.is_file() else ""
    return {
        "name": resolved.name,
        "sha256": f"sha256:{digest}" if digest else "",
        "path": str(resolved),
    }


def _select_truth_rows(truth_rows: list[dict[str, Any]] | None,
                       doc_keys: dict[str, str]) -> list[dict[str, Any]]:
    """选出属于本文档的真值行（document_id = 目录名 / sha256 键 / 绝对路径 / '*'）。"""
    if truth_rows is None:
        return []
    accepted = {TRUTH_GLOBAL_DOCUMENT} | {
        key for key in (doc_keys.get("name"), doc_keys.get("sha256"), doc_keys.get("path"))
        if key
    }
    selected: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, raw in enumerate(truth_rows):
        row = _validate_truth_row(raw, source=f"truth_rows[{index}]")
        if str(row["document_id"]).strip() not in accepted:
            continue
        if row["truth_id"] in seen:
            raise ValueError(f"truth selection has duplicate truth_id {row['truth_id']!r}")
        seen.add(row["truth_id"])
        selected.append(row)
    return selected


# ---------------------------------------------------------------------------
# §5.3 最终 XLSX 读取
# ---------------------------------------------------------------------------

def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().casefold())


def _locate_columns(header_row: tuple[Any, ...]) -> dict[str, int]:
    """表头行 → 逻辑列 → 1-based 列号（先命中先得，确定性）。"""
    mapping: dict[str, int] = {}
    for column_index, raw in enumerate(header_row, 1):
        header = _norm_header(raw)
        if not header:
            continue
        for logical, aliases in XLSX_COLUMN_ALIASES.items():
            if logical in mapping:
                continue
            if header in {_norm_header(alias) for alias in aliases}:
                mapping[logical] = column_index
                break
    return mapping


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _logical_cell(row: list[Any], columns: dict[str, int], logical: str) -> str:
    column = columns.get(logical)
    if column is None or column > len(row):
        return ""
    return _cell_text(row[column - 1])


def _writer_contract_columns(header_row: tuple[Any, ...]) -> dict[str, int] | None:
    """V2.3.x 需求 sheet 签名（表头含 序号+子模块）→ 写入器列契约列位。

    WS0 门禁复盘（2026-08-17）：计量需求 sheet 的「需求」列被电表类型列拆分
    （1P2W_SP/3P4W_DC/3P4W_LVCT，表头别名不命中），而 ``template_writer`` 对全部
    需求 sheet 按固定列位追加——读取侧直接采用写入器同一列位权威，读的正是写入
    器写的位置。非签名 sheet 或表头列数不足正文列位 → None（视为缺列，宁判坏）。
    """
    from template_writer import REQUIREMENT_SHEET_SIGNATURE, WRITER_COLUMN_CONTRACT

    headers = {_norm_header(cell) for cell in header_row}
    if not all(_norm_header(cell) in headers for cell in REQUIREMENT_SHEET_SIGNATURE):
        return None
    if len(header_row) < WRITER_COLUMN_CONTRACT["body"]:
        return None
    return dict(WRITER_COLUMN_CONTRACT)


def _load_template_extents(template_path: Path | None) -> dict[str, int] | None:
    """模板校准（§5.3）：sheet 名 → 模板内最后一个非空行号（1-based）。

    成文器只在模板末个非空行之后追加（``template_writer._next_seq`` 同一判定），
    因此最终 XLSX 中 ≤ 模板行界的行是模板自带内容（annex 清单 sheet / 需求 sheet
    样例行 / 留空行），行界之后才是本管线产物。模板缺席/不可读 → None（读取器退回
    无校准语义；生产路径模板不可读时链路本身已先行失败，不会静默放行）。
    """
    if template_path is None:
        return None
    path = Path(template_path)
    if not path.is_file():
        return None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — 校准缺席退回旧语义；模板坏在生产链路上先失败
        return None
    try:
        extents: dict[str, int] = {}
        for worksheet in workbook.worksheets:
            if worksheet is None:
                continue
            last_row = 0
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if row and any(_cell_text(cell) for cell in row):
                    last_row = row_number
            extents[worksheet.title] = last_row
        return extents
    finally:
        workbook.close()


def _read_final_xlsx_rows(xlsx_path: Path,
                          *, template_extents: dict[str, int] | None = None) -> dict[str, Any]:
    """读最终交付 软件需求列表-成文.xlsx（全 sheet、按别名/写入器契约定位列）。

    返回::

        {"ok", "error", "row_count", "rows", "sheets",
         "empty_body_rows", "unreadable_cells", "missing_body_sheets",
         "contract_body_sheets", "template_only_sheets", "template_rows_skipped"}

    - rows: ``{"sheet", "row_number", "body", "section", "module", "context"}``（context =
      body+条件/验收/说明/描述 拼接，保存率检查的检索面）；
    - 列定位：表头别名优先；签名 sheet（序号+子模块）body 别名不命中时按写入器
      列契约兜底（``contract_body_sheets`` 记录），计量需求 sheet 的拆分类型列由此
      可读；
    - 模板校准（``template_extents``，来自 ``_load_template_extents``）：≤ 模板行界
      的行是模板自带内容（annex 清单 sheet、需求 sheet 样例行），剥离进
      ``template_rows_skipped``、纯模板 sheet 记入 ``template_only_sheets``——不计
      produced、不触发缺列/空正文失败；行界之后才是管线产物；
    - 缺 body 列（有产物行的 sheet）、空正文行、不可读单元格均计入失败明细，由
      调用方判 FAIL——宁判坏不放过。
    """
    result: dict[str, Any] = {
        "ok": False, "error": "", "row_count": 0, "rows": [], "sheets": [],
        "empty_body_rows": [], "unreadable_cells": [], "missing_body_sheets": [],
        "contract_body_sheets": [], "template_only_sheets": [], "template_rows_skipped": 0,
    }
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — 文件级不可读由调用方 FAIL
        result["error"] = f"{type(exc).__name__}: {exc}"
        return result
    try:
        for sheet in workbook.worksheets:
            if sheet is None:
                continue
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            header_index: int | None = None
            columns: dict[str, int] = {}
            contract_sheet = False
            nonempty_seen = 0
            for index, row in enumerate(rows):
                if not any(_cell_text(cell) for cell in row):
                    continue
                nonempty_seen += 1
                candidate = _locate_columns(tuple(row))
                if candidate.get("body") is not None:
                    header_index = index
                    columns = candidate
                    break
                contract = _writer_contract_columns(tuple(row))
                if contract is not None:
                    header_index = index
                    columns = {**candidate, **contract}
                    contract_sheet = True
                    break
                if nonempty_seen >= 10:  # 前 10 个非空行内找不到表头 → 视为缺列
                    break
            first_nonempty = next(
                (index for index, row in enumerate(rows)
                 if any(_cell_text(cell) for cell in row)), None)
            if header_index is None:
                header_index = first_nonempty if first_nonempty is not None else 0
            extent = template_extents.get(sheet.title) if template_extents else None
            numbered = list(enumerate(rows[header_index + 1:], header_index + 2))
            nonempty_numbered = [
                (row_number, row) for row_number, row in numbered
                if any(_cell_text(cell) for cell in row)]
            counted = [
                (row_number, row) for row_number, row in nonempty_numbered
                if extent is None or row_number > extent]
            if extent is not None:
                result["template_rows_skipped"] += len(nonempty_numbered) - len(counted)
            if not counted:
                if nonempty_numbered and extent is not None:
                    result["template_only_sheets"].append(sheet.title)
                continue  # 无产物行的 sheet 不参与判定
            result["sheets"].append(sheet.title)
            if "body" not in columns:
                result["missing_body_sheets"].append(sheet.title)
            if contract_sheet:
                result["contract_body_sheets"].append(sheet.title)
            for row_number, row in counted:
                for logical in CONTEXT_COLUMNS:
                    value = _logical_cell(row, columns, logical)
                    if value.casefold() in _XLSX_ERROR_LITERALS:
                        result["unreadable_cells"].append(
                            f"{sheet.title}!{logical}@row{row_number}")
                body = _logical_cell(row, columns, "body")
                if body and body.casefold() not in _XLSX_ERROR_LITERALS:
                    context = " ; ".join(
                        part for part in
                        (_logical_cell(row, columns, logical) for logical in CONTEXT_COLUMNS)
                        if part and part.casefold() not in _XLSX_ERROR_LITERALS)
                    result["rows"].append({
                        "sheet": sheet.title,
                        "row_number": row_number,
                        "body": body,
                        "section": _logical_cell(row, columns, "section"),
                        "module": _logical_cell(row, columns, "module"),
                        "context": context,
                    })
                else:
                    result["empty_body_rows"].append(f"{sheet.title}@row{row_number}")
    finally:
        workbook.close()
    result["ok"] = True
    result["row_count"] = len(result["rows"])
    return result


def _duplicate_rate_texts(texts: list[str]) -> float | None:
    """重复率：objective 归一化后重复条数 / 总数（确定性，不依赖真值集）。"""
    from functional_extract import _normalize_key

    keys = [_normalize_key(text) for text in texts]
    keys = [key for key in keys if key]
    if not keys:
        return None
    seen: set[str] = set()
    duplicates = 0
    for key in keys:
        if key in seen:
            duplicates += 1
        seen.add(key)
    return duplicates / len(keys)


def _duplicate_rate(items: list[dict[str, Any]]) -> float | None:
    """functional JSON 诊断口径的重复率（最终门以最终 XLSX 行口径为准）。"""
    return _duplicate_rate_texts([str(item.get("objective") or "") for item in items])


# ---------------------------------------------------------------------------
# §5.2 truth × 最终 XLSX 行 一对一最大权重匹配 + 保存率
# ---------------------------------------------------------------------------

def _norm_section(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).casefold()


def _row_numbers(context: str) -> list[float]:
    """行文本中的数值（先剥离点分编码，避免 OBIS 链被拆成数字）。"""
    stripped = _DOTTED_CODE_RE.sub(" ", context)
    return [float(match) for match in _NUMBER_RE.findall(stripped)]


def _unit_present(unit: str, context: str) -> bool:
    """期望单位是否保留（字母边界 + 大小写不敏感；"min" 不命中 "minutes"，"15kWh" 命中）。"""
    pattern = rf"(?<![A-Za-z]){re.escape(unit.strip())}(?![A-Za-z])"
    return re.search(pattern, context, re.IGNORECASE) is not None


def _code_present(code: str, context: str) -> bool:
    """期望编码是否保留（去空白 + casefold 子串）。"""
    squashed = "".join(str(context or "").split()).casefold()
    return code.strip().casefold() in squashed


def _entry_preserved(entry: str, row_tokens: set[str]) -> bool:
    """条件/例外/否定条目保留判定：内容 token 覆盖 ≥0.6（与守恒口径一致）。"""
    from functional_extract import _content_tokens

    tokens = _content_tokens(entry)
    if not tokens:
        return False  # 无内容 token 的期望条目视为未保存（诚实）
    return len(tokens & row_tokens) / len(tokens) >= 0.6


def _evaluate_pair(truth: dict[str, Any], row: dict[str, Any]) -> dict[str, Any]:
    """一对 (truth, produced row) 的确定性打分与可匹配性。"""
    from functional_extract import _content_tokens

    expected_tokens = _content_tokens(truth["expected_text"])
    body_tokens = _content_tokens(row["body"])
    context_tokens = _content_tokens(row["context"]) | body_tokens
    coverage = (
        len(expected_tokens & body_tokens) / len(expected_tokens)
        if expected_tokens else 0.0)
    truth_section = _norm_section(truth["section_id"])
    row_section = _norm_section(row["section"])
    section_match = bool(truth_section) and truth_section == row_section
    cross_section = bool(truth_section) and bool(row_section) and truth_section != row_section

    text_entries = [*(truth["conditions"]), *(truth["exceptions"]), *(truth["negations"])]
    text_flags = [_entry_preserved(entry, context_tokens) for entry in text_entries]
    consistency = (sum(text_flags) / len(text_flags)) if text_flags else 1.0

    numbers = _row_numbers(row["context"])
    codes = _DOTTED_CODE_RE.findall(row["context"])
    has_unit = bool(_UNIT_TOKEN_RE.search(row["context"]))
    conflicts: list[str] = []
    for number in truth["numbers"]:
        if float(number) not in numbers and numbers:
            conflicts.append(f"number:{number}")
    for unit in truth["units"]:
        if not _unit_present(unit, row["context"]) and has_unit:
            conflicts.append(f"unit:{unit}")
    for code in truth["codes"]:
        if not _code_present(code, row["context"]) and codes:
            conflicts.append(f"code:{code}")
    conflict_free = not conflicts

    score = (0.40 * coverage + 0.20 * float(section_match)
             + 0.20 * consistency + 0.20 * float(conflict_free))
    eligible = (
        bool(expected_tokens)
        and coverage >= 0.5
        and conflict_free
        and not cross_section  # 双方 section 均非空且不相等 → 跨条款借位，禁配
    )
    return {
        "score": score, "coverage": coverage, "section_match": section_match,
        "consistency": consistency, "conflict_free": conflict_free,
        "conflicts": conflicts, "eligible": eligible,
    }


_PRESERVATION_KINDS = (
    ("condition", "conditions"), ("exception", "exceptions"), ("negation", "negations"),
    ("number", "numbers"), ("unit", "units"), ("code", "codes"),
)

# 匹配分数 → 整数权重的规模化精度（6 位小数）
_MATCH_SCORE_SCALE = 10 ** 6
# 复审 P1-3 二轮：几何决胜位（2^(E-1-r)）的整数位数随可配边数 E 增长——
# 全候选边时 500×500 的权重整数矩阵可达 GiB 级。双档：E ≤ 阈值用严格字典序
# 几何编码（位数有界）；超阈值降级为**有界加性决胜**（任何匹配的决胜总和
# < 1 个分数单位 → TP/分数目标严格优先；确定性 + 规范序稳定，非严格字典序）。
_LEX_EXACT_EDGE_LIMIT = 4096
# 匹配算法标识（入报告的诊断键；确定性最大权重二分匹配，见 match_truth_to_rows）
_MATCHER_VERSION = "hungarian-max-weight/v1"


def _hungarian_max_weight(weights: list[list[int]]) -> list[int]:
    """最大权重二分匹配（Kuhn-Munkres/Hungarian，O(n^3)，自实现零第三方依赖）。

    ``weights[i][j]`` = 行 i ↔ 列 j 的整数权重；不可配边给 0（任何正权可配边都严格
    更优，事后由调用方按边资格过滤）。矩阵内部补方（补 0 权），返回
    ``assignment[i] = j``（-1 = 该行未分到真实列）。实现为经典对偶势 + Dijkstra 式
    最短增广路（e-maxx 变体），全程整数运算；同一矩阵输出唯一确定——确定性由
    ``match_truth_to_rows`` 的权重编码（规范序 + 平局决胜项）保证。
    """
    n = len(weights)
    m = len(weights[0]) if n else 0
    size = max(n, m)
    cost = [[-(weights[i][j] if j < m else 0) if i < n else 0
             for j in range(size)] for i in range(size)]
    inf = float("inf")
    u = [0] * (size + 1)
    v = [0] * (size + 1)
    p = [0] * (size + 1)   # p[j] = 分到列 j 的行（1-based；0 = 空列，p[0] = 当前增广行）
    way = [0] * (size + 1)
    for i in range(1, size + 1):
        p[0] = i
        j0 = 0
        minv = [inf] * (size + 1)
        used = [False] * (size + 1)
        while True:
            used[j0] = True
            i0 = p[j0]
            delta = inf
            j1 = 0
            for j in range(1, size + 1):
                if used[j]:
                    continue
                current = cost[i0 - 1][j - 1] - u[i0] - v[j]
                if current < minv[j]:
                    minv[j] = current
                    way[j] = j0
                if minv[j] < delta:
                    delta = minv[j]
                    j1 = j
            for j in range(size + 1):
                if used[j]:
                    u[p[j]] += delta
                    v[j] -= delta
                else:
                    minv[j] -= delta
            j0 = j1
            if p[j0] == 0:
                break
        while j0:
            j1 = way[j0]
            p[j0] = p[j1]
            j0 = j1
    assignment = [-1] * n
    for j in range(1, size + 1):
        row = p[j]
        if 1 <= row <= n and j <= m:
            assignment[row - 1] = j - 1
    return assignment


def match_truth_to_rows(truth_rows: list[dict[str, Any]],
                        produced_rows: list[dict[str, Any]]) -> dict[str, Any]:
    """§5.2 一对一匹配（Hungarian 最大权重二分匹配）+ §5.3 保存率（全真值行口径）。

    匹配（P2-2）：只有可匹配对（覆盖 ≥0.5、无冲突、非跨条款）进图，权重 = 现有匹配
    分数。三层优化目标编码进同一整数权重一次求解：

    1. 先最大化匹配数（TP）：每条可配边带 ``cardinality_bonus``（大于任何总分数差，
       多一条可配边永远严格更优）；
    2. 再最大化总匹配分数（分数 ×10^6 取整）；
    3. 同分平局按 (truth_id, XLSX 行序) 稳定决胜（复审 P1-3 二轮改双档，规模安全）：
       可配边数 E ≤ ``_LEX_EXACT_EDGE_LIMIT`` 时用几何编码 ``2^(E-1-r)``（严格字典序，
       位数有界）；E 超阈值降级为有界加性决胜（任何匹配的决胜总和 < 1 个分数单位，
       TP/分数严格优先、规范序确定性保持，非严格字典序——如实降级并在
       ``matcher_tiebreak`` 诊断键标注）。

    真值行先按 (truth_id, section_id, expected_text) 规范排序再进矩阵（truth_id 由
    加载期强制唯一，重复时回退内容键）→ 结果对真值行的输入顺序稳定。

    保存率（P1-3）：分母 = **全部真值行**的期望条目——未匹配真值（FN）的期望条件/
    例外/否定/数值/单位/编码全部计为未保存（分子不增），即「整份文档的真值信息保存
    率」，与 recall 同口径的诚实分母；返回值 ``preservation_denominator_scope``
    说明口径。
    """
    from functional_extract import _content_tokens

    num_truth = len(truth_rows)
    num_rows = len(produced_rows)
    # 规范序：与真值输入顺序无关（truth_id 唯一；重复时回退内容键，保持确定性）
    canonical_order = sorted(
        range(num_truth),
        key=lambda index: (str(truth_rows[index]["truth_id"]),
                           str(truth_rows[index].get("section_id") or ""),
                           str(truth_rows[index].get("expected_text") or "")))
    pair_info: dict[tuple[int, int], dict[str, Any]] = {}
    scaled_score: dict[tuple[int, int], int] = {}
    for truth_index, truth in enumerate(truth_rows):
        for row_index, row in enumerate(produced_rows):
            pair = _evaluate_pair(truth, row)
            pair_info[(truth_index, row_index)] = pair
            if pair["eligible"]:
                scaled_score[(truth_index, row_index)] = int(
                    round(pair["score"] * _MATCH_SCORE_SCALE))

    # 可配边按 (truth 规范序, 行序) 全序排名 → 几何决胜位
    edge_rank: dict[tuple[int, int], int] = {}
    for rank, truth_index in enumerate(canonical_order):
        for row_index in range(num_rows):
            if (truth_index, row_index) in scaled_score:
                edge_rank[(truth_index, row_index)] = len(edge_rank)
    eligible_count = len(edge_rank)
    # cardinality_bonus > 任何规模下可配边分数总和的最大差 → 多一条可配边永远严格更优
    cardinality_bonus = min(num_truth, num_rows) * (_MATCH_SCORE_SCALE + 1) + 1
    # 复审 P1-3 二轮：决胜编码双档（规模安全）。
    use_geometric = eligible_count <= _LEX_EXACT_EDGE_LIMIT
    match_size = max(1, min(num_truth, num_rows))
    tie_budget = 0 if use_geometric else ((_MATCH_SCORE_SCALE - 1) // match_size)
    # 三轮复审 P1-3：有界档先前把决胜加成**直接加在 scaled 分数上**——scaled 的
    # 最小差是 1（不是 1e6），决胜总和虽 < 1e6 仍可能翻转 1 单位的分数差（复现：
    # 66 真值/65 行/4290 边，最优 52.0 被打成 51.999999）。改用乘法隔离层级：
    #   weight = (cardinality_bonus + scaled) * (max_tie_sum + 1) + tie_bonus
    # 任何匹配的决胜总和 ≤ max_tie_sum < 乘数 → TP → 总分 → 稳定决胜严格有序。
    max_tie_sum = tie_budget * match_size

    weights: list[list[int]] = []
    for rank, truth_index in enumerate(canonical_order):
        row_weights = []
        for row_index in range(num_rows):
            scaled = scaled_score.get((truth_index, row_index))
            if scaled is None:
                row_weights.append(0)  # 不可匹配对：永不优于任何可配边（事后亦过滤）
                continue
            if use_geometric:
                tiebreak = 1 << (eligible_count - 1 - edge_rank[(truth_index, row_index)])
                row_weights.append(((cardinality_bonus + scaled) << eligible_count) | tiebreak)
            else:
                # 有界加性决胜（乘法隔离）：第 r 名边附 max(0, B-r)，B·k ≤ max_tie_sum
                # ——TP 与总分数严格优先不被翻转，规范序确定性保持。
                bonus = max(0, tie_budget - edge_rank[(truth_index, row_index)])
                row_weights.append(
                    (cardinality_bonus + scaled) * (max_tie_sum + 1) + bonus)
        weights.append(row_weights)

    assignment = _hungarian_max_weight(weights)
    selected: list[tuple[int, int]] = []
    for rank, truth_index in enumerate(canonical_order):
        row_index = assignment[rank] if rank < len(assignment) else -1
        if row_index is not None and row_index >= 0 and (truth_index, row_index) in edge_rank:
            selected.append((truth_index, row_index))

    # 输出顺序与旧贪心契约一致：(-score, truth_id, 行序)
    selected_details: list[tuple[float, str, int, dict[str, Any], dict[str, Any], dict[str, Any]]] = []
    used_row_indexes: set[int] = set()
    row_by_truth_id: dict[Any, dict[str, Any]] = {}
    for truth_index, row_index in selected:
        truth = truth_rows[truth_index]
        row = produced_rows[row_index]
        pair = pair_info[(truth_index, row_index)]
        used_row_indexes.add(row_index)
        row_by_truth_id[truth["truth_id"]] = row
        selected_details.append(
            (pair["score"], str(truth["truth_id"]), row_index, truth, row, pair))
    selected_details.sort(key=lambda item: (-item[0], item[1], item[2]))
    matches = [
        {
            "truth_id": truth["truth_id"],
            "sheet": row["sheet"],
            "row_number": row["row_number"],
            "score": round(pair["score"], 6),
            "coverage": round(pair["coverage"], 6),
            "section_match": pair["section_match"],
        }
        for _score, _truth_id, _row_index, truth, row, pair in selected_details
    ]

    truth_count = len(truth_rows)
    produced_count = len(produced_rows)
    tp = len(matches)
    fn = truth_count - tp
    fp = produced_count - tp
    precision = tp / (tp + fp) if (tp + fp) else (1.0 if truth_count == 0 else 0.0)
    recall = tp / (tp + fn) if (tp + fn) else 1.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0

    # 保存率（P1-3 口径）：分母 = 全部真值行的期望条目；未匹配真值的期望条目计为
    # 未保存（分子不增）——指标语义 = 整份文档的真值信息保存率，与 recall 同口径。
    expected_counts = {kind: 0 for kind, _field in _PRESERVATION_KINDS}
    preserved_counts = {kind: 0 for kind, _field in _PRESERVATION_KINDS}
    for truth in truth_rows:
        row = row_by_truth_id.get(truth["truth_id"])
        context = row["context"] if row is not None else ""
        row_tokens = (
            _content_tokens(context) | _content_tokens(row["body"])
            if row is not None else set())
        numbers = _row_numbers(context) if row is not None else []
        for kind, field_name in _PRESERVATION_KINDS:
            for entry in truth[field_name]:
                expected_counts[kind] += 1
                if row is None:
                    continue  # 未匹配真值（FN）：期望进分母，保存记 0
                if kind in ("condition", "exception", "negation"):
                    preserved = _entry_preserved(entry, row_tokens)
                elif kind == "number":
                    preserved = float(entry) in numbers
                elif kind == "unit":
                    preserved = _unit_present(entry, context)
                else:  # code
                    preserved = _code_present(entry, context)
                preserved_counts[kind] += 1 if preserved else 0
    preservation = {
        kind: {
            "expected": expected_counts[kind],
            "preserved": preserved_counts[kind],
            "rate": (preserved_counts[kind] / expected_counts[kind])
            if expected_counts[kind] else 1.0,  # 无期望 → 空真（计数如实报告）
        }
        for kind, _field in _PRESERVATION_KINDS
    }

    # 拆分率（按最终 XLSX 行的 section 计数 vs 真值 section 计数）
    truth_sections = Counter(_norm_section(t["section_id"]) for t in truth_rows)
    produced_sections = Counter(
        _norm_section(r["section"]) for r in produced_rows if _norm_section(r["section"]))
    sections = sorted(truth_sections)
    oversplit_sections = [
        s for s in sections if produced_sections.get(s, 0) > truth_sections[s]]
    undersplit_sections = [
        s for s in sections if produced_sections.get(s, 0) < truth_sections[s]]

    unmatched_truth_ids = sorted(
        t["truth_id"] for t in truth_rows if t["truth_id"] not in {m["truth_id"] for m in matches})
    unmatched_rows = [
        {"sheet": r["sheet"], "row_number": r["row_number"],
         "body_head": r["body"][:60]}
        for i, r in enumerate(produced_rows) if i not in used_row_indexes][:20]

    return {
        "available": True,
        "truth_count": truth_count,
        "produced_count": produced_count,
        "tp": tp, "fn": fn, "fp": fp,
        "precision": precision, "recall": recall, "f1": f1,
        "matches": matches,
        "unmatched_truth_ids": unmatched_truth_ids,
        "unmatched_rows": unmatched_rows,
        "preservation": preservation,
        "preservation_denominator_scope": "all_truth_rows",
        "matcher": _MATCHER_VERSION,
        "matcher_tiebreak": (
            "lexicographic-geometric" if use_geometric else "bounded-additive"),
        "oversplit_rate": len(oversplit_sections) / len(sections) if sections else 0.0,
        "undersplit_rate": len(undersplit_sections) / len(sections) if sections else 0.0,
        "oversplit_sections": oversplit_sections,
        "undersplit_sections": undersplit_sections,
        "manual_action_estimate": sum(
            abs(produced_sections.get(s, 0) - truth_sections[s]) for s in sections),
    }


# ---------------------------------------------------------------------------
# §5.4 阈值解析与评估
# ---------------------------------------------------------------------------

def _resolve_thresholds(payload: dict[str, Any],
                        doc_keys: dict[str, str]) -> tuple[dict[str, Any], dict[str, Any], list[str]]:
    """按文档解析有效阈值：文档层覆盖默认层（顶层键）。

    返回 (effective thresholds, layering meta, missing required keys)。
    documents 层键可用 parsed-dir 名、``sha256:<hex>`` 或解析绝对路径。
    """
    documents = payload.get("documents")
    if "documents" in payload and not isinstance(documents, dict):
        raise ValueError("thresholds 'documents' layer must be a JSON object")
    defaults = {key: value for key, value in payload.items() if key != "documents"}
    override: dict[str, Any] = {}
    matched_key: str | None = None
    if isinstance(documents, dict):
        for key in (doc_keys.get("name"), doc_keys.get("sha256"), doc_keys.get("path")):
            if key and key in documents:
                layer = documents[key]
                if not isinstance(layer, dict):
                    raise ValueError(
                        f"thresholds documents[{key!r}] must be a JSON object")
                matched_key = key
                override = layer
                break
    effective = {**defaults, **override}
    missing = [key for key in REQUIRED_THRESHOLD_KEYS if key not in effective]
    meta = {
        "document_key_matched": matched_key,
        "has_default_layer": bool(defaults),
        "document_layer_keys": sorted(documents) if isinstance(documents, dict) else [],
        "document_keys": {k: doc_keys.get(k, "") for k in ("name", "sha256", "path")},
    }
    return effective, meta, missing


def _evaluate_thresholds(metrics: dict[str, Any], thresholds: dict[str, Any]) -> list[str]:
    """逐键评估阈值；返回违例描述（空 = 全部达标）。未知键报违例（响亮）。"""
    violations: list[str] = []
    truth = metrics.get("truth_metrics") if isinstance(metrics.get("truth_metrics"), dict) else {}
    truth_available = bool(truth.get("available"))
    preservation = truth.get("preservation") if isinstance(truth.get("preservation"), dict) else {}

    def _truth_value(key: str) -> Any:
        return truth.get(key) if truth_available else None

    min_lookups = {
        "min_truth_precision": lambda: _truth_value("precision"),
        "min_truth_recall": lambda: _truth_value("recall"),
        "min_truth_f1": lambda: _truth_value("f1"),
        "min_condition_preservation": lambda: (
            preservation.get("condition", {}).get("rate") if truth_available else None),
        "min_exception_preservation": lambda: (
            preservation.get("exception", {}).get("rate") if truth_available else None),
        "min_negation_preservation": lambda: (
            preservation.get("negation", {}).get("rate") if truth_available else None),
        "min_number_preservation": lambda: (
            preservation.get("number", {}).get("rate") if truth_available else None),
        "min_unit_preservation": lambda: (
            preservation.get("unit", {}).get("rate") if truth_available else None),
        "min_code_preservation": lambda: (
            preservation.get("code", {}).get("rate") if truth_available else None),
    }
    max_lookups = {
        "max_duplicate_rate": lambda: metrics.get("duplicate_rate"),
        "max_oversplit_rate": lambda: _truth_value("oversplit_rate"),
        "max_undersplit_rate": lambda: _truth_value("undersplit_rate"),
        "max_manual_action_estimate": lambda: _truth_value("manual_action_estimate"),
        "max_preservation_blocking_losses": lambda: metrics.get("preservation_blocking_losses"),
    }

    for key, value in sorted(thresholds.items()):
        if key not in THRESHOLD_KEYS:
            violations.append(f"unknown threshold key: {key}")
            continue
        try:
            limit = float(value)
        except (TypeError, ValueError):
            violations.append(f"threshold {key} is not a number: {value!r}")
            continue
        if key == "max_final_row_growth_ratio":
            b_rows = metrics.get("b_final_xlsx_rows")
            a_rows = metrics.get("a_final_xlsx_rows")
            if not isinstance(a_rows, int) or not isinstance(b_rows, int) or a_rows <= 0:
                violations.append(
                    f"{key}: final xlsx rows unavailable (a={a_rows!r}, b={b_rows!r})")
            elif b_rows / a_rows > limit:
                violations.append(
                    f"{key}: {b_rows}/{a_rows} = {b_rows / a_rows:.3f} > {limit}")
        elif key in min_lookups:
            actual = min_lookups[key]()
            if actual is None or actual < limit:
                violations.append(
                    f"{key}: {actual if actual is not None else 'unavailable'} < {limit}")
        else:  # max_* 家族
            actual = max_lookups[key]()
            if actual is None or actual > limit:
                violations.append(
                    f"{key}: {actual if actual is not None else 'unavailable'} > {limit}")
    return violations


def run_ab_for_document(
    parsed_dir: Path,
    *,
    route: str,
    template_path: Path,
    truth_rows: list[dict[str, Any]] | None = None,
    thresholds: dict[str, Any] | None = None,
    work_root: Path | None = None,
    chain_runner: Callable[..., dict[str, Any]] | None = None,
    keep_dirs: bool = False,
    warm_a_cache: Path | None = None,
) -> dict[str, Any]:
    """单文档 A/B：唯一差异 = 直抽开关。返回逐文档报告（PASS/FAIL/NO_GATE + 指标明细）。

    ``chain_runner`` 注入点是测试钩子（默认 ``desktop_tasks.chain_task``）——生产路径
    必须走真实 chain（完整 B 轨是本 runner 的存在意义）。

    判定语义（M3）：链失败/产物缺失/执行不完整/守恒未闭合/stub 降级/最终交付物缺席/
    缺必需列/空正文行/不可读单元格 → FAIL；缺真值集（含真值集无本文档行）、缺阈值文
    件或缺任一必需阈值 → **NO_GATE**（本报告不能作为 Go/No-Go 依据，不给 PASS）；两
    者齐备且全部阈值达标才 PASS。precision/recall/F1/保存率全部基于最终交付
    ``软件需求列表-成文.xlsx`` 的行（行口径 = 模板行界之后的追加行——模板校准
    ``_load_template_extents`` 剥离样例行/annex 清单 sheet）；functional JSON 指标
    仅诊断。
    """
    if route == "stub":
        raise ValueError(
            "ab_runner 拒绝 stub route——直抽 stub 产物按 §3.5 是不可发布占位，"
            "A/B 判定必须基于真实 LLM 路由"
        )
    if chain_runner is None:
        import desktop_tasks as chain_module

        chain_runner = chain_module.chain_task
    parsed_dir = Path(parsed_dir).expanduser().resolve()
    if not (parsed_dir / "blocks.jsonl").is_file():
        raise FileNotFoundError(f"{parsed_dir} 缺 blocks.jsonl——请先 parse 再跑 A/B")
    doc_keys = _document_keys(parsed_dir)

    # 模板校准（§5.3）：annex 清单 sheet 与需求 sheet 样例行 = 模板自带内容，行界
    # 之后才是本管线产物——读取器据此剥离，precision/recall 口径 = 真实追加行。
    template_extents = _load_template_extents(template_path)

    work_root = Path(work_root) if work_root else Path(tempfile.mkdtemp(prefix="ab-runner."))
    work_root.mkdir(parents=True, exist_ok=True)
    a_dir = work_root / "A_atoms"
    b_dir = work_root / "B_direct"
    for directory in (a_dir, b_dir):
        directory.mkdir(parents=True, exist_ok=True)
        _copy_parsed_artifacts(parsed_dir, directory)

    stages = ["ai-extract", "functional-synthesis", "requirements-analysis", "template-write"]

    def _run(label: str, out_dir: Path, switch: str) -> PathResult:
        result = PathResult(label=label, ok=False, env_switch=switch)
        try:
            with _switch_env(switch):
                payload = chain_runner(
                    out_dir, stages=stages, route=route, template_path=template_path,
                )
            result.chain_payload = {
                key: payload.get(key) for key in ("functional_extract", "analysis")
            }
            result.ok = True
        except Exception as exc:  # noqa: BLE001 — A/B 门：任何链异常即 FAIL，不吞
            result.error = f"{type(exc).__name__}: {exc}"
            result.ok = False
            return result
        # 产物与失败语义检查（B 路全查；A 路查成文交付物在场）
        product = _read_functional_product(out_dir)
        result.execution_status = str(product.get("execution_status") or "")
        conservation = product.get("conservation")
        result.conservation_ok = (
            bool(conservation.get("ok")) if isinstance(conservation, dict) else None
        )
        items = product.get("items") if isinstance(product.get("items"), list) else []
        result.functional_count = len(items) if items else None
        final_xlsx = out_dir / FINAL_XLSX
        struct = (_read_final_xlsx_rows(final_xlsx, template_extents=template_extents)
                  if final_xlsx.is_file() else None)
        if struct is None:
            result.ok = False
            result.error = result.error or f"最终交付物 {FINAL_XLSX} 缺失或不可读"
        elif not struct["ok"]:
            result.ok = False
            result.error = result.error or f"最终交付物 {FINAL_XLSX} 不可读: {struct['error']}"
        else:
            result.xlsx_rows = struct["row_count"]
            result.xlsx_struct = {
                "row_count": struct["row_count"],
                "sheets": struct["sheets"],
                "empty_body_row_count": len(struct["empty_body_rows"]),
                "unreadable_cell_count": len(struct["unreadable_cells"]),
                "missing_body_sheets": struct["missing_body_sheets"],
                "contract_body_sheets": struct["contract_body_sheets"],
                "template_only_sheets": struct["template_only_sheets"],
                "template_rows_skipped": struct["template_rows_skipped"],
            }
            if struct["missing_body_sheets"]:
                result.ok = False
                result.error = (result.error or
                                f"最终 XLSX 缺需求正文列（表头别名不命中）: sheets="
                                f"{struct['missing_body_sheets']}")
            elif struct["empty_body_rows"]:
                sample = ", ".join(struct["empty_body_rows"][:5])
                result.ok = False
                result.error = (result.error or
                                f"最终 XLSX 空正文行 {len(struct['empty_body_rows'])} 行（{sample}）")
            elif struct["unreadable_cells"]:
                sample = ", ".join(struct["unreadable_cells"][:5])
                result.ok = False
                result.error = (result.error or
                                f"最终 XLSX 不可读单元格 {len(struct['unreadable_cells'])} 个（{sample}）")
        if label == "B_direct":
            if not product:
                result.ok = False
                result.error = "B 路无 functional_requirements.json 产物"
            elif result.execution_status != "ok":
                result.ok = False
                result.error = f"B 路执行不完整：execution_status={result.execution_status}"
            elif result.conservation_ok is False:
                result.ok = False
                result.error = "B 路守恒未闭合"
            elif str(product.get("route") or "") == "stub":
                result.ok = False
                result.error = "B 路降级 stub（占位条目）——不可作为 A/B 依据"
            elif product.get("draft"):
                result.ok = False
                result.error = "B 路产物带 stub 草稿水印——不可作为 A/B 依据"
        if result.xlsx_rows is None:
            result.ok = False
            result.error = result.error or f"最终交付物 {FINAL_XLSX} 缺失或不可读"
        return result

    warm_cache_info: dict[str, Any] | None = None
    if warm_a_cache is not None:
        source = Path(warm_a_cache)
        cache_file = source / "ai_extract_cache.jsonl"
        if not cache_file.is_file():
            raise FileNotFoundError(
                f"--warm-a-cache 目录缺 ai_extract_cache.jsonl：{source}")
        import hashlib as _hashlib

        shutil.copy2(cache_file, a_dir / "ai_extract_cache.jsonl")
        warm_cache_info = {
            "source": str(source),
            "sha256": "sha256:" + _hashlib.sha256(
                cache_file.read_bytes()).hexdigest(),
            "note": ("A 轨整链仍真实执行（synthesis/analysis/template 全跑），仅抽取命中"
                     "缓存零付费；指纹含 route/model，不匹配自然全 miss"),
        }

    a_result = _run("A_atoms", a_dir, "0")
    b_result = _run("B_direct", b_dir, "1")

    failures = [
        f"{result.label}: {result.error}" for result in (a_result, b_result)
        if not result.ok
    ]

    # ---- M3 门输入：真值集（按文档选择）+ 阈值（按文档解析） ----
    doc_truth: list[dict[str, Any]] = []
    missing_gates: list[str] = []
    if truth_rows is None:
        missing_gates.append("truth set not provided (--truth)")
    else:
        doc_truth = _select_truth_rows(truth_rows, doc_keys)
        if not doc_truth:
            missing_gates.append(
                "truth set has no rows for document "
                f"{doc_keys['name']!r} (document_id must match the parsed-dir name, "
                f"{doc_keys['sha256'] or 'sha256:<blocks.jsonl sha256>'}, or be '*')")

    effective_thresholds: dict[str, Any] = {}
    threshold_meta: dict[str, Any] = {}
    if not thresholds:
        missing_gates.append("thresholds not provided (--thresholds)")
    else:
        effective_thresholds, threshold_meta, missing_threshold_keys = (
            _resolve_thresholds(thresholds, doc_keys))
        missing_gates.extend(f"missing required threshold: {key}" for key in missing_threshold_keys)

    # ---- 指标：最终 XLSX 行为门口径；functional JSON 仅诊断 ----
    metrics: dict[str, Any] = {"available": False}
    b_rows: list[dict[str, Any]] = []
    if b_result.ok:
        b_product = _read_functional_product(b_dir)
        b_items = b_product.get("items") if isinstance(b_product.get("items"), list) else []
        b_struct = _read_final_xlsx_rows(b_dir / FINAL_XLSX,
                                         template_extents=template_extents)
        b_rows = b_struct.get("rows") or []
        conservation = b_product.get("conservation") or {}
        checks = conservation.get("checks") if isinstance(conservation.get("checks"), dict) else {}
        preservation = checks.get("preservation") if isinstance(checks.get("preservation"), dict) else {}
        metrics = {
            "available": True,
            "functional_count": len(b_items) if b_items else 0,
            "a_final_xlsx_rows": a_result.xlsx_rows,
            "b_final_xlsx_rows": b_result.xlsx_rows,
            "duplicate_rate": _duplicate_rate_texts([row["body"] for row in b_rows]),
            "functional_duplicate_rate": _duplicate_rate(b_items),  # 诊断口径
            "preservation_blocking_losses": len(preservation.get("blocking_losses") or []),
            "preservation_warning_losses": len(preservation.get("warning_losses") or []),
            "conservation_warning_count": conservation.get("warning_count"),
            "final_xlsx": dict(b_result.xlsx_struct),
        }
        if doc_truth:
            metrics["truth_metrics"] = match_truth_to_rows(doc_truth, b_rows)
        else:
            metrics["truth_metrics"] = {
                "available": False, "reason": "no truth rows selected for this document"}

    threshold_violations = (
        _evaluate_thresholds(metrics, effective_thresholds)
        if effective_thresholds and not failures else []
    )
    if failures:
        verdict = "FAIL"
    elif missing_gates:
        verdict = "NO_GATE"
    elif threshold_violations:
        verdict = "FAIL"
    else:
        verdict = "PASS"

    report = {
        "schema": REPORT_SCHEMA,
        "parsed_dir": str(parsed_dir),
        "document_keys": {k: doc_keys.get(k, "") for k in ("name", "sha256", "path")},
        "route": route,
        "env_snapshot": {k: v for k, v in sorted(os.environ.items())
                         if k.startswith("RATOMIZER_")},
        "switch_env": SWITCH_ENV,
        "results": {
            "A_atoms": a_result.__dict__,
            "B_direct": b_result.__dict__,
        },
        "metrics": metrics,
        "thresholds": dict(effective_thresholds),
        "threshold_layering": threshold_meta,
        "verdict": verdict,
        "failures": failures,
        "missing_gates": missing_gates,
        "threshold_violations": threshold_violations,
        "work_dirs": {"A": str(a_dir), "B": str(b_dir)} if keep_dirs else {},
        "a_warm_cache": warm_cache_info,
    }
    if not keep_dirs:
        shutil.rmtree(work_root, ignore_errors=True)
    return report


def _load_thresholds(path: Path | None) -> dict[str, Any] | None:
    if path is None:
        return None
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"thresholds file must be a JSON object: {path}")
    return payload


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--parsed-dir", type=Path, required=True, action="append",
                        help="已 parse 的产物目录（含 blocks.jsonl/chunks.jsonl）；可重复传入多份语料")
    parser.add_argument("--template", type=Path, required=True,
                        help="公司需求列表模板 xlsx（A/B 必须同用一份空模板）")
    parser.add_argument("--route", default="openai_compatible",
                        help="LLM 路由（默认 openai_compatible；stub 被拒绝）")
    parser.add_argument("--truth", type=Path, default=None,
                        help="功能级人工真值集 JSONL（schemas/functional_truth.schema.json；"
                             "缺真值集或无本文档行 → NO_GATE，不作翻转依据）")
    parser.add_argument("--thresholds", type=Path, default=None,
                        help="门槛 JSON（REQUIRED_THRESHOLD_KEYS 14 项全必需；可含 "
                             "documents 按文档覆盖层；缺任一必需键 → NO_GATE）")
    parser.add_argument("--out", type=Path, default=None, help="报告输出路径（JSON）")
    parser.add_argument("--warm-a-cache", type=Path, default=None,
                        help="A 轨缓存暖启动目录（含 ai_extract_cache.jsonl，如上次 --keep-dirs 的 A_atoms）"
                             "——重跑门禁时 A 轨抽取零付费，整链仍真实执行")
    parser.add_argument("--keep-dirs", action="store_true",
                        help="保留 A/B 工作目录供人工复核（默认清理）")
    args = parser.parse_args(argv)

    try:
        truth_rows = _load_truth(args.truth)
        thresholds = _load_thresholds(args.thresholds)
    except (OSError, ValueError) as exc:  # 门输入坏 → 响亮失败，不出含糊报告
        print(f"[ab_runner] gate inputs failed to load: {exc}", file=sys.stderr)
        return 1  # NO_GATE 语义：门不完备，报告不可作为 Go/No-Go 依据
    reports = []
    for parsed_dir in args.parsed_dir:
        try:
            report = run_ab_for_document(
                parsed_dir,
                route=args.route,
                template_path=args.template,
                truth_rows=truth_rows,
                thresholds=thresholds,
                keep_dirs=args.keep_dirs,
                warm_a_cache=args.warm_a_cache,
            )
        except Exception as exc:  # noqa: BLE001 — 逐份判定：单份异常不掩盖其他份
            report = {
                "schema": REPORT_SCHEMA,
                "parsed_dir": str(parsed_dir),
                "verdict": "FAIL",
                "failures": [f"runner error: {type(exc).__name__}: {exc}"],
                "traceback": traceback.format_exc()[-2000:],
            }
        reports.append(report)
        print(f"[ab_runner] {parsed_dir}: {report.get('verdict')} "
              f"failures={report.get('failures')} "
              f"missing_gates={report.get('missing_gates')} "
              f"threshold_violations={report.get('threshold_violations')}",
              file=sys.stderr)

    verdict_priority = {"FAIL": 2, "NO_GATE": 1, "PASS": 0}
    overall = max(
        (r.get("verdict", "FAIL") for r in reports),
        key=lambda v: verdict_priority.get(v, 2),
    )
    summary = {
        "schema": REPORT_SCHEMA,
        "overall_verdict": overall,
        "document_count": len(reports),
        "documents": reports,
    }
    output = json.dumps(summary, ensure_ascii=False, indent=2)
    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_text(output + "\n", encoding="utf-8")
        print(f"[ab_runner] report written: {args.out}", file=sys.stderr)
    else:
        print(output)
    # 退出码：0=PASS（可作翻转依据）、1=NO_GATE（门不完备）、2=FAIL（含单份失败）
    return {"PASS": 0, "NO_GATE": 1, "FAIL": 2}[overall]


if __name__ == "__main__":
    sys.exit(main())
