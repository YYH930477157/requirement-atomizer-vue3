"""模板成文器（终局架构的第三步，用户两轮裁定后的最终形态）。

  理解（AI 抽取，按文档逻辑） → 分析（软件需求分析轨：软/硬/协同 + 可研发正文 + 研发指引）
  → **成文（本模块）：分析结果按公司标准化需求列表 V2.3.x 的格式写入对应模块 sheet**

模板 = 交付格式与风格样本，**不是问题库**——已有行原样保留（风格参照），每条分析后的客户
需求作为**新行追加**到对应模块 sheet 尾部：子模块/描述/需求/说明/是否客户需求=是/客户需求
章节/驱动硬件相关。纯确定性（零 LLM）：分析内容来自 analyze 轨，这里只做格式装配。

模块 → sheet 映射按模板实际 sheet 清单构建；没有对应 sheet 的模块（如 安全/环境可靠性）落到
新建的「其他需求(新增)」sheet——诚实存放，不硬塞。

用法：python -m template_writer --out <目录（含 engineering_analysis.json）>
      --template <V2.3.x.xlsx>
产物：<out>/软件需求列表-成文.xlsx + <out>/template_writer_report.json
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path
from typing import Any

from requirements_analysis_excel import _notes_text, _safe_cell
from requirements_analysis_schema import OWNERSHIP_CO_DESIGN, OWNERSHIP_HARDWARE
from compliance import is_compliance_requirement

LOGGER = logging.getLogger("requirement_atomizer")

WRITTEN_WORKBOOK = "软件需求列表-成文.xlsx"
WRITER_REPORT = "template_writer_report.json"
FALLBACK_SHEET = "其他需求(新增)"

# 模板统一列位（V2.3.x 实测 19 个需求 sheet 一致）
_COL_SEQ = 2
_COL_SUBMODULE = 3
_COL_QUESTION = 4    # 描述
_COL_ANSWER = 6      # 需求
_COL_NOTES = 7       # 说明、示例、注意事项
_COL_IS_CUSTOMER = 8
_COL_SECTION = 9
_COL_HW = 10         # 驱动/硬件相关

# 抽取轨模块名 → 模板 sheet 名（sheet 存在性运行时校验；缺的落 FALLBACK）
MODULE_TO_SHEET = {
    "计量": "计量需求", "时钟": "时钟需求", "费率": "费率需求", "显示": "显示需求",
    "需量": "需量需求", "结算": "结算需求", "曲线": "负荷曲线", "窃电": "报警窃电需求",
    "电网质量": "电网质量需求", "升级": "升级需求", "负控": "负控需求", "状态字": "状态字需求",
    "事件记录": "事件需求", "通信协议": "协议栈需求", "Push": "push需求", "预付费": "预付费需求",
    "计量精度": "计量需求", "门限范围": "系统需求", "数据存储": "系统需求",
    "CIU": FALLBACK_SHEET, "其它": FALLBACK_SHEET, "安全": FALLBACK_SHEET,
    "机械结构": FALLBACK_SHEET, "测试合规": FALLBACK_SHEET, "环境可靠性": FALLBACK_SHEET,
    "节假日": FALLBACK_SHEET, "附加功能": FALLBACK_SHEET,
}


def target_sheet(item: dict[str, Any], sheetnames: list[str]) -> str:
    # module 是模板词表归一后的值，submodule 保留抽取轨原模块名——两级尝试
    for key in ("module", "submodule"):
        sheet = MODULE_TO_SHEET.get(str(item.get(key) or "").strip(), "")
        if sheet in sheetnames:
            return sheet
    return FALLBACK_SHEET


def module_mapping_drift() -> tuple[list[str], list[str]]:
    """MODULE_TO_SHEET 与抽取轨 MODULE_VOCAB 的漂移检查（0711 评审）。

    返回 (vocab 里没有映射的模块, 映射表里多余的、不在 vocab 的键)。两份词表分头维护
    易漂移：新增模块没进映射表 → 静默落 FALLBACK_SHEET；映射表里写错/多写 → 永不命中。
    报告里显式列出，供人核，不阻断成文。
    """
    try:
        from ai_extract import MODULE_VOCAB
    except Exception:
        return ([], [])
    vocab = set(MODULE_VOCAB)
    mapped = set(MODULE_TO_SHEET)
    unmapped = sorted(vocab - mapped)
    extra = sorted(mapped - vocab)
    return (unmapped, extra)


def _next_seq(ws: Any) -> tuple[int, int]:
    """返回 (下一序号, 追加起始行)。序号接着表内最大数字序号继续。"""
    max_seq = 0
    last_row = 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row and any(c not in (None, "") for c in row):
            last_row = row_idx
        try:
            max_seq = max(max_seq, int(str(row[_COL_SEQ - 1]).strip()))
        except (TypeError, ValueError, IndexError):
            pass
    return max_seq + 1, last_row + 1


def build_row_values(item: dict[str, Any], seq: int) -> dict[int, Any]:
    """分析条目 → 模板行（列号→值）。内容全部来自 analyze 轨产物，此处零生成。"""
    hw = "是" if item.get("ownership") == OWNERSHIP_CO_DESIGN else ""
    return {
        _COL_SEQ: seq,
        _COL_SUBMODULE: item.get("submodule") or item.get("module") or "",
        _COL_QUESTION: item.get("description") or "",
        _COL_ANSWER: item.get("software_requirement_text") or item.get("requirement") or "",
        _COL_NOTES: _notes_text(item),
        _COL_IS_CUSTOMER: "是",
        _COL_SECTION: item.get("source_section") or "",
        _COL_HW: hw,
    }


def append_analysis_to_template(template_path: Path, items: list[dict[str, Any]],
                                out_path: Path) -> dict[str, Any]:
    """把分析条目按模块追加进模板副本。硬件独占项不进软件需求列表（走 hardware_items.md，
    与 analyze 轨一致）；协同项进列表并标「驱动/硬件相关=是」。"""
    from openpyxl import load_workbook
    wb = load_workbook(template_path)
    appended: dict[str, int] = {}
    skipped_hardware = 0
    skipped_compliance = 0

    software_items = []
    for item in items:
        # ``source_requirement_type`` is emitted by requirements_analysis, not by the narrative
        # LLM. Trust it only at this final defensive boundary so a stale analysis file cannot put
        # a known compliance row back into the software workbook.
        source_type = str(item.get("source_requirement_type") or "").casefold()
        if source_type == "compliance" or is_compliance_requirement(item):
            skipped_compliance += 1
            continue
        if item.get("ownership") == OWNERSHIP_HARDWARE:
            skipped_hardware += 1
            continue
        software_items.append(item)

    # 分组后逐 sheet 追加（保证同 sheet 序号连续）
    by_sheet: dict[str, list[dict[str, Any]]] = {}
    for item in software_items:
        by_sheet.setdefault(target_sheet(item, wb.sheetnames), []).append(item)

    for sheet, sheet_items in by_sheet.items():
        if sheet == FALLBACK_SHEET and sheet not in wb.sheetnames:
            ws = wb.create_sheet(FALLBACK_SHEET)
            ws.append(["关闭", "序号", "子模块", "描述", "需求模版", "需求",
                       "说明、示例、注意事项", "是否客户需求", "客户需求章节", "驱动/硬件相关"])
        ws = wb[sheet]
        seq, row_idx = _next_seq(ws)
        for item in sheet_items:
            for col, value in build_row_values(item, seq).items():
                ws.cell(row=row_idx, column=col, value=_safe_cell(value))
            seq += 1
            row_idx += 1
        appended[sheet] = len(sheet_items)

    from xlsx_io import safe_save_workbook
    out_path = safe_save_workbook(wb, out_path)
    return {"appended_by_sheet": dict(sorted(appended.items(), key=lambda x: -x[1])),
            "appended_total": sum(appended.values()),
            "skipped_hardware": skipped_hardware,
            "skipped_compliance": skipped_compliance,
            "workbook": out_path.name}


def run_writer(out_dir: Path, template_path: Path) -> dict[str, Any]:
    out_dir = Path(out_dir).expanduser().resolve()
    analysis_path = out_dir / "engineering_analysis.json"
    if not analysis_path.exists():
        raise FileNotFoundError(
            f"engineering_analysis.json not found in {out_dir} — 先跑「软件需求分析」（分析在前，成文在后）")
    payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    items = payload.get("items") or []
    out_path = out_dir / WRITTEN_WORKBOOK
    report = append_analysis_to_template(template_path, items, out_path)
    from requirement_record import provenance as _prov
    report["provenance"] = _prov("template_writer", "template_writer/v1")
    unmapped, extra = module_mapping_drift()
    if unmapped or extra:
        report["module_mapping_drift"] = {"unmapped_vocab": unmapped, "extra_keys": extra}
        LOGGER.warning("模块→sheet 映射与抽取词表漂移：未映射=%s 多余=%s（缺失项将落兜底 sheet）",
                       unmapped, extra)
    report.update({"items": len(items), "analysis_route": payload.get("route"),
                   "written": [report.get("workbook") or WRITTEN_WORKBOOK, WRITER_REPORT]})
    (out_dir / WRITER_REPORT).write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Write analyzed requirements into the company template workbook.")
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    args = parser.parse_args(argv)
    try:
        report = run_writer(args.out, args.template)
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
