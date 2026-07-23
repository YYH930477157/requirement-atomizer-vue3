"""requirements_analysis_excel 回归（unittest 风格——pytest 未装，模块级函数不会被 discover 收集）。"""
from __future__ import annotations

import tempfile
import unittest
import warnings
from pathlib import Path

from openpyxl import load_workbook

from requirements_analysis_excel import write_software_requirements_xlsx


class SheetTitleTests(unittest.TestCase):
    def test_sheet_titles_are_unique_and_excel_safe(self) -> None:
        long = "A" * 40
        items = [
            {"ownership": "software", "module": long + "x", "description": "one"},
            {"ownership": "software", "module": long + "y", "description": "two"},
        ]

        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                write_software_requirements_xlsx(items, tmp_path / "software.xlsx")

            assert not [w for w in caught if "Title is more than 31 characters" in str(w.message)]
            wb = load_workbook(tmp_path / "software.xlsx", data_only=True)
            assert len(set(wb.sheetnames)) == 2
            assert all(len(name) <= 31 for name in wb.sheetnames)

    def test_sheet_titles_are_unique_case_insensitively(self) -> None:
        items = [
            {"ownership": "software", "module": "a" * 31, "description": "one"},
            {"ownership": "software", "module": "A" * 31, "description": "two"},
        ]

        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                write_software_requirements_xlsx(items, tmp_path / "software.xlsx")

            assert not [w for w in caught if "Title is more than 31 characters" in str(w.message)]
            wb = load_workbook(tmp_path / "software.xlsx", data_only=True)
            lowered = [name.casefold() for name in wb.sheetnames]
            assert len(set(lowered)) == 2
            assert all(len(name) <= 31 for name in wb.sheetnames)


class ThresholdTableTests(unittest.TestCase):
    def test_threshold_table_rendered_into_notes_column(self) -> None:
        """真实反馈：粉尘粒径/成分数值清单必须出现在最终交付物——不能只留在中间产物。"""
        items = [{
            "ownership": "software", "module": "环境可靠性", "description": "尘埃测试用尘规格",
            "threshold_table": {"columns": ["批次", "粒径范围", "平均粒径"],
                                 "rows": [["1", "0-100 um", "(50 ± 10) um"],
                                          ["2", "100-200 um", "(150 ± 10) um"]]},
        }]
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "software.xlsx"
            write_software_requirements_xlsx(items, path)
            wb = load_workbook(path, data_only=True)
            texts = [str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.value]
        joined = "\n".join(texts)
        self.assertIn("参数表：批次 | 粒径范围 | 平均粒径", joined)
        self.assertIn("(50 ± 10) um", joined)
        self.assertIn("(150 ± 10) um", joined)


class CellSafetyTests(unittest.TestCase):
    def test_formula_injection_is_neutralized(self) -> None:
        """交付研发的工作簿零活公式（spec-data-integrity 同一红线；此前缺 formula_safe 被复引入）。"""
        items = [{
            "ownership": "software", "module": "计量",
            "description": '=HYPERLINK("http://evil.example/x","点我")',
            "requirement": "+A1+A2",
            "developer_guidance": ["-2+3+cmd|' /C calc'!A0"],
        }]
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "software.xlsx"
            write_software_requirements_xlsx(items, path)
            wb = load_workbook(path)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        assert cell.data_type != "f", f"live formula at {ws.title}!{cell.coordinate}"
                        if isinstance(cell.value, str):
                            assert not cell.value.startswith("=")

    def test_control_characters_are_scrubbed(self) -> None:
        """PDF 文本层常见控制字符（\\x0b 等）会让 openpyxl 抛 IllegalCharacterError——写入前剥离。"""
        items = [{
            "ownership": "software", "module": "计量",
            "description": "before\x0bafter\x00tail",
            "source_quote": "quote\x0ehere",
        }]
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "software.xlsx"
            write_software_requirements_xlsx(items, path)  # 不抛即胜一半
            wb = load_workbook(path)
            texts = [str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.value]
            assert any("beforeafter" in t for t in texts)  # 控制字符被剥、正文保留


class HardwareDependencyRenderTests(unittest.TestCase):
    """审计 P1-b：hardware_dependency（含 WP2 待澄清标注）必须落到交付物——
    xlsx 说明列直出；template_writer 走同一 _notes_text 自动获得。"""

    def test_notes_column_renders_hardware_dependency(self) -> None:
        from requirements_analysis_excel import _notes_text

        notes = _notes_text({
            "ownership": "co_design", "module": "计量",
            "hardware_dependency": "需计量芯片支持四费率寄存器",
        })
        self.assertIn("硬件依赖：需计量芯片支持四费率寄存器", notes)

    def test_clarified_hardware_dependency_renders_with_fallback_label(self) -> None:
        """待澄清的依赖带"未经依据校验 + 原始候选"标注透出（clarify_display_text 通道）。"""
        from requirements_analysis_excel import _notes_text

        notes = _notes_text({
            "ownership": "co_design", "module": "计量",
            "hardware_dependency": "待澄清",
            "clarify_fallback": {"hardware_dependency": "原始候选依赖内容"},
        })
        self.assertIn("硬件依赖：待澄清（未经依据校验，需专家核补）", notes)
        self.assertIn("原始候选（未经依据校验，仅供参考，不得作为实现依据）：原始候选依赖内容", notes)

    def test_empty_hardware_dependency_not_rendered(self) -> None:
        from requirements_analysis_excel import _notes_text

        notes = _notes_text({"ownership": "software", "module": "计量", "hardware_dependency": ""})
        self.assertNotIn("硬件依赖：", notes)

    def test_xlsx_notes_column_carries_hardware_dependency(self) -> None:
        items = [{
            "ownership": "co_design", "module": "计量", "description": "费控切换",
            "software_requirement_text": "按费率时段切换",
            "hardware_dependency": "需内置继电器",
        }]
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "software.xlsx"
            write_software_requirements_xlsx(items, path)
            wb = load_workbook(path, data_only=True)
            texts = [str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.value]
        self.assertTrue(any("硬件依赖：需内置继电器" in text for text in texts))


if __name__ == "__main__":
    unittest.main()
