from __future__ import annotations

import logging
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from atomize import DocumentProfile, build_table_artifacts, clean_text
from requirement_kb import KnowledgeRepository
from xlsx_region_detect import (
    SheetTableFingerprint,
    boundary_conflicts_to_audit,
    extract_obis_keys_from_matrix,
    link_multi_sheet_tables,
    link_result_to_audit,
    validate_region_boundaries,
)


LOGGER = logging.getLogger("requirement_atomizer")
MAX_SHEET_ROWS = 50_000
# S14：公式扫描的列维上限（xlsx 规格硬上限 16384 列）。行维有 MAX_SHEET_ROWS 截断
# 审计而列维无上限，口径不齐；超限必须 fail-closed 计入审计，超界列的无缓存公式格
# 不得静默逃逸。
MAX_SHEET_COLUMNS = 16_384


def extract_xlsx(
    input_path: Path,
    knowledge_bases: KnowledgeRepository | None = None,
    document_profile: DocumentProfile | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    del document_profile
    knowledge_bases = knowledge_bases or KnowledgeRepository.from_paths([])
    # read_only=True 时，部分含条件格式/数据校验的 xlsx 不预填 sheet.max_row/max_column
    # （返回 None），导致所有 sheet 被误判为空。改用 read_only=False 保证维度可靠，
    # 且 _merged_fill_values 需要 sheet.cell() 随机访问（read_only 下不可用）。
    workbook = load_workbook(input_path, data_only=True, read_only=False)
    # B7：公式双视图——data_only=True 对无缓存公式返回 None（`="The meter shall …"`
    # 单元格凭空消失且 parse_incomplete=False）；data_only=False 视图识别公式存在性，
    # 有公式但无缓存值 = 内容不可见，必须 fail-closed 并计入非空守恒
    formula_workbook = load_workbook(input_path, data_only=False, read_only=False)
    # S14：merge ranges 复用首次加载的 workbook（与 data_only 无关），省掉第三次整文件加载
    merge_ranges_by_sheet = _merged_ranges_by_sheet(workbook)
    blocks: list[dict[str, Any]] = []
    table_items: list[dict[str, Any]] = []
    table_cell_items: list[dict[str, Any]] = []
    order = 0
    table_count = 0
    # WS1 wk7: 收集每表 OBIS 指纹，供跨 sheet 关联一致性门（link_multi_sheet_tables）
    region_fingerprints: list[SheetTableFingerprint] = []

    try:
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                continue
            formula_sheet = next(
                (candidate for candidate in formula_workbook.worksheets
                 if candidate.title == sheet.title),
                None,
            )
            unavailable_formulas, formula_scan_truncated = (
                _formula_cells_without_cached_values(sheet, formula_sheet)
            )
            sheet_merges = merge_ranges_by_sheet.get(sheet.title, [])
            parse_audit: dict[str, Any] = {}
            if unavailable_formulas:
                parse_audit["parse_incomplete"] = True
                parse_audit["parse_incomplete_reason"] = {
                    "code": "xlsx_formula_value_unavailable",
                    "cell_count": len(unavailable_formulas),
                    "sample": [
                        f"R{row_index}C{column_index}"
                        for row_index, column_index in sorted(unavailable_formulas)[:10]
                    ],
                }
            if formula_scan_truncated:
                # 列超上限时公式扫描不完整（unavailable 集合只是部分结果）——
                # 以截断原因为准 fail-closed，不得拿部分扫描冒充完整
                parse_audit["parse_incomplete"] = True
                parse_audit["parse_incomplete_reason"] = {
                    "code": "xlsx_column_limit",
                    "observed_columns": int(sheet.max_column or 0),
                    "scanned_columns": MAX_SHEET_COLUMNS,
                    "limit": MAX_SHEET_COLUMNS,
                }
            regions, region_merges, region_headers = _sheet_table_regions(
                sheet, sheet_merges, audit=parse_audit,
                extra_non_empty=unavailable_formulas,
            )
            if not regions:
                continue

            # WS1 wk7: 区域边界确定性校验门（不重叠 / 不切断原生合并格）。校验只读
            # 既有 region + merge 模型，冲突如实进 parse_incomplete_reason 既有通道
            # （作为 additional_reasons 附加，不覆盖既有 code、不新造格式）。
            boundary_check = validate_region_boundaries(regions, sheet_merges)
            boundary_audit = boundary_conflicts_to_audit(boundary_check, sheet_name=sheet.title)
            if boundary_audit is not None:
                parse_audit["parse_incomplete"] = True
                existing_reason = parse_audit.get("parse_incomplete_reason")
                if existing_reason:
                    existing_reason.setdefault("additional_reasons", []).append(boundary_audit)
                else:
                    parse_audit["parse_incomplete_reason"] = boundary_audit

            order += 1
            section_path = [sheet.title]
            blocks.append(
                {
                    "block_id": f"BLK-{order:06d}",
                    "order": order,
                    "type": "heading",
                    "source_format": "xlsx",
                    "heading_level": 1,
                    "text": sheet.title,
                    "section_path": section_path,
                    "domain_tags": [],
                    "kb_matches": [],
                    "requirement_like": False,
                    "noise": False,
                }
            )

            for region_index, region in enumerate(regions):
                min_row, min_col, max_row, max_col = region
                matrix = _region_matrix(sheet, region, region_merges[region_index])
                if not matrix:
                    continue
                order += 1
                table_count += 1
                table_id = f"TBL-{table_count:06d}"
                # WS1 wk7: 收集本表 OBIS 指纹（确定性 extract_codes 扫全矩阵），
                # 供 return 前的跨 sheet 关联一致性门
                region_fingerprints.append(
                    SheetTableFingerprint(
                        sheet_name=sheet.title,
                        table_id=table_id,
                        obis_keys=extract_obis_keys_from_matrix(matrix),
                    )
                )
                # 标题识别交给 table_structure（全宽合并证据）；sheet 名作回退标题，
                # 删除"首行一个非空格即标题"硬规则
                table_block, new_table_items, new_cell_items = build_table_artifacts(
                    matrix,
                    table_id=table_id,
                    block_id=f"BLK-{order:06d}",
                    order=order,
                    table_title=sheet.title,
                    section_path=section_path,
                    knowledge_bases=knowledge_bases,
                    parse_incomplete=bool(parse_audit.get("parse_incomplete")),
                    parse_incomplete_reason=parse_audit.get("parse_incomplete_reason"),
                    merge_ranges=region_merges[region_index],
                    explicit_header_rows=region_headers[region_index],
                    source_format="xlsx",
                    sheet_name=sheet.title,
                    a1_origin=(min_row, min_col),
                )
                for item in new_table_items:
                    item["source_format"] = "xlsx"
                blocks.append(table_block)
                table_items.extend(new_table_items)
                table_cell_items.extend(new_cell_items)
    finally:
        workbook.close()
        formula_workbook.close()

    # WS1 wk7: 多 sheet OBIS 关联一致性门。extract_xlsx 本就不跨 sheet 合并
    # table_items（每区域独立成表），故"不静默合并"在抽取层天然成立；此门是未来
    # 跨 sheet 合并提交的确定性前置校验——键缺失/冲突时如实 LOGGER 报告，并把
    # 阻塞信息回溯标进相关表块的 parse_incomplete_reason 既有通道（不新造格式）。
    link_result = link_multi_sheet_tables(region_fingerprints)
    link_audit = link_result_to_audit(link_result)
    if link_audit is not None:
        LOGGER.warning(
            "multi-sheet OBIS link blocked: status=%s keyed_tables=%d conflicts=%d",
            link_result.status, link_result.keyed_table_count, len(link_result.conflicts),
        )
        keyed_table_ids = {fp.table_id for fp in region_fingerprints if fp.has_keys}
        for block in blocks:
            if block.get("type") != "table" or block.get("table_id") not in keyed_table_ids:
                continue
            block["parse_incomplete"] = True
            reason = block.get("parse_incomplete_reason")
            if isinstance(reason, dict):
                reason.setdefault("additional_reasons", []).append(link_audit)
            else:
                block["parse_incomplete_reason"] = link_audit

    return blocks, table_items, table_cell_items


def _formula_cells_without_cached_values(
    sheet: Any,
    formula_sheet: Any,
) -> tuple[set[tuple[int, int]], bool]:
    """有公式但无缓存值的单元格坐标（B7）+ 列扫描是否被上限截断（S14）。

    data_only=True 对无缓存公式返回 None——内容不可见；不得假装它是空格
    （守恒/连通性都看不到 = 静默丢失）。坐标计入 extra_non_empty 并由
    调用方 fail-closed（xlsx_formula_value_unavailable）。列维超
    MAX_SHEET_COLUMNS 时扫描不完整，第二个返回值置 True，调用方必须以
    xlsx_column_limit fail-closed——超界列的公式格不得静默逃逸。"""
    unavailable: set[tuple[int, int]] = set()
    if formula_sheet is None:
        return unavailable, False
    max_row = min(sheet.max_row or 0, MAX_SHEET_ROWS)
    observed_columns = sheet.max_column or 0
    max_column = min(observed_columns, MAX_SHEET_COLUMNS)
    truncated = observed_columns > MAX_SHEET_COLUMNS
    if max_row == 0 or max_column == 0:
        return unavailable, truncated
    for row in formula_sheet.iter_rows(
        min_row=1, max_row=max_row, max_col=max_column
    ):
        for cell in row:
            if getattr(cell, "data_type", None) != "f":
                continue
            cached = sheet.cell(row=cell.row, column=cell.column).value
            if cached is None or not str(cached).strip():
                unavailable.add((cell.row, cell.column))
    return unavailable, truncated


def _sheet_table_regions(
    sheet: Any,
    merge_ranges: list[tuple[int, int, int, int]],
    *,
    audit: dict[str, Any],
    extra_non_empty: set[tuple[int, int]] | None = None,
) -> tuple[
    list[tuple[int, int, int, int]],
    list[list[tuple[int, int, int, int]]],
    list[list[int]],
]:
    """同一 sheet 的表区域拆分：优先 Excel Table 定义，否则按非空连通区域。

    返回 (regions, per-region 相对 merge ranges, per-region 表头证据)。
    表头证据三态：list[int]（显式表头行，空列表=显式 headerless）| None（无证据，
    走推断）。区域保留左上角 A1 坐标（非 A1 起始区域不丢溯源）。
    extra_non_empty：值视图不可见但内容确实存在的坐标（B7 无缓存公式格）——
    连通性与守恒计数一律按非空对待。"""
    extra_non_empty = extra_non_empty or set()
    max_row = sheet.max_row or 0
    max_column = sheet.max_column or 0
    if max_row == 0 or max_column == 0:
        return [], [], []
    if max_row > MAX_SHEET_ROWS:
        LOGGER.warning("sheet %s has %s rows; truncating to %s", sheet.title, max_row, MAX_SHEET_ROWS)
        audit["parse_incomplete"] = True
        audit.setdefault(
            "parse_incomplete_reason",
            {
                "code": "xlsx_row_limit",
                "observed_rows": max_row,
                "parsed_rows": MAX_SHEET_ROWS,
                "limit": MAX_SHEET_ROWS,
            },
        )
        max_row = MAX_SHEET_ROWS

    table_defs = _excel_table_regions(sheet, max_row=max_row)
    if table_defs:
        regions = [entry[0] for entry in table_defs]
        header_rows: list[list[int] | None] = [entry[1] for entry in table_defs]
        # ListObject 之外仍须守恒：表定义覆盖不到的非空连通区域同样成表（否则表外
        # 需求彻底消失且任何下游计数器都看不到）
        covered_cells: set[tuple[int, int]] = set()
        for min_row, min_col, region_max_row, max_col in regions:
            for row_index in range(min_row, region_max_row + 1):
                for column_index in range(min_col, max_col + 1):
                    covered_cells.add((row_index, column_index))
        leftovers = _connected_regions(
            sheet, max_row=max_row, max_column=max_column, skip=covered_cells,
            extra_non_empty=extra_non_empty,
        )
        regions.extend(leftovers)
        header_rows.extend([None for _ in leftovers])
        regions, header_rows = zip(
            *sorted(zip(regions, header_rows), key=lambda entry: (entry[0][0], entry[0][1]))
        )
        regions, header_rows = list(regions), list(header_rows)
    else:
        regions = _connected_regions(
            sheet, max_row=max_row, max_column=max_column,
            extra_non_empty=extra_non_empty,
        )
        header_rows = [None for _ in regions]
    region_merges = [
        _clip_merges_to_region(merge_ranges, region) for region in regions
    ]
    # sheet 级守恒计数器：每个非空格必须落在恰好一个区域内。区域拆分（ListObject
    # 矩形 + 表外连通区域）理论上全覆盖，但覆盖盲区/区域重叠 = 静默丢内容——
    # 计数器把"相信全覆盖"变成 fail-closed 硬门（宁 parse_incomplete 不丢字）
    coverage: dict[tuple[int, int], int] = {}
    for min_row, min_col, region_max_row, region_max_col in regions:
        for row_index in range(min_row, region_max_row + 1):
            for column_index in range(min_col, region_max_col + 1):
                key = (row_index, column_index)
                coverage[key] = coverage.get(key, 0) + 1
    dropped_cells: list[tuple[int, int]] = []
    multi_covered_cells: list[tuple[int, int]] = []
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True),
        start=1,
    ):
        for column_index, value in enumerate(row, start=1):
            if value is None or not str(value).strip():
                continue
            count = coverage.get((row_index, column_index), 0)
            if count == 0:
                dropped_cells.append((row_index, column_index))
            elif count > 1:
                multi_covered_cells.append((row_index, column_index))
    # 无缓存公式格（值视图不可见）同样参与守恒：区域内不丢、区域外如实 dropped
    for coord in sorted(extra_non_empty):
        if coverage.get(coord, 0) == 0:
            dropped_cells.append(coord)
    if dropped_cells or multi_covered_cells:
        LOGGER.warning(
            "sheet %s region conservation violated: dropped=%s multi_covered=%s",
            sheet.title, len(dropped_cells), len(multi_covered_cells),
        )
        audit["parse_incomplete"] = True
        audit.setdefault(
            "parse_incomplete_reason",
            {
                "code": "xlsx_region_conservation",
                "dropped_cell_count": len(dropped_cells),
                "multi_covered_cell_count": len(multi_covered_cells),
                "dropped_sample": [
                    f"R{row_index}C{column_index}"
                    for row_index, column_index in dropped_cells[:10]
                ],
            },
        )
    return regions, region_merges, header_rows


def _excel_table_regions(
    sheet: Any, *, max_row: int
) -> list[tuple[tuple[int, int, int, int], list[int]]]:
    """Excel Table（ListObject）定义区域 + 显式表头行（相对区域 1-based）。"""
    tables = getattr(sheet, "tables", None) or {}
    regions: list[tuple[tuple[int, int, int, int], list[int]]] = []
    for table in tables.values():
        ref = getattr(table, "ref", None)
        if not ref:
            continue
        try:
            min_col, min_row, max_col, table_max_row = range_boundaries(str(ref))
        except ValueError:
            continue
        table_max_row = min(table_max_row, max_row)
        if table_max_row < min_row:
            continue
        header_count = getattr(table, "headerRowCount", None)
        if header_count is None:
            header_count = 1
        # headerRowCount=0 是显式 headerless 证据，不得抹成 None 走推断
        header_rows = list(range(1, max(0, int(header_count)) + 1))
        regions.append(((min_row, min_col, table_max_row, max_col), header_rows))
    regions.sort(key=lambda entry: (entry[0][0], entry[0][1]))
    return regions


def _connected_regions(
    sheet: Any, *, max_row: int, max_column: int,
    skip: set[tuple[int, int]] | None = None,
    extra_non_empty: set[tuple[int, int]] | None = None,
) -> list[tuple[int, int, int, int]]:
    """无 Excel Table 定义时按非空连通区域拆表（保留区域左上角坐标）。

    八连通：矩阵表的角落孤格（如仅对角相邻的 X marker）仍属同一张表；
    真正分立的表之间至少隔一整行/列空白，不会被粘连。skip 坐标（已被
    ListObject 覆盖）不参与扫描。extra_non_empty 坐标（无缓存公式格）
    值视图不可见但按非空参与连通。"""
    skip = skip or set()
    non_empty: set[tuple[int, int]] = set()
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column, values_only=True),
        start=1,
    ):
        for column_index, value in enumerate(row, start=1):
            if (row_index, column_index) in skip:
                continue
            if value is not None and str(value).strip():
                non_empty.add((row_index, column_index))
    non_empty.update(coord for coord in (extra_non_empty or set()) if coord not in skip)
    regions: list[tuple[int, int, int, int]] = []
    remaining = set(non_empty)
    while remaining:
        seed = remaining.pop()
        stack = [seed]
        component = {seed}
        while stack:
            row_index, column_index = stack.pop()
            for d_row in (-1, 0, 1):
                for d_col in (-1, 0, 1):
                    if d_row == 0 and d_col == 0:
                        continue
                    neighbor = (row_index + d_row, column_index + d_col)
                    if neighbor in remaining:
                        remaining.remove(neighbor)
                        component.add(neighbor)
                        stack.append(neighbor)
        rows = [coord[0] for coord in component]
        columns = [coord[1] for coord in component]
        regions.append((min(rows), min(columns), max(rows), max(columns)))
    regions.sort(key=lambda region: (region[0], region[1]))
    return regions


def _clip_merges_to_region(
    merge_ranges: list[tuple[int, int, int, int]],
    region: tuple[int, int, int, int],
) -> list[tuple[int, int, int, int]]:
    """sheet 级 merge ranges 裁剪进区域并转为区域相对（1-based）坐标。"""
    min_row, min_col, max_row, max_col = region
    clipped: list[tuple[int, int, int, int]] = []
    for merge_min_row, merge_min_col, merge_max_row, merge_max_col in merge_ranges:
        if merge_max_row < min_row or merge_min_row > max_row:
            continue
        if merge_max_col < min_col or merge_min_col > max_col:
            continue
        clipped.append(
            (
                max(merge_min_row, min_row) - min_row + 1,
                max(merge_min_col, min_col) - min_col + 1,
                min(merge_max_row, max_row) - min_row + 1,
                min(merge_max_col, max_col) - min_col + 1,
            )
        )
    return clipped


def _region_matrix(
    sheet: Any,
    region: tuple[int, int, int, int],
    merge_ranges: list[tuple[int, int, int, int]],
) -> list[list[str]]:
    min_row, min_col, max_row, max_col = region
    merged_values = _merged_fill_values(
        sheet,
        [
            (
                min_row + entry[0] - 1,
                min_col + entry[1] - 1,
                min_row + entry[2] - 1,
                min_col + entry[3] - 1,
            )
            for entry in merge_ranges
        ],
        max_row=max_row,
    )
    matrix: list[list[str]] = []
    for row_offset, row in enumerate(
        sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col,
            values_only=True,
        )
    ):
        row_index = min_row + row_offset
        matrix.append(
            [
                clean_text(
                    _stringify_cell_value(
                        merged_values.get((row_index, min_col + column_offset), value)
                    )
                )
                for column_offset, value in enumerate(row)
            ]
        )
    return matrix


def _merged_ranges_by_sheet(workbook: Any) -> dict[str, list[tuple[int, int, int, int]]]:
    """S14：merge ranges 与 data_only 无关，复用 extract_xlsx 首次加载的 workbook——
    不再整文件第三次加载（大 xlsx 上解析成本曾 ×3）。"""
    return {
        sheet.title: [
            (cell_range.min_row, cell_range.min_col, cell_range.max_row, cell_range.max_col)
            for cell_range in sheet.merged_cells.ranges
        ]
        for sheet in workbook.worksheets
    }


def _merged_fill_values(
    sheet: Any,
    merge_ranges: list[tuple[int, int, int, int]],
    *,
    max_row: int,
) -> dict[tuple[int, int], Any]:
    """covered 格填充 anchor 值——但先校验原文（B6）。

    xlsx 的 covered 格通常为空（MergedCell.value=None）；若某 covered 格带有
    与 anchor 不同的非空原文（脏文件/隐藏内容），填充会静默覆盖别人的内容。
    冲突 range 整体不填充（保留全部原值）——矩阵带着异文进入
    validate_merge_text，该 range 被拒收并标 dropped_text_conflict/needs_review。"""
    values: dict[tuple[int, int], Any] = {}
    for min_row, min_col, range_max_row, max_col in merge_ranges:
        if min_row > max_row:
            continue
        top_left = sheet.cell(row=min_row, column=min_col).value
        anchor_text = clean_text(_stringify_cell_value(top_left))
        conflict = False
        for row_index in range(min_row, min(range_max_row, max_row) + 1):
            if conflict:
                break
            for column_index in range(min_col, max_col + 1):
                if (row_index, column_index) == (min_row, min_col):
                    continue
                raw_value = sheet.cell(row=row_index, column=column_index).value
                covered_text = clean_text(_stringify_cell_value(raw_value))
                if covered_text and covered_text != anchor_text:
                    conflict = True
                    break
        if conflict:
            continue
        for row_index in range(min_row, min(range_max_row, max_row) + 1):
            for column_index in range(min_col, max_col + 1):
                values[(row_index, column_index)] = top_left
    return values


def _stringify_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal) and value == value.to_integral_value():
        return str(int(value))
    return str(value)
