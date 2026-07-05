"""requirements_analysis 编排端到端回归（unittest 风格——pytest 未装，模块级函数不会被 discover 收集）。"""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

import api_server
from requirements_analysis import run_requirements_analysis


def write_jsonl(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


class RequirementsAnalysisPipelineTests(unittest.TestCase):
    def test_writes_json_and_reports(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {
                    "ai_req_id": "AI-1",
                    "title": "Clock",
                    "description": "The meter shall support Clock object daylight saving time.",
                    "source_quote": "support Clock object daylight saving time",
                    "source_block_ids": ["B-1"],
                    "module": "时钟需求",
                },
                {
                    "ai_req_id": "AI-2",
                    "description": "计量芯片型号为 Att7022e。",
                    "source_quote": "计量芯片型号为 Att7022e",
                    "source_block_ids": ["B-2"],
                    "module": "计量需求",
                },
            ])
            write_jsonl(tmp_path / "ai_review_states.jsonl", [
                {"ai_req_id": "AI-2", "ownership_override": "co_design", "reason": "软件需适配驱动"}
            ])

            result = run_requirements_analysis(tmp_path, route="stub", template_path=None)

            assert result["analysis_count"] == 2
            assert (tmp_path / "engineering_analysis.json").exists()
            assert (tmp_path / "hardware_items.md").exists()
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            by_id = {row["source_requirement_ids"][0]: row for row in payload["items"]}
            assert by_id["AI-2"]["ownership"] == "co_design"
            assert by_id["AI-2"]["ownership_source"] == "reviewer_override"

    def test_fills_base_item_contract(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {
                    "stable_req_id": "STABLE-1",
                    "description": "The meter shall support this feature.",
                    "source_block_ids": [101],
                }
            ])

            run_requirements_analysis(tmp_path, route="stub", template_path=None)

            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            item = payload["items"][0]
            assert item["source_kind"] == "ai_requirement"
            assert item["source_requirement_ids"] == ["STABLE-1"]
            assert item["source_block_ids"] == ["101"]
            assert item["software_requirement_text"] == ""

    def test_applies_review_state_for_raw_ai_requirement(self) -> None:
        from ai_review_actions import ai_req_id

        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            req = {
                "description": "The meter shall support Clock object daylight saving time.",
                "source_quote": "support Clock object daylight saving time",
                "source_block_ids": ["B-1"],
                "module": "时钟需求",
            }
            computed_id = ai_req_id(req)
            write_jsonl(tmp_path / "ai_requirements.jsonl", [req])
            write_jsonl(tmp_path / "ai_review_states.jsonl", [
                {"ai_req_id": computed_id, "ownership_override": "co_design", "reason": "人工改为协同"}
            ])

            run_requirements_analysis(tmp_path, route="stub", template_path=None)

            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            item = payload["items"][0]
            assert item["source_requirement_ids"] == [computed_id]
            assert item["ownership"] == "co_design"
            assert item["ownership_source"] == "reviewer_override"

    def test_api_review_id_matches_analysis_id_for_explicit_raw_ai_requirement(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            write_jsonl(tmp_path / "blocks.jsonl", [
                {
                    "block_id": "B-1",
                    "order": 1,
                    "type": "paragraph",
                    "text": "The meter shall support Clock object daylight saving time.",
                    "section_path": ["4"],
                }
            ])
            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {
                    "ai_req_id": "AI-1",
                    "description": "The meter shall support Clock object daylight saving time.",
                    "source_quote": "support Clock object daylight saving time",
                    "source_block_ids": ["B-1"],
                    "module": "clock requirements",
                }
            ])
            rows = api_server.build_ai_requirements(tmp_path)
            review_id = rows[0]["ai_req_id"]
            write_jsonl(tmp_path / "ai_review_states.jsonl", [
                {"ai_req_id": review_id, "ownership_override": "hardware", "reason": "reviewed as hardware"}
            ])

            run_requirements_analysis(tmp_path, route="stub", template_path=None)

            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            item = payload["items"][0]
            assert review_id == "AI-1"
            assert item["source_requirement_ids"] == ["AI-1"]
            assert item["ownership"] == "hardware"
            assert item["ownership_source"] == "reviewer_override"

    def test_skips_rejected_review_states(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {
                    "ai_req_id": "AI-1",
                    "description": "The meter shall support push notification.",
                    "source_quote": "support push notification",
                    "source_block_ids": ["B-1"],
                    "module": "push requirements",
                },
                {
                    "ai_req_id": "AI-2",
                    "description": "The meter shall support tariff display.",
                    "source_quote": "support tariff display",
                    "source_block_ids": ["B-2"],
                    "module": "display requirements",
                },
            ])
            write_jsonl(tmp_path / "ai_review_states.jsonl", [
                {"ai_req_id": "AI-1", "status": "rejected", "reason": "not a product requirement"}
            ])

            result = run_requirements_analysis(tmp_path, route="stub", template_path=None)

            assert result["analysis_count"] == 1
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            source_ids = [row["source_requirement_ids"][0] for row in payload["items"]]
            assert source_ids == ["AI-2"]

            wb = load_workbook(tmp_path / "software_requirements.xlsx", data_only=True)
            descriptions = [
                str(row[3] or "")
                for ws in wb.worksheets
                for row in ws.iter_rows(min_row=2, values_only=True)
                if any(row)
            ]
            assert not any("push notification" in value for value in descriptions)
            assert any("tariff display" in value for value in descriptions)

    def test_applies_module_override_before_template_match(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            template = tmp_path / "template.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Reviewed Module"
            ws.append(["unused", "Submodule"])
            ws.append(["", "Reviewed Module"])
            wb.save(template)

            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {
                    "ai_req_id": "AI-1",
                    "description": "The meter shall support push notification.",
                    "source_quote": "support push notification",
                    "source_block_ids": ["B-1"],
                    "module": "Original Module",
                }
            ])
            write_jsonl(tmp_path / "ai_review_states.jsonl", [
                {"ai_req_id": "AI-1", "status": "accepted", "module_override": "Reviewed Module"}
            ])

            run_requirements_analysis(tmp_path, route="stub", template_path=template)

            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            item = payload["items"][0]
            assert item["module"] == "Reviewed Module"
            assert item["submodule"] == "Reviewed Module"
            assert item["template_match"] == "matched"

            workbook = load_workbook(tmp_path / "software_requirements.xlsx", data_only=True)
            assert "Reviewed Module" in workbook.sheetnames
            rows = list(workbook["Reviewed Module"].iter_rows(min_row=2, values_only=True))
            assert rows[0][2] == "Reviewed Module"

    def test_software_workbook_excludes_hardware_and_includes_codesign(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {
                    "ai_req_id": "AI-1",
                    "description": "The meter shall support push notification.",
                    "source_quote": "support push notification",
                    "source_block_ids": ["B-1"],
                    "module": "push需求",
                },
                {
                    "ai_req_id": "AI-2",
                    "description": "计量芯片型号为 Att7022e。",
                    "source_quote": "计量芯片型号为 Att7022e",
                    "source_block_ids": ["B-2"],
                    "module": "计量需求",
                },
                {
                    "ai_req_id": "AI-3",
                    "description": "波特率最大值与硬件相关，需要驱动适配。",
                    "source_quote": "波特率最大值与硬件相关",
                    "source_block_ids": ["B-3"],
                    "module": "协议栈需求",
                },
            ])

            run_requirements_analysis(tmp_path, route="stub", template_path=None)

            wb = load_workbook(tmp_path / "software_requirements.xlsx", data_only=True)
            values = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        values.append(row)

            descriptions = [str(row[3] or "") for row in values]
            hardware_flags = [str(row[9] or "") for row in values]
            assert any("push" in value.lower() for value in descriptions)
            assert not any("计量芯片" in value for value in descriptions)
            assert "是" in hardware_flags

    def test_parse_args_rejects_unknown_route(self) -> None:
        from requirements_analysis import parse_args

        with self.assertRaises(SystemExit) as ctx:
            parse_args(["--out", ".", "--route", "bad"])
        assert ctx.exception.code != 0


class PipelineHardeningTests(unittest.TestCase):
    """复查修复的回归：缺输入响亮失败 / 撕裂状态行容错 / 非法归属单条降级 / route 出处诚实。"""

    def _seed_one(self, tmp_path: Path) -> None:
        write_jsonl(tmp_path / "ai_requirements.jsonl", [
            {
                "ai_req_id": "AI-1",
                "description": "The meter shall support tariff display.",
                "source_quote": "support tariff display",
                "source_block_ids": ["B-1"],
                "module": "display",
            }
        ])

    def test_missing_ai_requirements_raises(self) -> None:
        """缺输入必须响亮失败——打错 --out 不能静默产出"0 条 0 问题"的空交付物。"""
        with tempfile.TemporaryDirectory() as td:
            with self.assertRaises(FileNotFoundError):
                run_requirements_analysis(Path(td), route="stub", template_path=None)

    def test_torn_review_state_line_does_not_kill_run(self) -> None:
        """追加式 jsonl 的撕裂行（半行 JSON）跳过即可，好行照常生效。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed_one(tmp_path)
            (tmp_path / "ai_review_states.jsonl").write_text(
                '{"ai_req_id": "AI-1", "ownership_override": "co_de'  # 撕裂行
                + "\n"
                + json.dumps({"ai_req_id": "AI-1", "ownership_override": "hardware"}) + "\n",
                encoding="utf-8")

            result = run_requirements_analysis(tmp_path, route="stub", template_path=None)

            assert result["analysis_count"] == 1
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            assert payload["items"][0]["ownership"] == "hardware"  # 好行生效

    def test_invalid_ownership_override_degrades_per_item(self) -> None:
        """手改状态行的非法归属值：单条记 issue、忽略覆盖，绝不整跑死（设计文档逐条继续）。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed_one(tmp_path)
            write_jsonl(tmp_path / "ai_review_states.jsonl", [
                {"ai_req_id": "AI-1", "ownership_override": "firmware"}
            ])

            result = run_requirements_analysis(tmp_path, route="stub", template_path=None)

            assert result["analysis_count"] == 1  # 条目保留（规则归属）
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            assert payload["items"][0]["ownership_source"] == "rule"
            assert any("ownership_override 非法" in issue
                       for row in payload["issues"] for issue in row["issues"])

    def test_route_provenance_is_honest_when_endpoint_unusable(self) -> None:
        """请求 openai_compatible 但端点不可用（无 API key）→ 如实降级并记录，不谎称跑过 LLM。"""
        import os
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed_one(tmp_path)
            saved = os.environ.pop("RATOMIZER_LLM_API_KEY", None)
            try:
                result = run_requirements_analysis(tmp_path, route="openai_compatible", template_path=None)
            finally:
                if saved is not None:
                    os.environ["RATOMIZER_LLM_API_KEY"] = saved

            assert result["route"] == "stub"                     # 实际执行的
            assert result["route_requested"] == "openai_compatible"
            assert result["enriched"] == 0
            assert result.get("note")
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            assert payload["route"] == "stub"
            assert payload["route_requested"] == "openai_compatible"


class TemplateKnowledgeInjectionTests(unittest.TestCase):
    """模板知识注入端到端：设置模板 → 富化 prompt 带该模块的公司标准做法参考。"""

    def _seed(self, tmp_path: Path) -> Path:
        write_jsonl(tmp_path / "ai_requirements.jsonl", [{
            "ai_req_id": "AI-1",
            "title": "时钟精度要求",
            "description": "时钟精度须优于每天 5 秒",
            "source_quote": "clock accuracy shall be within 5 s per day",
            "source_block_ids": ["B-1"],
            "module": "时钟",
        }])
        from openpyxl import Workbook
        template = tmp_path / "template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "时钟需求"
        ws.append(["关闭", "序号", "子模块", "描述", "需求模版", "需求", "说明、示例、注意事项"])
        ws.append(["", 1, "时钟", "时钟精度：", "±0.5s/天", "±0.5s/天", "对应程序中 RTC_CAL 宏定义"])
        wb.save(template)
        return template

    def test_prompt_receives_company_reference_and_key_depends_on_refs(self) -> None:
        from requirements_analysis import _enrich_key
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            template = self._seed(tmp_path)
            captured: list[str] = []

            def fake_chat(system: str, user: str) -> dict:
                captured.append(user)
                return {"items": [{"source_requirement_ids": ["AI-1"],
                                   "software_requirement_text": "维持时钟精度在每天 5 秒内。",
                                   "developer_guidance": ["公司通用做法：校准对应 RTC_CAL 宏定义"],
                                   "acceptance_criteria": [], "hardware_dependency": "",
                                   "open_questions": [], "ownership_reason": "软件时钟逻辑"}]}

            result = run_requirements_analysis(
                tmp_path, route="openai_compatible", chat=fake_chat, template_path=template)

            self.assertEqual(result["enriched"], 1)
            self.assertIn("公司标准做法参考", captured[0])          # 注入进了真实 prompt
            self.assertIn("RTC_CAL", captured[0])                    # 说明列（宏名）随之注入
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            self.assertIn("公司通用做法：", payload["items"][0]["developer_guidance"][0])

        req = {"source_quote": "q", "description": "d", "requirement": "r", "module": "时钟"}
        self.assertNotEqual(_enrich_key(req, "m"), _enrich_key(req, "m", "参考行内容"))  # 模板行变→缓存失效

    def test_without_template_no_reference_block(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed(tmp_path)
            captured: list[str] = []

            def fake_chat(system: str, user: str) -> dict:
                captured.append(user)
                return {"items": [{"source_requirement_ids": ["AI-1"],
                                   "software_requirement_text": "维持时钟精度。",
                                   "developer_guidance": [], "acceptance_criteria": [],
                                   "hardware_dependency": "", "open_questions": [],
                                   "ownership_reason": "软件"}]}

            run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat,
                                      template_path=None)
            self.assertNotIn("公司标准做法参考", captured[0])       # 不设模板 → 行为如旧


class LlmEnrichmentTests(unittest.TestCase):
    """LLM 分析层（注入 fake chat，零网络）：填叙述字段、冻结结构字段、编造即拒绝、幂等缓存。"""

    def _seed(self, tmp_path: Path) -> None:
        write_jsonl(tmp_path / "ai_requirements.jsonl", [
            {
                "ai_req_id": "AI-1",
                "description": "The meter shall log power-down events to OBIS 0-0:96.1.0.",
                "source_quote": "log power-down events at 0-0:96.1.0, keep 100 entries",
                "source_block_ids": ["B-1"],
                "module": "事件记录",
            }
        ])

    def test_enrichment_fills_narrative_and_freezes_structure(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed(tmp_path)
            calls: list[str] = []

            def fake_chat(system: str, user: str) -> dict:
                calls.append(user)
                return {"items": [{
                    "source_requirement_ids": ["AI-1"],
                    "ownership": "hardware",  # 越权字段：必须被忽略，归属仍走规则
                    "software_requirement_text": "监听掉电中断，将事件写入 0-0:96.1.0 日志，保留最近 100 条。",
                    "developer_guidance": ["订阅掉电中断", "环形缓冲 100 条"],
                    "acceptance_criteria": ["掉电后事件出现在 0-0:96.1.0，且不超过 100 条"],
                    "hardware_dependency": "",
                    "open_questions": [],
                    "ownership_reason": "纯软件事件记录逻辑",
                }]}

            result = run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat)

            assert result["route"] == "openai_compatible"   # 注入 chat 即真实执行
            assert result["enriched"] == 1
            assert result["enrich_degraded"] == 0
            assert len(calls) == 1
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            item = payload["items"][0]
            assert item["software_requirement_text"].startswith("监听掉电中断")
            assert item["developer_guidance"] == ["订阅掉电中断", "环形缓冲 100 条"]
            assert item["acceptance_criteria"]
            assert item["analysis_source"] == "llm"
            # 结构/归属字段冻结：LLM 给的 ownership=hardware 不得生效（规则判 software）
            assert item["ownership"] == "software"
            assert item["ownership_source"] == "rule"
            assert item["source_requirement_ids"] == ["AI-1"]

    def test_fabricated_code_rejects_enrichment_and_degrades(self) -> None:
        """LLM 编造原文没有的 OBIS（换位）→ 整条富化拒绝、item 保持确定性空值、记 issue。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed(tmp_path)

            def fake_chat(system: str, user: str) -> dict:
                return {"items": [{
                    "source_requirement_ids": ["AI-1"],
                    # 源文是 0-0:96.1.0；这里换成 0-0:96.1.7（错一位）——必须被拦
                    "software_requirement_text": "将事件写入 0-0:96.1.7 日志。",
                    "developer_guidance": ["写 0-0:96.1.7"],
                }]}

            result = run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat)

            assert result["enriched"] == 0
            assert result["enrich_degraded"] == 1
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            item = payload["items"][0]
            assert item["software_requirement_text"] == ""      # 未被污染
            assert item["developer_guidance"] == []
            assert item["analysis_source"] == "deterministic"
            assert any("编造结构编码" in msg
                       for row in payload["issues"] for msg in row["issues"])

    def test_fabricated_plain_integer_is_soft_not_rejected(self) -> None:
        """散文里的序号/步骤数（如"1. 2. 3."）非结构编码 → 软标记、保留富化（与 ai_extract 同纪律）。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed(tmp_path)

            def fake_chat(system: str, user: str) -> dict:
                return {"items": [{
                    "source_requirement_ids": ["AI-1"],
                    # 用步骤编号 1/2/3（源文没有），但不碰 OBIS 编码——应保留、只软标记
                    "software_requirement_text": "步骤：1. 监听掉电；2. 写入 0-0:96.1.0；3. 保留 100 条。",
                    "developer_guidance": ["1. 订阅中断", "2. 环形缓冲"],
                }]}

            result = run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat)

            assert result["enriched"] == 1        # 保留富化，不因整数降级
            assert result["enrich_degraded"] == 0
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            item = payload["items"][0]
            assert item["software_requirement_text"].startswith("步骤")
            assert item["analysis_source"] == "llm"
            # 数字漂移作为软提示被记录，供审查对照 source_quote
            assert any("软提示" in msg for row in payload["issues"] for msg in row["issues"])

    def test_enrichment_cache_is_idempotent_across_runs(self) -> None:
        """内容不变时二次运行不再调 LLM（命中 analyze_enrich_cache.json）。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed(tmp_path)
            count = {"n": 0}

            def fake_chat(system: str, user: str) -> dict:
                count["n"] += 1
                return {"items": [{"source_requirement_ids": ["AI-1"],
                                   "software_requirement_text": "缓存测试正文。"}]}

            run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat)
            run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat)

            assert count["n"] == 1                              # 第二次命中缓存，零新增调用
            assert (tmp_path / "analyze_enrich_cache.json").exists()

    def test_hardware_items_skip_enrichment_calls(self) -> None:
        """硬件项只需简要说明（不产 software_requirement_text）——不该为它烧真实 LLM 调用。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {"ai_req_id": "AI-HW", "description": "计量芯片型号为 Att7022e。",
                 "source_quote": "计量芯片型号为 Att7022e", "source_block_ids": ["B-1"], "module": "计量"},
            ])
            calls = {"n": 0}

            def fake_chat(system: str, user: str) -> dict:
                calls["n"] += 1
                return {"items": [{"software_requirement_text": "不该出现"}]}

            result = run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat)

            assert calls["n"] == 0                     # 硬件项零调用
            assert result["enriched"] == 0
            assert result["enrich_degraded"] == 0      # 跳过≠降级
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            assert payload["items"][0]["ownership"] == "hardware"
            assert payload["items"][0]["software_requirement_text"] == ""

    def test_concurrent_enrichment_is_correct_and_reports_progress(self) -> None:
        """并发富化（288 条规模的关键）：多条并发跑、每条落对自己的 item、逐条进度上报。"""
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            n = 12
            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {"ai_req_id": f"AI-{i}", "description": f"The meter shall do task {i}.",
                 "source_quote": f"shall do task {i}", "source_block_ids": [f"B-{i}"],
                 "module": "事件记录"} for i in range(n)
            ])

            def fake_chat(system: str, user: str) -> dict:
                # 从 prompt 里回读该条的任务号，验证并发下富化不串条
                import re as _re
                m = _re.search(r"shall do task (\d+)", user)
                idx = m.group(1) if m else "?"
                return {"items": [{"software_requirement_text": f"实现任务 {idx} 的处理逻辑。"}]}

            events: list[dict] = []
            result = run_requirements_analysis(
                tmp_path, route="openai_compatible", chat=fake_chat,
                concurrency=4, progress_callback=events.append)

            assert result["enriched"] == n
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            for item in payload["items"]:
                idx = item["source_requirement_ids"][0].split("-")[1]
                assert item["software_requirement_text"] == f"实现任务 {idx} 的处理逻辑。"  # 并发不串条
            # 进度：初始 0 + 每条一次，末次 100%
            assert len(events) == n + 1
            assert events[-1]["stage"] == "analyze"
            assert events[-1]["completed"] == n
            assert events[-1]["percent"] == 100

    def test_enrich_cache_flushes_incrementally(self) -> None:
        """增量缓存：跑够 10 条就落盘一次——中途被杀不丢已完成的真实调用（288 条跑挂的教训）。"""
        from unittest.mock import patch
        import requirements_analysis as ra

        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            n = 12
            write_jsonl(tmp_path / "ai_requirements.jsonl", [
                {"ai_req_id": f"AI-{i}", "description": f"The meter shall do task {i}.",
                 "source_quote": f"shall do task {i}", "source_block_ids": [f"B-{i}"],
                 "module": "事件记录"} for i in range(n)
            ])

            def fake_chat(system: str, user: str) -> dict:
                return {"items": [{"software_requirement_text": "正文。"}]}

            with patch.object(ra, "_save_enrich_cache", wraps=ra._save_enrich_cache) as save:
                run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat, concurrency=2)

            # 12 条 → 第 10 条时增量一次 + 收尾一次 ≥ 2 次落盘
            assert save.call_count >= 2

    def test_malformed_llm_response_degrades_non_fatally(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            tmp_path = Path(td)
            self._seed(tmp_path)

            def fake_chat(system: str, user: str) -> dict:
                return {"garbage": True}       # 无 items、非法形状

            result = run_requirements_analysis(tmp_path, route="openai_compatible", chat=fake_chat)

            assert result["analysis_count"] == 1               # 不整跑死
            assert result["enrich_degraded"] == 1
            payload = json.loads((tmp_path / "engineering_analysis.json").read_text(encoding="utf-8"))
            assert payload["items"][0]["analysis_source"] == "deterministic"


if __name__ == "__main__":
    unittest.main()
