"""批次三（0714 整体 review 落地,续批次一/二）回归：

- E6 裁决回灌抽取：样本库 few-shot 注入抽取 prompt;软背景不进指纹;护栏不放宽。
（E7 prompt v16 的锁在 test_ai_extract.PromptV5Tests;S7 余项/内容锁在各自实现后追加。）
"""
from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import ai_extract


class ExtractExemplarRenderTests(unittest.TestCase):
    def _bank(self) -> dict:
        accepted = {}
        for i in range(6):
            accepted[f"AIR-{i:03d}"] = {"module": "时钟" if i % 2 else "预付费",
                                        "title": f"范例标题{i}"}
        return {"accepted": accepted}

    def test_module_diversity_and_cap(self) -> None:
        text = ai_extract.render_extract_exemplars(self._bank())
        lines = text.splitlines()
        self.assertLessEqual(len(lines), ai_extract.EXTRACT_EXEMPLARS_MAX)
        self.assertEqual(sum(1 for l in lines if l.startswith("-【时钟】")), 2)   # 每模块 ≤2
        self.assertEqual(sum(1 for l in lines if l.startswith("-【预付费】")), 2)

    def test_deterministic_order(self) -> None:
        a = ai_extract.render_extract_exemplars(self._bank())
        b = ai_extract.render_extract_exemplars(self._bank())
        self.assertEqual(a, b)

    def test_empty_bank_renders_nothing(self) -> None:
        self.assertEqual(ai_extract.render_extract_exemplars({}), "")
        self.assertEqual(ai_extract.render_extract_exemplars({"accepted": {}}), "")


class ExtractExemplarInjectionTests(unittest.TestCase):
    SECTION = {"section_id": "S1", "heading": "4.6 XDEV2",
               "text": "The XDEV shall close the valve.", "block_ids": []}

    def _chat_capture(self, captured: list) -> object:
        def chat(system: str, user: str) -> dict:
            captured.append(user)
            return {"requirements": []}
        return chat

    def test_exemplars_injected_with_no_copy_instruction(self) -> None:
        captured: list[str] = []
        ai_extract.extract_section(self.SECTION, self._chat_capture(captured),
                                   doc_context="【文档背景】燃气表",
                                   exemplars="-【预付费】购电余额扣减")
        user = captured[0]
        self.assertIn("专家已验收范例", user)
        self.assertIn("-【预付费】购电余额扣减", user)
        self.assertIn("不得搬运", user)
        self.assertLess(user.index("【文档背景】"), user.index("专家已验收范例"))
        self.assertLess(user.index("专家已验收范例"), user.index("当前章节"))

    def test_no_exemplars_prompt_unchanged(self) -> None:
        with_ex: list[str] = []
        without: list[str] = []
        ai_extract.extract_section(self.SECTION, self._chat_capture(without),
                                   doc_context="【文档背景】燃气表")
        ai_extract.extract_section(self.SECTION, self._chat_capture(with_ex),
                                   doc_context="【文档背景】燃气表", exemplars="")
        self.assertEqual(without[0], with_ex[0])      # 空注入=旧 prompt 逐字节一致

    def test_exemplars_do_not_touch_section_fingerprint(self) -> None:
        # 软背景不进指纹（S3 同理）：样本库更新不报废抽取缓存
        fp = ai_extract.section_fingerprint(self.SECTION, "m", "ctx")
        self.assertEqual(fp, ai_extract.section_fingerprint(self.SECTION, "m", "ctx"))
        import inspect
        src = inspect.getsource(ai_extract.section_fingerprint)
        self.assertNotIn("exemplar", src)

    def test_exemplar_code_borrowing_still_hard_blocked(self) -> None:
        """范例里的 OBIS 被搬进抽取产物 → 漂移基线（章节原文）照常拦截。"""
        captured: list[str] = []

        def chat(system: str, user: str) -> dict:
            captured.append(user)
            return {"requirements": [{
                "title": "阀门关闭", "description": "写入对象 0-0:96.3.10.255 关闭阀门。",
                "source_quote": "The XDEV shall close the valve.",
                "type": "functional", "priority": "P1", "module": "阀门控制", "labels": ["阀门"],
            }]}

        results = ai_extract.extract_section(
            self.SECTION, chat,
            exemplars="-【阀门控制】阀门对象 0-0:96.3.10.255 关闭控制")
        self.assertEqual(len(results), 1)
        notes = str(results[0].get("notes") or "")
        self.assertIn("结构漂移已拦截", notes)          # 编码硬拦,不因范例注入而放行


class ApiPayloadMemoTests(unittest.TestCase):
    """S7b：GUI 装配路径按源文件签名 memo——刷新不再全量重读重 join;写入自然失效。"""

    def setUp(self) -> None:
        import api_server
        api_server._reset_payload_memo()

    tearDown = setUp

    def _seed(self, out: Path) -> None:
        (out / "blocks.jsonl").write_text(json.dumps({
            "block_id": "B1", "order": 1, "text": "The meter shall log events.",
            "requirement_like": True, "noise": False, "type": "paragraph",
            "doc_region": "body", "section_path": ["4"]}, ensure_ascii=False) + "\n",
            encoding="utf-8")
        (out / "ai_requirements.jsonl").write_text(json.dumps({
            "ai_req_id": "AIR-1", "title": "事件记录", "module": "事件记录",
            "source_quote": "The meter shall log events.", "source_block_ids": ["B1"]},
            ensure_ascii=False) + "\n", encoding="utf-8")

    def test_repeat_calls_hit_memo(self) -> None:
        from unittest.mock import patch

        import api_server
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out)
            with patch.object(api_server, "_build_ai_requirements_impl",
                              wraps=api_server._build_ai_requirements_impl) as impl:
                first = api_server.build_ai_requirements(out)
                second = api_server.build_ai_requirements(out)
            self.assertEqual(impl.call_count, 1)          # 第二次命中 memo
            self.assertEqual(first, second)

    def test_source_write_invalidates(self) -> None:
        from unittest.mock import patch

        import api_server
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out)
            with patch.object(api_server, "_build_ai_requirements_impl",
                              wraps=api_server._build_ai_requirements_impl) as impl:
                api_server.build_ai_requirements(out)
                with (out / "ai_review_states.jsonl").open("a", encoding="utf-8") as f:
                    f.write(json.dumps({"ai_req_id": "AIR-1", "status": "accepted"}) + "\n")
                rows = api_server.build_ai_requirements(out)
            self.assertEqual(impl.call_count, 2)          # 裁决写入 → 签名变化 → 重算
            self.assertEqual(rows[0]["status"], "accepted")

    def test_returned_payload_mutation_isolated(self) -> None:
        import api_server
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            self._seed(out)
            first = api_server.build_document_blocks(out)
            first["blocks"][0]["text"] = "被消费方原地改了"
            second = api_server.build_document_blocks(out)
            self.assertEqual(second["blocks"][0]["text"], "The meter shall log events.")


class PdfBlockZoneTests(unittest.TestCase):
    """影印全段落热区（0714「点一段出翻译和解析」）：kind 路由与重排块点击语义同源。"""

    GEOMETRY = {bid: [{"page_number": 1, "bbox": [10, 100 + i * 30, 500, 118 + i * 30],
                       "page_width": 595, "page_height": 842}]
                for i, bid in enumerate(["B1", "B2", "B3", "B4", "B5", "B6"])}

    BLOCKS = [
        {"block_id": "B1", "type": "paragraph", "text": "The meter shall measure volume.",
         "requirement_like": True, "noise": False, "doc_region": "body"},          # 锚点 → req
        {"block_id": "B2", "type": "paragraph", "text": "The XDEV shall stay closed forever.",
         "requirement_like": True, "noise": False, "doc_region": "body"},          # 未覆盖 → omission
        {"block_id": "B3", "type": "paragraph", "text": "Background prose paragraph.",
         "requirement_like": False, "noise": False, "doc_region": "body"},         # → context
        {"block_id": "B4", "type": "heading", "text": "4 General requirements",
         "requirement_like": False, "noise": False, "doc_region": "body"},         # 标题 → 无热区
        {"block_id": "B5", "type": "table", "text": "col | col",
         "requirement_like": False, "noise": False, "doc_region": "body"},         # 表格 → 无热区
        {"block_id": "B6", "type": "paragraph", "text": "Repeated footer noise",
         "requirement_like": False, "noise": True, "doc_region": "body"},          # 噪声 → 无热区
    ]
    REQS = [{"ai_req_id": "AIR-1", "anchor_block_id": "B1", "source_block_ids": ["B1"],
             "source_quote": "The meter shall measure volume."}]

    def _zones(self):
        import doc_annotation_export as dae
        covered = dae._covered_blocks(self.REQS)
        return dae._pdf_block_zones(self.BLOCKS, self.REQS, self.GEOMETRY, covered)

    def test_kind_routing_matches_reflow_semantics(self) -> None:
        zones = {z["block_id"]: z for z in self._zones()}
        self.assertEqual(zones["B1"]["kind"], "req")
        self.assertEqual(zones["B1"]["req_id"], "AIR-1")
        self.assertEqual(zones["B2"]["kind"], "omission")
        self.assertEqual(zones["B3"]["kind"], "context")
        for excluded in ("B4", "B5", "B6"):                  # 标题/表格/噪声不给热区
            self.assertNotIn(excluded, zones)
        for zone in zones.values():                          # 百分比矩形齐全
            self.assertEqual(zone["page"], 1)
            self.assertEqual(sorted(zone["rect"]), ["height", "left", "top", "width"])

    def test_production_coverage_does_not_trust_section_fallback_span(self) -> None:
        import doc_annotation_export as dae
        reqs = [{
            "ai_req_id": "AIR-1",
            "source_quote": "The meter shall measure volume.",
            "source_block_ids": ["B1", "B2"],
            "source_mapping": "section_fallback",
        }]

        covered = dae._covered_blocks(reqs, self.BLOCKS)

        self.assertEqual(covered, {"B1"})

    def test_echo_zone_preserves_all_requirement_links(self) -> None:
        import doc_annotation_export as dae
        reqs = [
            {"ai_req_id": "AIR-1", "anchor_block_id": "B1", "source_block_ids": ["B1"],
             "echo_block_ids": ["B3"]},
            {"ai_req_id": "AIR-2", "anchor_block_id": "B5", "source_block_ids": ["B5"],
             "echo_block_ids": ["B3"]},
        ]
        zones = dae._pdf_block_zones(self.BLOCKS, reqs, self.GEOMETRY,
                                     dae._covered_blocks(reqs))
        echo = next(zone for zone in zones if zone["block_id"] == "B3")

        self.assertEqual(echo["kind"], "echo")
        self.assertEqual(echo["req_ids"], ["AIR-1", "AIR-2"])

        records = dae._pdf_context_records(self.BLOCKS, zones)
        self.assertEqual(records["B3"]["echo_req_ids"], ["AIR-1", "AIR-2"])

        html_out = dae._render_pdf_page_stack(
            [{"page_number": 1, "href": "document_pages/page-0001.png",
              "width": 595, "height": 842}],
            reqs, [], {"AIR-1": 1, "AIR-2": 2}, self.GEOMETRY, block_zones=zones)
        self.assertIn('class="pdf-block-zone zone-echo"', html_out)
        self.assertIn('data-echo-reqs="AIR-1 AIR-2"', html_out)
        self.assertIn("重复·见01/02", html_out)

    def test_payload_carries_block_zones(self) -> None:
        import doc_annotation_export as dae
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            (out / "document_source.pdf").write_bytes(b"%PDF-1.4 fake")
            (out / "manifest.json").write_text(
                json.dumps({"input": "document_source.pdf"}), encoding="utf-8")
            pages_dir = out / dae.ANNOTATION_PAGES_DIR
            pages_dir.mkdir()
            (pages_dir / "page-0001.png").write_bytes(b"\x89PNG fake")
            (pages_dir / dae.ANNOTATION_PAGES_MANIFEST).write_text(json.dumps({
                "version": 1, "source_sha256": dae._file_sha256(out / "document_source.pdf"),
                "dpi": dae.PDF_PAGE_RENDER_DPI,
                "pages": [{"page_number": 1, "file": "page-0001.png",
                                          "width": 595, "height": 842}]}), encoding="utf-8")
            (out / dae.ANNOTATION_PDF_GEOMETRY).write_text(json.dumps({
                "version": 2, "blocks": {}}), encoding="utf-8")
            _write_blocks = [dict(b) for b in self.BLOCKS]
            (out / "blocks.jsonl").write_text(
                "\n".join(json.dumps(b, ensure_ascii=False) for b in _write_blocks) + "\n",
                encoding="utf-8")
            (out / "ai_requirements.jsonl").write_text(
                json.dumps(dict(self.REQS[0], title="计量", module="计量"), ensure_ascii=False) + "\n",
                encoding="utf-8")
            from unittest.mock import patch
            with patch.object(dae, "_resolve_pdf_geometry", return_value=self.GEOMETRY):
                payload = dae.build_pdf_annotation_payload(out)
        self.assertTrue(payload["available"])
        kinds = {z["block_id"]: z["kind"] for z in payload["block_zones"]}
        self.assertEqual(kinds.get("B1"), "req")
        self.assertEqual(kinds.get("B3"), "context")

    def test_static_pdf_stack_renders_zones_and_context_data(self) -> None:
        import doc_annotation_export as dae
        zones = self._zones()
        html_out = dae._render_pdf_page_stack(
            [{"page_number": 1, "href": "document_pages/page-0001.png", "width": 595, "height": 842}],
            self.REQS, [], {"AIR-1": 1}, self.GEOMETRY, block_zones=zones)
        self.assertIn('class="pdf-block-zone zone-req"', html_out)
        self.assertIn('class="pdf-block-zone zone-context"', html_out)
        self.assertIn('data-block-id="B3"', html_out)
        self.assertIn('data-req="AIR-1"', html_out)
        self.assertIn("查看该段翻译与解析", html_out)
        context = dae._pdf_context_records(self.BLOCKS, zones)
        self.assertIn("B3", context)                          # 背景卡数据随包
        self.assertEqual(context["B3"]["text"], "Background prose paragraph.")
        self.assertNotIn("B1", context)                       # req/omission 不进背景卡数据

    def test_static_template_wires_zone_click_and_context_card(self) -> None:
        import doc_annotation_export as dae
        self.assertIn("selectPdfContextRecord", dae._TEMPLATE)
        self.assertIn("selectPdfEchoRecord", dae._TEMPLATE)
        self.assertIn("PDF_CONTEXT", dae._TEMPLATE)
        self.assertIn('closest(".pdf-block-zone")', dae._TEMPLATE)
        self.assertIn(".pdf-block-zone:hover", dae._TEMPLATE)


class AnchorFallbackTests(unittest.TestCase):
    """锚定回退路径回归（review 测试缺口）：此前只有 exact 路径有测,前缀兜底/首块回退
    是最易错锚的路径却零覆盖。"""

    BLOCKS = {"B1": "Intro paragraph about scope.",
              "B2": "The XDEV shall close the valve within 5 s after detection.",
              "B3": "Following exposure the meter shall be inspected."}

    def _anchor(self, quote: str, span: list[str]) -> str:
        import api_server
        return api_server.anchor_block_id(
            {"source_quote": quote, "source_block_ids": span}, dict(self.BLOCKS))

    def test_exact_quote_lands_on_containing_block(self) -> None:
        self.assertEqual(
            self._anchor("The XDEV shall close the valve within 5 s after detection.",
                         ["B1", "B2", "B3"]),
            "B2")

    def test_whitespace_and_case_normalized(self) -> None:
        self.assertEqual(
            self._anchor("the xdev  SHALL close the valve within 5 s after detection.",
                         ["B1", "B2"]),
            "B2")

    def test_tail_deviation_falls_back_to_prefix(self) -> None:
        # LLM 引用尾部偏差（改写了句尾）→ 前 40 字前缀仍应锚对块
        self.assertEqual(
            self._anchor("The XDEV shall close the valve within 5 seconds of any tamper event",
                         ["B1", "B2", "B3"]),
            "B2")

    def test_unmatchable_quote_falls_back_to_first_span_block(self) -> None:
        self.assertEqual(self._anchor("完全对不上的引用文本", ["B3", "B2"]), "B3")

    def test_empty_span_returns_empty(self) -> None:
        self.assertEqual(self._anchor("anything", []), "")


class SoftwareXlsxContentLockTests(unittest.TestCase):
    """B 轨交付物内容锁（review 测试缺口）：富化字段必须真的落 software_requirements.xlsx
    的说明列——此前只锁版式/注入,\"软标随行\"等 B1 不变量在 xlsx 层无锁。"""

    ITEM = {
        "analysis_id": "SRA-001", "module": "阀门控制", "submodule": "阀门控制",
        "template_match": "matched", "ownership": "software",
        "ownership_reason": "阀门控制逻辑由软件实现", "ownership_reason_source": "llm",
        "description": "阀门关闭控制",
        "software_requirement_text": "检测到异常后 5 s 内驱动阀门关闭。",
        "developer_guidance": ["实现阀门驱动接口与超时监控"],
        "design_options": ["队列化关闭指令以支持重试"],
        "acceptance_criteria": ["注入异常后 5 s 内阀门状态变为关闭"],
        "assumptions": ["假定阀门驱动电路由硬件保证时序"],
        "open_questions": ["关闭失败的重试次数上限是多少？"],
        "enrichment_warnings": ["fabricated number in guidance: 99"],
        "source_quote": "The XDEV shall close the valve.",
        "source_requirement_ids": ["AIR-1"],
    }

    def test_enriched_fields_land_in_notes_column(self) -> None:
        from openpyxl import load_workbook

        from requirements_analysis_excel import write_software_requirements_xlsx
        with tempfile.TemporaryDirectory() as td:
            path = write_software_requirements_xlsx([dict(self.ITEM)], Path(td) / "s.xlsx")
            wb = load_workbook(path, read_only=True)
            try:
                text = "\n".join(
                    str(cell) for ws in wb.worksheets
                    for row in ws.iter_rows(values_only=True) for cell in row if cell)
            finally:
                wb.close()
        self.assertIn("检测到异常后 5 s 内驱动阀门关闭。", text)       # 富化正文
        self.assertIn("实现阀门驱动接口与超时监控", text)              # developer_guidance
        self.assertIn("设计候选（非规范约束）：队列化关闭指令以支持重试", text)
        self.assertIn("验收建议：注入异常后 5 s 内阀门状态变为关闭", text)
        self.assertIn("假设：假定阀门驱动电路由硬件保证时序", text)
        self.assertIn("待确认：关闭失败的重试次数上限是多少？", text)   # open_questions
        self.assertIn("⚠ 富化待核：fabricated number in guidance: 99", text)   # 软标随行(B1)
        self.assertIn("归属判定：软件（依据：阀门控制逻辑由软件实现，LLM 判定）", text)
        self.assertIn("原文：The XDEV shall close the valve.", text)   # 溯源随行


if __name__ == "__main__":
    unittest.main()
