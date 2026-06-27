"""终点交付物端到端回归（Phase 3）。

锁定从 atomizer 输出 → 装配实现规格 → Excel 的关键不变量，防止 Phase 1+2 的
数据完整性修复（H1 OBIS 拆码护栏、H4 访问权限不错列、公式注入守卫）被未来改动
静默回退。用仓库内合成 fixture（含三类触发场景），fresh-clone 可跑、确定性。
"""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from assemble_spec import assemble
from cosem_object_model import access_cells, attr_access_unresolved, build_object_model
from spec_excel import write_xlsx

try:
    from openpyxl import load_workbook
    HAVE_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAVE_OPENPYXL = False


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in rows) + "\n", encoding="utf-8")


def build_deliverable_fixture(out: Path) -> None:
    """合成一份 atomizer 输出，刻意覆盖三类回归触发场景：
    - H1: 不完整的 2 位值组 OBIS `0-0:96.1.0`（历史会被拆成 9.6.1.0）；空格 `1-0:0 4.5.255`
    - H4: 非 4 段访问串 `-/-/-A--`（历史会整串塞进 RC 列）
    - 公式注入: 属性类型以 `=` 开头（历史会变成 Excel 活公式）
    """
    write_jsonl(out / "atomic_requirements.jsonl", [
        {"stable_req_id": "O-CLOCK", "requirement_type": "cosem_object_instance",
         "object": "Clock", "domain": "time", "source_refs": ["TBL-1-R1"], "confidence": 0.9},
        {"stable_req_id": "O-ENERGY", "requirement_type": "cosem_object_instance",
         "object": "Active energy import total", "domain": "metering", "source_refs": ["TBL-1-R2"], "confidence": 0.9},
        # H1: 缺一段、含 2 位值组的不完整 OBIS — 必须原样保留，不得拆成 0-0:9.6.1.0
        {"stable_req_id": "O-MERGED", "requirement_type": "cosem_object_instance",
         "object": "Device ID 1", "domain": "abstract", "source_refs": ["TBL-1-R3"], "confidence": 0.8},
        # H1: 空格仍应还原为分隔点
        {"stable_req_id": "O-SPACED", "requirement_type": "cosem_object_instance",
         "object": "Transformer ratio", "domain": "metering", "source_refs": ["TBL-1-R4"], "confidence": 0.8},
        # 父对象供 add_user 挂载
        {"stable_req_id": "O-ASSOC", "requirement_type": "cosem_object_instance",
         "object": "Association LN", "domain": "association", "source_refs": ["TBL-1-R5"], "confidence": 0.9},
        {"stable_req_id": "A-CLOCK-TIME", "requirement_type": "cosem_attribute_access",
         "object": "Clock.time", "verification_method": "configuration_check",
         "source_refs": ["TBL-2-R1"], "confidence": 0.9, "ambiguity": False},
        # H4: 非 4 段访问串
        {"stable_req_id": "A-ADDUSER", "requirement_type": "cosem_attribute_access",
         "object": "Association LN.add_user", "source_refs": ["TBL-2-R2"], "confidence": 0.7},
        # 公式注入: 属性类型以 = 开头
        {"stable_req_id": "A-INJECT", "requirement_type": "cosem_attribute_access",
         "object": "Clock.danger", "source_refs": ["TBL-2-R3"], "confidence": 0.6},
    ])
    write_jsonl(out / "table_items.jsonl", [
        {"item_id": "TBL-1-R1", "table_id": "TBL-1", "row_index": 1,
         "fields": {"Object/attribute name": "Clock", "CL": "8", "Value": "0-0:1.0.0.255"}},
        {"item_id": "TBL-1-R2", "table_id": "TBL-1", "row_index": 2,
         "fields": {"Object/attribute name": "Active energy import total", "CL": "3", "Value": "1-0:1.8.0.255"}},
        {"item_id": "TBL-1-R3", "table_id": "TBL-1", "row_index": 3,
         "fields": {"Object/attribute name": "Device ID 1", "CL": "1", "Value": "0-0:96.1.0"}},
        {"item_id": "TBL-1-R4", "table_id": "TBL-1", "row_index": 4,
         "fields": {"Object/attribute name": "Transformer ratio", "CL": "1", "Value": "1-0:0 4.5.255"}},
        {"item_id": "TBL-1-R5", "table_id": "TBL-1", "row_index": 5,
         "fields": {"Object/attribute name": "Association LN", "CL": "15", "Value": "0-0:40.0.0.255"}},
        {"item_id": "TBL-2-R1", "table_id": "TBL-2", "row_index": 1,
         "fields": {"#": "2", "Object/attribute name": "time", "Type": "octet-string",
                    "Value": "00", "Access rights RC/PC/SC/LC": "R-/RW/--/R-"}},
        {"item_id": "TBL-2-R2", "table_id": "TBL-2", "row_index": 2,
         "fields": {"#": "1", "Object/attribute name": "add_user", "Type": "date",
                    "Access rights RC/PC/SC/LC": "-/-/-A--"}},
        {"item_id": "TBL-2-R3", "table_id": "TBL-2", "row_index": 3,
         "fields": {"#": "3", "Object/attribute name": "danger", "Type": "=cmd|' /C calc'!A0",
                    "Value": "0", "Access rights RC/PC/SC/LC": "R-/R-/R-/R-"}},
    ])
    write_jsonl(out / "review_states.jsonl", [])
    (out / "manifest.json").write_text(json.dumps({"source_files": ["synthetic.docx"]}), encoding="utf-8")


class SpecDeliverableRegressionTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.out = Path(self._tmp.name)
        build_deliverable_fixture(self.out)
        self.model = build_object_model(self.out)
        self.by_name = {o["object"]: o for o in self.model["objects"]}

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_obis_values_are_stable_no_silent_split(self) -> None:
        # H1 回归：2 位值组不被拆，空格还原为点
        self.assertEqual(self.by_name["Device ID 1"]["obis"], "0-0:96.1.0")
        self.assertEqual(self.by_name["Transformer ratio"]["obis"], "1-0:0.4.5.255")
        self.assertEqual(self.by_name["Clock"]["obis"], "0-0:1.0.0.255")
        # 整个交付物里不得出现被拆出来的 9.6 痕迹
        all_obis = [o["obis"] for o in self.model["objects"]]
        self.assertFalse([o for o in all_obis if ":9.6." in o])

    def test_access_matrix_never_leaks_raw_into_columns(self) -> None:
        # H4 回归：任何属性的访问四列都不得携带原始串；非 4 段串进未解析附录
        all_attrs = [a for o in self.model["objects"] for a in o["attributes"]] + self.model["orphan_attributes"]
        for attr in all_attrs:
            for cell in access_cells(attr):
                self.assertNotIn("/", cell)
                self.assertLessEqual(len(cell), 3)
        unresolved = {u["attribute"] for u in self.model["unresolved_access"]}
        self.assertIn("add_user", unresolved)
        self.assertEqual(self.model["counts"]["unresolved_access"], 1)
        # 正常 4 段属性照常解析
        time_attr = next(a for a in self.by_name["Clock"]["attributes"] if a["name"] == "time")
        self.assertEqual(access_cells(time_attr), ("R-", "RW", "--", "R-"))
        self.assertFalse(attr_access_unresolved(time_attr))

    @unittest.skipUnless(HAVE_OPENPYXL, "openpyxl required")
    def test_assembled_excel_has_no_live_formula_cells(self) -> None:
        # 公式注入回归：端到端 assemble → Excel，注入的 `=cmd...` 必须被中和为文本
        doc, _ = assemble(self.out, None, source="regression", extracted_at="2026-01-01T00:00:00")
        xlsx = self.out / "dlms_cosem_spec.xlsx"
        write_xlsx(doc, xlsx)
        wb = load_workbook(xlsx)
        live = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value[:1] in ("=", "+", "@") and "cmd" in c.value:
                        live.append((ws.title, c.coordinate, c.value))
        self.assertEqual(live, [], f"发现未中和的活公式单元格: {live}")

    def test_deliverable_doc_counts_are_stable(self) -> None:
        # 装配产出的需求文档结构稳定（防 P1 装配回退）
        doc, breakdown = assemble(self.out, None, source="regression", extracted_at="2026-01-01T00:00:00")
        # 5 个对象实例 → 至少 5 条 P1 对象需求
        self.assertEqual(breakdown["p1_object_requirements"], 5)
        self.assertGreaterEqual(breakdown["total"], 5)
        self.assertTrue(doc["requirements"])
        # 每条需求都有稳定 id 与域标签
        for req in doc["requirements"]:
            self.assertTrue(req.get("id"))
            self.assertTrue(req.get("labels"))


if __name__ == "__main__":
    unittest.main()
