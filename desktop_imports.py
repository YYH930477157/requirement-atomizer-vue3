"""desktop_tasks 导入/回灌任务族（M9 第 6 刀，2026-08-17）。

从 ``desktop_tasks.py`` 逐字搬运的线下导入家族：澄清清单工作簿回灌、verification
六列工作簿回灌与手工写入、HTML 裁决 JSON 回灌。``desktop_tasks`` 原名重导出，
CLI 分发与测试的 ``desktop_tasks.X`` 调用面零变化。

选族纪律（M9 蓝图红线）：本族不含任何测试 patch 目标
（``out/m9-patch-targets.json`` desktop_tasks 32 个全部留守）；重依赖全部函数内
惰性导入（openpyxl/clarification_report/requirement_schema/api_server/
omission_actions），顶层只依赖 ai_extract 与 requirements_analysis_schema
——不反向依赖 desktop_tasks，无环。
"""
from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

import ai_extract
from requirements_analysis_schema import normalize_ownership

LOGGER = logging.getLogger("requirement_atomizer")


def import_clarification_workbook_task(
    out_dir: Path,
    workbook_path: Path,
    *,
    actor: str = "desktop-import",
) -> dict[str, Any]:
    """Import customer answers and internal acknowledgements from one report workbook."""
    from openpyxl import load_workbook
    from clarification_report import import_answers, import_internal_checks, run_report

    out_dir = out_dir.expanduser().resolve()
    workbook_path = workbook_path.expanduser().resolve()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        has_internal_sheet = "必答-内部核对" in workbook.sheetnames
        internal_headers = (
            [str(cell.value or "").strip() for cell in workbook["必答-内部核对"][1]]
            if has_internal_sheet else []
        )
    finally:
        workbook.close()
    if not has_internal_sheet:
        raise ValueError("工作簿缺少「必答-内部核对」sheet，请重新生成澄清清单")
    required = {"澄清ID", "证据指纹", "阻塞级", "模块", "信号", "来源需求", "核对人", "备注"}
    missing = sorted(required.difference(internal_headers))
    if not any(value.startswith("新处置") for value in internal_headers):
        missing.append("新处置(确认无误/确认有问题/暂缓)")
    if missing:
        raise ValueError(f"「必答-内部核对」sheet 缺少列：{', '.join(missing)}")
    answers = import_answers(out_dir, workbook_path)
    checks = import_internal_checks(out_dir, workbook_path, actor=actor)
    report = run_report(out_dir)
    return {
        "kind": "clarification_answers",
        "out_dir": str(out_dir),
        "imported": int(answers.get("imported") or 0),
        "internal_imported": int(checks.get("imported") or 0),
        "readiness": report.get("readiness") or {},
        "questions": int(report.get("questions") or 0),
        "written": list(dict.fromkeys([
            *(str(value) for value in (answers.get("written") or [])),
            *(str(value) for value in (checks.get("written") or [])),
            *(str(value) for value in (report.get("written") or [])),
        ])),
    }


# ---------------------------------------------------------------------------
# WS4 能力补齐：verification 回写-回灌、手工入口、状态机回退、需求库、依赖推荐
# 全程零 LLM 调用；共享状态文件写走锁 + 原子替换（review_state）。
# ---------------------------------------------------------------------------
_WS4_TRACE_ID_RE = re.compile(r"需求追溯ID[：:]\s*([^\n\r]+)")


def import_verification_workbook_task(
    out_dir: Path,
    workbook_path: Path,
    *,
    actor: str = "desktop-verification",
) -> dict[str, Any]:
    """回灌线下改过的 software_requirements.xlsx 六列 → verification_states.jsonl。

    复用 import-clarification-answers 的解析模式：按需求追溯ID（notes 列）定位行，
    读六列单元格 → verification 子对象，CAS 指纹失配（结构字段漂移）拒绝自动合入转人工。

    T3-2 CAS 分桶：回灌闸只比对**结构列**（子模块 + 客户需求章节）——``description``（叙述）
    不再进闸，叙述措辞变化不再误拒回灌。结构漂移行进 ``rejected``；叙述漂移（结构匹配、
    描述变化）进 ``narrative_review`` 清单——状态仍回灌（不吊销），仅提示专家复核措辞。
    S1-10b：``rejected`` 每条含 requirement_id + xlsx 物理行号 + sheet + 原因。
    """
    from openpyxl import load_workbook
    from requirement_schema import parse_verification_columns, structural_fingerprint_from_cells
    from requirements_analysis_rules import apply_verification_override, load_requirement_index

    root = out_dir.expanduser().resolve()
    index = load_requirement_index(root)
    workbook_path = workbook_path.expanduser().resolve()
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    # (rid, six, structural_cells, description_cell, row_number, sheet_title)
    harvested: list[tuple[str, list[Any], tuple[Any, ...], Any, int, str]] = []
    try:
        for sheet in wb.worksheets:
            header = [str(cell.value or "").strip() for cell in next(
                sheet.iter_rows(min_row=1, max_row=1), [])]
            if "项目负责人确认" not in header:
                continue
            col = {name: idx for idx, name in enumerate(header)}
            # row_number = xlsx 物理行号（min_row=2 → 首条数据行是第 2 行）
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue
                notes_idx = col.get("说明、示例、注意事项", 6)
                notes = str(row[notes_idx] if len(row) > notes_idx else "")
                match = _WS4_TRACE_ID_RE.search(notes)
                if not match:
                    continue
                rid = match.group(1).strip()
                six = [
                    row[col.get(name, idx)] if len(row) > col.get(name, idx) else ""
                    for name, idx in (
                        ("项目负责人确认", 10), ("测试负责人确认", 11), ("研发测试确认", 12),
                        ("功能是否实现", 13), ("测试用例号", 14), ("测试是否完成", 15),
                    )
                ]
                # T3-2：结构闸只用子模块 + 客户需求章节（描述=叙述，降级为复核提示）
                structural_cells = (
                    row[col.get("子模块", 2)] if len(row) > col.get("子模块", 2) else "",
                    row[col.get("客户需求章节", 8)] if len(row) > col.get("客户需求章节", 8) else "",
                )
                description_cell = row[col.get("描述", 3)] if len(row) > col.get("描述", 3) else ""
                harvested.append((rid, six, structural_cells, description_cell, row_number, sheet.title))
    finally:
        wb.close()

    imported = stale = missing = 0
    rejected: list[dict[str, Any]] = []
    narrative_review: list[dict[str, Any]] = []
    for rid, six, structural_cells, description_cell, row_number, sheet_title in harvested:
        entry = index.get(rid)
        if not entry:
            missing += 1
            rejected.append({
                "requirement_id": rid, "row": row_number, "sheet": sheet_title,
                "reason": "需求追溯ID不在当前索引（需求可能已删除或尚未生成）",
            })
            continue
        # CAS 结构闸：子模块 + 客户需求章节 指纹必须与当前需求一致（结构漂移=转人工）
        if structural_fingerprint_from_cells(*structural_cells) != entry["cell_fingerprint"]:
            stale += 1
            rejected.append({
                "requirement_id": rid, "row": row_number, "sheet": sheet_title,
                "reason": "结构字段失配（子模块/客户需求章节已变化，请人工核对后再回灌）",
            })
            continue
        # 叙述复核提示：结构匹配但描述变化 → 不吊销，进 narrative_review 提示专家复核
        item_description = str((entry.get("item") or {}).get("description") or "")
        if str(description_cell or "").strip() and item_description and \
                _backfill_description_drifted(description_cell, item_description):
            narrative_review.append({
                "requirement_id": rid, "row": row_number, "sheet": sheet_title,
                "reason": "描述（叙述）变化——状态已回灌，请复核措辞是否仍准确",
            })
        verification = parse_verification_columns(six, actor_fallback=actor)
        # 仅当六列至少有一项非默认值才写（避免空行覆盖既有状态）
        if verification == default_verification_for_check():
            continue
        apply_verification_override(root, rid, verification, actor=actor,
                                    evidence_fingerprint=entry["fingerprint"])
        imported += 1
    return {
        "kind": "verification_import",
        "out_dir": str(root),
        "imported": imported,
        "stale": stale,
        "missing": missing,
        # S1-10b：拒绝清单精确到 requirement_id + xlsx 行号 + sheet + 原因
        "rejected": rejected,
        # T3-2：叙述复核清单（状态已回灌不吊销，仅提示）
        "narrative_review": narrative_review,
        "written": ["verification_states.jsonl"] if imported else [],
    }


def _backfill_description_drifted(cell_description: Any, item_description: str) -> bool:
    """回灌描述（叙述）漂移判定：折叠空白 + 去控制字符后逐字不等即视为叙述变化。

    仅用于 ``narrative_review`` 复核提示，不参与 CAS 吊销；宽松归一避免无害空白差异误报。
    """
    import re as _re
    norm_cell = _re.sub(r"\s+", " ", str(cell_description or "")).strip()
    norm_item = _re.sub(r"\s+", " ", str(item_description or "")).strip()
    return bool(norm_cell) and bool(norm_item) and norm_cell != norm_item


def default_verification_for_check() -> Any:
    """空 verification（用于回灌跳过全空行）。延迟导入避免顶层依赖。"""
    from requirement_schema import default_verification
    return default_verification()


def set_verification_task(
    out_dir: Path,
    requirement_id: str,
    *,
    implemented: str | None = None,
    test_completed: bool | None = None,
    test_case_ids: str | None = None,
    confirm_pm: bool | None = None,
    confirm_tl: bool | None = None,
    confirm_dt: bool | None = None,
    actor: str = "desktop-verification",
) -> dict[str, Any]:
    """直接写入一条 verification 覆盖（CLI 数据入口；六列字段分散为 flag）。"""
    from requirement_schema import IMPLEMENTED_VALUES, default_verification, normalize_verification
    from requirements_analysis_rules import apply_verification_override

    root = out_dir.expanduser().resolve()
    patch = default_verification()
    if implemented is not None:
        if implemented not in IMPLEMENTED_VALUES:
            raise ValueError(f"非法 implemented 值：{implemented}（可选 {IMPLEMENTED_VALUES}）")
        patch["implemented"] = implemented
    if test_completed is not None:
        patch["test_completed"] = bool(test_completed)
    if test_case_ids is not None:
        patch["test_case_ids"] = [item.strip() for item in re.split(r"[;\n,、 ]+", test_case_ids) if item.strip()]
    if confirm_pm is not None:
        patch["project_manager_confirm"] = {"confirmed": bool(confirm_pm), "by": actor, "at": ""}
    if confirm_tl is not None:
        patch["test_lead_confirm"] = {"confirmed": bool(confirm_tl), "by": actor, "at": ""}
    if confirm_dt is not None:
        patch["dev_test_confirm"] = {"confirmed": bool(confirm_dt), "by": actor, "at": ""}
    record = apply_verification_override(root, requirement_id, normalize_verification(patch), actor=actor)
    return {
        "kind": "verification_set",
        "out_dir": str(root),
        "requirement_id": requirement_id,
        "verification": record.get("verification"),
        "lifecycle_state": record.get("lifecycle_state"),
        "written": ["verification_states.jsonl"],
    }


def import_ai_decisions_task(out_dir: Path, decisions_file: Path) -> dict[str, Any]:
    """把 HTML 导出的裁决 JSON 回灌到 ai_review_states.jsonl（合进交付物）。"""
    import ai_review_actions
    out_dir = out_dir.expanduser().resolve()
    data = json.loads(Path(decisions_file).expanduser().read_text(encoding="utf-8"))
    decisions = data.get("decisions") if isinstance(data, dict) else data
    applied = 0
    skipped = 0
    needs_reconfirmation = 0
    conflicts = 0
    ownership_skipped = 0
    for d in (decisions or []):
        rid = str((d or {}).get("ai_req_id") or "").strip()
        status = str((d or {}).get("status") or "").strip()
        if not rid or not status:
            skipped += 1
            continue
        submitted_source = str(d.get("source_fingerprint") or "").strip()
        submitted_subject = str(d.get("review_subject_fingerprint") or "").strip()
        expected_target_fingerprint = str(
            d.get("expected_target_fingerprint") or ""
        ).strip()
        expected_target_publication_revision = str(
            d.get("expected_target_publication_revision") or ""
        ).strip()
        expected_target_authority_write_revision = str(
            d.get("expected_target_authority_write_revision") or ""
        ).strip()
        if not all((
            submitted_source,
            submitted_subject,
            expected_target_fingerprint,
            expected_target_publication_revision,
            expected_target_authority_write_revision,
        )):
            skipped += 1
            needs_reconfirmation += 1
            continue
        # 归属值单独校验：仅归属非法时丢归属、保留整行裁决（status/模块/意见不陪葬）
        ownership = str(d.get("ownership_override") or "").strip() or None
        if ownership:
            try:
                normalize_ownership(ownership)
            except ValueError:
                ownership = None
                ownership_skipped += 1
        try:
            from api_server import find_current_ai_requirement
            from omission_actions import extraction_operation_lock

            with extraction_operation_lock(out_dir, operation="import-ai-decision"):
                current = find_current_ai_requirement(out_dir, rid)
                if current is None:
                    raise ai_review_actions.AIReviewAuthorityConflict(
                        "AI requirement is not present in the current run",
                        current_revision="",
                    )
                current_cas = (
                    str(current.get("source_fingerprint") or ""),
                    str(current.get("review_subject_fingerprint") or ""),
                    str(current.get("target_fingerprint") or ""),
                    str(current.get("target_publication_revision") or ""),
                    str(current.get("target_authority_write_revision") or ""),
                )
                submitted_cas = (
                    submitted_source,
                    submitted_subject,
                    expected_target_fingerprint,
                    expected_target_publication_revision,
                    expected_target_authority_write_revision,
                )
                if submitted_cas != current_cas:
                    raise ai_review_actions.AIReviewAuthorityConflict(
                        "AI requirement or review authority changed",
                        current_revision=current_cas[-1],
                    )
                ai_review_actions.apply_ai_review_action(
                    out_dir,
                    rid,
                    status,
                    module_override=(d.get("module_override") or None),
                    ownership_override=ownership,
                    reason=(d.get("reason") or ""),
                    actor="html-import",
                    source_fingerprint_value=current_cas[0],
                    review_subject_fingerprint_value=current_cas[1],
                    review_anchor_fingerprint_value=(
                        ai_review_actions.review_anchor_fingerprint(current)
                    ),
                    expected_target_authority_write_revision=current_cas[-1],
                )
            applied += 1
        except ai_review_actions.AIReviewAuthorityConflict:
            skipped += 1
            conflicts += 1
        except ValueError:
            skipped += 1
    payload: dict[str, Any] = {"kind": "ai_decisions_import", "out_dir": str(out_dir),
                               "applied": applied, "skipped": skipped}
    if needs_reconfirmation:
        payload["needs_reconfirmation"] = needs_reconfirmation
    if conflicts:
        payload["conflicts"] = conflicts
    if ownership_skipped:
        payload["ownership_skipped"] = ownership_skipped
    # 裁决回流交付物：导入后立即重建 merged_spec（免 LLM）
    if applied and (out_dir / "ai_requirements.jsonl").exists():
        rebuilt = ai_extract.rebuild_merged_spec(out_dir)
        payload["rebuilt"] = rebuilt
        try:
            from adjudication_bank import resolve_bank_path, update_bank

            bank_path = resolve_bank_path()
            if bank_path is not None:
                payload["adjudication_bank"] = update_bank(bank_path, out_dir)
        except Exception as exc:  # 导入/重建已成功；学习资产失败只留告警
            LOGGER.warning("HTML 裁决导入后样本库收割失败（忽略）：%s", exc)
    return payload
