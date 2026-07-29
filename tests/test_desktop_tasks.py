from __future__ import annotations

import json
import io
import multiprocessing
import tempfile
import threading
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from unittest import mock
from unittest.mock import ANY, patch

import desktop_tasks

from llm_pipeline import write_jsonl


def _update_manifest_process(out_dir: str, stage: str, start_event) -> None:
    start_event.wait(10)
    desktop_tasks.update_run_manifest(Path(out_dir), stage, "running")


def _hold_manifest_lock(out_dir: str, ready_event, release_event) -> None:
    with desktop_tasks._run_manifest_lock(Path(out_dir)):
        ready_event.set()
        if not release_event.wait(10):
            raise RuntimeError("test did not release run manifest lock")


class ResolveKbPathsTests(unittest.TestCase):
    """锁定 desktop_tasks.resolve_kb_paths：前端预设送相对 --kb 路径，打包后端 cwd=resources/backend
    命中不到时必须按 package_root() 解析（否则报 'No such file: …/backend/knowledge_bases/…json'）。"""

    def test_none_uses_default_kb_paths(self) -> None:
        from desktop_tasks import resolve_kb_paths

        sentinel = [Path("X") / "default.json"]
        with patch("desktop_tasks.default_kb_paths", return_value=sentinel) as default_kb_paths:
            self.assertEqual(resolve_kb_paths(None), sentinel)
            default_kb_paths.assert_called_once()

    def test_absolute_paths_pass_through(self) -> None:
        from desktop_tasks import resolve_kb_paths

        absolute = (Path(tempfile.gettempdir()).resolve() / "abs_kb.json")
        self.assertEqual(resolve_kb_paths([absolute]), [absolute])

    def test_relative_missing_in_cwd_resolves_against_package_root(self) -> None:
        from desktop_tasks import resolve_kb_paths

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            # 唯一名，保证不在测试 cwd 命中 -> 走 package_root 兜底
            rel = Path("knowledge_bases") / "__resolve_kb_probe__.json"
            with patch("desktop_tasks.package_root", return_value=root):
                self.assertEqual(resolve_kb_paths([rel]), [root / rel])

    def test_resolve_bundled_path_none_returns_none(self) -> None:
        from desktop_tasks import resolve_bundled_path

        self.assertIsNone(resolve_bundled_path(None))

    def test_resolve_bundled_path_relative_domain_pack_uses_package_root(self) -> None:
        from desktop_tasks import resolve_bundled_path

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            # domain pack 预设相对路径（cwd 命中不到）-> package_root 兜底（打包后即 resources/）
            rel = Path("domain_packs") / "__resolve_pack_probe__"
            with patch("desktop_tasks.package_root", return_value=root):
                self.assertEqual(resolve_bundled_path(rel), root / rel)


class RunLoggingTests(unittest.TestCase):
    """GUI 路径日志：run.log 跟着输出目录走；任务结束关句柄（不锁目录）；幂等不重复挂 handler。"""

    def tearDown(self) -> None:
        import desktop_tasks
        desktop_tasks.teardown_run_logging()

    def test_setup_writes_run_log_and_is_idempotent(self) -> None:
        import logging
        from desktop_tasks import setup_run_logging, teardown_run_logging

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            setup_run_logging(out)
            setup_run_logging(out)   # 幂等：不重复挂 handler
            logger = logging.getLogger("requirement_atomizer")
            tags = [getattr(h, "_ratomizer_tag", None) for h in logger.handlers]
            self.assertEqual(tags.count("runlog"), 1)
            logger.info("marker-line-123")
            teardown_run_logging()   # 关句柄，tmp 目录才能删（Windows）
            content = (out / "run.log").read_text(encoding="utf-8")
        self.assertIn("marker-line-123", content)

    def test_main_produces_run_log_for_summary_command(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "o"
            out.mkdir()
            write_jsonl(out / "atomic_requirements.jsonl", [])
            with redirect_stdout(io.StringIO()):
                exit_code = desktop_tasks.main(["summary", "--out", str(out)])
            self.assertEqual(exit_code, 0)
            log_text = (out / "run.log").read_text(encoding="utf-8")
        self.assertIn("desktop task 开始：summary", log_text)


class DesktopTaskEncodingTests(unittest.TestCase):
    def test_main_writes_json_payload_on_gbk_stdout(self) -> None:
        """Packaged Windows stdout may be GBK; non-GBK payload chars must not crash."""
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "o"
            out.mkdir()
            write_jsonl(out / "atomic_requirements.jsonl", [])
            raw = io.BytesIO()
            gbk_stdout = io.TextIOWrapper(raw, encoding="gbk", errors="strict")
            with (
                mock.patch("sys.stdout", gbk_stdout),
                mock.patch("desktop_tasks.build_output_summary", return_value={"text": "A\u0300"}),
            ):
                exit_code = desktop_tasks.main(["summary", "--out", str(out)])
            gbk_stdout.flush()

        self.assertEqual(exit_code, 0)
        payload = json.loads(raw.getvalue().decode("gbk"))
        self.assertEqual(payload["kind"], "summary")

    def test_emit_progress_writes_json_on_gbk_stdout(self) -> None:
        import desktop_tasks

        raw = io.BytesIO()
        gbk_stdout = io.TextIOWrapper(raw, encoding="gbk", errors="strict")
        with mock.patch("sys.stdout", gbk_stdout):
            desktop_tasks.emit_progress({"stage": "probe", "text": "A\u0300"})
        gbk_stdout.flush()

        text = raw.getvalue().decode("gbk")
        self.assertIn(desktop_tasks.PROGRESS_PREFIX, text)
        self.assertEqual(json.loads(text.split(desktop_tasks.PROGRESS_PREFIX, 1)[1])["text"], "A\u0300")


class DesktopTaskTests(unittest.TestCase):
    def test_run_pipeline_task_uses_default_kbs_when_not_supplied(self) -> None:
        from desktop_tasks import run_pipeline_task

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "input.docx"
            out_dir = root / "out"
            input_path.write_text("placeholder", encoding="utf-8")
            out_dir.mkdir()

            with (
                patch("desktop_tasks.default_kb_paths") as default_kb_paths,
                patch("desktop_tasks.run_atomizer_pipeline") as atomize,
            ):
                default_kb_paths.return_value = [root / "default-a.json", root / "default-b.json"]
                atomize.return_value = {"counts": {"atomic_requirements": 0}}
                write_jsonl(out_dir / "atomic_requirements.jsonl", [])

                run_pipeline_task(input_path, out_dir, skip_review=True)

        atomize.assert_called_once()
        self.assertEqual(atomize.call_args.kwargs["kb_paths"], [root / "default-a.json", root / "default-b.json"])

    def test_run_pipeline_task_writes_outputs_and_review_summary(self) -> None:
        from desktop_tasks import run_pipeline_task

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "input.docx"
            out_dir = root / "out"
            input_path.write_text("placeholder", encoding="utf-8")
            out_dir.mkdir()

            with patch("desktop_tasks.run_atomizer_pipeline") as atomize, patch("desktop_tasks.run_review_pipeline") as review:
                atomize.return_value = {
                    "input": str(input_path),
                    "output_dir": str(out_dir),
                    "counts": {"atomic_requirements": 2},
                }
                review.return_value = {"reviews": 2, "accepted": 1, "expert_pending": 1}
                write_jsonl(
                    out_dir / "atomic_requirements.jsonl",
                    [
                        {"stable_req_id": "SREQ-1", "requirement_type": "functional", "confidence": 0.9},
                        {"stable_req_id": "SREQ-2", "requirement_type": "security", "confidence": 0.7},
                    ],
                )
                write_jsonl(
                    out_dir / "review_states.jsonl",
                    [
                        {"requirement_id": "SREQ-1", "status": "accepted"},
                        {"requirement_id": "SREQ-2", "status": "expert_pending"},
                    ],
                )

                payload = run_pipeline_task(input_path, out_dir, skip_review=False)

        self.assertEqual(payload["kind"], "pipeline")
        self.assertEqual(payload["manifest"]["counts"]["atomic_requirements"], 2)
        self.assertEqual(payload["review"]["reviews"], 2)
        self.assertEqual(payload["summary"]["counts"]["requirements"], 2)
        self.assertEqual(payload["summary"]["status_counts"]["accepted"], 1)
        atomize.assert_called_once()
        review.assert_called_once_with(out_dir.resolve(), route=None, scope=None, llm_review_limit=0, progress_callback=ANY, kb_paths=None, domain_pack_path=desktop_tasks.DEFAULT_DOMAIN_PACK_PATH)

    def test_run_pipeline_task_passes_explicit_kb_to_review(self) -> None:
        """审计 P1-d①：--kb 贯通到审查阶段——此前只传 atomize，review 工具落回默认 KB。"""
        from desktop_tasks import run_pipeline_task

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "input.docx"
            out_dir = root / "out"
            input_path.write_text("placeholder", encoding="utf-8")
            out_dir.mkdir()
            kb = root / "kb.json"
            kb.write_text('{"kb_id": "k", "entries": []}', encoding="utf-8")

            with patch("desktop_tasks.run_atomizer_pipeline") as atomize, patch("desktop_tasks.run_review_pipeline") as review:
                atomize.return_value = {
                    "input": str(input_path),
                    "output_dir": str(out_dir),
                    "counts": {"atomic_requirements": 0},
                }
                review.return_value = {"reviews": 0}
                write_jsonl(out_dir / "atomic_requirements.jsonl", [])
                write_jsonl(out_dir / "review_states.jsonl", [])

                run_pipeline_task(input_path, out_dir, skip_review=False, kb_paths=[kb])

        review.assert_called_once_with(out_dir.resolve(), route=None, scope=None, llm_review_limit=0, progress_callback=ANY, kb_paths=[kb], domain_pack_path=desktop_tasks.DEFAULT_DOMAIN_PACK_PATH)

    def test_run_pipeline_task_forwards_domain_pack_to_review(self) -> None:
        """审计 R2-H3：desktop run 的 --domain-pack 不再只喂 atomize——审查按同一包的
        pack.yaml 合并 review_policy；未传时回落默认捆绑包（见上两个测试）。"""
        from desktop_tasks import run_pipeline_task

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "input.docx"
            out_dir = root / "out"
            input_path.write_text("placeholder", encoding="utf-8")
            out_dir.mkdir()
            pack = root / "domain_packs" / "dlms_cosem"
            pack.mkdir(parents=True)
            (pack / "pack.yaml").write_text("review_policy: {}\n", encoding="utf-8")

            with patch("desktop_tasks.run_atomizer_pipeline") as atomize, patch("desktop_tasks.run_review_pipeline") as review:
                atomize.return_value = {
                    "input": str(input_path),
                    "output_dir": str(out_dir),
                    "counts": {"atomic_requirements": 0},
                }
                review.return_value = {"reviews": 0}
                write_jsonl(out_dir / "atomic_requirements.jsonl", [])
                write_jsonl(out_dir / "review_states.jsonl", [])

                run_pipeline_task(input_path, out_dir, skip_review=False, domain_pack_dir=pack)

        review.assert_called_once_with(
            out_dir.resolve(), route=None, scope=None, llm_review_limit=0,
            progress_callback=ANY, kb_paths=None, domain_pack_path=pack / "pack.yaml")

    def test_llm_review_producer_tracks_evidence_content(self) -> None:
        """llm-review 阶段 producer 纳入证据指纹：改 KB 内容阶段戳变；无 out_dir 保持基础戳+代码版本。"""
        from llm_pipeline import LLM_REVIEW_CACHE_VERSION, PROMPT_VERSION
        from review_tools import REVIEW_TOOLS_VERSION

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            out = root / "out"
            out.mkdir()
            (out / "blocks.jsonl").write_text("{}\n", encoding="utf-8")
            (out / "atomic_requirements.jsonl").write_text("{}\n", encoding="utf-8")
            kb = root / "kb.json"
            kb.write_text('{"kb_id": "k", "entries": []}', encoding="utf-8")

            first = desktop_tasks.stage_producer("llm-review", out_dir=out, kb_paths=[kb])
            kb.write_text('{"kb_id": "k", "entries": [{"id": "x"}]}', encoding="utf-8")
            second = desktop_tasks.stage_producer("llm-review", out_dir=out, kb_paths=[kb])

        self.assertTrue(first.startswith("review/v1+"))
        self.assertIn("+evidence-", first)
        self.assertNotEqual(first, second)
        base = desktop_tasks.stage_producer("llm-review")
        self.assertEqual(
            base,
            f"review/v1+{PROMPT_VERSION}+{LLM_REVIEW_CACHE_VERSION}+{REVIEW_TOOLS_VERSION}",
        )

    def test_llm_review_producer_tracks_code_versions(self) -> None:
        """审计 R2-H2：prompt/cache/tools 任一 bump 都必须让 llm-review 阶段戳变化
        （此前三者均不在戳内，bump 后 stage_is_reusable 仍判 True 整阶段复用旧结果）。"""
        import llm_pipeline
        import review_tools

        current = desktop_tasks.stage_producer("llm-review")
        for version in (
            llm_pipeline.PROMPT_VERSION,
            llm_pipeline.LLM_REVIEW_CACHE_VERSION,
            review_tools.REVIEW_TOOLS_VERSION,
        ):
            self.assertIn(version, current)
        with patch.object(review_tools, "REVIEW_TOOLS_VERSION", "review-tools-vNEXT"):
            changed = desktop_tasks.stage_producer("llm-review")
        with patch.object(llm_pipeline, "LLM_REVIEW_CACHE_VERSION", "llm-review-cache-vNEXT"):
            changed_cache = desktop_tasks.stage_producer("llm-review")

        self.assertNotEqual(current, changed)
        self.assertIn("review-tools-vNEXT", changed)
        self.assertNotEqual(current, changed_cache)
        self.assertIn("llm-review-cache-vNEXT", changed_cache)

    def test_llm_review_fingerprint_tracks_scope_and_limit(self) -> None:
        """审计 R2-H2：review_scope/llm_review_limit 进阶段指纹——先 targeted 后 all、
        先限量后全量时指纹必须不同，否则阶段被整体跳过。"""
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            targeted = desktop_tasks.stage_input_fingerprint(
                out, "llm-review", config={"review_scope": "targeted", "llm_review_limit": 0})
            full = desktop_tasks.stage_input_fingerprint(
                out, "llm-review", config={"review_scope": "all", "llm_review_limit": 0})
            limited = desktop_tasks.stage_input_fingerprint(
                out, "llm-review", config={"review_scope": "all", "llm_review_limit": 5})

        self.assertNotEqual(targeted, full)
        self.assertNotEqual(full, limited)

    def test_llm_review_stage_not_reusable_after_scope_change(self) -> None:
        """审计 R2-H2：先 targeted 后 all——scope 进指纹，旧审查产物不得复用。"""
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            out = root / "out"
            out.mkdir()
            for name in ["blocks.jsonl", "atomic_requirements.jsonl", "llm_tasks.jsonl",
                         "llm_review_results.jsonl", "review_states.jsonl"]:
                (out / name).write_text("{}\n", encoding="utf-8")
            targeted = {"review_scope": "targeted", "llm_review_limit": 0}
            full = {"review_scope": "all", "llm_review_limit": 0}

            desktop_tasks.update_run_manifest(out, "llm-review", "ok", route="stub", config=targeted)
            reusable_same = desktop_tasks.stage_is_reusable(out, "llm-review", route="stub", config=targeted)
            reusable_changed = desktop_tasks.stage_is_reusable(out, "llm-review", route="stub", config=full)

        self.assertTrue(reusable_same)
        self.assertFalse(reusable_changed)

    def test_llm_review_fingerprint_tracks_domain_pack_contents(self) -> None:
        """审计 R2-H2：llm-review 合并 domain-pack 的 review_policy——包内容变，阶段指纹变。"""
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            pack = root / "domain_packs" / "dlms_cosem"
            pack.mkdir(parents=True)
            pack_yaml = pack / "pack.yaml"
            pack_yaml.write_text("review_policy: {}\n", encoding="utf-8")
            config = {"review_scope": None, "llm_review_limit": 0, "domain_pack_dir": str(pack)}

            first = desktop_tasks.stage_input_fingerprint(root, "llm-review", config=config)
            pack_yaml.write_text("review_policy:\n  low_confidence_threshold: 0.5\n", encoding="utf-8")
            second = desktop_tasks.stage_input_fingerprint(root, "llm-review", config=config)
            no_pack = desktop_tasks.stage_input_fingerprint(
                root, "llm-review", config={"review_scope": None, "llm_review_limit": 0})

        self.assertNotEqual(first, second)
        self.assertNotEqual(first, no_pack)

    def test_llm_review_stage_not_reusable_after_kb_change(self) -> None:
        """阶段复用同样吃证据指纹：KB 内容变 → llm-review 旧产物不再复用。"""
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            out = root / "out"
            out.mkdir()
            for name in ["blocks.jsonl", "atomic_requirements.jsonl", "llm_tasks.jsonl",
                         "llm_review_results.jsonl", "review_states.jsonl"]:
                (out / name).write_text("{}\n", encoding="utf-8")
            kb = root / "kb.json"
            kb.write_text('{"kb_id": "k", "entries": []}', encoding="utf-8")
            config = {"kb_paths": [str(kb)]}

            desktop_tasks.update_run_manifest(out, "llm-review", "ok", route="openai_compatible", config=config)
            reusable_before = desktop_tasks.stage_is_reusable(
                out, "llm-review", route="openai_compatible", config=config)
            kb.write_text('{"kb_id": "k", "entries": [{"id": "x"}]}', encoding="utf-8")
            reusable_after = desktop_tasks.stage_is_reusable(
                out, "llm-review", route="openai_compatible", config=config)

        self.assertTrue(reusable_before)
        self.assertFalse(reusable_after)

    def test_main_run_command_passes_kb_and_domain_pack_to_pipeline(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "input.docx"
            out_dir = root / "out"
            kb_path = root / "kb.json"
            domain_pack = root / "domain_packs" / "dlms_cosem"
            input_path.write_text("placeholder", encoding="utf-8")
            kb_path.write_text("{}", encoding="utf-8")
            domain_pack.mkdir(parents=True)

            with patch("desktop_tasks.run_pipeline_task") as run_pipeline:
                run_pipeline.return_value = {"kind": "pipeline", "out_dir": str(out_dir), "summary": {}}

                with redirect_stdout(io.StringIO()):
                    exit_code = desktop_tasks.main([
                        "run",
                        "--input",
                        str(input_path),
                        "--out",
                        str(out_dir),
                        "--chunk-chars",
                        "1200",
                        "--kb",
                        str(kb_path),
                        "--domain-pack",
                        str(domain_pack),
                    ])

        self.assertEqual(exit_code, 0)
        run_pipeline.assert_called_once_with(
            input_path,
            out_dir,
            skip_review=False,
            llm_route=None,
            review_scope=None,
            llm_review_limit=0,
            chunk_chars=1200,
            kb_paths=[kb_path],
            domain_pack_dir=domain_pack,
        )

    def test_export_task_returns_written_files(self) -> None:
        from desktop_tasks import export_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            with patch("desktop_tasks.export_requirements") as export_requirements:
                export_requirements.return_value = ["requirements_export.csv", "requirements_export.md"]

                payload = export_task(out_dir, ["csv", "md"])

        self.assertEqual(payload["kind"], "export")
        self.assertEqual(payload["written"], ["requirements_export.csv", "requirements_export.md"])
        export_requirements.assert_called_once_with(out_dir.resolve(), formats=["csv", "md"])

    def test_assemble_task_writes_json_and_exports_formats(self) -> None:
        from desktop_tasks import ASSEMBLED_JSON, assemble_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            with patch("desktop_tasks.assemble") as assemble, patch("desktop_tasks.export_spec") as export_spec:
                assemble.return_value = ({"requirements": [{"id": "REQ-1"}], "analysis": {"total_count": 1}}, {"安全": 1})
                export_spec.return_value = ["dlms_cosem_spec_requirements.md"]

                payload = assemble_task(out_dir, formats=["md"])

            assembled = json.loads((out_dir / ASSEMBLED_JSON).read_text(encoding="utf-8"))

        self.assertEqual(payload["kind"], "assemble")
        self.assertEqual(payload["count"], 1)
        self.assertEqual(assembled["requirements"][0]["id"], "REQ-1")
        self.assertIn(str(out_dir / ASSEMBLED_JSON), payload["written"])
        self.assertIn(str(out_dir / "dlms_cosem_spec_requirements.md"), payload["written"])
        export_spec.assert_called_once()

    def test_assemble_task_passes_blue_book_index_to_assemble(self) -> None:
        from desktop_tasks import assemble_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            index_path = out_dir / "blue_book_index.json"
            index_path.write_text("{}", encoding="utf-8")
            with patch("desktop_tasks.assemble") as assemble:
                assemble.return_value = ({"requirements": [], "analysis": {}}, {"total": 0})

                assemble_task(out_dir, formats=[], enrich_route="openai_compatible", blue_book_index_path=index_path)

        assemble.assert_called_once()
        self.assertEqual(assemble.call_args.kwargs["blue_book_index_path"], index_path)

    def test_assemble_task_autodetects_index_in_out_dir(self) -> None:
        """桌面「运行」链没有索引输入口——out_dir 下有编译好的索引就自动带上（零配置）。"""
        from desktop_tasks import assemble_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            index_path = out_dir / "blue_book_index.json"
            index_path.write_text("{}", encoding="utf-8")
            with patch("desktop_tasks.assemble") as assemble:
                assemble.return_value = ({"requirements": [], "analysis": {}}, {"total": 0})

                payload = assemble_task(out_dir, formats=[], enrich_route="openai_compatible")

        self.assertEqual(assemble.call_args.kwargs["blue_book_index_path"], index_path)
        self.assertEqual(payload["blue_book_index"], str(index_path))   # 载荷追溯用了哪个索引

    def test_assemble_task_autodetect_falls_back_to_package_root(self) -> None:
        """dev 仓库编译位置 out/bluebook/ 作最后候选（打包后 resources/ 无 out/，自然探测不到）。"""
        from desktop_tasks import assemble_task

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "repo"
            out_dir = Path(tmp) / "out"
            out_dir.mkdir()
            index_path = root / "out" / "bluebook" / "blue_book_index.json"
            index_path.parent.mkdir(parents=True)
            index_path.write_text("{}", encoding="utf-8")
            with (
                patch("desktop_tasks.assemble") as assemble,
                patch("desktop_tasks.package_root", return_value=root),
            ):
                assemble.return_value = ({"requirements": [], "analysis": {}}, {"total": 0})

                assemble_task(out_dir, formats=[], enrich_route="openai_compatible")

        self.assertEqual(assemble.call_args.kwargs["blue_book_index_path"], index_path)

    def test_assemble_task_env_var_overrides_autodetect(self) -> None:
        import os
        from desktop_tasks import BLUE_BOOK_INDEX_ENV, assemble_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            (out_dir / "blue_book_index.json").write_text("{}", encoding="utf-8")  # 候选存在
            env_path = out_dir / "elsewhere.json"
            prior = os.environ.get(BLUE_BOOK_INDEX_ENV)
            os.environ[BLUE_BOOK_INDEX_ENV] = str(env_path)
            try:
                with patch("desktop_tasks.assemble") as assemble:
                    assemble.return_value = ({"requirements": [], "analysis": {}}, {"total": 0})
                    assemble_task(out_dir, formats=[], enrich_route="openai_compatible")
            finally:
                if prior is None:
                    os.environ.pop(BLUE_BOOK_INDEX_ENV, None)
                else:
                    os.environ[BLUE_BOOK_INDEX_ENV] = prior

        self.assertEqual(assemble.call_args.kwargs["blue_book_index_path"], env_path)   # env 优先于探测

    def test_assemble_task_no_index_anywhere_passes_none(self) -> None:
        """哪都没有索引 → None，行为与 P2 之前完全一致（默认零变化）。"""
        from desktop_tasks import assemble_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp) / "out"
            out_dir.mkdir()
            with (
                patch("desktop_tasks.assemble") as assemble,
                patch("desktop_tasks.package_root", return_value=Path(tmp) / "repo"),
            ):
                assemble.return_value = ({"requirements": [], "analysis": {}}, {"total": 0})

                payload = assemble_task(out_dir, formats=[], enrich_route="openai_compatible")

        self.assertIsNone(assemble.call_args.kwargs["blue_book_index_path"])
        self.assertIsNone(payload["blue_book_index"])

    def test_main_assemble_accepts_blue_book_index_argument(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp) / "out"
            index_path = Path(tmp) / "blue_book_index.json"
            index_path.write_text("{}", encoding="utf-8")
            with patch("desktop_tasks.assemble_task") as assemble:
                assemble.return_value = {"kind": "assemble", "out_dir": str(out_dir), "written": []}
                stdout = io.StringIO()
                with redirect_stdout(stdout):
                    exit_code = desktop_tasks.main([
                        "assemble",
                        "--out",
                        str(out_dir),
                        "--formats",
                        "",
                        "--enrich-route",
                        "openai_compatible",
                        "--blue-book-index",
                        str(index_path),
                    ])

        self.assertEqual(exit_code, 0)
        assemble.assert_called_once_with(
            out_dir,
            formats=[],
            enrich_route="openai_compatible",
            blue_book_index_path=index_path,
        )

    def test_compose_task_writes_engineering_requirement_outputs(self) -> None:
        from desktop_tasks import compose_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            with (
                patch("desktop_tasks.compose_engineering_requirements") as compose,
                patch("desktop_tasks.write_engineering_requirements") as write_outputs,
            ):
                compose.return_value = {
                    "analysis": {"requirement_functions": 2, "dlms_objects": 3},
                    "requirement_functions": [{}, {}],
                    "dlms_objects": [{}, {}, {}],
                }
                write_outputs.return_value = [
                    "engineering_requirements/engineering_requirements.json",
                    "engineering_requirements/requirement_functions.md",
                    "engineering_requirements/dlms_objects.md",
                ]

                payload = compose_task(out_dir)

        self.assertEqual(payload["kind"], "compose")
        self.assertEqual(payload["count"], 2)
        self.assertEqual(payload["analysis"]["dlms_objects"], 3)
        self.assertIn("engineering_requirements/requirement_functions.md", payload["written"])
        compose.assert_called_once_with(out_dir.resolve())
        write_outputs.assert_called_once_with(out_dir.resolve(), compose.return_value)

    def test_main_compose_command_runs_engineering_composer(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            with patch("desktop_tasks.compose_task") as compose:
                compose.return_value = {"kind": "compose", "out_dir": str(out_dir), "count": 1, "written": []}

                stdout = io.StringIO()
                with redirect_stdout(stdout):
                    exit_code = desktop_tasks.main(["compose", "--out", str(out_dir)])

        self.assertEqual(exit_code, 0)
        compose.assert_called_once_with(out_dir)
        self.assertEqual(json.loads(stdout.getvalue())["kind"], "compose")

    def test_requirements_analysis_task_wraps_run_requirements_analysis(self) -> None:
        from desktop_tasks import requirements_analysis_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp) / "out"
            template = Path(tmp) / "template.xlsx"
            out_dir.mkdir()
            template.write_bytes(b"placeholder")
            # written 只报真实存在的产物：只落盘其中 2 个，xlsx/co_design 缺席
            (out_dir / "engineering_analysis.json").write_text("{}", encoding="utf-8")
            (out_dir / "hardware_items.md").write_text("# Hardware Items\n", encoding="utf-8")
            with patch("desktop_tasks.run_requirements_analysis") as run_analysis:
                run_analysis.return_value = {"kind": "requirements_analysis", "analysis_count": 1, "issues": 0}

                payload = requirements_analysis_task(out_dir, route="stub", template_path=template)

        from desktop_tasks import emit_progress
        run_analysis.assert_called_once_with(out_dir.resolve(), route="stub", template_path=template,
                                             progress_callback=emit_progress)
        self.assertEqual(payload["kind"], "requirements_analysis")
        self.assertEqual(payload["analysis"]["analysis_count"], 1)
        self.assertEqual(
            sorted(payload["written"]),
            sorted([
                str(out_dir.resolve() / "engineering_analysis.json"),
                str(out_dir.resolve() / "hardware_items.md"),
            ]),
        )
        self.assertEqual(payload["summary"]["counts"]["requirements"], 0)

    def test_requirements_analysis_task_rejects_missing_explicit_template(self) -> None:
        from desktop_tasks import requirements_analysis_task

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp) / "out"
            out_dir.mkdir()
            missing_template = Path(tmp) / "missing-template.xlsx"

            with self.assertRaises(FileNotFoundError) as ctx:
                requirements_analysis_task(out_dir, route="stub", template_path=missing_template)

        self.assertIn("Template file does not exist", str(ctx.exception))

    def test_main_requirements_analysis_command_dispatches(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp) / "out"
            template = Path(tmp) / "template.xlsx"
            with patch("desktop_tasks.requirements_analysis_task") as task:
                task.return_value = {
                    "kind": "requirements_analysis",
                    "out_dir": str(out_dir),
                    "analysis": {"analysis_count": 1},
                    "written": [],
                    "summary": {},
                }
                stdout = io.StringIO()
                with redirect_stdout(stdout):
                    exit_code = desktop_tasks.main([
                        "requirements-analysis",
                        "--out",
                        str(out_dir),
                        "--llm-route",
                        "stub",
                        "--template",
                        str(template),
                    ])

        self.assertEqual(exit_code, 0)
        task.assert_called_once_with(out_dir, route="stub", template_path=template)
        self.assertEqual(json.loads(stdout.getvalue())["kind"], "requirements_analysis")

    def test_functional_synthesis_cli_records_actual_degraded_route(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            (out / "ai_requirements.jsonl").write_text(
                json.dumps({"ai_req_id": "AI-1", "title": "事件管理", "module": "事件"}) + "\n",
                encoding="utf-8",
            )
            stdout = io.StringIO()
            with redirect_stdout(stdout), mock.patch.dict("os.environ", {"RATOMIZER_LLM_API_KEY": ""}):
                rc = desktop_tasks.main([
                    "functional-synthesis", "--out", str(out),
                    "--llm-route", "openai_compatible",
                ])
            payload = json.loads(stdout.getvalue())
            manifest = json.loads((out / desktop_tasks.RUN_MANIFEST).read_text(encoding="utf-8"))

        self.assertEqual(rc, 0)
        self.assertEqual(payload["route"], "stub")
        self.assertEqual(payload["route_requested"], "openai_compatible")
        self.assertEqual(manifest["stages"]["functional-synthesis"]["route"], "stub")

    def test_functional_synthesis_task_forwards_route(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            with mock.patch.object(
                desktop_tasks, "run_functional_synthesis",
                return_value={"kind": "functional_synthesis", "written": ["functional_requirements.json"]},
            ) as run:
                payload = desktop_tasks.functional_synthesis_task(out, route="openai_compatible")

        run.assert_called_once_with(out.resolve(), route="openai_compatible")
        self.assertEqual(payload["kind"], "functional_synthesis")

    def test_ai_extract_task_wraps_run_ai_extract(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            with patch("ai_extract.run_ai_extract") as run_ai:
                run_ai.return_value = {
                    "route": "openai_compatible", "requirements": 3,
                    "merged": {"total": 10, "ai_behavioral": 3, "deterministic_structural": 7},
                    "code_drift_flagged": 0, "int_drift_flagged": 1,
                    "written": ["merged_spec.xlsx", "merged_spec_requirements.json"],
                }
                payload = desktop_tasks.ai_extract_task(out_dir, route="openai_compatible")

        run_ai.assert_called_once_with(out_dir.resolve(), route="openai_compatible",
                                       merge_deterministic=True,
                                       progress_callback=desktop_tasks.emit_progress,
                                       limit_sections=None, sample_ratio=None)
        self.assertEqual(payload["kind"], "ai_extract")
        self.assertEqual(payload["count"], 3)
        self.assertEqual(payload["merged"]["total"], 10)
        self.assertIn(str(out_dir.resolve() / "merged_spec.xlsx"), payload["written"])

    def test_main_ai_extract_command_dispatches(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out_dir = Path(tmp)
            with patch("desktop_tasks.ai_extract_task") as task:
                task.return_value = {"kind": "ai_extract", "out_dir": str(out_dir), "count": 0}
                stdout = io.StringIO()
                with redirect_stdout(stdout):
                    exit_code = desktop_tasks.main(["ai-extract", "--out", str(out_dir), "--llm-route", "stub"])

        self.assertEqual(exit_code, 0)
        task.assert_called_once_with(out_dir, route="stub", limit_sections=None, sample_ratio=None)
        self.assertEqual(json.loads(stdout.getvalue())["kind"], "ai_extract")

    def test_export_annotation_html_and_import_round_trip(self) -> None:
        import desktop_tasks
        import ai_review_actions
        from doc_annotation_export import build_ai_requirements

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            (out / "blocks.jsonl").write_text(
                json.dumps({"block_id": "B2", "order": 2, "text": "The meter shall measure volume.",
                            "section_path": ["4"], "requirement_like": True, "noise": False,
                            "type": "paragraph"}) + "\n", encoding="utf-8")
            doc = {"requirements": [{"id": "REQ-001", "title": "体积计量", "description": "应计量体积",
                    "module": "计量", "source_section": "4", "source_quote": "The meter shall measure volume.",
                    "source_block_ids": ["B2"], "acceptance_criteria": ["按 4.2 测试"], "labels": ["计量"]}]}
            (out / "merged_spec_requirements.json").write_text(json.dumps(doc, ensure_ascii=False), encoding="utf-8")

            # 导出 HTML
            stdout = io.StringIO()
            with redirect_stdout(stdout):
                rc = desktop_tasks.main(["export-annotation-html", "--out", str(out)])
            self.assertEqual(rc, 0)
            self.assertEqual(json.loads(stdout.getvalue())["kind"], "annotation_html")
            self.assertTrue((out / "document_annotation.html").exists())

            # 导入裁决回灌
            rid = build_ai_requirements(out)[0]["ai_req_id"]
            (out / "dec.json").write_text(json.dumps(
                {"decisions": [{"ai_req_id": rid, "status": "accepted", "module_override": "计量精度"}]}),
                encoding="utf-8")
            stdout = io.StringIO()
            with redirect_stdout(stdout):
                rc = desktop_tasks.main(["import-ai-decisions", "--out", str(out), "--file", str(out / "dec.json")])
            self.assertEqual(rc, 0)
            self.assertEqual(json.loads(stdout.getvalue())["applied"], 1)
            states = ai_review_actions.read_ai_review_states(out)
            self.assertEqual(states[rid]["status"], "accepted")
            self.assertEqual(states[rid]["module_override"], "计量精度")

    def test_export_annotation_html_cli_forwards_layout_mode(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            with patch("desktop_tasks.export_annotation_html_task") as task:
                task.return_value = {
                    "kind": "annotation_html",
                    "out_dir": str(out),
                    "path": str(out / "document_annotation.html"),
                    "written": [],
                }
                stdout = io.StringIO()
                with redirect_stdout(stdout):
                    exit_code = desktop_tasks.main([
                        "export-annotation-html",
                        "--out",
                        str(out),
                        "--layout-mode",
                        "optimized",
                    ])

        self.assertEqual(exit_code, 0)
        task.assert_called_once_with(out, route=None, layout_mode="optimized")

    def test_export_annotation_html_cli_defaults_to_original_pdf_layout(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            with patch("desktop_tasks.export_annotation_html_task") as task:
                task.return_value = {
                    "kind": "annotation_html",
                    "out_dir": str(out),
                    "path": str(out / "document_annotation.html"),
                    "written": [],
                }
                with redirect_stdout(io.StringIO()):
                    exit_code = desktop_tasks.main(["export-annotation-html", "--out", str(out)])

        self.assertEqual(exit_code, 0)
        task.assert_called_once_with(out, route=None, layout_mode="pdf_original")

    def test_import_ai_decisions_preserves_ownership_override(self) -> None:
        import desktop_tasks
        import ai_review_actions

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            decisions_file = out / "decisions.json"
            decisions_file.write_text(json.dumps({
                "decisions": [{
                    "ai_req_id": "AI-1",
                    "status": "accepted",
                    "module_override": "时钟需求",
                    "ownership_override": "co_design",
                    "reason": "硬件 RTC 依赖",
                }]
            }, ensure_ascii=False), encoding="utf-8")

            result = desktop_tasks.import_ai_decisions_task(out, decisions_file)
            states = ai_review_actions.read_ai_review_states(out)

        self.assertEqual(result["applied"], 1)
        self.assertEqual(states["AI-1"]["ownership_override"], "co_design")

    def test_import_decisions_rebuilds_merged_spec(self) -> None:
        """P0 裁决回流：导入裁决后交付物自动重建，rejected 不再出现在 merged_spec。"""
        import desktop_tasks
        import ai_review_actions

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            req = {"title": "体积计量", "description": "应计量体积", "type": "functional",
                   "priority": "P1", "module": "计量", "labels": ["计量"],
                   "source_section": "4", "source_quote": "The meter shall measure volume.",
                   "source_block_ids": ["B2"], "acceptance_criteria": [], "notes": "",
                   "status": "draft", "dependencies": [], "parent": None, "children": []}
            (out / "ai_requirements.jsonl").write_text(
                json.dumps(req, ensure_ascii=False) + "\n", encoding="utf-8")
            (out / "dlms_cosem_spec_requirements.json").write_text('{"requirements": []}', encoding="utf-8")
            rid = ai_review_actions.ai_req_id(req)
            (out / "dec.json").write_text(
                json.dumps({"decisions": [{"ai_req_id": rid, "status": "rejected"}]}), encoding="utf-8")

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                rc = desktop_tasks.main(["import-ai-decisions", "--out", str(out), "--file", str(out / "dec.json")])
            self.assertEqual(rc, 0)
            payload = json.loads(stdout.getvalue())
            self.assertEqual(payload["applied"], 1)
            self.assertIn("rebuilt", payload)                          # 交付物已重建
            merged = json.loads((out / "merged_spec_requirements.json").read_text(encoding="utf-8"))
            self.assertEqual(merged["requirements"], [])               # rejected 已从交付物剔除



class ClarificationWorkbookImportTests(unittest.TestCase):
    def test_legacy_workbook_is_rejected_before_any_answer_is_written(self) -> None:
        import clarification_report
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            (out / "ai_requirements.jsonl").write_text(json.dumps({
                "ai_req_id": "AIR-1",
                "title": "Limit",
                "source_section": "4",
                "source_quote": "The meter shall expose the configured limit.",
                "suspicion_reasons": ["原文数值未带全", "引用非逐字"],
            }, ensure_ascii=False) + "\n", encoding="utf-8")
            (out / "ai_extract_quality.json").write_text(json.dumps({
                "failed_sections": 0, "coverage_pct": 80.0,
            }), encoding="utf-8")
            clarification_report.run_report(out)
            workbook = load_workbook(out / clarification_report.REPORT_XLSX)
            workbook["必答-问客户"].cell(2, 8, "25")
            internal = workbook["必答-内部核对"]
            columns = {str(cell.value): cell.column for cell in internal[1]}
            internal.cell(
                1,
                columns["新处置(确认无误/确认有问题/暂缓)"],
                "旧版处置列",
            )
            legacy = out / "legacy-filled.xlsx"
            workbook.save(legacy)
            workbook.close()

            with self.assertRaisesRegex(ValueError, "缺少列"):
                desktop_tasks.import_clarification_workbook_task(out, legacy)

            self.assertFalse((out / clarification_report.ANSWERS_FILE).exists())
            self.assertFalse((out / "clarification_check_states.jsonl").exists())

    def test_one_import_closes_customer_and_internal_rows_then_recomputes_readiness(self) -> None:
        import clarification_report
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            (out / "ai_requirements.jsonl").write_text(json.dumps({
                "ai_req_id": "AIR-1",
                "title": "Limit",
                "module": "计量",
                "source_section": "4",
                "source_quote": "The meter shall expose the configured limit.",
                "suspicion_reasons": ["原文数值未带全", "引用非逐字"],
            }, ensure_ascii=False) + "\n", encoding="utf-8")
            (out / "ai_extract_quality.json").write_text(json.dumps({
                "failed_sections": 0, "coverage_pct": 80.0,
            }), encoding="utf-8")
            first = clarification_report.run_report(out)
            workbook = load_workbook(out / clarification_report.REPORT_XLSX)
            customer = workbook["必答-问客户"]
            customer.cell(2, 8, "25")
            customer.cell(2, 9, "是")
            internal = workbook["必答-内部核对"]
            columns = {str(cell.value): cell.column for cell in internal[1]}
            internal.cell(2, columns["新处置(确认无误/确认有问题/暂缓)"], "确认无误")
            internal.cell(2, columns["核对人"], "reviewer")
            filled = out / "filled.xlsx"
            workbook.save(filled)
            workbook.close()

            result = desktop_tasks.import_clarification_workbook_task(out, filled)

        self.assertEqual(first["questions"], 2)
        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["internal_imported"], 1)
        self.assertEqual(result["questions"], 0)
        self.assertEqual(result["readiness"]["verdict"], "READY")


class ChainAndManifestTests(unittest.TestCase):
    """F1+F7：后端链编排 + run_manifest 显式状态账本。"""

    def test_affected_stage_producers_include_implementation_revision(self) -> None:
        from parsers.pdf_parser import PDF_TEXT_REPAIR_VERSION, text_repair_vocabulary_fingerprint

        # Future agent stages have a separate policy suffix. Keep this current-stage snapshot
        # unchanged so adding the Phase 0 anchor cannot invalidate existing cached outputs.
        expected = {
            "atomize": (
                f"atomize+{PDF_TEXT_REPAIR_VERSION}"
                f"+repair-vocab-{text_repair_vocabulary_fingerprint()}+impl-v7"
            ),
            # 专家审核 0715:版本戳必须覆盖全部影响产物的代码层——guards/verify 版本
            # 缺席使护栏与复核升级后 chain 续跑直接跳过 ai-extract
            "ai-extract": (
                "ai-extract-v21+guards-v16+ai-verify-v2"
                "+merged-consistency/v3-noise-tolerant-window"
                "+ai-supplement-v3-identity-preconditions+impl-v4"
            ),
            "assemble": "assemble_spec/v1+enrich-v3+enrich-guards-v1+ai-supplement-v3-identity-preconditions+impl-v2",
            "functional-synthesis": "functional-synthesis-v7+ai-supplement-v3-identity-preconditions+impl-v3",
            "requirements-analysis": "analyze-llm-v7+analyze-unfounded-v3+ai-supplement-v3-identity-preconditions+impl-v6",
            "template-write": "template_writer/v1+ai-supplement-v3-identity-preconditions+impl-v4",
            "clarification-report": "clarification/v6-coverage-basis+ai-supplement-v3-identity-preconditions+impl-v5",
            "export-annotation-html": (
                "doc_annotation_export/v12+annotation-translation-v2-segment-fallback"
                "+annotation-translation-guards-v1+doc-facsimile-v1"
                "+ai-supplement-v3-identity-preconditions"
            ),
        }
        self.assertEqual(
            {stage: desktop_tasks.stage_producer(stage) for stage in expected},
            expected,
        )

    def test_atomize_fingerprint_tracks_kb_and_domain_pack_contents(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            kb = root / "kb.json"
            domain = root / "domain"
            domain.mkdir()
            rules = domain / "rules.yaml"
            kb.write_text('{"entries": []}', encoding="utf-8")
            rules.write_text("version: 1\n", encoding="utf-8")
            config = {"kb_paths": [str(kb)], "domain_pack_dir": str(domain)}

            first = desktop_tasks.stage_input_fingerprint(root, "atomize", config=config)
            kb.write_text('{"entries": [{"id": "changed"}]}', encoding="utf-8")
            second = desktop_tasks.stage_input_fingerprint(root, "atomize", config=config)
            rules.write_text("version: 2\n", encoding="utf-8")
            third = desktop_tasks.stage_input_fingerprint(root, "atomize", config=config)

        self.assertNotEqual(first, second)
        self.assertNotEqual(second, third)

    def test_clarification_stage_tracks_all_resolution_sidecars(self) -> None:
        inputs = set(desktop_tasks.STAGE_INPUTS["clarification-report"])
        self.assertTrue({
            "clarification_answers.jsonl",
            "clarification_check_states.jsonl",
            "omission_states.jsonl",
            "ai_review_states.jsonl",
        }.issubset(inputs))

    def test_assemble_producer_tracks_enrich_guards_version(self) -> None:
        import spec_enrich

        current = desktop_tasks.stage_producer("assemble")
        with patch.object(spec_enrich, "ENRICH_GUARDS_VERSION", "enrich-guards-vNEXT"):
            changed = desktop_tasks.stage_producer("assemble")

        self.assertNotEqual(current, changed)
        self.assertIn("enrich-guards-vNEXT", changed)

    def test_ai_extract_producer_tracks_consistency_version(self) -> None:
        import merged_consistency

        current = desktop_tasks.stage_producer("ai-extract")
        with patch.object(merged_consistency, "MERGED_CONSISTENCY_VERSION", "merged-consistency/vNEXT"):
            changed = desktop_tasks.stage_producer("ai-extract")

        self.assertNotEqual(current, changed)
        self.assertIn("merged-consistency/vNEXT", changed)

    def test_annotation_producer_tracks_translation_guards_version(self) -> None:
        import doc_annotation_export

        current = desktop_tasks.stage_producer("export-annotation-html")
        with patch.object(
                doc_annotation_export,
                "ANNOTATION_TRANSLATION_GUARDS_VERSION",
                "annotation-translation-guards-vNEXT"):
            changed = desktop_tasks.stage_producer("export-annotation-html")

        self.assertNotEqual(current, changed)
        self.assertIn("annotation-translation-guards-vNEXT", changed)

    def test_legacy_v7_annotation_export_is_not_reusable(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            (out / "document_annotation.html").write_text("legacy", encoding="utf-8")
            fingerprint = desktop_tasks.stage_input_fingerprint(
                out,
                "export-annotation-html",
                route="openai_compatible",
                config={"layout_mode": "pdf_original"},
            )
            (out / desktop_tasks.RUN_MANIFEST).write_text(json.dumps({
                "manifest_version": 2,
                "stages": {
                    "export-annotation-html": {
                        "status": "ok",
                        "producer": "doc_annotation_export/v7",
                        "route": "openai_compatible",
                        "input_fingerprint": fingerprint,
                        "outputs": ["document_annotation.html"],
                    },
                },
            }), encoding="utf-8")

            reusable = desktop_tasks.stage_is_reusable(
                out,
                "export-annotation-html",
                route="openai_compatible",
                config={"layout_mode": "pdf_original"},
            )

        self.assertFalse(reusable)

    def test_ai_extract_fingerprint_tracks_verify_toggles(self) -> None:
        # 专家审核 0715:verify 开关/轮数不进阶段指纹 → 改设置后 chain 续跑
        # 直接跳过 ai-extract,新设置静默零生效
        import os
        import unittest.mock as mock
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            with mock.patch.dict(os.environ, {"RATOMIZER_AI_VERIFY": "1",
                                              "RATOMIZER_AI_VERIFY_ROUNDS": "2"}):
                base = desktop_tasks.stage_input_fingerprint(out, "ai-extract")
            with mock.patch.dict(os.environ, {"RATOMIZER_AI_VERIFY": "0",
                                              "RATOMIZER_AI_VERIFY_ROUNDS": "2"}):
                toggled = desktop_tasks.stage_input_fingerprint(out, "ai-extract")
            with mock.patch.dict(os.environ, {"RATOMIZER_AI_VERIFY": "1",
                                              "RATOMIZER_AI_VERIFY_ROUNDS": "3"}):
                rounds = desktop_tasks.stage_input_fingerprint(out, "ai-extract")
        self.assertNotEqual(base, toggled)
        self.assertNotEqual(base, rounds)

    def test_requirements_analysis_fingerprint_tracks_term_map(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            for name in desktop_tasks.STAGE_INPUTS["requirements-analysis"]:
                (out / name).write_text("{}\n", encoding="utf-8")
            term_map = out / "term_map.json"
            term_map.write_text(
                json.dumps({"terms": [{"source": "meter", "target": "电表"}]}),
                encoding="utf-8",
            )
            first = desktop_tasks.stage_input_fingerprint(out, "requirements-analysis")
            term_map.write_text(
                json.dumps({"terms": [{"source": "meter", "target": "表计"}]}),
                encoding="utf-8",
            )
            second = desktop_tasks.stage_input_fingerprint(out, "requirements-analysis")

        self.assertNotEqual(first, second)

    def test_requirements_analysis_fingerprint_tracks_enrichment_toggle(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            with mock.patch.dict(desktop_tasks.os.environ, {
                    "RATOMIZER_REQUIREMENTS_ANALYSIS_ENRICH": "0"}):
                disabled = desktop_tasks.stage_input_fingerprint(out, "requirements-analysis")
            with mock.patch.dict(desktop_tasks.os.environ, {
                    "RATOMIZER_REQUIREMENTS_ANALYSIS_ENRICH": "1"}):
                enabled = desktop_tasks.stage_input_fingerprint(out, "requirements-analysis")

        self.assertNotEqual(disabled, enabled)

    def test_ai_extract_fingerprint_tracks_verify_settings(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            with mock.patch.dict(desktop_tasks.os.environ, {
                    "RATOMIZER_AI_VERIFY": "0", "RATOMIZER_AI_VERIFY_ROUNDS": "1"}):
                disabled = desktop_tasks.stage_input_fingerprint(out, "ai-extract")
            with mock.patch.dict(desktop_tasks.os.environ, {
                    "RATOMIZER_AI_VERIFY": "1", "RATOMIZER_AI_VERIFY_ROUNDS": "3"}):
                enabled = desktop_tasks.stage_input_fingerprint(out, "ai-extract")

        self.assertNotEqual(disabled, enabled)

    def test_verify_toggle_is_not_claimed_as_gui_exposed(self) -> None:
        import config
        verify = next(item for item in config.ENV_REGISTRY if item.name == "RATOMIZER_AI_VERIFY")
        self.assertFalse(verify.gui_exposed)

    def test_annotation_fingerprint_tracks_layout_mode(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            optimized = desktop_tasks.stage_input_fingerprint(
                out, "export-annotation-html", config={"layout_mode": "optimized"})
            original = desktop_tasks.stage_input_fingerprint(
                out, "export-annotation-html", config={"layout_mode": "pdf_original"})

        self.assertNotEqual(optimized, original)

    def test_manifest_with_old_producer_is_not_reusable(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            (out / "engineering_analysis.json").write_text("{}", encoding="utf-8")
            for name in desktop_tasks.STAGE_REQUIRED_OUTPUTS["template-write"]:
                (out / name).write_bytes(b"old-output")
            old_producer = "template_writer/v1"
            with mock.patch.object(desktop_tasks, "stage_producer", return_value=old_producer):
                fingerprint = desktop_tasks.stage_input_fingerprint(out, "template-write")
                desktop_tasks.update_run_manifest(
                    out, "template-write", "ok",
                    outputs=desktop_tasks.STAGE_REQUIRED_OUTPUTS["template-write"],
                    input_fingerprint=fingerprint,
                )

            self.assertFalse(desktop_tasks.stage_is_reusable(out, "template-write"))

    def test_clarification_fingerprint_tracks_functional_conflicts(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            for name in ("ai_requirements.jsonl", "engineering_analysis.json",
                         "consistency_report.json", "blocks.jsonl"):
                (out / name).write_text("{}\n", encoding="utf-8")
            synthesis = out / desktop_tasks.FUNCTIONAL_REQUIREMENTS
            synthesis.write_text(
                json.dumps({"items": [{"conflict_flags": ["30 min"]}]}), encoding="utf-8")
            first = desktop_tasks.stage_input_fingerprint(out, "clarification-report")
            synthesis.write_text(
                json.dumps({"items": [{"conflict_flags": ["60 min"]}]}), encoding="utf-8")
            second = desktop_tasks.stage_input_fingerprint(out, "clarification-report")

        self.assertNotEqual(first, second)

    def test_chain_manifest_records_actual_functional_synthesis_route(self) -> None:
        import desktop_tasks

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            (out / "ai_requirements.jsonl").write_text(
                json.dumps({"ai_req_id": "AI-1", "title": "事件管理", "module": "事件"}) + "\n",
                encoding="utf-8",
            )
            with mock.patch.object(
                desktop_tasks, "functional_synthesis_task",
                return_value={
                    "kind": "functional_synthesis", "route_requested": "openai_compatible",
                    "route": "stub", "written": ["functional_requirements.json"],
                },
            ):
                (out / "functional_requirements.json").write_text("{}", encoding="utf-8")
                desktop_tasks.chain_task(
                    out, stages=["functional-synthesis"], route="openai_compatible")

            manifest = json.loads((out / desktop_tasks.RUN_MANIFEST).read_text(encoding="utf-8"))

        self.assertEqual(manifest["stages"]["functional-synthesis"]["route"], "stub")

    def test_chain_orders_dedupes_and_aggregates(self) -> None:
        calls: list[str] = []
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            with (mock.patch.object(desktop_tasks, "clarification_report_task",
                                    side_effect=lambda o: (calls.append("clarification-report") or
                                                           {"kind": "clarification_report", "questions": 7,
                                                            "readiness": {"verdict": "READY", "reasons": []},
                                                            "summary": {"big": 1}})),
                  mock.patch.object(desktop_tasks, "ai_extract_task",
                                    side_effect=lambda o, **kw: (calls.append("ai-extract") or
                                                                 {"kind": "ai_extract", "count": 3,
                                                                  "consistency": {"duplicate_groups": 1},
                                                                  "summary": {"big": 1}}))):
                payload = desktop_tasks.chain_task(
                    out, stages=["clarification-report", "ai-extract", "ai-extract"],
                    route="openai_compatible")

            self.assertEqual(calls, ["ai-extract", "clarification-report"])   # 依赖序 + 去重
            self.assertEqual(payload["stages"], ["ai-extract", "clarification-report"])
            self.assertEqual(payload["questions"], 7)                          # 顶层聚合
            self.assertEqual(payload["consistency"], {"duplicate_groups": 1})
            self.assertNotIn("summary", payload["results"]["ai-extract"])      # 阶段 summary 被剥离
            manifest = json.loads((out / desktop_tasks.RUN_MANIFEST).read_text(encoding="utf-8"))
            self.assertEqual(manifest["stages"]["ai-extract"]["status"], "ok")
            self.assertEqual(manifest["stages"]["ai-extract"]["producer"],
                             desktop_tasks.stage_producer("ai-extract"))

    def test_chain_forwards_annotation_layout_mode(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            with mock.patch.object(
                desktop_tasks,
                "export_annotation_html_task",
                return_value={
                    "kind": "annotation_html",
                    "path": str(out / "document_annotation.html"),
                    "written": [],
                },
            ) as export_task:
                desktop_tasks.chain_task(
                    out,
                    stages=["export-annotation-html"],
                    route="stub",
                    annotation_layout_mode="optimized",
                )

        export_task.assert_called_once_with(
            out.resolve(), route="stub", layout_mode="optimized")

    def test_chain_defaults_annotation_export_to_original_pdf_layout(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            with mock.patch.object(
                desktop_tasks,
                "export_annotation_html_task",
                return_value={"kind": "annotation_html", "path": str(out / "document_annotation.html"), "written": []},
            ) as export_task:
                desktop_tasks.chain_task(out, stages=["export-annotation-html"], route="stub")

        export_task.assert_called_once_with(
            out.resolve(), route="stub", layout_mode="pdf_original")

    def test_chain_retries_partial_translation_export_until_clean_run(self) -> None:
        summaries = [
            {"unresolved": 1, "failed_calls": 0},
            {"unresolved": 0, "failed_calls": 1},
            {"unresolved": 0, "failed_calls": 0},
        ]

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)

            def run_export(*_args, **_kwargs):
                path = out / "document_annotation.html"
                path.write_text("annotation", encoding="utf-8")
                return {
                    "kind": "annotation_html",
                    "path": str(path),
                    "route": "openai_compatible",
                    "written": [str(path)],
                    "translations": summaries.pop(0),
                }

            with mock.patch.object(
                    desktop_tasks, "export_annotation_html_task",
                    side_effect=run_export) as export_task:
                desktop_tasks.chain_task(
                    out, stages=["export-annotation-html"], route="openai_compatible")
                first = desktop_tasks.read_run_manifest(out)["stages"]["export-annotation-html"]
                desktop_tasks.chain_task(
                    out, stages=["export-annotation-html"], route="openai_compatible")
                second = desktop_tasks.read_run_manifest(out)["stages"]["export-annotation-html"]
                desktop_tasks.chain_task(
                    out, stages=["export-annotation-html"], route="openai_compatible")
                third = desktop_tasks.read_run_manifest(out)["stages"]["export-annotation-html"]

        self.assertEqual(first["status"], "partial")
        self.assertEqual(second["status"], "partial")
        self.assertEqual(third["status"], "ok")
        self.assertEqual(export_task.call_count, 3)

    def test_ai_extract_with_failed_sections_is_non_reusable(self) -> None:
        self.assertEqual(
            desktop_tasks._stage_completion_status("ai-extract", {"failed_sections": 1}),
            "partial",
        )
        self.assertEqual(
            desktop_tasks._stage_completion_status(
                "ai-extract", {"quality": {"failed_sections": 1}}
            ),
            "partial",
        )
        self.assertEqual(
            desktop_tasks._stage_completion_status("ai-extract", {"failed_sections": 0}),
            "ok",
        )

    def test_single_annotation_command_records_incomplete_translation_as_partial(self) -> None:
        for translations in (
                {"unresolved": 1, "failed_calls": 0},
                {"unresolved": 0, "failed_calls": 1}):
            with self.subTest(translations=translations), tempfile.TemporaryDirectory() as td:
                out = Path(td)
                path = out / "document_annotation.html"
                payload = {
                    "kind": "annotation_html",
                    "path": str(path),
                    "route": "openai_compatible",
                    "written": [str(path)],
                    "translations": translations,
                }
                with mock.patch.object(
                        desktop_tasks, "export_annotation_html_task",
                        return_value=payload), redirect_stdout(io.StringIO()):
                    exit_code = desktop_tasks.main([
                        "export-annotation-html", "--out", str(out),
                        "--route", "openai_compatible",
                    ])

                entry = desktop_tasks.read_run_manifest(out)["stages"]["export-annotation-html"]
                self.assertEqual(exit_code, 0)
                self.assertEqual(entry["status"], "partial")

    def test_chain_unknown_stage_and_missing_template_fail_fast(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            with self.assertRaises(ValueError):
                desktop_tasks.chain_task(out, stages=["nope"], route="stub")
            with self.assertRaises(ValueError):
                desktop_tasks.chain_task(out, stages=["template-write"], route="stub")

    def test_chain_stage_failure_recorded_and_raises(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            with mock.patch.object(desktop_tasks, "ai_extract_task",
                                   side_effect=RuntimeError("boom")):
                with self.assertRaises(RuntimeError):
                    desktop_tasks.chain_task(out, stages=["ai-extract"], route="stub")
            manifest = json.loads((out / desktop_tasks.RUN_MANIFEST).read_text(encoding="utf-8"))
            entry = manifest["stages"]["ai-extract"]
            self.assertEqual(entry["status"], "failed")
            self.assertIn("boom", entry["error"])

    def test_update_run_manifest_lifecycle(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            desktop_tasks.update_run_manifest(out, "assemble", "running")
            desktop_tasks.update_run_manifest(out, "assemble", "ok")
            data = json.loads((out / desktop_tasks.RUN_MANIFEST).read_text(encoding="utf-8"))
            entry = data["stages"]["assemble"]
            self.assertEqual(entry["status"], "ok")
            self.assertIn("started", entry)
            self.assertIn("finished", entry)
            self.assertEqual(data["manifest_version"], 2)

    def test_concurrent_process_updates_preserve_both_stages(self) -> None:
        context = multiprocessing.get_context("spawn")
        with tempfile.TemporaryDirectory() as td:
            start_event = context.Event()
            processes = [
                context.Process(
                    target=_update_manifest_process,
                    args=(td, stage, start_event),
                )
                for stage in ("concurrent-a", "concurrent-b")
            ]
            for process in processes:
                process.start()
            start_event.set()
            for process in processes:
                process.join(20)
                self.assertEqual(process.exitcode, 0)

            data = json.loads((Path(td) / desktop_tasks.RUN_MANIFEST).read_text(encoding="utf-8"))
            stage_a = json.loads(
                desktop_tasks._stage_manifest_path(Path(td), "concurrent-a").read_text(encoding="utf-8")
            )
            stage_b = json.loads(
                desktop_tasks._stage_manifest_path(Path(td), "concurrent-b").read_text(encoding="utf-8")
            )

        self.assertEqual(set(data["stages"]), {"concurrent-a", "concurrent-b"})
        self.assertEqual(stage_a["stage"], "concurrent-a")
        self.assertEqual(stage_b["stage"], "concurrent-b")

    def test_manifest_transaction_waits_for_cross_process_lock(self) -> None:
        context = multiprocessing.get_context("spawn")
        with tempfile.TemporaryDirectory() as td:
            ready_event = context.Event()
            release_event = context.Event()
            holder = context.Process(
                target=_hold_manifest_lock,
                args=(td, ready_event, release_event),
            )
            holder.start()
            self.assertTrue(ready_event.wait(10))

            started = threading.Event()
            finished = threading.Event()

            def update_manifest() -> None:
                started.set()
                desktop_tasks.update_run_manifest(Path(td), "waiting-stage", "running")
                finished.set()

            writer = threading.Thread(target=update_manifest)
            writer.start()
            self.assertTrue(started.wait(2))
            self.assertFalse(finished.wait(0.2))
            release_event.set()
            writer.join(10)
            holder.join(10)

            self.assertFalse(writer.is_alive())
            self.assertEqual(holder.exitcode, 0)
            self.assertTrue(finished.is_set())
            data = desktop_tasks.read_run_manifest(Path(td))
            self.assertEqual(data["stages"]["waiting-stage"]["status"], "running")

    def test_atomic_json_writes_use_unique_temporary_files(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "shared.json"
            start = threading.Barrier(6)
            errors: list[BaseException] = []

            def write_payload(index: int) -> None:
                try:
                    start.wait(5)
                    desktop_tasks._atomic_write_json(
                        path,
                        json.dumps({"writer": index, "payload": "x" * 100_000}),
                    )
                except BaseException as exc:  # capture worker failures for the assertion thread
                    errors.append(exc)

            threads = [threading.Thread(target=write_payload, args=(index,)) for index in range(6)]
            for thread in threads:
                thread.start()
            for thread in threads:
                thread.join(10)

            self.assertEqual(errors, [])
            self.assertIn(json.loads(path.read_text(encoding="utf-8"))["writer"], range(6))
            self.assertEqual(list(Path(td).glob(".shared.json.*.tmp")), [])

    def test_atomic_json_replace_retries_windows_sharing_violation(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "manifest.json"
            real_replace = desktop_tasks.os.replace
            attempts = 0

            def flaky_replace(source, target) -> None:
                nonlocal attempts
                attempts += 1
                if attempts < 3:
                    raise PermissionError("temporarily locked")
                real_replace(source, target)

            with patch.object(desktop_tasks.os, "replace", side_effect=flaky_replace):
                desktop_tasks._atomic_write_json(path, '{"ok": true}\n')

            self.assertEqual(attempts, 3)
            self.assertEqual(json.loads(path.read_text(encoding="utf-8")), {"ok": True})

    def test_run_pipeline_task_reuses_completed_atomize_and_review_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            input_path = root / "input.docx"
            input_path.write_text("doc", encoding="utf-8")
            out = root / "out"
            out.mkdir()
            (out / "manifest.json").write_text(json.dumps({
                "input": str(input_path.resolve()),
                "counts": {"atomic_requirements": 1},
            }), encoding="utf-8")
            for name in [
                "blocks.jsonl",
                "chunks.jsonl",
                "table_items.jsonl",
                "atomic_requirements.jsonl",
                "llm_tasks.jsonl",
                "quality_report.json",
                "summary.md",
                "llm_review_results.jsonl",
                "review_states.jsonl",
            ]:
                (out / name).write_text("{}\n", encoding="utf-8")
            atomize_config = {
                "chunk_chars": 3500,
                "kb_paths": [str(path) for path in desktop_tasks.resolve_kb_paths(None)],
                "domain_pack_dir": "",
            }
            desktop_tasks.update_run_manifest(
                out, "atomize", "ok", input_path=input_path, config=atomize_config)
            # config 形状与 run_pipeline_task 当前生成的一致（scope/limit 已入指纹，R2-H2）
            desktop_tasks.update_run_manifest(
                out, "llm-review", "ok", route="stub",
                config={"review_scope": None, "llm_review_limit": 0})

            with (mock.patch("desktop_tasks.run_atomizer_pipeline") as atomize,
                  mock.patch("desktop_tasks.run_review_pipeline") as review):
                payload = desktop_tasks.run_pipeline_task(input_path, out, llm_route="stub")

            atomize.assert_not_called()
            review.assert_not_called()
            self.assertEqual(payload["manifest"]["resume_action"], "skipped")
            self.assertEqual(payload["review"]["resume_action"], "skipped")
            self.assertEqual(payload["summary"]["run_manifest"]["stages"]["atomize"]["status"], "ok")

    def test_chain_skips_completed_stage_with_existing_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            (out / "ai_requirements.jsonl").write_text("{}\n", encoding="utf-8")
            (out / "merged_spec_requirements.json").write_text('{"requirements":[]}', encoding="utf-8")
            (out / "merged_spec.xlsx").write_bytes(b"xlsx")
            desktop_tasks.update_run_manifest(
                out,
                "ai-extract",
                "ok",
                route="stub",
                outputs=["ai_requirements.jsonl", "merged_spec_requirements.json", "merged_spec.xlsx"],
                config={"sample_ratio": None, "limit_sections": None},
            )

            with mock.patch.object(desktop_tasks, "ai_extract_task") as task:
                payload = desktop_tasks.chain_task(out, stages=["ai-extract"], route="stub")

            task.assert_not_called()
            self.assertEqual(payload["results"]["ai-extract"]["resume_action"], "skipped")
            self.assertIn("ai-extract", payload["skipped_stages"])

    def test_chain_reruns_legacy_outputs_without_run_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            (out / "ai_requirements.jsonl").write_text("{}\n", encoding="utf-8")
            (out / "merged_spec_requirements.json").write_text('{"requirements":[]}', encoding="utf-8")

            with mock.patch.object(
                desktop_tasks, "ai_extract_task",
                return_value={"written": ["ai_requirements.jsonl", "merged_spec_requirements.json"]},
            ) as task:
                payload = desktop_tasks.chain_task(out, stages=["ai-extract"], route="stub")

            task.assert_called_once()
            self.assertNotIn("ai-extract", payload["skipped_stages"])
            manifest = json.loads((out / desktop_tasks.RUN_MANIFEST).read_text(encoding="utf-8"))
            self.assertEqual(manifest["stages"]["ai-extract"]["last_action"], "ran")


if __name__ == "__main__":
    unittest.main()
