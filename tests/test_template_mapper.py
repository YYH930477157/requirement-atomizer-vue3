"""模板映射器回归（理解在前、映射在后；封闭槽位 + 程序护栏）。"""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import template_mapper as tm


def make_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "时钟需求"
    ws.append(["关闭", "序号", "子模块", "描述", "需求模版", "需求", "说明、示例、注意事项",
               "是否客户需求", "客户需求章节", "驱动/硬件相关"])
    ws.append(["", 1, "时钟", "历法：", "公历", "公历", "", "", "", ""])
    ws.append(["", 2, "时钟", "夏令时：", "支持", "支持", "", "", "", ""])
    ws.append(["", 3, "时钟", "时钟精度：", "±0.5s/天", "±0.5s/天", "", "", "", ""])
    wb.create_sheet("其他sheet").append(["占位"])
    wb.save(path)


def req(rid: str, title: str, quote: str, module: str = "时钟", section: str = "7.9") -> dict:
    return {"ai_req_id": rid, "title": title, "description": title, "source_quote": quote,
            "source_section": section, "module": module, "labels": [module]}


class QuestionBankTests(unittest.TestCase):
    def test_loads_questions_with_real_rows(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "t.xlsx"
            make_template(path)
            bank = tm.load_question_bank(path, "时钟需求")
        self.assertEqual(len(bank), 3)
        self.assertEqual(bank[0], {"xlsx_row": 2, "seq": "1", "submodule": "时钟",
                                   "question": "历法：", "default": "公历"})

    def test_unknown_sheet_raises(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "t.xlsx"
            make_template(path)
            with self.assertRaises(ValueError):
                tm.load_question_bank(path, "不存在需求")

    def test_sheet_module_mapping(self) -> None:
        self.assertEqual(tm.sheet_module("时钟需求"), "时钟")
        self.assertEqual(tm.sheet_module("负荷曲线"), "曲线")
        self.assertEqual(tm.sheet_module("P1需求"), "通信协议")


class MappingGuardTests(unittest.TestCase):
    def _bank(self) -> list[dict]:
        return [{"xlsx_row": 2, "seq": "1", "submodule": "时钟", "question": "历法：", "default": "公历"},
                {"xlsx_row": 3, "seq": "2", "submodule": "时钟", "question": "夏令时：", "default": "支持"},
                {"xlsx_row": 4, "seq": "3", "submodule": "时钟", "question": "时钟精度：", "default": "±0.5s/天"}]

    def test_valid_assignment_accepted_with_sections(self) -> None:
        reqs = [req("AI-1", "时钟精度要求", "clock accuracy shall be within 5 s per day", section="7.9")]

        def chat(system: str, user: str) -> dict:
            return {"assignments": [{"seq": "3", "answer": "时钟精度 5 s/天以内",
                                     "source_req_ids": ["AI-1"]}],
                    "unmatched_req_ids": []}

        result = tm.map_requirements(self._bank(), reqs, chat)
        self.assertEqual(len(result["assignments"]), 1)
        a = result["assignments"][0]
        self.assertEqual(a["seq"], "3")
        self.assertEqual(a["customer_sections"], ["7.9"])
        self.assertEqual(result["unmatched_requirements"], [])

    def test_fabricated_seq_rejected(self) -> None:
        reqs = [req("AI-1", "无关需求", "something")]

        def chat(system: str, user: str) -> dict:
            return {"assignments": [{"seq": "99", "answer": "x", "source_req_ids": ["AI-1"]}]}

        result = tm.map_requirements(self._bank(), reqs, chat)
        self.assertEqual(result["assignments"], [])
        self.assertIn("编造槽位", result["rejected"][0]["reason"])
        # 指派被拒 → 该需求回到未匹配（程序核算，不信 LLM 自报）
        self.assertEqual(result["unmatched_requirements"][0]["ai_req_id"], "AI-1")

    def test_fabricated_answer_value_rejected(self) -> None:
        """答案里的数字必须有据（来源需求文本）——LLM 编个 ±0.1s 出来即弃。"""
        reqs = [req("AI-1", "时钟精度要求", "clock accuracy shall be within 5 s per day")]

        def chat(system: str, user: str) -> dict:
            return {"assignments": [{"seq": "3", "answer": "时钟精度 ±0.1s/天",
                                     "source_req_ids": ["AI-1"]}]}

        result = tm.map_requirements(self._bank(), reqs, chat)
        self.assertEqual(result["assignments"], [])
        self.assertIn("无据", result["rejected"][0]["reason"])

    def test_unknown_source_req_id_rejected(self) -> None:
        reqs = [req("AI-1", "t", "q")]

        def chat(system: str, user: str) -> dict:
            return {"assignments": [{"seq": "1", "answer": "公历", "source_req_ids": ["AI-GHOST"]}]}

        result = tm.map_requirements(self._bank(), reqs, chat)
        self.assertEqual(result["assignments"], [])


class WorkbookWriteTests(unittest.TestCase):
    def test_fills_only_target_cells_and_neutralizes_formulas(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "t.xlsx"
            out = Path(td) / "mapped.xlsx"
            make_template(template)
            bank = tm.load_question_bank(template, "时钟需求")
            assignments = [{"seq": "2", "answer": "=1+1 支持夏令时",   # 注入尝试
                            "source_req_ids": ["AI-1"], "customer_sections": ["4.2", "7.9"]}]
            tm.apply_to_workbook(template, "时钟需求", bank, assignments, out)

            wb = load_workbook(out)
            ws = wb["时钟需求"]
            self.assertNotEqual(ws.cell(row=3, column=6).data_type, "f")     # 公式已中和
            self.assertIn("支持夏令时", str(ws.cell(row=3, column=6).value))
            self.assertEqual(ws.cell(row=3, column=8).value, "是")
            self.assertEqual(ws.cell(row=3, column=9).value, "4.2、7.9")
            self.assertEqual(ws.cell(row=2, column=6).value, "公历")          # 未指派行原样
            self.assertIn("其他sheet", wb.sheetnames)                          # 其余 sheet 保留


class RunMappingTests(unittest.TestCase):
    def test_end_to_end_writes_workbook_and_report(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            template = out / "t.xlsx"
            make_template(template)
            with (out / "ai_requirements.jsonl").open("w", encoding="utf-8") as f:
                f.write(json.dumps(req("AI-1", "时钟精度要求",
                                       "clock accuracy shall be within 5 s per day"),
                                   ensure_ascii=False) + "\n")
                f.write(json.dumps(req("AI-2", "显示轮显", "display scroll", module="显示"),
                                   ensure_ascii=False) + "\n")   # 异模块 → 不进候选

            def chat(system: str, user: str) -> dict:
                assert "AI-2" not in user                          # 模块预筛生效
                return {"assignments": [{"seq": "3", "answer": "时钟精度 5 s/天以内",
                                         "source_req_ids": ["AI-1"]}]}

            report = tm.run_mapping(out, template, "时钟需求", chat=chat)

            self.assertEqual(report["questions"], 3)
            self.assertEqual(report["candidate_requirements"], 1)
            self.assertEqual(report["answered"], 1)
            self.assertEqual(report["unanswered_seqs"], ["1", "2"])
            self.assertTrue((out / "软件需求列表-映射.xlsx").exists())
            data = json.loads((out / "mapping_report.json").read_text(encoding="utf-8"))
            self.assertEqual(data["assignments"][0]["seq"], "3")

    def test_missing_requirements_raises(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "t.xlsx"
            make_template(template)
            with self.assertRaises(FileNotFoundError):
                tm.run_mapping(Path(td), template, "时钟需求", chat=lambda s, u: {})


if __name__ == "__main__":
    unittest.main()
