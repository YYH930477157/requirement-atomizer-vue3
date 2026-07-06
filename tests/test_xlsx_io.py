"""xlsx 安全保存回归（F8）：目标被 Excel 锁住不弄死整链，另存并如实返回实际路径。"""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from xlsx_io import safe_save_workbook


class SafeSaveTests(unittest.TestCase):
    def test_normal_save_returns_target(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            target = Path(td) / "a.xlsx"
            result = safe_save_workbook(Workbook(), target)
        self.assertEqual(result, target)

    def test_locked_target_falls_back_to_suffixed_copy(self) -> None:
        wb = Workbook()
        calls = {"n": 0}
        real_save = wb.save

        def flaky_save(path):
            calls["n"] += 1
            if calls["n"] <= 2:                      # 首写 + 重试都被锁
                raise PermissionError("locked by Excel")
            real_save(path)

        with tempfile.TemporaryDirectory() as td:
            target = Path(td) / "b.xlsx"
            with mock.patch.object(wb, "save", side_effect=flaky_save),                  mock.patch("xlsx_io.time.sleep"):
                result = safe_save_workbook(wb, target)
            self.assertNotEqual(result, target)       # 另存副本
            self.assertTrue(result.name.startswith("b-"))
            self.assertTrue(result.exists())

    def test_retry_succeeds_on_second_attempt(self) -> None:
        wb = Workbook()
        calls = {"n": 0}
        real_save = wb.save

        def flaky_save(path):
            calls["n"] += 1
            if calls["n"] == 1:
                raise PermissionError("locked")
            real_save(path)

        with tempfile.TemporaryDirectory() as td:
            target = Path(td) / "c.xlsx"
            with mock.patch.object(wb, "save", side_effect=flaky_save),                  mock.patch("xlsx_io.time.sleep"):
                result = safe_save_workbook(wb, target)
            self.assertEqual(result, target)
