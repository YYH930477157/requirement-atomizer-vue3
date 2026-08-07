"""WS4 能力补齐六项能力的回归测试（全程零 LLM 调用）。

覆盖：
1. verification 子对象 + 六列回写/回灌（含 CAS）
2. 四态状态机（draft→confirmed→implemented→verified，前进由 verification 驱动，回退仅人工留痕）
3. 弱词扫描器进澄清报告（汇入模糊/缺失四分类，就绪判定同步变化）
4. 手工建需求入口（manual provenance，追溯列留空）
5. 需求库词面检索（Jaccard，零向量零 LLM，reviewer_override 修正归属优先）
6. dependencies/parent/children 半自动推荐（候选接受才写库，拒绝不落库）

只使用 unittest.TestCase（pytest 未安装，模块级 test_* 会被静默跳过）。
"""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path


def _write_analysis_item(out_dir: Path, **overrides) -> dict:
    """写一条最小 engineering_analysis.json item 并返回它（带稳定 functional_requirement_id）。"""
    item = {
        "analysis_id": "AN-001",
        "functional_requirement_id": "FREQ-test0001",
        "module": "升级",
        "submodule": "升级",
        "description": "表具应支持固件远程升级",
        "requirement": "表具应支持固件远程升级",
        "software_requirement_text": "表具应支持固件远程升级",
        "source_section": "4.1.1",
        "source_requirement_ids": ["AIR-1"],
        "ownership": "software",
        "ownership_source": "rule",
        "labels": ["升级"],
        "notes": [],
        "objective": "实现固件远程升级",
        "behaviors": ["校验镜像签名后写入"],
    }
    item.update(overrides)
    items = []
    payload = {"items": items}
    existing = out_dir / "engineering_analysis.json"
    if existing.exists():
        try:
            payload = json.loads(existing.read_text(encoding="utf-8"))
            items = payload.get("items") or []
        except json.JSONDecodeError:
            items = []
    # 替换或追加
    items = [it for it in items if it.get("functional_requirement_id") != item.get("functional_requirement_id")]
    items.append(item)
    payload["items"] = items
    existing.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return item


def _write_functional_item(out_dir: Path, **overrides) -> dict:
    """写一条 functional_requirements.json item（functional_extract 直抽形态，FRE-* 前缀）。

    与 _write_analysis_item 不同——该产物只落 functional_requirements.json，不进
    engineering_analysis.json。F1 修复前 load_requirement_index 漏读此文件，FRE-* 条目
    的 verification CAS 形同虚设（current_fingerprint 恒 ""）。
    """
    item = {
        "functional_requirement_id": "FRE-direct0001",
        "objective": "表具应支持掉电事件记录",
        "behaviors": ["检测到掉电时写日志"],
        "description": "表具应支持掉电事件记录",
        "source_section": "4.2.1",
        "source_quote": "The meter shall log power failure events.",
        "source_block_ids": ["BLK-001"],
    }
    item.update(overrides)
    target = out_dir / "functional_requirements.json"
    payload = {"producer": "functional-extract-v1", "items": [item]}
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return item


# ===========================================================================
# Cap1/2 纯契约：verification 子对象 + 四态生命周期
# ===========================================================================
class VerificationContractTests(unittest.TestCase):
    def test_default_verification_shape(self) -> None:
        from requirement_schema import (
            CONFIRM_ROLES, IMPLEMENTED_NOT_STARTED, default_verification,
        )
        v = default_verification()
        for role in CONFIRM_ROLES:
            triple = v[f"{role}_confirm"]
            self.assertFalse(triple["confirmed"])
            self.assertEqual(triple["by"], "")
            self.assertEqual(triple["at"], "")
        self.assertEqual(v["implemented"], IMPLEMENTED_NOT_STARTED)
        self.assertEqual(v["test_case_ids"], [])
        self.assertFalse(v["test_completed"])

    def test_normalize_accepts_partial_and_scalar(self) -> None:
        from requirement_schema import normalize_verification
        v = normalize_verification({
            "test_lead_confirm": "已确认",
            "dev_test_confirm": True,
            "implemented": "done",
            "test_case_ids": ["TC-1", "TC-2", "TC-3"],
            "test_completed": "true",
        })
        self.assertTrue(v["test_lead_confirm"]["confirmed"])
        self.assertTrue(v["dev_test_confirm"]["confirmed"])
        self.assertFalse(v["project_manager_confirm"]["confirmed"])
        self.assertEqual(v["test_case_ids"], ["TC-1", "TC-2", "TC-3"])
        self.assertTrue(v["test_completed"])

    def test_excel_columns_and_parse_round_trip(self) -> None:
        from requirement_schema import (
            normalize_verification, parse_verification_columns, verification_excel_columns,
        )
        v = normalize_verification({
            "project_manager_confirm": {"confirmed": True, "by": "张三", "at": "2026-08-06T10:00"},
            "test_lead_confirm": True,
            "dev_test_confirm": True,
            "implemented": "done",
            "test_case_ids": ["TC-1", "TC-2"],
            "test_completed": True,
        })
        cols = verification_excel_columns(v)
        self.assertIn("张三", cols[0])  # 项目负责人确认
        self.assertEqual(cols[3], "已完成")  # 功能是否实现
        self.assertEqual(cols[4], "TC-1; TC-2")  # 测试用例号
        self.assertEqual(cols[5], "是")  # 测试是否完成
        parsed = parse_verification_columns(cols, actor_fallback="评审人")
        self.assertTrue(parsed["project_manager_confirm"]["confirmed"])
        self.assertEqual(parsed["project_manager_confirm"]["by"], "张三")
        self.assertTrue(all(parsed[f"{r}_confirm"]["confirmed"] for r in
                            ("project_manager", "test_lead", "dev_test")))
        self.assertEqual(parsed["implemented"], "done")
        self.assertEqual(parsed["test_case_ids"], ["TC-1", "TC-2"])
        self.assertTrue(parsed["test_completed"])

    def test_derived_lifecycle_four_states(self) -> None:
        from requirement_schema import (
            LIFECYCLE_CONFIRMED, LIFECYCLE_DRAFT, LIFECYCLE_IMPLEMENTED, LIFECYCLE_VERIFIED,
            default_verification, derived_lifecycle, normalize_verification,
        )
        self.assertEqual(derived_lifecycle(default_verification()), LIFECYCLE_DRAFT)
        three_confirms = {f"{r}_confirm": True for r in
                          ("project_manager", "test_lead", "dev_test")}
        self.assertEqual(derived_lifecycle(three_confirms), LIFECYCLE_CONFIRMED)
        self.assertEqual(derived_lifecycle({**three_confirms, "implemented": "done"}),
                         LIFECYCLE_IMPLEMENTED)
        self.assertEqual(derived_lifecycle({**three_confirms, "implemented": "done",
                                            "test_completed": True}), LIFECYCLE_VERIFIED)

    def test_advance_never_auto_downgrades(self) -> None:
        """自动降级不存在：verification 字段回退后，已达到的生命周期态不下落。"""
        from requirement_schema import (
            LIFECYCLE_DRAFT, LIFECYCLE_VERIFIED, advance_lifecycle, default_verification,
        )
        # 已 verified，verification 清空 → 仍保留 verified（不自动降级）
        self.assertEqual(advance_lifecycle(LIFECYCLE_VERIFIED, default_verification()),
                         LIFECYCLE_VERIFIED)
        # draft 起，verification 推到顶 → 前进到 verified
        full = {"project_manager_confirm": True, "test_lead_confirm": True,
                "dev_test_confirm": True, "implemented": "done", "test_completed": True}
        self.assertEqual(advance_lifecycle(LIFECYCLE_DRAFT, full), LIFECYCLE_VERIFIED)


# ===========================================================================
# Cap1/2 I/O：verification 覆盖 + 状态机 + CAS（锁 + 原子替换）
# ===========================================================================
class VerificationStateMachineIOTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.out = Path(self._tmp.name)
        self.item = _write_analysis_item(self.out)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def _override(self, **verification_kw) -> dict:
        from requirements_analysis_rules import apply_verification_override
        return apply_verification_override(self.out, self.item["functional_requirement_id"],
                                           verification_kw, actor="tester")

    def test_forward_migration_driven_by_verification(self) -> None:
        from requirement_schema import (
            LIFECYCLE_CONFIRMED, LIFECYCLE_DRAFT, LIFECYCLE_IMPLEMENTED, LIFECYCLE_VERIFIED,
        )
        self.assertEqual(self._override(project_manager_confirm=True, test_lead_confirm=True,
                                        dev_test_confirm=True)["lifecycle_state"],
                         LIFECYCLE_CONFIRMED)
        self.assertEqual(self._override(project_manager_confirm=True, test_lead_confirm=True,
                                        dev_test_confirm=True, implemented="done")["lifecycle_state"],
                         LIFECYCLE_IMPLEMENTED)
        rec = self._override(project_manager_confirm=True, test_lead_confirm=True,
                             dev_test_confirm=True, implemented="done", test_completed=True)
        self.assertEqual(rec["lifecycle_state"], LIFECYCLE_VERIFIED)

    def test_cas_rejects_on_content_drift(self) -> None:
        from requirements_analysis_rules import apply_verification_override
        from review_state import VerificationStateConflict
        with self.assertRaises(VerificationStateConflict):
            apply_verification_override(self.out, self.item["functional_requirement_id"],
                                        {"implemented": "done"}, actor="t",
                                        expected_evidence_fingerprint="stale-fingerprint")

    def test_cas_accepts_when_fingerprint_matches(self) -> None:
        from requirement_schema import requirement_structural_fingerprint
        from requirements_analysis_rules import apply_verification_override
        fp = requirement_structural_fingerprint(self.item)
        rec = apply_verification_override(self.out, self.item["functional_requirement_id"],
                                          {"implemented": "done"}, actor="t",
                                          expected_evidence_fingerprint=fp)
        self.assertEqual(rec["lifecycle_state"], "draft")  # 无三确认→仍 draft

    def test_rollback_manual_only_and_logs_event(self) -> None:
        from requirement_schema import LIFECYCLE_DRAFT
        from requirements_analysis_rules import (
            current_lifecycle, rollback_requirement_lifecycle,
        )
        from review_state import read_lifecycle_events
        # 先前进到 verified
        self._override(project_manager_confirm=True, test_lead_confirm=True,
                       dev_test_confirm=True, implemented="done", test_completed=True)
        self.assertEqual(current_lifecycle(self.out, self.item["functional_requirement_id"]), "verified")
        # 人工回退到 implemented
        rollback_requirement_lifecycle(self.out, self.item["functional_requirement_id"],
                                       "implemented", actor="专家", reason="测试发现缺陷需返工")
        self.assertEqual(current_lifecycle(self.out, self.item["functional_requirement_id"]), "implemented")
        # 回退事件 append-only 留痕
        events = read_lifecycle_events(self.out)
        rollback_events = [e for e in events if e.get("kind") == "rollback"]
        self.assertEqual(len(rollback_events), 1)
        self.assertEqual(rollback_events[0]["from_state"], "verified")
        self.assertEqual(rollback_events[0]["to_state"], "implemented")
        self.assertEqual(rollback_events[0]["actor"], "专家")
        # 回退到 draft 后再回退非法（必须低于当前态）
        rollback_requirement_lifecycle(self.out, self.item["functional_requirement_id"],
                                       LIFECYCLE_DRAFT, actor="专家", reason="回到起点")
        with self.assertRaises(ValueError):
            rollback_requirement_lifecycle(self.out, self.item["functional_requirement_id"],
                                           "draft", actor="专家", reason="不是回退")

    def test_forward_migration_appends_advance_events(self) -> None:
        """S1-10a：前进迁移同样 append 事件——与回退事件同构同流。

        draft→confirmed→implemented→verified 三次升态写各产一条 advance 事件，
        from_state/to_state 完整覆盖四个生命周期态；事件携带 trigger=verification-driven
        （与回退事件同构）。重复保存（无升态）不追加事件——append-only 流不被重复保存污染。
        """
        from review_state import read_lifecycle_events

        rid = self.item["functional_requirement_id"]
        # draft → confirmed（三确认位全部确认）
        self._override(project_manager_confirm=True, test_lead_confirm=True, dev_test_confirm=True)
        # confirmed → implemented
        self._override(project_manager_confirm=True, test_lead_confirm=True,
                       dev_test_confirm=True, implemented="done")
        # implemented → verified
        self._override(project_manager_confirm=True, test_lead_confirm=True,
                       dev_test_confirm=True, implemented="done", test_completed=True)

        events = read_lifecycle_events(self.out)
        advance = [e for e in events if e.get("kind") == "advance"]
        # 三次升态 → 三条 advance 事件，from/to 序列精确还原四态链
        self.assertEqual([(e.get("from_state"), e.get("to_state")) for e in advance],
                         [("draft", "confirmed"), ("confirmed", "implemented"),
                          ("implemented", "verified")])
        # 与回退事件同构：含 trigger 字段，标记为 verification 驱动
        self.assertTrue(all(e.get("trigger") == "verification-driven" for e in advance))
        self.assertTrue(all(e.get("actor") == "tester" for e in advance))

        # 重复保存（已在 verified，无升态）不追加事件——避免 append-only 流被重复保存污染
        self._override(project_manager_confirm=True, test_lead_confirm=True,
                       dev_test_confirm=True, implemented="done", test_completed=True)
        advance_after_resave = [e for e in read_lifecycle_events(self.out)
                                if e.get("kind") == "advance"]
        self.assertEqual(len(advance_after_resave), 3)

    def test_rollback_event_carries_trigger_field(self) -> None:
        """S1-10a：回退事件与前进事件同构——同样携带 trigger 字段（manual）。"""
        from requirements_analysis_rules import (
            apply_verification_override, rollback_requirement_lifecycle,
        )
        from review_state import read_lifecycle_events

        rid = self.item["functional_requirement_id"]
        apply_verification_override(self.out, rid, {
            "project_manager_confirm": True, "test_lead_confirm": True, "dev_test_confirm": True,
            "implemented": "done", "test_completed": True}, actor="t")
        rollback_requirement_lifecycle(self.out, rid, "implemented", actor="专家", reason="返工")
        rollback_events = [e for e in read_lifecycle_events(self.out)
                           if e.get("kind") == "rollback"]
        self.assertEqual(rollback_events[0].get("trigger"), "manual")

    def test_overlay_persisted_and_readable(self) -> None:
        from requirement_schema import normalize_verification
        from review_state import read_verification_states
        self._override(implemented="done", test_completed=True)
        states = read_verification_states(self.out)
        record = states[self.item["functional_requirement_id"]]
        self.assertEqual(record["source"], "reviewer_override")
        self.assertEqual(normalize_verification(record["verification"])["implemented"], "done")


# ===========================================================================
# F1：functional_requirements.json 纳入 requirement 索引——FRE-* 直抽条目 CAS 真实生效
# ---------------------------------------------------------------------------
# 旧 load_requirement_index 只读 engineering_analysis.json + manual_requirements.jsonl，
# functional_extract 直抽的 FRE-* 条目只落 functional_requirements.json，索引漏读使其
# current_fingerprint 恒 ""，任何 expected_evidence_fingerprint 都“匹配”→ CAS 形同虚设。
# ===========================================================================
class FunctionalRequirementIndexCasTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.out = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_fre_narrative_drift_keeps_confirmation_with_review_hint(self) -> None:
        """T3-2：纯叙述字段变化（objective/behaviors/description）→ 确认保留 + 复核提示，
        不吊销。结构指纹不变，故 CAS 闸放行；叙述指纹变化 → ``narrative_drift_hint=True``。"""
        from requirement_schema import (
            requirement_narrative_fingerprint, requirement_structural_fingerprint,
        )
        from requirements_analysis_rules import apply_verification_override

        item = _write_functional_item(self.out)
        rid = item["functional_requirement_id"]
        struct_fp = requirement_structural_fingerprint(item)

        # 首次回写：结构指纹一致 → 通过，建立叙述基线
        rec = apply_verification_override(self.out, rid, {"implemented": "done"},
                                          actor="tester",
                                          expected_evidence_fingerprint=struct_fp)
        self.assertEqual(rec["evidence_fingerprint"], struct_fp)
        self.assertFalse(rec["narrative_drift_hint"])

        # 需求重新生成：只改叙述（objective/behaviors/description）——结构指纹不变
        narr_mutated = _write_functional_item(
            self.out, objective="表具应支持全新不同的掉电记录能力",
            behaviors=["全新行为"], description="全新的掉电记录能力描述")
        self.assertEqual(requirement_structural_fingerprint(narr_mutated), struct_fp)
        self.assertNotEqual(
            requirement_narrative_fingerprint(narr_mutated),
            requirement_narrative_fingerprint(item),
        )

        # 二次回写携结构指纹（客户端 round-trip 它）→ 通过（不吊销），且带复核提示
        rec2 = apply_verification_override(self.out, rid, {"implemented": "done"},
                                           actor="tester",
                                           expected_evidence_fingerprint=struct_fp)
        self.assertEqual(rec2["evidence_fingerprint"], struct_fp)
        self.assertTrue(rec2["narrative_drift_hint"])

    def test_fre_structural_drift_revokes_confirmation(self) -> None:
        """T3-2：结构字段变化（模块/归属/OBIS/来源 block_ids）→ 结构指纹漂移 → CAS 吊销。"""
        from requirement_schema import requirement_structural_fingerprint
        from requirements_analysis_rules import apply_verification_override
        from review_state import VerificationStateConflict

        item = _write_functional_item(self.out)
        rid = item["functional_requirement_id"]
        struct_fp = requirement_structural_fingerprint(item)
        apply_verification_override(self.out, rid, {"implemented": "done"}, actor="tester",
                                    expected_evidence_fingerprint=struct_fp)

        # 结构漂移：改模块 + 来源 block_ids（物质性变化）
        struct_mutated = _write_functional_item(
            self.out, module="事件记录", source_block_ids=["BLK-999"],
            source_quote="The meter shall log events. OBIS 1-1:1.1.1",
            related_dlms_objects=["1-1:1.1.1"],
        )
        self.assertNotEqual(requirement_structural_fingerprint(struct_mutated), struct_fp)

        # 携旧结构指纹回写 → 必须被拒（吊销转人工）
        with self.assertRaises(VerificationStateConflict) as ctx:
            apply_verification_override(self.out, rid, {"implemented": "done"}, actor="tester",
                                        expected_evidence_fingerprint=struct_fp)
        self.assertEqual(ctx.exception.requirement_id, rid)
        self.assertEqual(ctx.exception.current_fingerprint,
                         requirement_structural_fingerprint(struct_mutated))
        self.assertNotEqual(ctx.exception.current_fingerprint, "")

    def test_fre_item_indexed_from_governed_pipeline_path(self) -> None:
        """package_v1 下 functional_requirements.json 在 .ratomizer/pipeline/ 也能索引到。"""
        from result_package import governed_artifact_path, initialize_result_package
        from requirements_analysis_rules import load_requirement_index

        # 真实 package_v1 marker：governed_artifact_path 此时解析到 .ratomizer/pipeline/
        source = self.out / "input.docx"
        source.write_bytes(b"fixture")
        initialize_result_package(self.out, input_path=source, requested_stages=["atomize"])
        governed = governed_artifact_path(self.out, "functional_requirements.json",
                                          category="pipeline", for_write=True)
        self.assertEqual(governed, self.out / ".ratomizer" / "pipeline" / "functional_requirements.json")
        governed.write_text(json.dumps({
            "producer": "functional-extract-v1",
            "items": [{"functional_requirement_id": "FRE-pkg0001",
                       "objective": "应支持时钟同步", "behaviors": ["周期校时"]}],
        }, ensure_ascii=False), encoding="utf-8")
        index = load_requirement_index(self.out)
        self.assertIn("FRE-pkg0001", index)
        self.assertTrue(index["FRE-pkg0001"]["fingerprint"])


# ===========================================================================
# Cap1：六列回写（导出侧读 verification）+ 回灌闭环 + 追溯ID
# ===========================================================================
class ExcelWritebackTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.out = Path(self._tmp.name)
        self.item = _write_analysis_item(self.out)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def _export(self) -> Path:
        from requirements_analysis_excel import write_software_requirements_xlsx
        path = self.out / "software_requirements.xlsx"
        return write_software_requirements_xlsx([dict(self.item)], path)

    def test_six_columns_empty_without_overlay(self) -> None:
        from openpyxl import load_workbook
        path = self._export()
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        header = [c.value for c in ws[1]]
        col = {name: idx for idx, name in enumerate(header)}
        row = [c.value for c in ws[2]]
        # 六列全空/未开始（openpyxl 把空串存为 None）
        self.assertIn(row[col["项目负责人确认"]], ("", None))
        self.assertEqual(row[col["功能是否实现"]], "未开始")
        self.assertIn(row[col["测试是否完成"]], ("", None))
        # 追溯ID 落进 notes
        self.assertIn("需求追溯ID：FREQ-test0001", row[col["说明、示例、注意事项"]])

    def test_six_columns_populated_from_overlay(self) -> None:
        from openpyxl import load_workbook
        from requirements_analysis_rules import apply_verification_override
        apply_verification_override(self.out, "FREQ-test0001", {
            "project_manager_confirm": {"confirmed": True, "by": "李四", "at": "2026-08-06"},
            "test_lead_confirm": True, "dev_test_confirm": True,
            "implemented": "done", "test_case_ids": ["TC-9"], "test_completed": True,
        }, actor="李四")
        path = self._export()
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        header = [c.value for c in ws[1]]
        col = {name: idx for idx, name in enumerate(header)}
        row = [c.value for c in ws[2]]
        self.assertIn("李四", row[col["项目负责人确认"]])
        self.assertEqual(row[col["功能是否实现"]], "已完成")
        self.assertEqual(row[col["测试用例号"]], "TC-9")
        self.assertEqual(row[col["测试是否完成"]], "是")

    def test_export_does_not_mutate_input_items(self) -> None:
        """verification/_trace_id 不能泄进 engineering_analysis.json 的 item dict。"""
        self._export()
        payload = json.loads((self.out / "engineering_analysis.json").read_text(encoding="utf-8"))
        item = payload["items"][0]
        self.assertNotIn("verification", item)
        self.assertNotIn("_trace_id", item)

    def test_backfill_round_trip(self) -> None:
        """导出→线下改六列→导回→状态可追踪。"""
        from openpyxl import load_workbook
        from desktop_tasks import import_verification_workbook_task
        from review_state import read_verification_states
        path = self._export()
        # 线下编辑六列
        wb = load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        col = {name: idx for idx, name in enumerate(header)}
        row_idx = 2
        ws.cell(row=row_idx, column=col["项目负责人确认"] + 1, value="已确认（王五）")
        ws.cell(row=row_idx, column=col["测试负责人确认"] + 1, value="是")
        ws.cell(row=row_idx, column=col["研发测试确认"] + 1, value="是")
        ws.cell(row=row_idx, column=col["功能是否实现"] + 1, value="已完成")
        ws.cell(row=row_idx, column=col["测试用例号"] + 1, value="TC-1; TC-2")
        ws.cell(row=row_idx, column=col["测试是否完成"] + 1, value="是")
        wb.save(path)
        wb.close()
        result = import_verification_workbook_task(self.out, path, actor="王五")
        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["stale"], 0)
        record = read_verification_states(self.out)["FREQ-test0001"]
        self.assertTrue(record["verification"]["project_manager_confirm"]["confirmed"])
        self.assertEqual(record["verification"]["implemented"], "done")
        self.assertEqual(record["verification"]["test_case_ids"], ["TC-1", "TC-2"])
        self.assertTrue(record["verification"]["test_completed"])

    def test_backfill_narrative_drift_tolerated_with_review_hint(self) -> None:
        """T3-2：纯叙述变化（description）不再拒回灌——结构列（子模块+客户需求章节）匹配即
        合入，描述变化进 ``narrative_review`` 复核提示（状态不吊销）。"""
        from openpyxl import load_workbook
        from desktop_tasks import import_verification_workbook_task
        path = self._export()
        wb = load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        col = {name: idx for idx, name in enumerate(header)}
        ws.cell(row=2, column=col["功能是否实现"] + 1, value="已完成")
        wb.save(path)
        wb.close()
        # 模拟需求重新生成：只改 description/requirement/soft_text（叙述）——结构列不变
        _write_analysis_item(self.out, description="表具应支持全新不同的固件升级能力",
                             requirement="表具应支持全新不同的固件升级能力",
                             software_requirement_text="表具应支持全新不同的固件升级能力")
        result = import_verification_workbook_task(self.out, path, actor="x")
        self.assertEqual(result["imported"], 1)   # 结构匹配 → 回灌（状态保留）
        self.assertEqual(result["stale"], 0)      # 叙述变化不再拒
        narrative_review = result.get("narrative_review") or []
        self.assertEqual(len(narrative_review), 1)
        self.assertEqual(narrative_review[0]["requirement_id"], "FREQ-test0001")

    def test_backfill_structural_drift_rejects_regenerated_content(self) -> None:
        """T3-2：结构列（子模块/客户需求章节）漂移→回灌 CAS 失配→转人工（不自动合入）。"""
        from openpyxl import load_workbook
        from desktop_tasks import import_verification_workbook_task
        path = self._export()
        wb = load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        col = {name: idx for idx, name in enumerate(header)}
        ws.cell(row=2, column=col["功能是否实现"] + 1, value="已完成")
        wb.save(path)
        wb.close()
        # 模拟需求重新生成：改 module/submodule + source_section（结构漂移）
        _write_analysis_item(self.out, module="事件记录", submodule="事件记录",
                             source_section="9.9.9")
        result = import_verification_workbook_task(self.out, path, actor="x")
        self.assertEqual(result["imported"], 0)
        self.assertGreater(result["stale"], 0)

    def test_backfill_tolerates_clarify_mark_software_text(self) -> None:
        """CAS 只比对稳定身份列，不比渲染列：software_requirement_text=待澄清 不应造成 false-stale。"""
        from openpyxl import load_workbook
        from desktop_tasks import import_verification_workbook_task
        from requirement_schema import requirement_content_fingerprint
        # 待澄清的 software_requirement_text，但 submodule/description/source_section 稳定
        self.item["software_requirement_text"] = "待澄清"
        path = self._export()
        wb = load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        col = {name: idx for idx, name in enumerate(header)}
        ws.cell(row=2, column=col["功能是否实现"] + 1, value="进行中")
        wb.save(path)
        wb.close()
        result = import_verification_workbook_task(self.out, path, actor="x")
        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["stale"], 0)

    def test_backfill_rejection_list_names_drifting_rows(self) -> None:
        """S1-10b：CAS 拒绝清单精确到 requirement_id + xlsx 行号 + 原因。

        T3-2：构造两行**结构**漂移（子模块/客户需求章节已变化）回灌：报告 rejected 列表含
        两条，各自带 rid、行号、原因；imported==0、stale==2。行号与 rid 一一对应可定位到
        具体物理行。（叙述漂移不再拒——见 test_backfill_narrative_drift_tolerated_with_review_hint。）
        """
        from openpyxl import load_workbook
        from desktop_tasks import import_verification_workbook_task
        from requirements_analysis_excel import write_software_requirements_xlsx

        # 两条需求（独立 functional_requirement_id）
        item_a = self.item
        item_b = _write_analysis_item(
            self.out, analysis_id="AN-002", functional_requirement_id="FREQ-test0002",
            module="事件记录", submodule="事件记录",
            description="表具应记录掉电事件", requirement="表具应记录掉电事件",
            software_requirement_text="表具应记录掉电事件", source_section="4.2.1",
            source_requirement_ids=["AIR-2"], labels=["事件记录"],
            objective="记录掉电事件", behaviors=["写入事件日志"],
        )
        path = self.out / "software_requirements.xlsx"
        write_software_requirements_xlsx([dict(item_a), dict(item_b)], path)

        # 线下编辑两行的六列
        wb = load_workbook(path)
        import re as _re
        trace_re = _re.compile(r"需求追溯ID[：:]\s*([^\n\r]+)")
        # 导出按模块分 sheet——两条在不同 sheet，逐 sheet 扫描建 rid→(row, sheet) 映射
        rid_to_loc: dict[str, tuple[int, str]] = {}
        cell_by_loc: dict[str, tuple[Any, int]] = {}
        for sheet in wb.worksheets:
            header = [c.value for c in sheet[1]]
            if "项目负责人确认" not in header:
                continue
            col = {name: idx for idx, name in enumerate(header)}
            impl_col = col["功能是否实现"] + 1
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                notes = str(row[col["说明、示例、注意事项"]] or "")
                m = trace_re.search(notes)
                if not m:
                    continue
                rid = m.group(1).strip()
                rid_to_loc[rid] = (row_number, sheet.title)
                cell_by_loc[rid] = (sheet, impl_col)
        # 两行都填上「已完成」
        for rid, (sheet, impl_col) in cell_by_loc.items():
            _row_number = rid_to_loc[rid][0]
            sheet.cell(row=_row_number, column=impl_col, value="已完成")
        wb.save(path)
        wb.close()

        # 模拟需求重新生成：两条都漂移**结构列**（子模块 + 客户需求章节）
        _write_analysis_item(self.out, module="升级V2", submodule="升级V2",
                             source_section="4.1.1X")
        _write_analysis_item(self.out, analysis_id="AN-002",
                             functional_requirement_id="FREQ-test0002",
                             module="事件记录V2", submodule="事件记录V2",
                             source_section="4.2.1X")

        result = import_verification_workbook_task(self.out, path, actor="x")
        # 两行均被 CAS 拒绝
        self.assertEqual(result["imported"], 0)
        self.assertEqual(result["stale"], 2)
        rejected = result.get("rejected") or []
        self.assertEqual(len(rejected), 2)
        # 每条含 rid + 行号 + sheet + 原因，且 (row, sheet) 与导出时一一对应
        by_rid = {row["requirement_id"]: row for row in rejected}
        self.assertEqual(set(by_rid), {"FREQ-test0001", "FREQ-test0002"})
        for rid, row in by_rid.items():
            expected_row, expected_sheet = rid_to_loc[rid]
            self.assertEqual(row["row"], expected_row)
            self.assertEqual(row["sheet"], expected_sheet)
            self.assertTrue(str(row["reason"]).strip())
        # 精确到行的证明：两条 (row, sheet) 各不相同（两条物理行可区分定位）
        locs = {(r["row"], r["sheet"]) for r in rejected}
        self.assertEqual(len(locs), 2)


# ===========================================================================
# Cap4：手工建需求入口（manual provenance，追溯列留空不伪引）
# ===========================================================================
class ManualEntryTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.out = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_manual_requirement_empty_traceability(self) -> None:
        from requirement_schema import PROVENANCE_MANUAL, build_manual_requirement
        record = build_manual_requirement(objective="表具应支持本地显示切换",
                                          behaviors=["按键切换显示页"], module="显示", actor="工程师")
        self.assertEqual(record["source_kind"], PROVENANCE_MANUAL)
        self.assertEqual(record["source_quote"], "")  # 留空不伪引
        self.assertEqual(record["source_section"], "")
        self.assertEqual(record["source_block_ids"], [])
        self.assertTrue(record["functional_requirement_id"].startswith("FREQ-MANUAL-"))

    def test_manual_requirement_full_pipeline_to_export(self) -> None:
        from desktop_tasks import add_manual_requirement_task
        from requirements_analysis_excel import write_software_requirements_xlsx
        from review_state import read_manual_requirements
        add_manual_requirement_task(self.out, objective="表具应支持事件主动上报",
                                    behaviors=["掉电事件上报"], module="事件记录", actor="工程师")
        manual = read_manual_requirements(self.out)
        self.assertEqual(len(manual), 1)
        # 导出：手工条目出现，追溯列（客户需求章节）为空，notes 标手工来源
        path = write_software_requirements_xlsx([], self.out / "software_requirements.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        found = False
        for sheet in wb.worksheets:
            header = [c.value for c in sheet[1]]
            if "项目负责人确认" not in header:
                continue
            col = {name: idx for idx, name in enumerate(header)}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                notes = str(row[col["说明、示例、注意事项"]] or "")
                if "手工录入" in notes and "事件主动上报" in str(row[col["描述"]] or ""):
                    found = True
                    self.assertIn(row[col["客户需求章节"]], ("", None))  # 追溯列明示为空
                    self.assertIn(row[col["是否客户需求"]], ("", None))  # 手工条目标空
                    self.assertIn("需求追溯ID：", notes)
        self.assertTrue(found, "手工条目未出现在导出 xlsx 中")
        # 手工条目同样可写 verification（走相同状态机下游）
        from requirements_analysis_rules import apply_verification_override
        rid = manual[0]["functional_requirement_id"]
        rec = apply_verification_override(self.out, rid, {"implemented": "done"}, actor="x")
        self.assertEqual(rec["lifecycle_state"], "draft")  # 无三确认仍 draft


# ===========================================================================
# Cap3：弱词扫描器进澄清报告
# ===========================================================================
class WeakWordScannerTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.out = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def _seed_functional(self, *, objective: str, acceptance=None) -> None:
        payload = {"producer": "functional-synthesis-v8", "items": [{
            "functional_requirement_id": "FREQ-weak1",
            "title": objective[:40],
            "objective": objective,
            "behaviors": [],
            "description": objective,
            "source_section": "5.1",
            "source_quote": "",
            "acceptance_criteria": acceptance or [],
        }]}
        (self.out / "functional_requirements.json").write_text(
            json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    def test_weak_word_hits_merge_into_ambiguous_category(self) -> None:
        from clarification_report import CAT_AMBIGUOUS, _weakness_scan_entries
        self._seed_functional(objective="系统应适当处理告警并尽快恢复")
        entries = _weakness_scan_entries(self.out, json.loads(
            (self.out / "functional_requirements.json").read_text(encoding="utf-8"))["items"])
        weak_entries = [e for e in entries if e["signal"] == "weakness:vague_word"]
        self.assertTrue(weak_entries)
        self.assertEqual(weak_entries[0]["category"], CAT_AMBIGUOUS)
        self.assertEqual(weak_entries[0]["audience"], "内部核对")

    def test_untestable_acceptance_merges_into_missing_category(self) -> None:
        from clarification_report import CAT_MISSING, _weakness_scan_entries
        self._seed_functional(objective="系统应记录事件",
                              acceptance=["事件处理应符合要求"])
        entries = _weakness_scan_entries(self.out, json.loads(
            (self.out / "functional_requirements.json").read_text(encoding="utf-8"))["items"])
        untestable = [e for e in entries if e["signal"] == "weakness:untestable"]
        self.assertTrue(untestable)
        self.assertEqual(untestable[0]["category"], CAT_MISSING)

    def test_readiness_flips_with_weak_word_entry(self) -> None:
        """就绪判定同步变化：含不可测验收的功能需求使 unresolved_blocking>0 → NEEDS WORK。"""
        from clarification_report import unresolved_hard_questions
        # 需要一份 ai_requirements.jsonl 让 collect_questions 能跑（弱词来自 functional_requirements）
        (self.out / "ai_requirements.jsonl").write_text(
            json.dumps({"ai_req_id": "AIR-1", "title": "x", "source_section": "5.1"}) + "\n",
            encoding="utf-8")
        self._seed_functional(objective="系统应记录事件",
                              acceptance=["功能应正常工作"])
        unresolved, counts = unresolved_hard_questions(self.out)
        signals = {e.get("signal") for e in unresolved}
        self.assertIn("weakness:untestable", signals)
        self.assertGreater(counts["blocking"], 0)


# ===========================================================================
# Cap5：需求库词面检索
# ===========================================================================
class RequirementLibraryTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.out = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_build_and_search_word_surface(self) -> None:
        from desktop_tasks import build_requirement_library_task, search_requirements_task
        with tempfile.TemporaryDirectory() as proj:
            proj = Path(proj)
            (proj / "functional_requirements.json").write_text(json.dumps({
                "source": "DocA",
                "provenance": {"generated_at": "2026-08-06"},
                "items": [
                    {"functional_requirement_id": "FREQ-a", "objective": "collect load profile",
                     "behaviors": ["archive 15 min interval"], "module": "曲线"},
                    {"functional_requirement_id": "FREQ-b", "objective": "manage security keys",
                     "behaviors": ["rotate cipher keys"], "module": "安全"},
                ],
            }, ensure_ascii=False), encoding="utf-8")
            library = self.out / "library.jsonl"
            # S1-10d：默认质量门只收 lifecycle>=confirmed；此用例验证检索机制，显式收录未确认
            build = build_requirement_library_task([proj], library, include_unconfirmed=True)
            self.assertEqual(build["entries"], 2)
        results = search_requirements_task(library, "load profile archive", limit=5)
        self.assertGreater(results["matches"], 0)
        self.assertEqual(results["results"][0]["functional_requirement_id"], "FREQ-a")

    def test_build_library_quality_gate_excludes_draft_by_default(self) -> None:
        """S1-10d：默认仅收录 lifecycle>=confirmed 的条目；draft 不入库。"""
        from desktop_tasks import build_requirement_library_task
        from requirements_analysis_rules import apply_verification_override
        with tempfile.TemporaryDirectory() as proj:
            proj = Path(proj)
            (proj / "functional_requirements.json").write_text(json.dumps({
                "source": "DocA",
                "items": [
                    {"functional_requirement_id": "FREQ-draft", "objective": "草稿需求",
                     "behaviors": ["x"], "module": "X"},
                    {"functional_requirement_id": "FREQ-conf", "objective": "已确认需求",
                     "behaviors": ["y"], "module": "Y"},
                    {"functional_requirement_id": "FREQ-impl", "objective": "已实现需求",
                     "behaviors": ["z"], "module": "Z"},
                ],
            }, ensure_ascii=False), encoding="utf-8")
            # 仅把 FREQ-conf 推到 confirmed、FREQ-impl 推到 implemented；FREQ-draft 留 draft
            apply_verification_override(proj, "FREQ-conf", {
                "project_manager_confirm": True, "test_lead_confirm": True,
                "dev_test_confirm": True}, actor="t")
            apply_verification_override(proj, "FREQ-impl", {
                "project_manager_confirm": True, "test_lead_confirm": True,
                "dev_test_confirm": True, "implemented": "done"}, actor="t")
            library = self.out / "library.jsonl"
            build = build_requirement_library_task([proj], library)
            # 默认门：draft 被排除，confirmed/implemented 收录
            self.assertEqual(build["entries"], 2)
            self.assertEqual(build.get("skipped_unconfirmed"), 1)
            entries = [json.loads(line) for line in library.read_text(encoding="utf-8").splitlines() if line.strip()]
            rids = {e["functional_requirement_id"] for e in entries}
            self.assertEqual(rids, {"FREQ-conf", "FREQ-impl"})
            self.assertNotIn("FREQ-draft", rids)
            # 收录条目携带 lifecycle_state，供采纳 UI 默认隐藏未确认
            by_rid = {e["functional_requirement_id"]: e for e in entries}
            self.assertEqual(by_rid["FREQ-conf"]["lifecycle_state"], "confirmed")
            self.assertEqual(by_rid["FREQ-impl"]["lifecycle_state"], "implemented")

    def test_build_library_include_unconfirmed_switch_admits_draft(self) -> None:
        """S1-10d：--include-unconfirmed 显式开关收录 draft 条目。"""
        from desktop_tasks import build_requirement_library_task
        with tempfile.TemporaryDirectory() as proj:
            proj = Path(proj)
            (proj / "functional_requirements.json").write_text(json.dumps({
                "source": "DocA",
                "items": [
                    {"functional_requirement_id": "FREQ-draft", "objective": "草稿需求",
                     "behaviors": ["x"], "module": "X"},
                ],
            }, ensure_ascii=False), encoding="utf-8")
            library = self.out / "library.jsonl"
            build = build_requirement_library_task([proj], library, include_unconfirmed=True)
            self.assertEqual(build["entries"], 1)
            entries = [json.loads(line) for line in library.read_text(encoding="utf-8").splitlines() if line.strip()]
            self.assertEqual(entries[0]["functional_requirement_id"], "FREQ-draft")
            self.assertEqual(entries[0]["lifecycle_state"], "draft")

    def test_corrected_ownership_ranked_first_on_tie(self) -> None:
        from requirement_schema import (
            library_entry_from_requirement, search_requirement_library,
        )
        library = [
            library_entry_from_requirement(
                {"functional_requirement_id": "F1", "objective": "event log record",
                 "behaviors": [], "module": "事件"}, project="P", doc_source="d", created_at=""),
            library_entry_from_requirement(
                {"functional_requirement_id": "F2", "objective": "event log record",
                 "behaviors": [], "module": "事件", "ownership_override": "事件记录"},
                project="P", doc_source="d", created_at=""),
        ]
        # 强制相同 token 集合 → 同分；F2 被 reviewer_override 修正过应靠前
        for entry in library:
            entry["tokens"] = sorted({"event", "log", "record"})
        results = search_requirement_library("event log record", library)
        self.assertEqual(results[0]["functional_requirement_id"], "F2")
        self.assertTrue(results[0]["ownership_corrected"])


# ===========================================================================
# Cap6：dependencies/parent/children 半自动推荐
# ===========================================================================
class DependencyRecommendTests(unittest.TestCase):
    def test_candidates_depend_exclude_refine(self) -> None:
        from requirement_schema import (
            DEPENDENCY_DEPEND, DEPENDENCY_EXCLUDE, DEPENDENCY_REFINE,
            recommend_dependency_candidates,
        )
        reqs = [
            {"functional_requirement_id": "F1", "title": "镜像升级", "source_section": "4.1.1",
             "related_dlms_objects": ["0-0:96.3.0"]},
            {"functional_requirement_id": "F2", "title": "镜像校验更具体一些", "source_section": "4.1.2",
             "related_dlms_objects": ["0-0:96.3.0"]},
        ]
        kinds = {c["kind"] for c in recommend_dependency_candidates(reqs)}
        self.assertIn(DEPENDENCY_DEPEND, kinds)
        self.assertIn(DEPENDENCY_EXCLUDE, kinds)
        self.assertIn(DEPENDENCY_REFINE, kinds)

    def test_accept_lands_edge_reject_keeps_record(self) -> None:
        """T3-1 RTM 边：accept 落边（物化库 + 事件流）；reject 留记录（只事件流，不落物化库）。

        旧「拒绝不落库=拒绝无声消失」升格为「拒绝进 append-only 事件流」，回放可重建裁决史。
        """
        from desktop_tasks import decide_dependency_task
        from review_state import (
            read_dependency_decisions, read_rtm_edge_events, replay_rtm_edges,
        )
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            accepted = decide_dependency_task(out, frm="F1", to="F2", kind="depend",
                                              accepted=True, actor="专家", reason="升级依赖校验")
            # accept：物化库 + 事件流都写
            self.assertIn("dependency_decisions.jsonl", accepted["written"])
            self.assertIn("requirement_rtm_edges.jsonl", accepted["written"])
            self.assertEqual(len(read_dependency_decisions(out)), 1)

            rejected = decide_dependency_task(out, frm="F1", to="F2", kind="exclude",
                                              accepted=False, actor="专家", reason="非互斥")
            # reject：只写事件流（留记录），不落物化库
            self.assertIn("requirement_rtm_edges.jsonl", rejected["written"])
            self.assertNotIn("dependency_decisions.jsonl", rejected["written"])
            self.assertEqual(len(read_dependency_decisions(out)), 1)  # 物化库仍只 accept 那条

            # 事件流含 accept + reject 两条；回放重建 accepted=1 / rejected=1
            events = read_rtm_edge_events(out)
            self.assertEqual(len(events), 2)
            replay = replay_rtm_edges(events)
            self.assertEqual(replay["accepted_count"], 1)
            self.assertEqual(replay["rejected_count"], 1)
            # 边带 kind/from/to/decision/actor/recorded_at
            edge = next(e for e in replay["edges"].values() if e["decision"] == "reject")
            self.assertEqual(edge["kind"], "exclude")
            self.assertEqual(edge["actor"], "专家")

    def test_rtm_edge_replay_last_decision_wins_and_is_idempotent(self) -> None:
        """T3-1：同一 edge accept→reject→accept 序列回放取末尾；两次回放逐字节一致。"""
        from review_state import replay_rtm_edges
        events = [
            {"edge_id": "E1", "kind": "depend", "from": "F1", "to": "F2",
             "decision": "accept", "actor": "a", "reason": "", "recorded_at": "t1"},
            {"edge_id": "E1", "kind": "depend", "from": "F1", "to": "F2",
             "decision": "reject", "actor": "a", "reason": "重审", "recorded_at": "t2"},
            {"edge_id": "E1", "kind": "depend", "from": "F1", "to": "F2",
             "decision": "accept", "actor": "a", "reason": "再确认", "recorded_at": "t3"},
        ]
        replay1 = replay_rtm_edges(events)
        replay2 = replay_rtm_edges(events)
        self.assertEqual(replay1, replay2)  # 幂等
        self.assertEqual(replay1["accepted_count"], 1)
        self.assertEqual(replay1["rejected_count"], 0)
        self.assertEqual(replay1["edges"]["E1"]["reason"], "再确认")  # 末尾胜出

    def test_project_candidates_read_functional_requirements(self) -> None:
        from requirements_analysis_rules import dependency_candidates_for_project
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            (out / "functional_requirements.json").write_text(json.dumps({
                "producer": "functional-synthesis-v8",
                "items": [
                    {"functional_requirement_id": "F1", "title": "A", "source_section": "4.1.1",
                     "related_dlms_objects": ["0-0:96.3.0"]},
                    {"functional_requirement_id": "F2", "title": "B", "source_section": "4.1.2",
                     "related_dlms_objects": ["0-0:96.3.3"]},
                ],
            }, ensure_ascii=False), encoding="utf-8")
            candidates = dependency_candidates_for_project(out)
            self.assertTrue(any(c["kind"] == "refine" for c in candidates))

    def test_project_candidates_read_governed_pipeline_path_under_package_v1(self) -> None:
        """S1-10c：package_v1 布局下 functional_requirements.json 在 .ratomizer/pipeline/ 也能读到候选。

        旧实现裸拼 root/"functional_requirements.json"——package_v1 下该文件不在根目录，
        候选恒空（B1 类寻址失守的重现）。改 governed 双路径探测后，根目录无该文件时仍非空。
        """
        from result_package import governed_artifact_path, initialize_result_package
        from requirements_analysis_rules import dependency_candidates_for_project
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            source = out / "input.docx"
            source.write_bytes(b"fixture")
            initialize_result_package(out, input_path=source, requested_stages=["atomize"])
            governed = governed_artifact_path(out, "functional_requirements.json",
                                              category="pipeline", for_write=True)
            self.assertEqual(governed, out / ".ratomizer" / "pipeline" / "functional_requirements.json")
            governed.write_text(json.dumps({
                "producer": "functional-extract-v1",
                "items": [
                    {"functional_requirement_id": "F1", "title": "镜像升级",
                     "source_section": "4.1.1", "related_dlms_objects": ["0-0:96.3.0"]},
                    {"functional_requirement_id": "F2", "title": "镜像校验",
                     "source_section": "4.1.2", "related_dlms_objects": ["0-0:96.3.0"]},
                ],
            }, ensure_ascii=False), encoding="utf-8")
            # 根目录【没有】functional_requirements.json——证明候选来自 governed 双路径探测
            self.assertFalse((out / "functional_requirements.json").exists())
            candidates = dependency_candidates_for_project(out)
            self.assertTrue(candidates, "package_v1 布局下候选应非空（governed 双路径探测）")
            # 共享 OBIS → depend/exclude 候选；相邻章节 → refine 候选
            self.assertTrue(any(c["kind"] == "refine" for c in candidates))
            self.assertTrue(any(c["kind"] in ("depend", "exclude") for c in candidates))


# ===========================================================================
# API 处理器（object.__new__ + monkeypatch 模式）
# ===========================================================================
class ApiHandlerTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.out = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def _handler(self, body: dict | None, output_dir: Path | None = None):
        import api_server
        handler = object.__new__(api_server.RequirementAPIHandler)
        handler.output_dir = output_dir if output_dir is not None else self.out
        handler.read_json_body = lambda: body
        responses: list[tuple[int, dict]] = []
        handler.send_json = lambda payload, status=200: responses.append((status, payload))
        handler.send_error = lambda code, msg="": responses.append((code, {"error": msg}))
        return handler, responses

    def test_verification_action_write(self) -> None:
        item = _write_analysis_item(self.out)
        handler, responses = self._handler({
            "requirement_id": item["functional_requirement_id"],
            "verification": {"project_manager_confirm": True, "test_lead_confirm": True,
                             "dev_test_confirm": True, "implemented": "done",
                             "test_completed": True},
            "actor": "api-test",
        })
        handler.handle_verification_action()
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["lifecycle_state"], "verified")

    def test_verification_action_cas_conflict(self) -> None:
        item = _write_analysis_item(self.out)
        handler, responses = self._handler({
            "requirement_id": item["functional_requirement_id"],
            "verification": {"implemented": "done"},
            "expected_evidence_fingerprint": "wrong",
        })
        handler.handle_verification_action()
        self.assertEqual(responses[0][0], 409)
        self.assertTrue(responses[0][1]["needs_reconfirmation"])

    def test_rollback_handler(self) -> None:
        from requirements_analysis_rules import apply_verification_override
        item = _write_analysis_item(self.out)
        apply_verification_override(self.out, item["functional_requirement_id"], {
            "project_manager_confirm": True, "test_lead_confirm": True, "dev_test_confirm": True,
            "implemented": "done", "test_completed": True}, actor="t")
        handler, responses = self._handler({
            "requirement_id": item["functional_requirement_id"], "target": "implemented",
            "actor": "专家", "reason": "返工",
        })
        handler.handle_requirement_rollback()
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["lifecycle_state"], "implemented")

    def test_rollback_handler_rejects_missing_fields(self) -> None:
        handler, responses = self._handler({"requirement_id": "F1", "target": "draft"})
        handler.handle_requirement_rollback()
        self.assertEqual(responses[0][0], 400)

    def test_manual_requirement_handler(self) -> None:
        handler, responses = self._handler({"objective": "应支持事件上报", "module": "事件记录"})
        handler.handle_manual_requirement()
        self.assertEqual(responses[0][0], 200)
        self.assertTrue(responses[0][1]["functional_requirement_id"].startswith("FREQ-MANUAL-"))

    def test_manual_requirement_requires_objective(self) -> None:
        handler, responses = self._handler({"module": "x"})
        handler.handle_manual_requirement()
        self.assertEqual(responses[0][0], 400)

    def test_dependency_decision_handler(self) -> None:
        handler, responses = self._handler(
            {"from": "F1", "to": "F2", "kind": "depend", "accept": True})
        handler.handle_dependency_decision()
        self.assertEqual(responses[0][0], 200)
        self.assertTrue(responses[0][1]["written"])

    def test_verification_states_get(self) -> None:
        from requirements_analysis_rules import apply_verification_override
        item = _write_analysis_item(self.out)
        apply_verification_override(self.out, item["functional_requirement_id"],
                                    {"implemented": "done"}, actor="t")
        handler, responses = self._handler(None)
        handler.handle_verification_states_get()
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["total"], 1)

    def test_library_search_without_config_returns_empty(self) -> None:
        import os
        from unittest.mock import patch
        with patch.dict(os.environ, {}, clear=False):
            os.environ.pop("RATOMIZER_REQUIREMENT_LIBRARY", None)
            handler, responses = self._handler(None)
            handler.handle_requirement_library_search({"q": ["test"]})
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["matches"], 0)

    def test_functional_requirements_get(self) -> None:
        _write_functional_item(self.out)
        handler, responses = self._handler(None)
        handler.handle_functional_requirements_get()
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["schema"], "functional-requirements/v1")
        self.assertEqual(responses[0][1]["total"], 1)
        self.assertEqual(responses[0][1]["items"][0]["functional_requirement_id"], "FRE-direct0001")

    def test_functional_requirements_get_empty_when_missing(self) -> None:
        handler, responses = self._handler(None)
        handler.handle_functional_requirements_get()
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["total"], 0)
        self.assertEqual(responses[0][1]["items"], [])

    def test_manual_requirements_get(self) -> None:
        from desktop_tasks import add_manual_requirement_task
        add_manual_requirement_task(self.out, objective="应支持事件上报", module="事件记录",
                                    actor="工程师")
        handler, responses = self._handler(None)
        handler.handle_manual_requirements_get()
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["schema"], "manual-requirements/v1")
        self.assertEqual(responses[0][1]["total"], 1)

    def test_lifecycle_events_get(self) -> None:
        from requirements_analysis_rules import apply_verification_override, rollback_requirement_lifecycle
        item = _write_analysis_item(self.out)
        rid = item["functional_requirement_id"]
        apply_verification_override(self.out, rid, {
            "project_manager_confirm": True, "test_lead_confirm": True, "dev_test_confirm": True,
            "implemented": "done", "test_completed": True}, actor="t")
        rollback_requirement_lifecycle(self.out, rid, "implemented", actor="专家", reason="返工")
        handler, responses = self._handler(None)
        handler.handle_lifecycle_events_get()
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["schema"], "requirement-lifecycle-events/v1")
        # S1-10a：前进迁移（draft→verified）与回退（verified→implemented）同流 append，
        # 共 2 条事件；GET 仍是只读投影，不改写历史。
        events = responses[0][1]["events"]
        self.assertEqual(responses[0][1]["total"], 2)
        self.assertEqual([e["kind"] for e in events].count("rollback"), 1)
        self.assertEqual([e["kind"] for e in events].count("advance"), 1)
        rollback = next(e for e in events if e["kind"] == "rollback")
        self.assertEqual(rollback["from_state"], "verified")
        self.assertEqual(rollback["to_state"], "implemented")

    def test_requirement_library_adopt_writes_via_reviewer_override(self) -> None:
        from review_state import read_verification_states
        item = _write_functional_item(self.out)
        rid = item["functional_requirement_id"]
        handler, responses = self._handler({
            "requirement_id": rid, "ownership": "software", "module": "事件记录",
            "actor": "专家A", "reason": "采纳历史条目归属",
        })
        handler.handle_requirement_library_adopt()
        self.assertEqual(responses[0][0], 200)
        self.assertEqual(responses[0][1]["written"], ["verification_states.jsonl"])
        # 留痕落在既有 reviewer_override 通道（verification_states.jsonl）
        record = read_verification_states(self.out)[rid]
        self.assertEqual(record["ownership_override"], "software")
        self.assertEqual(record["module_override"], "事件记录")
        self.assertEqual(record["adopt_source"], "requirement_library")
        self.assertEqual(record["adopt_actor"], "专家A")
        self.assertEqual(record["adopt_reason"], "采纳历史条目归属")

    def test_requirement_library_adopt_requires_actor_and_reason(self) -> None:
        item = _write_functional_item(self.out)
        handler, responses = self._handler({
            "requirement_id": item["functional_requirement_id"],
            "ownership": "software", "actor": "", "reason": "",
        })
        handler.handle_requirement_library_adopt()
        self.assertEqual(responses[0][0], 400)
        self.assertIn("actor and reason", responses[0][1]["error"])


if __name__ == "__main__":
    unittest.main()
