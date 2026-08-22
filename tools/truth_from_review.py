"""人工核对结果 → WS0 真值集转换器（runbook docs/ws0-truth-flip-runbook-2026-08-17.md 配套）。

把既有人工核对产物一步转成 ab_runner 可用的 ``functional_truth`` JSONL：
- XLSX：列名经 ab_runner.XLSX_COLUMN_ALIASES 同一权威别名表识别（中英文模板
  V2.3.x 均可直接用）；body 列必需，section/condition 列可选映射；
- JSON：条目数组，键名经 ``--json-*-key`` 显式指定（不做猜测）。

确定性规则（与 ab_runner 判定同口径）：
- 数值/单位/编码（OBIS 点分链）用 ab_runner 的正则从正文+条件文本抽取填充——
  可被 ``--no-extract`` 关闭（留空数组，门禁按无约束处理）；
- 输出前逐行过 ``schemas/functional_truth.schema.json`` + 必需键校验，任何一行
  不合法即整体拒绝写盘（不产半份真值）；
- ``document_id`` 由 ``--document-id`` 显式注入（匹配解析目录名/sha256/``*``）。

用法::

    python tools/truth_from_review.py \\
        --input "C:/.../人工核对需求表.xlsx" --document-id abnt_nbr_16968 \\
        --output golden_sets/ws0_human_v1/truth.jsonl
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
if str(REPO_ROOT / "tools") not in sys.path:
    sys.path.insert(0, str(REPO_ROOT / "tools"))

import ab_runner as ab  # noqa: E402  列别名/正则/校验的既有权威

SCHEMA_PATH = REPO_ROOT / "schemas" / "functional_truth.schema.json"


def configure_stdio() -> None:
    """stdout/stderr UTF-8（对齐 docs/cli-contract.md JSON 信封口径）。"""
    for stream, kwargs in (
        (sys.stdout, {"encoding": "utf-8"}),
        (sys.stderr, {"encoding": "utf-8", "errors": "replace"}),
    ):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure:
            reconfigure(**kwargs)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in ab._XLSX_ERROR_LITERALS else text


def _read_xlsx_rows(path: Path, *, body_col: str = "", section_col: str = "",
                    condition_col: str = "") -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header = [_cell_text(cell) for cell in rows[0]]
    col_by_alias: dict[str, int] = {}
    for logical, aliases in ab.XLSX_COLUMN_ALIASES.items():
        for index, name in enumerate(header):
            if name and name.casefold() in {a.casefold() for a in aliases}:
                col_by_alias[logical] = index
                break
    # 显式列名优先（任意语言表头）；随后回退别名表（含葡语便利项）
    if body_col:
        for index, name in enumerate(header):
            if name and name.casefold() == body_col.casefold():
                col_by_alias["body"] = index
                break
        else:
            raise SystemExit(f"--body-col {body_col!r} 不在表头 {header} 中")
    for logical, explicit in (("section", section_col), ("condition", condition_col)):
        if not explicit:
            continue
        for index, name in enumerate(header):
            if name and name.casefold() == explicit.casefold():
                col_by_alias[logical] = index
                break
    pt_aliases = {"body": ("conteúdo", "conteudo", "texto", "requisito"),
                  "section": ("seção", "secao", "cláusula", "clausula", "título", "titulo")}
    for logical, aliases in pt_aliases.items():
        if logical in col_by_alias:
            continue
        for index, name in enumerate(header):
            if name and name.casefold() in aliases:
                col_by_alias[logical] = index
                break
    if "body" not in col_by_alias:
        raise SystemExit(
            f"输入表缺需求正文列（别名：{'/'.join(ab.XLSX_COLUMN_ALIASES['body'][:4])}…或用 --body-col 显式指定）；"
            f"表头={header}")
    out = []
    for row in rows[1:]:
        def _get(logical: str) -> str:
            index = col_by_alias.get(logical)
            return _cell_text(row[index]) if index is not None and index < len(row) else ""
        body = _get("body")
        if not body:
            continue
        out.append({
            "expected_text": body,
            "section_id": _get("section") or "*",
            "condition_text": _get("condition"),
        })
    return out


def _read_json_rows(path: Path, *, text_key: str, section_key: str,
                    condition_key: str) -> list[dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    items = payload if isinstance(payload, list) else (
        payload.get("items") if isinstance(payload, dict) else None)
    if not isinstance(items, list):
        raise SystemExit("JSON 输入必须是条目数组，或含 items 数组的对象")
    out = []
    for index, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        body = str(item.get(text_key) or "").strip()
        if not body:
            raise SystemExit(f"第 {index} 条缺正文键 {text_key!r}")
        section = str(item.get(section_key) or "").strip() if section_key else ""
        condition = str(item.get(condition_key) or "").strip() if condition_key else ""
        out.append({"expected_text": body, "section_id": section or "*",
                    "condition_text": condition})
    return out


def _truth_row(document_id: str, seq: int, source: dict,
               *, extract: bool) -> dict:
    import re

    from ab_runner import _DOTTED_CODE_RE, _NUMBER_RE

    text = source["expected_text"]
    condition = source.get("condition_text") or ""
    haystack = f"{text}\n{condition}" if condition else text
    # OBIS 连字格式（1-1:32.7.0）先剥离，剩余再走点分码/数值——避免把 OBIS 拆碎
    obis_re = re.compile(r"\b\d+-\d+:[0-9A-Za-z*]+(?:\.[0-9A-Za-z*]+)*\b")
    codes = [m.group(0) for m in obis_re.finditer(haystack)]
    remainder = obis_re.sub(" ", haystack)
    codes += [m.group(1) for m in _DOTTED_CODE_RE.finditer(remainder)]
    numbers = [m.group(0) for m in _NUMBER_RE.finditer(
        " ".join(part for part in _DOTTED_CODE_RE.split(remainder)))]
    units = sorted({u for u in ab._UNIT_WORDS if u.lower() in haystack.lower()})
    return {
        "truth_id": f"T-{seq:04d}",
        "document_id": document_id,
        "section_id": source["section_id"],
        "expected_text": text,
        "conditions": [condition] if condition else [],
        "exceptions": [],
        "negations": [],
        "numbers": sorted(set(numbers)) if extract else [],
        "units": units if extract else [],
        "codes": sorted(set(codes)) if extract else [],
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="人工核对结果 → WS0 真值集（schema 校验后写盘）")
    parser.add_argument("--input", type=Path, required=True, help="人工核对 XLSX 或 JSON")
    parser.add_argument("--document-id", required=True,
                        help="真值 document_id（匹配解析目录名 / blocks sha256 / *）")
    parser.add_argument("--output", type=Path, required=True, help="truth.jsonl 输出路径")
    parser.add_argument("--json-text-key", default="expected_text")
    parser.add_argument("--json-section-key", default="section_id")
    parser.add_argument("--json-condition-key", default="")
    parser.add_argument("--body-col", default="",
                        help="显式正文列名（覆盖别名表，任意语言表头可用）")
    parser.add_argument("--section-col", default="", help="显式源条款列名")
    parser.add_argument("--condition-col", default="", help="显式条件列名")
    parser.add_argument("--no-extract", action="store_true",
                        help="不自动抽取 numbers/units/codes（留空，门禁按无约束处理）")
    args = parser.parse_args(argv)
    configure_stdio()

    source = args.input
    if source.suffix.lower() in (".xlsx", ".xlsm"):
        rows = _read_xlsx_rows(source, body_col=args.body_col,
                               section_col=args.section_col,
                               condition_col=args.condition_col)
    elif source.suffix.lower() == ".json":
        rows = _read_json_rows(source, text_key=args.json_text_key,
                               section_key=args.json_section_key,
                               condition_key=args.json_condition_key)
    else:
        raise SystemExit(f"不支持的输入类型：{source.suffix}（支持 .xlsx/.json）")

    truth = [_truth_row(args.document_id, i, row, extract=not args.no_extract)
             for i, row in enumerate(rows, start=1)]
    if not truth:
        raise SystemExit("转换结果为 0 行——拒绝写盘（请检查输入列/键映射）")

    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    try:
        from jsonschema import Draft202012Validator

        validator = Draft202012Validator(schema)
        for row in truth:
            for error in validator.iter_errors(row):
                raise SystemExit(
                    f"真值行校验失败 {row['truth_id']}: {error.message}")
    except ImportError:  # pragma: no cover - jsonschema 为运行依赖
        for row in truth:
            for key in ab.TRUTH_REQUIRED_KEYS:
                if key not in row:
                    raise SystemExit(f"真值行缺必需键 {key}: {row['truth_id']}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        "\n".join(json.dumps(row, ensure_ascii=False) for row in truth) + "\n",
        encoding="utf-8")
    print(json.dumps({
        "ok": True, "rows": len(truth), "document_id": args.document_id,
        "output": str(args.output),
        "hint": "下一步：按 runbook 运行 tools/ab_runner.py 真值门禁",
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
