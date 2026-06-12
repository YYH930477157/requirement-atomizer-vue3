from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook


def write_synthetic_xlsx(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    requirements = workbook.create_sheet("Requirements")
    requirements.append(["Req ID", "Requirement", "Priority"])
    requirements.append(["REQ-1", "The meter shall support xDLMS GET service.", "High"])
    requirements.append(["REQ-2", "The display shall show all segments.", "Medium"])

    matrix = workbook.create_sheet("Capability Matrix")
    matrix.append(["Customer application process", "xDLMS Service", None])
    matrix.append(["Customer application process", "GET", "ACTION"])
    matrix.append(["Public customer", "X", ""])
    matrix.append(["Management client", "", "X"])
    matrix.merge_cells("B1:C1")

    mixed = workbook.create_sheet("Mixed Types")
    mixed.append(["Mixed Type Values"])
    mixed.append(["Label", "Value", "Formula"])
    mixed.append(["Integer float", 42.0, None])
    mixed.append(["Date value", date(2026, 6, 12), None])
    mixed.append(["Boolean value", True, None])
    mixed.append(["Formula value", 2, "=B5+40"])
    mixed["C6"].value = 42

    hidden = workbook.create_sheet("Hidden Sheet")
    hidden.sheet_state = "hidden"
    hidden.append(["Requirement"])
    hidden.append(["This hidden sheet shall not be parsed."])

    workbook.save(path)
