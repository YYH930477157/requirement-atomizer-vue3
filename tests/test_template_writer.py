"""模板成文器回归：分析结果按 V2.3.x 格式追加进对应模块 sheet（模板=格式，非问题库）。"""
from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import template_writer as tw


def make_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "时钟需求"
    ws.append(["关闭", "序号", "子模块", "描述", "需求模版", "需求", "说明、示例、注意事项",
               "是否客户需求", "客户需求章节", "驱动/硬件相关"])
    ws.append(["", 1, "时钟", "历法：", "公历", "公历", "", "", "", ""])
    ws.append(["", 2, "时钟", "夏令时：", "支持", "支持", "", "", "", ""])
    ws2 = wb.create_sheet("事件需求")
    ws2.append(["关闭", "序号", "子模块", "描述", "需求模版", "需求", "说明、示例、注意事项",
                "是否客户需求", "客户需求章节", "驱动/硬件相关"])
    wb.save(path)


def item(module: str, desc: str, *, ownership: str = "software", seq_hint: int = 0,
         text: str = "", section: str = "7.9") -> dict:
    return {
        "analysis_id": f"AN-{seq_hint:03d}", "module": module, "submodule": module,
        "description": desc, "requirement": text or f"{desc} 原始正文",
        "software_requirement_text": text, "developer_guidance": ["实现要点甲"],
        "acceptance_criteria": [], "open_questions": [], "notes": [],
        "threshold_table": None, "ownership": ownership, "source_quote": "quoted words",
        "source_section": section,
    }


class TargetSheetTests(unittest.TestCase):
    def test_module_routes_to_sheet(self) -> None:
        names = ["时钟需求", "事件需求"]
        self.assertEqual(tw.target_sheet({"module": "时钟"}, names), "时钟需求")
        self.assertEqual(tw.target_sheet({"module": "事件记录"}, names), "事件需求")

    def test_submodule_fallback_then_fallback_sheet(self) -> None:
        names = ["时钟需求"]
        self.assertEqual(tw.target_sheet({"module": "未映射", "submodule": "时钟"}, names), "时钟需求")
        self.assertEqual(tw.target_sheet({"module": "安全"}, names), tw.FALLBACK_SHEET)


class AppendTests(unittest.TestCase):
    def test_appends_rows_with_continued_seq_and_full_columns(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "t.xlsx"
            out = Path(td) / "written.xlsx"
            make_template(template)
            items = [item("时钟", "时钟精度要求", text="时钟精度须优于 5 s/天", seq_hint=1)]
            report = tw.append_analysis_to_template(template, items, out)

            wb = load_workbook(out)
            ws = wb["时钟需求"]
            self.assertEqual(ws.cell(row=4, column=2).value, 3)              # 序号接着 2 继续
            self.assertEqual(ws.cell(row=4, column=3).value, "时钟")
            self.assertEqual(ws.cell(row=4, column=4).value, "时钟精度要求")
            self.assertEqual(ws.cell(row=4, column=6).value, "时钟精度须优于 5 s/天")
            self.assertIn("实现要点甲", str(ws.cell(row=4, column=7).value))  # 说明含研发指引
            self.assertEqual(ws.cell(row=4, column=8).value, "是")
            self.assertEqual(ws.cell(row=4, column=9).value, "7.9")
            self.assertEqual(ws.cell(row=2, column=6).value, "公历")          # 模板已有行原样
            self.assertEqual(report["appended_by_sheet"], {"时钟需求": 1})

    def test_hardware_skipped_codesign_marked(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "t.xlsx"
            out = Path(td) / "written.xlsx"
            make_template(template)
            items = [item("时钟", "硬件 RTC 选型", ownership="hardware", seq_hint=1),
                     item("时钟", "温补校准", ownership="co_design", seq_hint=2)]
            report = tw.append_analysis_to_template(template, items, out)

            self.assertEqual(report["skipped_hardware"], 1)                  # 硬件独占不进列表
            ws = load_workbook(out)["时钟需求"]
            self.assertEqual(ws.cell(row=4, column=4).value, "温补校准")
            self.assertEqual(ws.cell(row=4, column=10).value, "是")          # 协同标驱动/硬件相关

    def test_compliance_is_defensively_excluded_from_software_template(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "t.xlsx"
            out = Path(td) / "written.xlsx"
            make_template(template)
            compliance = item("测试合规", "型式证书", seq_hint=1)
            compliance["source_requirement_type"] = "compliance"

            report = tw.append_analysis_to_template(template, [compliance], out)

            self.assertEqual(report["appended_total"], 0)
            self.assertEqual(report["skipped_compliance"], 1)
            self.assertNotIn(tw.FALLBACK_SHEET, load_workbook(out).sheetnames)

    def test_unmapped_module_lands_in_fallback_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "t.xlsx"
            out = Path(td) / "written.xlsx"
            make_template(template)
            tw.append_analysis_to_template(template, [item("安全", "HLS 认证", seq_hint=1)], out)
            wb = load_workbook(out)
            self.assertIn(tw.FALLBACK_SHEET, wb.sheetnames)
            ws = wb[tw.FALLBACK_SHEET]
            self.assertEqual(ws.cell(row=2, column=4).value, "HLS 认证")

    def test_formula_neutralized(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "t.xlsx"
            out = Path(td) / "written.xlsx"
            make_template(template)
            evil = item("时钟", "=1+1 注入", text="=HYPERLINK() 也不行", seq_hint=1)
            tw.append_analysis_to_template(template, [evil], out)
            ws = load_workbook(out)["时钟需求"]
            self.assertNotEqual(ws.cell(row=4, column=4).data_type, "f")
            self.assertNotEqual(ws.cell(row=4, column=6).data_type, "f")


class RunWriterTests(unittest.TestCase):
    def test_end_to_end_reads_analysis_and_writes_report(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            template = out / "t.xlsx"
            make_template(template)
            payload = {"route": "openai_compatible",
                       "items": [item("时钟", "时钟精度要求", text="精度 5 s/天", seq_hint=1),
                                 item("事件记录", "失压事件记录", seq_hint=2)]}
            (out / "engineering_analysis.json").write_text(
                json.dumps(payload, ensure_ascii=False), encoding="utf-8")

            report = tw.run_writer(out, template)

            self.assertEqual(report["appended_total"], 2)
            self.assertEqual(report["appended_by_sheet"],
                             {"时钟需求": 1, "事件需求": 1})
            self.assertTrue((out / tw.WRITTEN_WORKBOOK).exists())
            saved = json.loads((out / tw.WRITER_REPORT).read_text(encoding="utf-8"))
            self.assertEqual(saved["analysis_route"], "openai_compatible")

    def test_missing_analysis_raises(self) -> None:
        with tempfile.TemporaryDirectory() as td:
            template = Path(td) / "t.xlsx"
            make_template(template)
            with self.assertRaises(FileNotFoundError):
                tw.run_writer(Path(td), template)


if __name__ == "__main__":
    unittest.main()
